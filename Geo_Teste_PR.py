import io
import os
import re
import utm
import yaml
import math
import folium
import pycountry
import numpy as np
import pandas as pd
from PIL import Image
from io import BytesIO
import streamlit as st
from textwrap import wrap
import statsmodels.api as sm
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import plotly.graph_objects as go
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
import matplotlib.patheffects as pe
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from streamlit_folium import st_folium
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
from streamlit_pdf_viewer import pdf_viewer
from scipy.interpolate import griddata, Rbf
from scipy.ndimage import gaussian_filter1d
import matplotlib.patheffects as path_effects
from matplotlib.offsetbox import OffsetImage, AnnotationBbox
from statsmodels.nonparametric.smoothers_lowess import lowess

@st.cache_data(show_spinner=False)
def carregar_workbook(file_bytes):
    import io
    buffer = io.BytesIO(file_bytes)
    wb = load_workbook(buffer, data_only=True)
    return wb

#  Configurações da página web
logo = 'logo.png'
img_logo = Image.open(logo)
cab = 'logo_syng.png'
img_cab = Image.open(cab)
image = Image.open(logo)
PAGE_CONFIG = {"page_title": "SYGA",
               "page_icon": image,
               "layout": "wide",
               "initial_sidebar_state": "auto",
               }
st.set_page_config(**PAGE_CONFIG)
st.image(img_cab, width=1500)
st.markdown(
    "<div style='text-align: left; font-size: 16px; color: gray;'>"
    "Desenvolvido por: Adriel Oliveira - 2025"
    "</div>",
    unsafe_allow_html=True
)

hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)

if "pdf_bytes" not in st.session_state:
    st.session_state.pdf_bytes = None

if "pdf_ready" not in st.session_state:
    st.session_state.pdf_ready = False

if "pdf_params_hash" not in st.session_state:
    st.session_state.pdf_params_hash = None

if "pdf_view_open" not in st.session_state:
    st.session_state.pdf_view_open = False

paises = {
    "Afeganistão": "af",
    "África do Sul": "za",
    "Albânia": "al",
    "Alemanha": "de",
    "Andorra": "ad",
    "Angola": "ao",
    "Antígua e Barbuda": "ag",
    "Arábia Saudita": "sa",
    "Argélia": "dz",
    "Argentina": "ar",
    "Armênia": "am",
    "Austrália": "au",
    "Áustria": "at",
    "Azerbaijão": "az",
    "Bahamas": "bs",
    "Bahrein": "bh",
    "Bangladesh": "bd",
    "Barbados": "bb",
    "Bélgica": "be",
    "Belize": "bz",
    "Benim": "bj",
    "Bielorrússia": "by",
    "Bolívia": "bo",
    "Bósnia e Herzegovina": "ba",
    "Botsuana": "bw",
    "Brasil": "br",
    "Brunei": "bn",
    "Bulgária": "bg",
    "Burkina Faso": "bf",
    "Burundi": "bi",
    "Butão": "bt",
    "Cabo Verde": "cv",
    "Camarões": "cm",
    "Camboja": "kh",
    "Canadá": "ca",
    "Catar": "qa",
    "Cazaquistão": "kz",
    "Chade": "td",
    "Chile": "cl",
    "China": "cn",
    "Chipre": "cy",
    "Colômbia": "co",
    "Comores": "km",
    "Congo": "cg",
    "Coreia do Norte": "kp",
    "Coreia do Sul": "kr",
    "Costa do Marfim": "ci",
    "Costa Rica": "cr",
    "Croácia": "hr",
    "Cuba": "cu",
    "Dinamarca": "dk",
    "Djibuti": "dj",
    "Dominica": "dm",
    "Egito": "eg",
    "El Salvador": "sv",
    "Emirados Árabes Unidos": "ae",
    "Equador": "ec",
    "Eritreia": "er",
    "Eslováquia": "sk",
    "Eslovênia": "si",
    "Espanha": "es",
    "Estados Unidos": "us",
    "Estônia": "ee",
    "Etiópia": "et",
    "Fiji": "fj",
    "Filipinas": "ph",
    "Finlândia": "fi",
    "França": "fr",
    "Gabão": "ga",
    "Gâmbia": "gm",
    "Gana": "gh",
    "Geórgia": "ge",
    "Granada": "gd",
    "Grécia": "gr",
    "Guatemala": "gt",
    "Guiana": "gy",
    "Guiné": "gn",
    "Guiné-Bissau": "gw",
    "Guiné Equatorial": "gq",
    "Haiti": "ht",
    "Honduras": "hn",
    "Hungria": "hu",
    "Iêmen": "ye",
    "Ilhas Marshall": "mh",
    "Ilhas Salomão": "sb",
    "Índia": "in",
    "Indonésia": "id",
    "Irã": "ir",
    "Iraque": "iq",
    "Irlanda": "ie",
    "Islândia": "is",
    "Israel": "il",
    "Itália": "it",
    "Jamaica": "jm",
    "Japão": "jp",
    "Jordânia": "jo",
    "Kiribati": "ki",
    "Kuwait": "kw",
    "Laos": "la",
    "Lesoto": "ls",
    "Letônia": "lv",
    "Líbano": "lb",
    "Libéria": "lr",
    "Líbia": "ly",
    "Liechtenstein": "li",
    "Lituânia": "lt",
    "Luxemburgo": "lu",
    "Macedônia do Norte": "mk",
    "Madagascar": "mg",
    "Malásia": "my",
    "Malawi": "mw",
    "Maldivas": "mv",
    "Mali": "ml",
    "Malta": "mt",
    "Marrocos": "ma",
    "Maurício": "mu",
    "Mauritânia": "mr",
    "México": "mx",
    "Micronésia": "fm",
    "Moçambique": "mz",
    "Moldávia": "md",
    "Mônaco": "mc",
    "Mongólia": "mn",
    "Montenegro": "me",
    "Myanmar": "mm",
    "Namíbia": "na",
    "Nauru": "nr",
    "Nepal": "np",
    "Nicarágua": "ni",
    "Níger": "ne",
    "Nigéria": "ng",
    "Noruega": "no",
    "Nova Zelândia": "nz",
    "Omã": "om",
    "Países Baixos": "nl",
    "Palau": "pw",
    "Panamá": "pa",
    "Papua-Nova Guiné": "pg",
    "Paquistão": "pk",
    "Paraguai": "py",
    "Peru": "pe",
    "Polônia": "pl",
    "Portugal": "pt",
    "Quênia": "ke",
    "Quirguistão": "kg",
    "Reino Unido": "gb",
    "República Centro-Africana": "cf",
    "República Dominicana": "do",
    "Romênia": "ro",
    "Ruanda": "rw",
    "Rússia": "ru",
    "Samoa": "ws",
    "San Marino": "sm",
    "Santa Lúcia": "lc",
    "São Cristóvão e Nevis": "kn",
    "São Tomé e Príncipe": "st",
    "São Vicente e Granadinas": "vc",
    "Senegal": "sn",
    "Serra Leoa": "sl",
    "Sérvia": "rs",
    "Seychelles": "sc",
    "Singapura": "sg",
    "Síria": "sy",
    "Somália": "so",
    "Sri Lanka": "lk",
    "Sudão": "sd",
    "Sudão do Sul": "ss",
    "Suécia": "se",
    "Suíça": "ch",
    "Suriname": "sr",
    "Tailândia": "th",
    "Tajiquistão": "tj",
    "Tanzânia": "tz",
    "Timor-Leste": "tl",
    "Togo": "tg",
    "Tonga": "to",
    "Trinidad e Tobago": "tt",
    "Tunísia": "tn",
    "Turcomenistão": "tm",
    "Turquia": "tr",
    "Tuvalu": "tv",
    "Ucrânia": "ua",
    "Uganda": "ug",
    "Uruguai": "uy",
    "Uzbequistão": "uz",
    "Vanuatu": "vu",
    "Vaticano": "va",
    "Venezuela": "ve",
    "Vietnã": "vn",
    "Zâmbia": "zm",
    "Zimbábue": "zw"
}


def numpy_to_python(obj):
    # converte recursivamente numpy -> python nativos
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    if isinstance(obj, (np.generic,)):
        return obj.item()
    if isinstance(obj, dict):
        return {k: numpy_to_python(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [numpy_to_python(i) for i in obj]
    return obj


@st.dialog("Remove well")
def remove():
    st.markdown('##### Well removal')
    if st.session_state.pocos:
        w = st.session_state.pocos.keys()
        st.multiselect('Select well to remove', list(w), key='well_to_remove')
    else:
        st.markdown('##### No well founded')

    if st.button('Remove selected wells', key='selected_bt'):
        if st.session_state.well_to_remove:
            for j in st.session_state.well_to_remove:
                del st.session_state.pocos[j]
            st.rerun()
        else:
            st.warning('No well was selected to remove')


def highlight_active(val):
    if val == st.session_state.well_selected:
        return "font-weight: bold; color: green;"  # destaque
    return ""  # os outros ficam normais


def gerar_perfil_sintetico(poco_origem, poco_destino, s_logs):
    """
    Gera perfil sintético no poço destino com base nas formações do poço origem,
    usando regra de 3 para reescalar profundidades.
    """

    profundidades_origem = poco_origem["profundidade"]
    formacoes_origem = poco_origem["formation"]

    profundidades_destino = poco_destino["profundidade"]
    formacoes_destino = poco_destino["formation"]

    df_sintetico = pd.DataFrame()

    for logs in s_logs:
        sint_prof_dest = []
        sint_prof_orig = []
        sint_valor = []
        sint_fm = []

        # converter listas do dicionário de perfil em arrays NumPy
        prof_array = np.array(poco_origem["perfil"]["Profundidade"])

        val_array = np.array(poco_origem["perfil"][logs])

        for i, fm in enumerate(formacoes_destino):
            if i < len(profundidades_destino) - 1:
                z_top_dest = profundidades_destino[i]
                z_base_dest = profundidades_destino[i + 1]

                if fm in formacoes_origem:
                    idx = formacoes_origem.index(fm)
                    z_top_orig = profundidades_origem[idx]
                    z_base_orig = profundidades_origem[idx + 1]

                    # perfis no intervalo da formação no poço origem
                    mask = (prof_array >= z_top_orig) & (prof_array < z_base_orig)

                    prof_fm = prof_array[mask]
                    val_fm = val_array[mask]

                    # cálculo da fração dentro da formação (α)
                    frac = (prof_fm - z_top_orig) / (z_base_orig - z_top_orig)

                    # mapeamento para poço destino
                    prof_esc = z_top_dest + frac * (z_base_dest - z_top_dest)
                    sint_prof_dest.extend(prof_esc)
                    sint_prof_orig.extend(prof_fm)
                    sint_valor.extend(val_fm)
                    sint_fm.extend([fm] * len(prof_fm))

        # df_sintetico[f'Formation'] = sint_fm
        # df_sintetico[f'Depth Original log'] = sint_prof_orig
        df_sintetico[f'Depth synthetic log'] = sint_prof_dest
        df_sintetico[f'Synthetic {logs}'] = sint_valor

        # df_sintetico = pd.DataFrame({
        #     "Formação": sint_fm,
        #     "Depth Origin": sint_prof_orig,
        #     "Depth Correlation": sint_prof_dest,
        #     f"Synthetic {logs}": sint_valor
        # })

    return df_sintetico


def plot_correlacao_com_logs(pocos, logs, description, lines, v_selected, tvd_labels, escala=(30, 300)):
    """
    Plota correlação entre múltiplos poços, cada um com seu log track no estilo clássico.
    """
    fig = go.Figure()
    paleta = {
        "Argilito": {"bg": "#9ACD32", "simbol": "|"},
        "Folhelho": {"bg": "#9ACD32", "simbol": "-"},
        "Siltito": {"bg": "#A67B5B", "simbol": "-"},
        "Arenito": {"bg": "#FFD580", "simbol": "."},
        "Diamictito": {"bg": "#E97451", "simbol": "."},
        "Conglomerado": {"bg": "#CD853F", "simbol": "."},
        "Anidrita / Gipsita": {"bg": "#E6E6FA", "simbol": "/"},
        "Halita": {"bg": "#FFFFFF", "simbol": "."},
        "Marga": {"bg": "#EEE8AA", "simbol": "-"},
        "Calcário": {"bg": "#B0C4DE", "simbol": "."},
        "Calcissiltito": "#D8BFD8",
        "Calcarenito": "#F5DEB3",
        "Calcirrudito": "#4682B4",
        "Coquina": "#FFDEAD",
        "Dolomito": "#C2B280",
        "Silexito": "#808080",
        "Basalto": {"bg": "#2F4F4F", "simbol": "+"},
        "Diabásio": "#556B2F",
        "Ígnea Ácida": "#CD5C5C",
        "Ígnea Alcalina": "#B22222",
        "Ígnea Não Especificada": "#A9A9A9",
        "Metamórfica Não Especificada": "#696969",
        "SDR": "#FF7F50",
        "Crosta Oceânica": "#000080"
    }

    nomes = list(pocos.keys())
    n_pocos = len(nomes)
    espacamento = 2  # espaço lateral entre poços

    for idx, nome in enumerate(nomes):
        if nome not in v_selected:
            pass
        else:
            x_center = idx * espacamento
            poco = pocos[nome]

            profundidade = poco["profundidade"]
            fm = poco['formation']
            litologias = poco["litologia"]
            perfil = pd.DataFrame(poco['perfil'])
            tvd = poco['tvd']
            syn = False
            if 'Syn Log' in poco:
                syn = True

            # === 1) Faixas de formation ===
            for i in range(len(profundidade)):
                if i + 1 == len(profundidade):
                    z_top, z_base = profundidade[i], tvd
                else:
                    z_top, z_base = profundidade[i], profundidade[i + 1]
                lit = litologias[i]

                fig.add_trace(go.Scatter(
                    x=[x_center - 0.5, x_center + 0.5, x_center + 0.5, x_center - 0.5, x_center - 0.5],
                    y=[z_top, z_top, z_base, z_base, z_top],
                    fill="toself",
                    line=dict(color="black", width=1),
                    fillpattern=dict(
                        shape=paleta[lit]["simbol"],  # padrão diagonal
                        fgcolor="black",  # cor do padrão
                        bgcolor=paleta[lit]["bg"],
                        size=4,  # <<< controla o tamanho/espacamento
                        solidity=0.05  # densidade do padrão (0 = ralo, 1 = cheio)
                    ),
                    showlegend=False
                ))

                if tvd_labels:
                    fig.add_annotation(
                        x=x_center + 0.6,  # posição no eixo x
                        y=z_top,  # posição no eixo y (topo da formação)
                        text=f'{profundidade[i]} m',  # nome da formação
                        showarrow=False,  # sem seta
                        font=dict(size=12, color="black", family="Arial"),
                        align="left",
                        bordercolor="black",
                        borderwidth=1,
                        borderpad=2,
                        bgcolor="white",  # fundo da caixa
                        opacity=0.8
                    )

                if description:
                    # nome da formation
                    fig.add_trace(go.Scatter(
                        x=[x_center],
                        y=[(z_top + z_base) / 2],
                        text=[f"<b>{fm[i]}</b> - {lit}"],  # <<< negrito com HTML
                        mode="text",
                        textfont=dict(size=12, color="black", family="Arial"),  # define a fonte
                        showlegend=False,
                        hoverinfo="skip"
                    ))

            if syn:
                k = poco['Syn Log'].keys()
                x = list(k)
                x.remove('Depth synthetic log')
                for line in range(3):
                    if logs[line]:
                        if line == 0:
                            log = 'Synthetic Sonic log'
                            color = 'red'
                            scale = st.session_state.s_scale
                        elif line == 1:
                            log = 'Synthetic Density log'
                            color = 'purple'
                            scale = st.session_state.d_scale
                        else:
                            log = 'Synthetic Gamma ray log'
                            color = 'black'
                            scale = st.session_state.g_scale
                        try:
                            # === 2) Curva de log ===
                            z_log = poco['Syn Log']["Depth synthetic log"]
                            valores = np.array(poco['Syn Log'][log])
                            # normalizar valores para caber dentro da faixa do poço
                            vmin = min(poco['Syn Log'][log])
                            vmax = max(poco['Syn Log'][log])
                            x_log = x_center - 0.5 + (valores - vmin) / (vmax - vmin) * 1.0

                            fig.add_trace(go.Scatter(
                                x=x_log,
                                y=z_log,
                                mode="lines",
                                line=dict(color=color, width=1),
                                name=f"Log {nome}",
                                showlegend=False,
                                opacity=st.session_state.op
                            ))
                            # ['Above', 'Below', 'Remove']
                            if scale == 'Above':
                                y_pos = -50
                                top = y_pos - 70
                            elif scale == 'Below':
                                y_pos = tvd + 50
                                top = y_pos + 70
                            else:
                                y_pos = False
                                top = False

                            if y_pos:
                                # === 3) Escala do log logo acima do topo (y=0) ===
                                for v in np.linspace(vmin, vmax, 6):
                                    xpos = x_center - 0.5 + (v - vmin) / (vmax - vmin) * 1.0
                                    fig.add_trace(go.Scatter(
                                        x=[xpos],
                                        y=[y_pos],
                                        text=[str(int(v))],
                                        mode="text",
                                        textfont=dict(size=9, color=color),
                                        showlegend=False,
                                        hoverinfo="skip"
                                    ))

                                    # título do perfil
                                    fig.add_trace(go.Scatter(
                                        x=[x_center],
                                        y=[top],
                                        text=log,
                                        mode="text",
                                        textfont=dict(size=11, color=color),
                                        showlegend=False,
                                        hoverinfo="skip"
                                    ))
                        except KeyError:
                            pass

            for line in range(3):
                if logs[line]:
                    if line == 0:
                        log = 'Sonic log'
                        color = 'red'
                        scale = st.session_state.s_scale
                    elif line == 1:
                        log = 'Density log'
                        color = 'purple'
                        scale = st.session_state.d_scale
                    else:
                        log = 'Gamma ray log'
                        color = 'black'
                        scale = st.session_state.g_scale

                    try:
                        # === 2) Curva de log ===
                        z_log = perfil["Profundidade"]
                        valores = perfil[log]
                        # normalizar valores para caber dentro da faixa do poço
                        vmin = min(perfil[log])
                        vmax = max(perfil[log])
                        x_log = x_center - 0.5 + (valores - vmin) / (vmax - vmin) * 1.0

                        fig.add_trace(go.Scatter(
                            x=x_log,
                            y=z_log,
                            mode="lines",
                            line=dict(color=color, width=1),
                            name=f"Log {nome}",
                            showlegend=False,
                            opacity=st.session_state.op
                        ))
                        # ['Above', 'Below', 'Remove']
                        if scale == 'Above':
                            y_pos = -50
                            top = y_pos - 70
                        elif scale == 'Below':
                            y_pos = tvd + 50
                            top = y_pos + 70
                        else:
                            y_pos = False
                            top = False

                        if y_pos:
                            # === 3) Escala do log logo acima do topo (y=0) ===
                            for v in np.linspace(vmin, vmax, 6):
                                xpos = x_center - 0.5 + (v - vmin) / (vmax - vmin) * 1.0
                                fig.add_trace(go.Scatter(
                                    x=[xpos],
                                    y=[y_pos],
                                    text=[str(int(v))],
                                    mode="text",
                                    textfont=dict(size=9, color=color),
                                    showlegend=False,
                                    hoverinfo="skip"
                                ))

                                # título do perfil
                                fig.add_trace(go.Scatter(
                                    x=[x_center],
                                    y=[top],
                                    text=log,
                                    mode="text",
                                    textfont=dict(size=11, color=color),
                                    showlegend=False,
                                    hoverinfo="skip"
                                ))
                    except KeyError:
                        pass

            if lines:
                if idx < n_pocos - 1:
                    prox = pocos[nomes[idx + 1]]
                    if nomes[idx + 1] in v_selected:
                        # cria dicionários formacao->profundidade
                        mapa_atual = dict(zip(pocos[nomes[idx]]["formation"], pocos[nomes[idx]]["profundidade"]))
                        mapa_prox = dict(zip(prox["formation"], prox["profundidade"]))

                        # percorre formações em comum
                        for formacao, z_top in mapa_atual.items():
                            if formacao in mapa_prox:
                                z_top_prox = mapa_prox[formacao]

                                # desenha linha tracejada
                                fig.add_trace(go.Scatter(
                                    x=[x_center + 0.5, (idx + 1) * espacamento - 0.5],
                                    y=[z_top, z_top_prox],
                                    mode="lines",
                                    line=dict(color="black", width=1, dash="dash"),
                                    showlegend=False
                                ))
                            if z_top == profundidade[-1]:
                                z_top = pocos[nomes[idx]]['tvd']
                                z_top_prox = prox['tvd']
                                fig.add_trace(go.Scatter(
                                    x=[x_center + 0.5, (idx + 1) * espacamento - 0.5],
                                    y=[z_top, z_top_prox],
                                    mode="lines",
                                    line=dict(color="black", width=1, dash="dash"),
                                    showlegend=False
                                ))

            # Well name above lithology layers
            fig.add_trace(go.Scatter(
                x=[x_center],
                y=[-250],
                text=[nome],
                mode="text",
                textfont=dict(size=14, color="blue"),
                showlegend=False,
                hoverinfo="skip"
            ))

    # === Ajustes ===
    fig.update_yaxes(dtick=200)  # passo da escala do eixo da profundidade

    fig.update_yaxes(
        autorange="reversed",
        title="Profundidade (m)",
        range=[0, max(max(p["profundidade"]) for p in pocos.values())]  # começa em 0
    )
    fig.update_xaxes(visible=False)  # remove eixo X

    fig.update_layout(
        title="Correlação estratigráfica",
        plot_bgcolor="white",
        width=1000,
        height=800,
    )

    return fig


@st.dialog("Parâmetros da Correlação de Miller")
def parametros_miller():
    st.write("Insira os parâmetros para a correlação de **Miller**:")

    st.number_input('Insira o valor da ***Porosidade a Grandes Profundidades***', help='Entre 0,3 e 0,4',
                    step=1.0, format='%f', key='pa', min_value=0.0)
    st.number_input('Insira o valor da ***Parâmetro de Ajuste***', help='Entre 0,3 e 0,5',
                    step=1.0, format='%f', key='pb', min_value=0.0)
    st.number_input('Insira o valor da ***Taxa de Declínio da Porosidade***', help='Entre 0,002 e 0,004',
                    step=1.0, format='%f', key='k', min_value=0.0)
    st.number_input('Insira o valor da ***Parâmetro de Curvatura***', help='Entre 1,0 e 1,3',
                    step=1.0, format='%f', key='n', min_value=0.0)
    st.number_input('Insira o valor da ***Densidade da Matriz***', help='Entre 2,0 e 3,0',
                    step=1.0, format='%f', key='dm', min_value=0.0)
    st.number_input('Insira o valor da ***Densidade da Água***', help='Entre 1,0 e 1,2',
                    step=1.0, format='%f', key='dw', min_value=0.0)

    if st.button("Salvar Parâmetros"):
        st.session_state.parametros_miller = {
            "Porosidade a Grandes Profundidades": st.session_state.pa,
            "Parâmetro de Ajuste": st.session_state.pb,
            "Taxa de Declínio da Porosidade": st.session_state.k,
            "Parâmetro de Curvatura": st.session_state.n,
            "Densidade da Matriz": st.session_state.dm,
            "Densidade da Água": st.session_state.dw,
        }
        st.success("Parâmetros salvos com sucesso!")
        return ()


def rft(fig):
    if "df_mud" in st.session_state and isinstance(st.session_state["df_mud"], pd.DataFrame):
        st.markdown("Dados importados da aba **Geopressões** (B/F/G).")
        st.dataframe(st.session_state["df_mud"], use_container_width=True, hide_index=True)
    else:
        st.info("Ainda não foi possível importar o peso do fluido do Excel.")


@st.dialog("Direções das Tensões In Situ")
def tensoes():
    with st.container(border=True):
        # Dados iniciais vazios ou com exemplo
        tis = pd.DataFrame({
            "Profundidade (m)": [],
            "Direção SH": []
        })

        # Tabela editável
        st.session_state.tise = st.data_editor(
            tis,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
        )

        # Botão para confirmar e fechar a aba
        if st.button("Inserir Direção das Tensões In Situ", use_container_width=True, type="primary"):
            st.session_state.direct = True
            st.rerun()


@st.dialog("Relação das tensões horizontais", width="large")
def rel_hor():
    with st.container(border=True):
        tis = pd.DataFrame({
            "Profundidade (m)": [],
            "SH% Sobrecarga": [],
            "Sh% Sobrecarga": []
        })

        st.session_state.rel_hor_df = st.data_editor(
            tis,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            key="rel_hor_editor"
        )

        if st.button(
            "Inserir Relação das Tensões Horizontais",
            use_container_width=True,
            type="primary"
        ):
            st.session_state.rel_hor = True
            st.rerun()

if "frac" not in st.session_state:
    st.session_state.frac = 0.1

if "gauss" not in st.session_state:
    st.session_state.gauss = 50

def suavizar(x, y):
    x = np.asarray(x)
    y = np.asarray(y)

    # Proteção básica
    if len(y) < 3:
        return y

    frac = st.session_state.frac
    sigma = st.session_state.gauss

    # Detecta patamar
    razao_unicos = np.unique(y).size / len(y)
    dy = np.abs(np.diff(y))
    patamar = np.mean(dy < 1e-3)

    if razao_unicos < 0.1 or patamar > 0.7:
        y_suav = gaussian_filter1d(y, sigma=sigma)
    else:
        y_suav = lowess(y, x, frac=frac, return_sorted=False)

    return y_suav

def suavizar_2(x, y, grad_medio):

    x = np.asarray(x)
    y = np.asarray(y)

    # Array de saída
    y_out = y.copy()

    z_ini = st.session_state.anormal

    # --- parte rasa: valor constante ---
    mask_raso = x < z_ini

    if np.isscalar(grad_medio):
        y_out[mask_raso] = grad_medio
    else:
        grad_medio = np.asarray(grad_medio)
        y_out[mask_raso] = grad_medio[mask_raso]

    # --- parte profunda: suavização ---
    mask_prof = x >= z_ini

    # Proteção
    if mask_prof.sum() < 3:
        return y_out

    x_seg = x[mask_prof]
    y_seg = y[mask_prof]

    frac = st.session_state.frac
    sigma = st.session_state.gauss

    # Detecta patamar (mesma lógica da sua suavizar)
    razao_unicos = np.unique(y_seg).size / len(y_seg)
    dy = np.abs(np.diff(y_seg))
    patamar = np.mean(dy < 1e-3)

    if razao_unicos < 0.1 or patamar > 0.7:
        y_out[mask_prof] = gaussian_filter1d(y_seg, sigma=sigma)
    else:
        y_out[mask_prof] = lowess(
            y_seg,
            x_seg,
            frac=frac,
            return_sorted=False
        )

    return y_out

# Constantes globais
RAIO = 1
ALTURA = 2
DESLOCAMENTO_TEXTO = 0.1
VETOR_TAU_LABEL = "τ_rθ"


def gerar_cilindro_inclinado(r=1, h=5, n_theta=50, n_z=50, incl=30, azi=45, angulo_central=0):
    theta = np.linspace(0, 2 * np.pi, n_theta)
    z = np.linspace(-h / 2, h / 2, n_z)
    theta, z = np.meshgrid(theta, z)
    x = r * np.cos(theta)
    y = r * np.sin(theta)

    incl_rad = np.deg2rad(incl)
    azi_rad = np.deg2rad(azi + 180)

    azi_rad2 = np.deg2rad(azi + 180)

    # Vetor direção (unitário)
    u = np.array([
        np.sin(incl_rad) * np.cos(azi_rad),
        np.sin(incl_rad) * np.sin(azi_rad),
        np.cos(incl_rad)
    ])

    c = np.array([
        np.sin(incl_rad) * np.cos(azi_rad2),
        np.sin(incl_rad) * np.sin(azi_rad2),
        np.cos(incl_rad)
    ])

    v0 = np.array([0, 0, 1])
    v1 = u
    axis = np.cross(v0, v1)
    norm_axis = np.linalg.norm(axis)
    if norm_axis != 0:
        axis = axis / norm_axis
        angle = np.arccos(np.clip(np.dot(v0, v1), -1, 1))
        K = np.array([[0, -axis[2], axis[1]],
                      [axis[2], 0, -axis[0]],
                      [-axis[1], axis[0], 0]])
        R = np.eye(3) + np.sin(angle) * K + (1 - np.cos(angle)) * (K @ K)
    else:
        R = np.eye(3)

    # Rotacionar cilindro
    xyz = np.stack([x.flatten(), y.flatten(), z.flatten()], axis=0)
    xyz_rot = R @ xyz
    x_rot = xyz_rot[0].reshape(x.shape)
    y_rot = xyz_rot[1].reshape(y.shape)
    z_rot = xyz_rot[2].reshape(z.shape)

    # Diferença angular em relação ao ângulo central
    diff = np.angle(np.exp(1j * (theta - angulo_central)))  # entre -pi e pi
    diff = np.abs(diff)  # distância angular positiva

    # Normalizar (0 até 180 graus -> 0 até 1)
    norm_diff = diff / np.pi

    return x_rot, y_rot, z_rot, u, R, norm_diff


def curva_tau_rtheta():
    theta = np.linspace(0, np.pi / 2, 20)
    r = 1.0
    z = np.ones_like(theta)
    x = r * np.cos(theta)
    y = r * np.sin(theta)
    return x, y, z


def ponta_seta(xf, yf, zf):
    dx, dy = 0.1, 0.1
    x_v = [dx - 0.1, xf + 0.2, None, dx - 0.1, xf + 0.2]
    y_v = [yf, yf - dy, None, yf, yf + dy]
    z_v = [zf, zf, None, zf, zf]
    return x_v, y_v, z_v


def adicionar_linha(fig, x_linha, y_linha, z_linha):
    fig.add_trace(go.Scatter3d(
        x=x_linha, y=y_linha, z=z_linha,
        mode='lines',
        line=dict(color='black', width=3, dash='dash'),
        showlegend=False
    ))


def adicionar_setor_circular(fig, angulo, R, origem=(0, 0, 0)):
    """
    Adiciona um setor circular que acompanha a rotação do cilindro.

    - angulo: ângulo em radianos
    - R: matriz de rotação do cilindro
    - origem: ponto de onde o setor parte (ex: base ou centro do cilindro)
    """
    ox, oy, oz = origem

    # ----- Superfície do setor -----
    n = 50
    theta = np.linspace(0, angulo, n)
    r = 1
    x = r * np.cos(theta)
    y = r * np.sin(theta)
    z = np.zeros_like(x)

    # fechar setor
    x = np.append(x, 0)
    y = np.append(y, 0)
    z = np.append(z, 0)

    # rotacionar pontos
    pts = np.vstack([x, y, z])  # (3, N)
    pts_rot = R @ pts  # aplicar rotação
    x_rot, y_rot, z_rot = pts_rot[0] + ox, pts_rot[1] + oy, pts_rot[2] + oz

    # mesh do setor
    i = list(range(n - 1))
    j = list(range(1, n))
    k = [n] * (n - 1)
    fig.add_trace(
        go.Mesh3d(x=x_rot, y=y_rot, z=z_rot,
                  i=i, j=j, k=k,
                  color='black', opacity=0.4,
                  name='Ângulo θ', showlegend=True)
    )

    # ----- Linha circular (raio interno) -----
    n = 100
    theta = np.linspace(0, angulo, n)
    x = 0.5 * np.cos(theta)
    y = 0.5 * np.sin(theta)
    z = np.zeros_like(x)

    pts = np.vstack([x, y, z])
    pts_rot = R @ pts
    x_rot, y_rot, z_rot = pts_rot[0] + ox, pts_rot[1] + oy, pts_rot[2] + oz

    fig.add_trace(go.Scatter3d(
        x=x_rot, y=y_rot, z=z_rot,
        mode="lines",
        line=dict(color="black", width=6),
        name="Dir. Tensões Principais",
        showlegend=True
    ))

    # ----- Label do ângulo -----
    theta_label = angulo / 2
    x = (0.5 + 0.05) * np.cos(theta_label)
    y = (0.5 + 0.05) * np.sin(theta_label)
    z = 0

    pts = np.array([[x], [y], [z]])
    pts_rot = R @ pts
    x_rot, y_rot, z_rot = pts_rot[0] + ox, pts_rot[1] + oy, pts_rot[2] + oz
    if st.session_state.show:
        fig.add_trace(go.Scatter3d(
            x=[0], y=[0], z=[0],
            mode="text",
            text=[f"θ = {np.rad2deg(angulo):.1f}º"],
            textfont=dict(size=16, color="black"),
            showlegend=False
        ))


def gerar_setor_fora_inclinado(fig, angulo, R, r_int=1, r_ext=2, origem=(0, 0, 0)):
    """
    Gera um setor circular externo que acompanha a rotação do poço/cilindro.

    - angulo: ângulo do setor (rad)
    - R: matriz de rotação do poço/cilindro
    - r_int: raio interno (normalmente o raio do cilindro)
    - r_ext: raio externo (até onde o setor se estende)
    - origem: posição da base do cilindro/poço
    """
    ox, oy, oz = origem
    n = 100
    theta = np.linspace(0, angulo, n)

    # Arco externo
    x_ext = r_ext * np.cos(theta)
    y_ext = r_ext * np.sin(theta)
    z_ext = np.zeros_like(x_ext)
    r_ext_arr = np.full_like(x_ext, r_ext)

    # Arco interno
    x_int = r_int * np.cos(theta)
    y_int = r_int * np.sin(theta)
    z_int = np.zeros_like(x_int)
    r_int_arr = np.full_like(x_int, r_int)

    # Junta arcos
    x = np.concatenate([x_ext, x_int])
    y = np.concatenate([y_ext, y_int])
    z = np.concatenate([z_ext, z_int])
    r = np.concatenate([r_ext_arr, r_int_arr])

    # --- aplica rotação ---
    pts = np.vstack([x, y, z])  # shape (3, 2*n)
    pts_rot = R @ pts
    x_rot = pts_rot[0] + ox
    y_rot = pts_rot[1] + oy
    z_rot = pts_rot[2] + oz

    # Triangulação
    i, j, k = [], [], []
    for t in range(n - 1):
        i.append(t)
        j.append(t + 1)
        k.append(n + t)

        i.append(t + 1)
        j.append(n + t + 1)
        k.append(n + t)

    # Mesh3D com escala de cor
    fig.add_trace(go.Mesh3d(
        x=x_rot, y=y_rot, z=z_rot,
        i=i, j=j, k=k,
        intensity=r,
        colorscale="Jet",
        reversescale=True,
        showscale=False,
        opacity=0.5,
        name="Setor sombreado externo"
    ))


def adicionar_eixos(fig):
    # Eixo X
    fig.add_trace(go.Scatter3d(
        x=[-2, 2], y=[0, 0], z=[0, 0],
        mode="lines",
        line=dict(color="red", width=4, dash="dot"),
        name="N/S"
    ))
    fig.add_trace(go.Scatter3d(
        x=[2, 2], y=[0, 0], z=[0, 0],
        mode="text",
        text=["N"],  # aqui você pode colocar 'θ', 'σθ' etc.
        textfont=dict(size=16, color="black"),
        showlegend=False
    ))
    fig.add_trace(go.Scatter3d(
        x=[-2, -2], y=[0, 0], z=[0, 0],
        mode="text",
        text=["S"],  # aqui você pode colocar 'θ', 'σθ' etc.
        textfont=dict(size=16, color="black"),
        showlegend=False
    ))

    # Eixo Y
    fig.add_trace(go.Scatter3d(
        x=[0, 0], y=[-2, 2], z=[0, 0],
        mode="lines",
        line=dict(color="green", width=4, dash="dot"),
        name="E/W"
    ))

    # Eixo Z
    fig.add_trace(go.Scatter3d(
        x=[0, 0], y=[0, 0], z=[-2, 2],
        mode="lines",
        line=dict(color="blue", width=4, dash="dot"),
        name="Eixo Z"
    ))

    fig.add_trace(go.Scatter3d(
        x=[0, 0], y=[2, 2], z=[0, 0],
        mode="text",
        text=["E"],  # aqui você pode colocar 'θ', 'σθ' etc.
        textfont=dict(size=16, color="black"),
        showlegend=False
    ))
    fig.add_trace(go.Scatter3d(
        x=[0, 0], y=[-2, -2], z=[0, 0],
        mode="text",
        text=["W"],  # aqui você pode colocar 'θ', 'σθ' etc.
        textfont=dict(size=16, color="black"),
        showlegend=False
    ))


def adicionar_cone_setor(fig, angulo_setor, R, z_local=0, r=1.0, orient='axial',color='Reds', size_ref=0.3, nome='Cone',desloc_radial=0.0, desloc_tang=0.0, desloc_axial=0.0, valor=0.0):
    """
    desloc_radial: empurra para fora/pra dentro do poço
    desloc_tang: gira tangencialmente (ao redor do eixo)
    desloc_axial: sobe/baixa ao longo do eixo do poço
    """

    th = angulo_setor
    # --- posição local antes da rotação ---
    x0 = (r + desloc_radial) * np.cos(th) - desloc_tang * np.sin(th)
    y0 = (r + desloc_radial) * np.sin(th) + desloc_tang * np.cos(th)
    z0 = z_local + desloc_axial

    pos_local = np.array([[x0], [y0], [z0]])
    pos_rot = R @ pos_local
    x_rot, y_rot, z_rot = pos_rot[:, 0]

    # --- direção ---
    radial = np.array([[np.cos(th)], [np.sin(th)], [0]])
    tangential = np.array([[-np.sin(th)], [np.cos(th)], [0]])
    axial = np.array([[0], [0], [1]])
    d = ''
    if orient.lower() in ['radial']:
        d = radial
    elif orient.lower() in ['tangential', 'tangencial']:
        d = tangential
    elif orient.lower() in ['axial', 'vertical']:
        d = axial

    d_rot = R @ d
    if orient == "axial":
        u, v, w = -d_rot[:, 0]
    else:
        u, v, w = d_rot[:, 0]

    # --- adiciona cone ---
    fig.add_trace(go.Cone(
        x=[x_rot], y=[y_rot], z=[z_rot],
        u=[u], v=[v], w=[w],
        sizemode="absolute", sizeref=size_ref,
        colorscale=color, showscale=False,
        name=nome, anchor="tail"
    ))

    if st.session_state.show:
        fig.add_trace(go.Scatter3d(
            x=[x_rot], y=[y_rot], z=[z_rot],
            mode="text",
            text=[f"<b><i>{nome}={valor:.1f} psi</i></b>"],  # HTML para negrito/itálico
            textposition="top center",
            textfont=dict(
                family="Arial, sans-serif",
                size=16,
                color="black"
            )
        ))


def adicionar_setas_parede(fig, raio, angulo, z, sig_h, sig_H, mostrar_texto=True):
    # Corpo da seta da esquerda (do ponto (0, 1.5, 1) até (0, 1, 1))
    # Base da seta (na parede do cilindro)
    x_base = raio * np.cos(angulo) + 0.5 * np.cos(angulo)
    y_base = raio * np.sin(angulo) + 0.5 * np.sin(angulo)

    # Ponta da seta (um pouco mais para dentro)
    x_ponta = (raio - 0.5) * np.cos(angulo) + 0.5 * np.cos(angulo)
    y_ponta = (raio - 0.5) * np.sin(angulo) + 0.5 * np.sin(angulo)

    # Corpo da seta
    fig.add_trace(go.Scatter3d(
        x=[x_base, x_ponta],
        y=[y_base, y_ponta],
        z=[z, z],
        mode="lines",
        line=dict(color="blue", width=6),
        showlegend=False
    ))

    # Ponta da seta (abertura)
    desloc = 0.1
    fig.add_trace(go.Scatter3d(
        x=[1.1 * x_ponta - desloc * np.sin(angulo),
           x_ponta,
           1.1 * x_ponta + desloc * np.sin(angulo)],
        y=[1.1 * y_ponta + desloc * np.cos(angulo),
           y_ponta,
           1.1 * y_ponta - desloc * np.cos(angulo)],
        z=[z, z, z],
        mode="lines",
        line=dict(color="blue", width=6),
        showlegend=mostrar_texto,
        name="σH" if mostrar_texto else None
    ))

    if st.session_state.show:
        fig.add_trace(go.Scatter3d(
            x=[x_base, x_ponta],
            y=[y_base, y_ponta],
            z=[z, z],
            mode="text",
            text=[f"<b><i>σH={sig_H:.1f} psi</i></b>"],  # HTML para negrito/itálico
            textposition="top center",
            textfont=dict(
                family="Arial, sans-serif",
                size=16,
                color="black"
            )
        ))

    x_base = raio * np.cos(angulo + np.pi / 2) + 0.5 * np.cos(angulo + np.pi / 2)
    y_base = raio * np.sin(angulo + np.pi / 2) + 0.5 * np.sin(angulo + np.pi / 2)

    # Ponta da seta (um pouco mais para dentro)
    x_ponta = (raio - 0.5) * np.cos(angulo + np.pi / 2) + 0.5 * np.cos(angulo + np.pi / 2)
    y_ponta = (raio - 0.5) * np.sin(angulo + np.pi / 2) + 0.5 * np.sin(angulo + np.pi / 2)

    # Corpo da seta 2
    fig.add_trace(go.Scatter3d(
        x=[x_base, x_ponta],
        y=[y_base, y_ponta],
        z=[z, z],
        mode="lines",
        line=dict(color="red", width=6),
        showlegend=False
    ))

    # Ponta da seta (abertura)
    desloc = 0.1
    fig.add_trace(go.Scatter3d(
        x=[1.1 * x_ponta - desloc * np.sin(angulo + np.pi / 2),
           x_ponta,
           1.1 * x_ponta + desloc * np.sin(angulo + np.pi / 2)],
        y=[1.1 * y_ponta + desloc * np.cos(angulo + np.pi / 2),
           y_ponta,
           1.1 * y_ponta - desloc * np.cos(angulo + np.pi / 2)],
        z=[z, z, z],
        mode="lines",
        line=dict(color="red", width=6),
        showlegend=mostrar_texto,
        name="σh" if mostrar_texto else None
    ))

    if st.session_state.show:
        fig.add_trace(go.Scatter3d(
            x=[x_base, x_ponta],
            y=[y_base, y_ponta],
            z=[z, z],
            mode="text",
            text=[f"<b><i>σh={sig_h:.1f} psi</i></b>"],  # HTML para negrito/itálico
            textposition="top center",
            textfont=dict(
                family="Arial, sans-serif",
                size=16,
                color="black"
            )
        ))


def criar_grafico(parametros, sr, sig_t, sig_a, sig_h, sig_H, ang_theta, tvd, ang_azi, ang_inc, ra, op, ang_horizontal, view, lg):
    fig = go.Figure()

    x, y, z, u, R, cont = gerar_cilindro_inclinado(r=1, h=2, incl=ang_inc, azi=ang_azi,
                                                   angulo_central=np.deg2rad(ang_theta))

    if st.session_state.ff:
        colorscale = [
            [0.0, "red"],  # no ponto central
            [0.15, "yellow"],
            [0.6, "green"],
            [0.85, "yellow"],
            [1.0, "red"]  # oposto
        ]

        fig.add_trace(go.Surface(
            x=x, y=y, z=z,
            surfacecolor=cont,
            colorscale=colorscale,
            opacity=op,
            showscale=False,
            hoverinfo="skip",
            name="Paredes do Poço"
        ))
    else:
        fig.add_trace(go.Surface(x=x, y=y, z=z,
                                 opacity=0.5,
                                 colorscale="Greys",
                                 showscale=False))

    if "Sistemas de coordenadas do poço" in parametros:
        eixos = np.eye(3)
        # Rotacionar eixos
        eixos_rot = R @ eixos.T

        cores = ["black", "black", "black"]
        nomes = ["x", "y", "z"]

        for i in range(3):
            fig.add_trace(go.Scatter3d(
                x=[0, eixos_rot[0, i]],
                y=[0, eixos_rot[1, i]],
                z=[0, eixos_rot[2, i]],
                mode="lines+text",
                line=dict(color=cores[i], width=6, dash="solid"),
                text=[None, nomes[i]],
                textfont=dict(size=12, color="black", family="Arial Black"),
                showlegend=False,
                textposition="top center",
                name=nomes[i]
            ))

    if st.session_state.arr:
        # Gera setor sombreado
        gerar_setor_fora_inclinado(fig, np.deg2rad(360), R=R, r_ext=ra)
    if "Profundidade" in parametros:
        fig.add_trace(go.Scatter3d(
            x=[0, 0], y=[0, 0], z=[1, 0],
            mode="text",
            text=[f"TVD={tvd:.1f} m"],  # aqui você pode colocar 'θ', 'σθ' etc.
            textfont=dict(size=16, color="black"),
            showlegend=False
        ))
    if "σθ" in parametros:
        adicionar_cone_setor(fig, np.deg2rad(ang_theta), R=R, r=1.0,
                             orient='tangential', color='Blues', nome='σθ', desloc_tang=-0.3, valor=sig_t)
    if "σr" in parametros:
        adicionar_cone_setor(fig, np.deg2rad(ang_theta), R=R, r=1.0,
                             orient='radial', color='Reds', nome='σr', desloc_radial=-0.3, valor=sr)
    if "σa" in parametros:
        adicionar_cone_setor(fig, np.deg2rad(ang_theta), R=R, r=1.0,
                             orient='axial', color='Greens', nome='σa', desloc_axial=0.3, valor=sig_a)
    if "Direção do poço" in parametros:
        x_end, y_end = u[0] * 1, u[1] * 1
        fig.add_trace(go.Scatter3d(
            x=[0, -x_end],
            y=[0, -y_end],
            z=[0, 0],
            mode="lines",
            line=dict(color="brown", width=6),
            name="Azimute"
        ))

        desloc = 0.1
        fig.add_trace(go.Scatter3d(
            x=[(x_end + 1.3 * np.cos(np.deg2rad(ang_azi))) + desloc * u[1],
               -x_end,
               (x_end + 1.3 * np.cos(np.deg2rad(ang_azi))) - desloc * u[1]],
            y=[(y_end + 1.3 * np.sin(np.deg2rad(ang_azi))) - desloc * u[0],
               -y_end,
               (y_end + 1.3 * np.sin(np.deg2rad(ang_azi))) + desloc * u[0]],
            z=[0, 0, 0],
            mode="lines",
            line=dict(color="brown", width=6),
            showlegend=False,
            name="AZI"
        ))
        if ang_inc > 0:
            fig.add_trace(go.Scatter3d(
                x=[-x_end, -x_end],
                y=[-y_end, -y_end],
                z=[0, 0],
                mode="text",
                text=[F"AZI={ang_azi:.1f}º"],
                textfont=dict(size=16, color="black"),
                showlegend=False
            ))
    if "Coordenadas geográficas" in parametros:
        adicionar_eixos(fig)
    if "Tensões Horizontais" in parametros:
        adicionar_setas_parede(fig, 2, np.radians(ang_horizontal), 0, sig_h, sig_H)
    if "Dir. Tensões principais" in parametros:
        adicionar_setor_circular(fig, angulo=np.deg2rad(ang_theta), R=R, origem=(0, 0, 0))

    if lg == "Sim":
        ax = True
    else:
        ax = False
    if st.session_state.arr:
        margem = ra
    else:
        margem = 0

    fig.update_layout(
        showlegend=False,
        legend=dict(
            orientation='h',
            x=0.5,
            xanchor='center',
            y=0,
            yanchor='bottom'
        )
    )

    if view == "Vista de planta":
        fig.update_layout(
            scene_camera=dict(
                eye=dict(x=0, y=0, z=1),
                up=dict(x=-1, y=0, z=0)
            ),

        )
    elif view == "Vista de seção N/S":
        fig.update_layout(
            scene_camera=dict(
                eye=dict(x=0, y=1, z=0)  # posição do observador
            )
        )
    elif view == "Vista de seção E/W":
        fig.update_layout(

            scene_camera=dict(
                eye=dict(x=1, y=0, z=0)  # posição do observador
            )
        )
    elif view == "Vista axial do poço":
        L = 1
        fig.update_layout(
            scene_dragmode='turntable',
            scene_camera=dict(
                eye=dict(x=-u[0] * L, y=u[1] * L, z=u[2] * L),
                center=dict(x=0, y=0, z=0),  # centro da vista
                up=dict(x=-1, y=0, z=0)
            ),
        )

    elif view == "Dentro do Poço":
        fig.update_layout(
            scene_dragmode='turntable',
            scene_camera=dict(
                eye=dict(x=0.01, y=0, z=0.01),
                center=dict(x=0, y=0, z=0.01),  # centro da vista
                up=dict(x=0, y=0, z=0)
            )
        )

    fig.update_layout(
        scene=dict(xaxis=dict(range=[margem + 3, -margem - 3], visible=ax),  # inverte eixo X
                   yaxis=dict(range=[-margem - 3, margem + 3], visible=ax),  # inverte eixo Y
                   zaxis=dict(range=[-margem - 3, margem + 3], visible=ax),
                   aspectmode="cube"),
        margin=dict(l=0, r=0, t=0, b=0),
        showlegend=False
    )

    return fig


def normal(df):
    prof_ini = st.session_state.rtkb
    prof_fim = float(df["Profundidade"].min())

    val_ini = 0.0
    val_fim = float(st.session_state.gn)

    # 🔹 alpha vindo do usuário
    alpha = float(st.session_state.get("alfa_pp", 10))

    n = len(df)
    profundidade = np.linspace(prof_ini, prof_fim, n)

    denom = (prof_fim - prof_ini)
    if denom == 0:
        valores = np.full(n, np.nan)
    else:
        t = (profundidade - prof_ini) / denom
        t = np.clip(t, 0.0, 1.0)

        valores = val_ini + (val_fim - val_ini) * (
            (1 - np.exp(-alpha * t)) / (1 - np.exp(-alpha))
        )

    df_gfs = pd.DataFrame({
        "Profundidade (m)": profundidade,
        "Gradiente de Pressão de Poros (lb/gal)": valores
    })

    # Até nf = None
    if prof_fim > prof_ini:
        mask_none = df_gfs["Profundidade (m)"] <= prof_ini
    else:
        mask_none = df_gfs["Profundidade (m)"] >= prof_ini

    df_gfs.loc[mask_none, "Gradiente de Pressão de Poros (lb/gal)"] = None

    st.session_state.df_gfs = df_gfs
    return df_gfs


def lito(ax1, df_pp, profundidades, litologias, bases):
    try:
        label = True
        line_w = 0.8
        if "s_gr" not in st.session_state:
            st.session_state.s_gr = False
        if not st.session_state.s_gr:
            curva = df_pp['Perfil Raio Gama']
        else:
            curva = df_pp['Raio Gama Suavizado']
        # Define apenas os topos e os nomes das litologias
        if not profundidades:
            prof = [0]
            lito_t = ['Fm. Permeável']
            if curva.iloc[0] >= df_pp['LBF_calc'].iloc[0]:
                aux = 'Fm. Permeável'
            else:
                aux = 'Folhelho'

            for i, line in enumerate(df_pp['Profundidade (m)']):
                if curva.iloc[i] >= df_pp['LBF_calc'].iloc[
                    i] and aux == 'Fm. Permeável':
                    prof.append(line)
                    lito_t.append('Folhelho')
                    aux = 'Folhelho'
                elif curva.iloc[i] <= df_pp['LBF_calc'].iloc[
                    i] and aux == 'Folhelho':
                    prof.append(line)
                    lito_t.append('Fm. Permeável')
                    aux = 'Fm. Permeável'

            litho_tops = [[x, y] for x, y in zip(prof, lito_t)]
            label = False
            line_w = 0

        else:
            litho_tops = [[x, y] for x, y in zip(profundidades, litologias)]
            label = True
            line_w = 0.8

        # Calcula as bases automaticamente
        litho_intervals = []
        for i, (top, lit) in enumerate(litho_tops):
            if i < len(litho_tops) - 1:
                base = litho_tops[i + 1][0]  # base = topo da próxima
            else:
                base = bases  # última formação vai até o fundo
            litho_intervals.append((top, base, lit))

        # Estilos de preenchimento
        litho_styles = {
            "Arenito": {"color": "#fff7a1", "hatch": "...", "edgecolor": "black"},
            "Fm. Permeável": {"color": "#fff7a1", "hatch": "...", "edgecolor": "black"},
            "Folhelho": {"color": "#2f4f4f", "hatch": None, "edgecolor": "black"},
            "Calcário": {"color": "#a7c7e7", "hatch": "///", "edgecolor": "#003366"},
            "Carbonato": {"color": "#cfe8f3", "hatch": "xx", "edgecolor": "#003366"},
            "Siltito": {"color": "#8b4513", "hatch": None, "edgecolor": "black"},
            "Basalto": {"color": "#2b2b2b", "hatch": None, "edgecolor": "black"},
            "Conglomerado": {"color": "#ffb347", "hatch": "oo", "edgecolor": "black"},
            "Halita": {"color": "#ffffff", "hatch": None, "edgecolor": "black"},
        }

        # Configurações do eixo da litologia
        ax1.set_xlim(0, 0.5)
        ax1.set_ylim(st.session_state.y_max_pp, st.session_state.y_min_pp)
        ax1.set_xticks([])
        ax1.tick_params(axis='y', which='both', left=False, labelleft=False)
        ax1.set_title("Litologia", fontsize=10, fontweight='bold')

        # Desenha as camadas automaticamente
        for top, base, lit in litho_intervals:
            style = litho_styles.get(lit,
                                     {"color": "gray", "hatch": None, "edgecolor": "black"})
            rect = mpatches.Rectangle(
                (0, top),
                width=0.5,
                height=base - top,
                facecolor=style["color"],
                edgecolor=style["edgecolor"],
                hatch=style["hatch"],
                linewidth=line_w,
            )
            ax1.add_patch(rect)

            mid = (top + base) / 2
            if label:
                txt = ax1.text(
                    0.25, mid, lit,
                    ha="center", va="center",
                    fontsize=9, fontweight="bold", color="black",
                    zorder=5,
                    rotation = 0
                )
                txt.set_path_effects([
                    path_effects.Stroke(linewidth=1.5, foreground='white'),
                    path_effects.Normal()
                ])
    except Exception as e:
        pass


def idade_formacao(ax, df_idade, y_max):

    ax.set_xlim(0, 1)
    ax.set_xticks([])
    ax.set_ylim(y_max, 0)

    idades_unicas = df_idade["Idade"].unique()
    cmap = plt.cm.get_cmap("Accent", len(idades_unicas))

    cores = {
        idade: cmap(i)
        for i, idade in enumerate(idades_unicas)
    }

    for _, row in df_idade.iterrows():
        ax.fill_betweenx(
            [row["Topo (m)"], row["Base (m)"]],
            0, 1,
            color=cores[row["Idade"]],
            edgecolor="black"
        )

        ax.text(
            0.5,
            (row["Topo (m)"] + row["Base (m)"]) / 2,
            row["Idade"],
            ha="center",
            va="center",
            fontsize=8,
            rotation=90,
            fontweight="bold",
            color="black",
            path_effects=[
                pe.Stroke(linewidth=3, foreground="white"),
                pe.Normal()
            ]
        )

    ax.set_title("Idade", fontsize=10, fontweight="bold")


def add_watermark(ax, logo_path="logo.png", xy=(0.80, 0.25), zoom=0.20, alpha=0.10, zorder=0):
    """
    Marca d'água robusta:
    - NÃO distorce em eixo log (usa ax.transAxes)
    - transparência SEMPRE muda (ajusta canal alpha do RGBA)
    """
    if not os.path.exists(logo_path):
        return

    # lê com PIL e força RGBA
    im = Image.open(logo_path).convert("RGBA")
    arr = np.asarray(im).astype(np.float32)

    # aplica alpha global multiplicando o canal A (0..255)
    arr[..., 3] = arr[..., 3] * float(alpha)
    arr[..., 3] = np.clip(arr[..., 3], 0, 255)

    arr = arr.astype(np.uint8)

    oi = OffsetImage(arr, zoom=zoom)  # aqui já vai com alpha aplicado
    ab = AnnotationBbox(
        oi,
        xy=xy,
        xycoords=ax.transAxes,   # independe de log/linear
        frameon=False,
        box_alignment=(0.5, 0.5),
        zorder=zorder
    )
    ax.add_artist(ab)


def draw_footer(c, width, footer_y):
    c.line(30, footer_y, width - 30, footer_y)


def salvar_fig_para_pdf(fig, path, dpi=200):
    # salva com boa resolução e recorte certinho (evita “distorção”/borrado)
    fig.savefig(
        path,
        dpi=dpi,
        bbox_inches="tight",
        facecolor=fig.get_facecolor()
    )


def desenhar_imagem_no_pdf(c, img_path, left, right, top, bottom, titulo=None):
    img = Image.open(img_path)
    img_w_px, img_h_px = img.size

    available_width = right - left
    available_height = top - bottom

    scale = min(available_width / img_w_px, available_height / img_h_px)

    img_w = img_w_px * scale
    img_h = img_h_px * scale

    x_pos = left + (available_width - img_w) / 2
    y_pos = top - img_h

    if titulo:
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(left + available_width / 2, top + 6, titulo)

    c.drawImage(
        img_path,
        x_pos,
        y_pos,
        width=img_w,
        height=img_h,
        preserveAspectRatio=True,
        mask="auto"
    )

    return y_pos - 20


def desenhar_fig_plotly_no_pdf(c, fig_plotly, left, right, top, bottom, titulo=None, scale=1.5):
    import io

    img_bytes = fig_plotly.to_image(format="png", scale=scale)
    img_buffer = io.BytesIO(img_bytes)
    img_reader = ImageReader(img_buffer)

    img_w_px, img_h_px = img_reader.getSize()
    available_width = right - left
    available_height = top - bottom

    scale_factor = min(available_width / img_w_px, available_height / img_h_px)

    img_w = img_w_px * scale_factor
    img_h = img_h_px * scale_factor

    x_pos = left + (available_width - img_w) / 2
    y_pos = top - img_h

    if titulo:
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(left + available_width / 2, top + 6, titulo)

    c.drawImage(
        img_reader,
        x_pos,
        y_pos,
        width=img_w,
        height=img_h,
        preserveAspectRatio=True,
        mask="auto"
    )

    return y_pos - 20


def desenhar_fig_mpl_no_pdf(c, fig, left, right, top, bottom, titulo=None, dpi=130):
    import io
    from reportlab.lib.utils import ImageReader

    img_buffer = io.BytesIO()
    fig.savefig(
        img_buffer,
        format="png",
        dpi=dpi,
        facecolor=fig.get_facecolor(),
        bbox_inches="tight"
    )
    img_buffer.seek(0)

    img_reader = ImageReader(img_buffer)
    img_width_px, img_height_px = img_reader.getSize()

    available_width = right - left
    available_height = top - bottom

    scale = min(available_width / img_width_px, available_height / img_height_px)

    img_width = img_width_px * scale
    img_height = img_height_px * scale

    x_pos = left + (available_width - img_width) / 2
    y_pos = top - img_height

    if titulo:
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(left + available_width / 2, top + 6, titulo)

    c.drawImage(
        img_reader,
        x_pos,
        y_pos,
        width=img_width,
        height=img_height,
        preserveAspectRatio=True,
        mask="auto"
    )

    return y_pos - 20


def draw_wrapped_text(c, text, x, y, max_width, line_height=12, font_name="Helvetica", font_size=10):
    if not text:
        return y

    c.setFont(font_name, font_size)

    # normaliza quebras de linha
    paragraphs = str(text).replace("\r\n", "\n").replace("\r", "\n").split("\n")

    for p in paragraphs:
        if p.strip() == "":
            y -= line_height  # linha em branco
            continue

        words = p.split()
        line = ""
        for w in words:
            test = (line + " " + w).strip()
            if c.stringWidth(test, font_name, font_size) <= max_width:
                line = test
            else:
                c.drawString(x, y, line)
                y -= line_height
                line = w

        if line:
            c.drawString(x, y, line)
            y -= line_height

    return y


def _classificar_falha_no_limite(row, ppg, lado):
                if row is None or ppg is None or pd.isna(ppg):
                    return ""

                if lado == "inferior":
                    ti = row.get("Tração Inferior", None)
                    ciA = row.get("Comp Inferior σθA", None)
                    ciB = row.get("Comp Inferior σθB", None)

                    if _aprox(ppg, ti):
                        return "Modo de falha no limite: Tração (limite inferior)"
                    if _aprox(ppg, ciA):
                        return "Modo de falha no limite: Colapso (σθA) – limite inferior"
                    if _aprox(ppg, ciB):
                        return "Modo de falha no limite: Colapso (σθB) – limite inferior"
                    return "Modo de falha no limite: Limite inferior atingido"

                # superior
                tsA = row.get("Tração Superior (σθA)", None)
                tsB = row.get("Tração Superior (σθB)", None)
                csA = row.get("Comp Superior σθA", None)
                csB = row.get("Comp Superior σθB", None)

                if _aprox(ppg, tsA):
                    return "Modo de falha no limite: Tração (σθA) – limite superior"
                if _aprox(ppg, tsB):
                    return "Modo de falha no limite: Tração (σθB) – limite superior"
                if _aprox(ppg, csA):
                    return "Modo de falha no limite: Compressão (σθA) – limite superior"
                if _aprox(ppg, csB):
                    return "Modo de falha no limite: Compressão (σθB) – limite superior"
                return "Modo de falha no limite: Limite superior atingido"


def _ler_trajetoria_do_xlsm(wb, modo: str) -> pd.DataFrame:
    """
    Lê diretamente por célula:
    - Planejada: header em B5:C5:D5 e dados em B6:C6:D...
    - Executada: header em K5:L5:M5 e dados em K6:L6:M...
    Retorna DF padronizado com colunas MD, Inc, Azi.
    """

    if "Trajetória" not in wb.sheetnames:
        raise ValueError("A aba 'Trajetória' não existe no arquivo.")

    ws = wb["Trajetória"]

    if modo == "Executada":
        col_md, col_inc, col_azi = "K", "L", "M"
    else:
        col_md, col_inc, col_azi = "B", "C", "D"

    h_md = ws[f"{col_md}6"].value
    h_inc = ws[f"{col_inc}6"].value
    h_azi = ws[f"{col_azi}6"].value

    def _norm(x):
        return str(x).strip().lower() if x is not None else ""

    if not (
        "md" in _norm(h_md)
        and ("incl" in _norm(h_inc) or "inc" in _norm(h_inc))
        and ("az" in _norm(h_azi) or "azim" in _norm(h_azi))
    ):
        raise ValueError(
            f"Header não bate no esperado. Lido em {col_md}5:{col_azi}5 -> "
            f"{[h_md, h_inc, h_azi]}. "
            f"Confirme se o header está mesmo na linha 5 (Excel)."
        )

    rows = []
    r = 6
    while r <= ws.max_row:
        md = ws[f"{col_md}{r}"].value
        inc = ws[f"{col_inc}{r}"].value
        azi = ws[f"{col_azi}{r}"].value

        if md is None or md == "":
            break

        rows.append((md, inc, azi))
        r += 1

    df = pd.DataFrame(rows, columns=["MD", "Inc", "Azi"])

    df["MD"] = pd.to_numeric(df["MD"], errors="coerce")
    df["Inc"] = pd.to_numeric(df["Inc"], errors="coerce")
    df["Azi"] = pd.to_numeric(df["Azi"], errors="coerce")

    df = df.dropna(subset=["MD", "Inc", "Azi"]).sort_values("MD").reset_index(drop=True)

    if df.empty:
        raise ValueError(f"Nenhum dado direcional encontrado em {col_md}6:{col_azi}...")

    if (df["MD"].diff().fillna(1) <= 0).any():
        raise ValueError("Coluna MD deve ser estritamente crescente.")

    return df


def _gerar_df_interp_a_partir_df1_df2(df1: pd.DataFrame, df2: pd.DataFrame) -> pd.DataFrame:
    """
    Replica a lógica do seu direcional(): interpola Inc/Azi da df2 para as profundidades de df1.
    Usa a mesma regra de expandir do zero via st.session_state.ex == 'Ativada'
    """
    # mesma lógica que você já usa dentro do direcional() :contentReference[oaicite:2]{index=2}
    expand_from_zero = (st.session_state.get("ex", "Desativada") == "Ativada")

    col_md1 = [col for col in df1.columns if "md" in col.lower()]
    col_md2 = [col for col in df2.columns if "md" in col.lower()]

    if not col_md1 or not col_md2:
        raise ValueError("Colunas de profundidade MD não encontradas em df1 ou df2.")
    if "Inc" not in df2.columns or "Azi" not in df2.columns:
        raise ValueError("df2 precisa conter colunas 'Inc' e 'Azi'.")

    md1_col = col_md1[0]
    md2_col = col_md2[0]

    md2 = df2[md2_col].astype(float).values
    inc2 = df2["Inc"].astype(float).values
    azi2 = df2["Azi"].astype(float).values

    sort_idx = np.argsort(md2)
    md2_sorted = md2[sort_idx]
    inc2_sorted = inc2[sort_idx]
    azi2_sorted = azi2[sort_idx]

    md1_original = df1[md1_col].astype(float).values
    md1_original = np.sort(md1_original)

    if expand_from_zero:
        first_depth = md1_original[0]
        md1_extra = np.arange(0, first_depth, 1.0)
        md1 = np.concatenate((md1_extra, md1_original))
    else:
        md1 = md1_original.copy()

    inc_interp = np.interp(md1, md2_sorted, inc2_sorted)
    azi_interp = np.interp(md1, md2_sorted, azi2_sorted)

    df_interp = pd.DataFrame({
        "Profundidade": md1,
        "Inc (°)": inc_interp,
        "Azi (°)": azi_interp
    })
    return df_interp


def _ler_litologia_do_xlsm(wb) -> pd.DataFrame:
    if "Litologia" not in wb.sheetnames:
        raise ValueError("A aba 'Litologia' não existe no arquivo.")

    ws = wb["Litologia"]

    h_fm = ws["C3"].value
    h_lit = ws["E3"].value
    h_top = ws["F3"].value
    h_bas = ws["G3"].value

    def _norm(x):
        return str(x).strip().lower() if x is not None else ""

    if not any(k in _norm(h_fm) for k in ["forma", "formação", "formacao", "fm"]):
        raise ValueError(f"Header de Formação inválido em C3: {h_fm}")
    if "litolog" not in _norm(h_lit):
        raise ValueError(f"Header de Litologia inválido em E3: {h_lit}")
    if "top" not in _norm(h_top):
        raise ValueError(f"Header de Topo inválido em F3: {h_top}")
    if "bas" not in _norm(h_bas):
        raise ValueError(f"Header de Base inválido em G3: {h_bas}")

    rows = []
    r = 4
    while r <= ws.max_row:
        fm = ws[f"C{r}"].value
        lit = ws[f"E{r}"].value
        topo = ws[f"F{r}"].value
        base = ws[f"G{r}"].value

        if (fm in (None, "") and lit in (None, "") and topo in (None, "") and base in (None, "")):
            break

        rows.append((fm, lit, topo, base))
        r += 1

    df = pd.DataFrame(rows, columns=["Formação", "Litologia", "Topo", "Base"])

    df["Formação"] = df["Formação"].astype(str).str.strip()
    df["Litologia"] = df["Litologia"].astype(str).str.strip()
    df["Topo"] = pd.to_numeric(df["Topo"], errors="coerce")
    df["Base"] = pd.to_numeric(df["Base"], errors="coerce")

    df = df.dropna(subset=["Topo"]).reset_index(drop=True)
    df = df[(df["Litologia"] != "") | (df["Formação"] != "")].reset_index(drop=True)

    if df.empty:
        raise ValueError("Nenhuma linha válida encontrada na aba 'Litologia'.")

    return df


def _aplicar_litologia_no_state(poco_nome: str, df_lito: pd.DataFrame):
    if "pocos" not in st.session_state:
        st.session_state.pocos = {}
    if poco_nome not in st.session_state.pocos:
        st.session_state.pocos[poco_nome] = {}

    st.session_state.pocos[poco_nome]["profundidade"] = df_lito["Topo"].astype(float).tolist()
    st.session_state.pocos[poco_nome]["litologia"] = df_lito["Litologia"].astype(str).tolist()

    # ✅ agora vem do Excel
    st.session_state.pocos[poco_nome]["formation"] = df_lito["Formação"].astype(str).tolist()

    st.session_state.n_fm = len(df_lito)

    # Pré-preenche widgets
    for i in range(st.session_state.n_fm):
        st.session_state[f"prof_{i}"] = float(st.session_state.pocos[poco_nome]["profundidade"][i])
        st.session_state[f"fm_{i}"] = st.session_state.pocos[poco_nome]["formation"][i]

        lit = st.session_state.pocos[poco_nome]["litologia"][i]
        mapa_lit = {
            "arenito": "Arenito",
            "folhelho": "Folhelho",
            "calcario": "Calcário",
            "calcário": "Calcário",
            "carbonato": "Carbonato",
            "siltito": "Siltito",
            "conglomerado": "Conglomerado",
            "halita": "Halita",
            "basalto": "Basalto",
        }
        st.session_state[f"lit_{i}"] = mapa_lit.get(str(lit).strip().lower(), "Arenito")


def _ler_inicio_do_xlsm(wb) -> dict:
    """
    Lê valores fixos na aba 'Início' via células:
      - D5  -> Nome do poço
      - D10 -> Objetivo (comments)
      - D7  -> Easting
      - D6  -> Northing
    """
    if "Início" not in wb.sheetnames and "Inicio" not in wb.sheetnames:
        raise ValueError("A aba 'Início' (ou 'Inicio') não existe no arquivo.")

    ws = wb["Início"] if "Início" in wb.sheetnames else wb["Inicio"]

    nome_poco = ws["D5"].value
    objetivo = ws["D10"].value
    easting = ws["D7"].value
    northing = ws["D6"].value

    def _to_float(v):
        try:
            if v is None or v == "":
                return None
            return float(v)
        except Exception:
            return None

    return {
        "poco": None if nome_poco in (None, "") else str(nome_poco).strip(),
        "comments": "" if objetivo in (None, "") else str(objetivo).strip(),
        "easting": _to_float(easting),
        "northing": _to_float(northing),
    }


def _ler_peso_fluido_do_xlsm(wb) -> pd.DataFrame:
    if "Fluido" in wb.sheetnames:
        ws = wb["Fluido"]
    elif "Geopressões" in wb.sheetnames:
        ws = wb["Geopressões"]
    elif "Geopressoes" in wb.sheetnames:
        ws = wb["Geopressoes"]
    else:
        raise ValueError("A aba 'Fluido' não existe no arquivo.")

    h_md = ws["B5"].value
    h_wp = ws["C5"].value
    h_we = ws["D5"].value

    def _norm(x):
        return str(x).strip().lower() if x is not None else ""

    if "prof" not in _norm(h_md):
        raise ValueError(f"Header inesperado em B5: {h_md}")
    if "planej" not in _norm(h_wp):
        raise ValueError(f"Header inesperado em C5: {h_wp}")
    if "execut" not in _norm(h_we):
        raise ValueError(f"Header inesperado em D5: {h_we}")

    rows = []
    vazio_seguidos = 0

    for md, wp, we in ws.iter_rows(
        min_row=6,
        max_row=min(ws.max_row, 10000),
        min_col=2,   # B
        max_col=4,   # D
        values_only=True
    ):
        if md in (None, ""):
            vazio_seguidos += 1
            if vazio_seguidos >= 3:
                break
            continue

        vazio_seguidos = 0
        rows.append((md, wp, we))

    if not rows:
        raise ValueError("Não encontrei dados de peso do fluido em B6/C6/D6...")

    df = pd.DataFrame(
        rows,
        columns=[
            "Profundidade (m)",
            "Peso do Fluido Planejado (lb/gal)",
            "Peso do Fluido Executado (lb/gal)",
        ],
    )

    df = df.apply(pd.to_numeric, errors="coerce")
    df = df.dropna(subset=["Profundidade (m)"]).sort_values("Profundidade (m)").reset_index(drop=True)

    if df.empty:
        raise ValueError("Não encontrei dados válidos de peso do fluido em B6/C6/D6...")

    return df


def _ler_sapatas_do_xlsm(wb) -> pd.DataFrame:
    """
    Lê as sapatas na aba 'Início' (ou 'Inicio').

    Células:
      - TVD: E13, E14, E15 e opcionalmente E16
      - OD : F13, F14, F15 e opcionalmente F16

    Retorna DataFrame com:
      - Ordem
      - Fase
      - Profundidade da sapata (m)
    """
    if "Início" not in wb.sheetnames and "Inicio" not in wb.sheetnames:
        raise ValueError("A aba 'Início' (ou 'Inicio') não existe no arquivo.")

    ws = wb["Início"] if "Início" in wb.sheetnames else wb["Inicio"]

    def _to_float(v):
        try:
            if v is None or str(v).strip() == "":
                return None
            return float(v)
        except Exception:
            return None

    rows = []

    for ordem, r in enumerate(range(13, 17), start=1):
        tvd = _to_float(ws[f"E{r}"].value)
        od = _to_float(ws[f"F{r}"].value)

        if tvd is not None and tvd > 0:
            rows.append((ordem, od, tvd))

    if not rows:
        raise ValueError("Nenhuma sapata válida foi encontrada em E13:E16 / F13:F16 da aba 'Início'.")

    sapatas_df = pd.DataFrame(
        rows,
        columns=["Ordem", "Fase", "Profundidade da sapata (m)"]
    )

    sapatas_df["Ordem"] = pd.to_numeric(sapatas_df["Ordem"], errors="coerce").astype(int)
    sapatas_df["Fase"] = pd.to_numeric(sapatas_df["Fase"], errors="coerce")
    sapatas_df["Profundidade da sapata (m)"] = pd.to_numeric(
        sapatas_df["Profundidade da sapata (m)"], errors="coerce"
    )

    sapatas_df = sapatas_df.dropna(subset=["Profundidade da sapata (m)"])
    sapatas_df = sapatas_df[sapatas_df["Profundidade da sapata (m)"] > 0].copy()

    sapatas_df["Fase"] = sapatas_df["Fase"].fillna(0).astype(float).round(3)
    sapatas_df["Profundidade da sapata (m)"] = sapatas_df["Profundidade da sapata (m)"].astype(float)

    return sapatas_df.reset_index(drop=True)


def _ler_fases_do_xlsm(wb) -> pd.DataFrame:
    """
    Lê as fases (diâmetro das brocas) na aba 'Início' (ou 'Inicio').

    Células:
      - C21, C22, C23, C24

    Retorna DataFrame com:
      - Ordem
      - Fase
    """
    if "Início" not in wb.sheetnames and "Inicio" not in wb.sheetnames:
        raise ValueError("A aba 'Início' (ou 'Inicio') não existe no arquivo.")

    ws = wb["Início"] if "Início" in wb.sheetnames else wb["Inicio"]

    def _to_float(v):
        try:
            if v is None or str(v).strip() == "":
                return None
            return float(v)
        except Exception:
            return None

    rows = []

    for ordem, r in enumerate(range(21, 25), start=1):
        fase = _to_float(ws[f"C{r}"].value)

        if fase is not None and fase > 0:
            rows.append((ordem, fase))

    if not rows:
        raise ValueError("Nenhuma fase válida foi encontrada em C21:C24 da aba 'Início'.")

    fases_df = pd.DataFrame(rows, columns=["Ordem", "Fase"])
    fases_df["Ordem"] = pd.to_numeric(fases_df["Ordem"], errors="coerce").astype(int)
    fases_df["Fase"] = pd.to_numeric(fases_df["Fase"], errors="coerce")

    fases_df = fases_df.dropna(subset=["Fase"])
    fases_df = fases_df[fases_df["Fase"] > 0].copy()
    fases_df["Fase"] = fases_df["Fase"].astype(float).round(3)

    return fases_df.reset_index(drop=True)


def _fmt_polegada(v):
    try:
        if v is None or pd.isna(v):
            return "—"
        v = float(v)
        if v.is_integer():
            return f'{int(v)}"'
        return f'{v:g}"'
    except Exception:
        return str(v)


def _montar_paginas_fases(fases_df, sapatas_df):
    """
    Retorna uma lista de dicts, um por página.

    Regras:
    - associa por ordem
    - se houver sapata correspondente:
        'Fase de X, Revestimento de Y'
    - se não houver:
        'Fase de X, Poço aberto'
    """
    paginas = []

    if fases_df is None or fases_df.empty:
        return paginas

    sap_map = {}
    if sapatas_df is not None and not sapatas_df.empty:
        for _, row in sapatas_df.iterrows():
            ordem = int(row["Ordem"])
            sap_map[ordem] = {
                "fase_revestimento": row.get("Fase", None),
                "prof_sapata": row.get("Profundidade da sapata (m)", None),
            }

    for _, row in fases_df.iterrows():
        ordem = int(row["Ordem"])
        fase_broca = row.get("Fase", None)

        if ordem in sap_map:
            fase_revest = sap_map[ordem]["fase_revestimento"]
            prof_sapata = sap_map[ordem]["prof_sapata"]
            titulo = f'Análise da Fase de {_fmt_polegada(fase_broca)}, Revestimento de {_fmt_polegada(fase_revest)}'
        else:
            fase_revest = None
            prof_sapata = None
            titulo = f'Fase de {_fmt_polegada(fase_broca)}, Poço aberto'

        paginas.append({
            "ordem": ordem,
            "fase_broca": fase_broca,
            "fase_revestimento": fase_revest,
            "prof_sapata": prof_sapata,
            "titulo": titulo,
        })

    return paginas


def _to_num(s):
    return pd.to_numeric(s, errors="coerce")


def _fmt2(v, suf=""):
    return f"{float(v):.2f}{suf}" if v is not None and pd.notna(v) else "—"


def _filtrar_intervalo(df, col_prof, prof_ini, prof_fim, incluir_fim=False):
    if df is None or df.empty or col_prof not in df.columns:
        return pd.DataFrame()

    s = pd.to_numeric(df[col_prof], errors="coerce")

    if prof_ini is None or prof_fim is None or pd.isna(prof_ini) or pd.isna(prof_fim):
        return pd.DataFrame()

    if incluir_fim:
        mask = (s >= float(prof_ini)) & (s <= float(prof_fim))
    else:
        mask = (s >= float(prof_ini)) & (s < float(prof_fim))

    return df.loc[mask].copy()


def _seg_contiguos(df_base, mask_bool, col_prof, col_exec, col_req):
    segs = []
    if df_base.empty:
        return segs

    mask = mask_bool.fillna(False).astype(bool).values
    prof = df_base[col_prof].astype(float).values
    execv = df_base[col_exec].astype(float).values
    reqv = df_base[col_req].astype(float).values

    i = 0
    n = len(df_base)
    while i < n:
        if not mask[i]:
            i += 1
            continue

        j = i
        while j + 1 < n and mask[j + 1]:
            j += 1

        prof_ini = prof[i]
        prof_fim = prof[j]

        exec_seg = execv[i:j + 1]
        req_seg = reqv[i:j + 1]

        diffs = np.abs(req_seg - exec_seg)
        if np.isfinite(diffs).any():
            k_rel = int(np.nanargmax(diffs))
            k_abs = i + k_rel
            diff_max = float(np.nanmax(diffs))
            prof_pior = float(prof[k_abs])
        else:
            diff_max = np.nan
            prof_pior = np.nan

        segs.append({
            "pi": prof_ini,
            "pf": prof_fim,
            "prof_pior": prof_pior,
            "exec_min": np.nanmin(exec_seg) if np.isfinite(exec_seg).any() else np.nan,
            "exec_max": np.nanmax(exec_seg) if np.isfinite(exec_seg).any() else np.nan,
            "req_min": np.nanmin(req_seg) if np.isfinite(req_seg).any() else np.nan,
            "req_max": np.nanmax(req_seg) if np.isfinite(req_seg).any() else np.nan,
            "diff_max": diff_max
        })

        i = j + 1

    return segs


def _draw_centered_text(c, x_center, y, text, font_name="Helvetica", font_size=9):
    c.setFont(font_name, font_size)
    tw = c.stringWidth(str(text), font_name, font_size)
    c.drawString(x_center - tw / 2, y, str(text))


def desenhar_mapa_folium_no_pdf(c, mapa_folium, left, right, top, bottom):
    mapa_folium.save('filename.png')
    if mapa_folium is None:
        raise ValueError("Mapa Folium não informado.")

    largura = right - left
    altura = top - bottom

    if largura <= 0 or altura <= 0:
        raise ValueError("Área inválida para desenhar o mapa.")

    try:
        png_data = mapa_folium._to_png(2)
    except Exception as e:
        raise RuntimeError(
            "Falha ao converter o mapa Folium para PNG. "
            "O método _to_png depende de Selenium + Firefox headless + GeckoDriver. "
            f"Erro original: {e}"
        )

    img = ImageReader(BytesIO(png_data))

    c.drawImage(
        img,
        left,
        bottom,
        width=largura,
        height=altura,
        preserveAspectRatio=True,
        mask='auto'
    )


def desenhar_tabela_segmentos(c, left_margin, right_margin, y_top, titulo, segs, col_balanco_label):
    box_width = right_margin - left_margin
    box_x = left_margin

    pad_left = 6
    pad_right = 6

    col_w = [0.20, 0.20, 0.21, 0.21, 0.18]
    inner_w = box_width - (pad_left + pad_right) * 2
    widths = [inner_w * w for w in col_w]

    x0 = box_x + pad_left
    xs = [x0]
    for w in widths:
        xs.append(xs[-1] + w)

    linha_altura = 15
    header_h = 18
    n_rows = max(1, len(segs))
    box_height = 30 + header_h + n_rows * linha_altura
    box_y = y_top - box_height

    c.setFillColorRGB(0.95, 0.95, 0.95)
    c.rect(box_x, box_y, box_width, box_height, fill=1, stroke=0)
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.8)
    c.rect(box_x, box_y, box_width, box_height, fill=0, stroke=1)

    c.line(box_x, box_y + box_height - 18, box_x + box_width, box_y + box_height - 18)
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(box_x + 10, box_y + box_height - 14, titulo)

    texto_y = box_y + box_height - 32

    if not segs:
        c.setFont("Helvetica", 10)
        c.drawString(box_x + 12, texto_y, "Não foram identificados trechos para esta condição.")
        return box_y - 14

    headers = ["Prof. inic.", "Prof. final", "Requerido", "Executado", col_balanco_label]
    c.setLineWidth(0.5)
    c.line(box_x + 8, texto_y - 4, box_x + box_width - 8, texto_y - 4)

    for i, htxt in enumerate(headers):
        xc = (xs[i] + xs[i + 1]) / 2
        _draw_centered_text(c, xc, texto_y, htxt, font_name="Helvetica-Bold", font_size=10)

    texto_y -= header_h

    for s in segs:
        row = [
            _fmt2(s["pi"], " m"),
            _fmt2(s["pf"], " m"),
            f'{_fmt2(s["req_min"])}–{_fmt2(s["req_max"])} lb/gal',
            f'{_fmt2(s["exec_min"])}–{_fmt2(s["exec_max"])} lb/gal',
            f'{_fmt2(s["diff_max"])} lb/gal',
        ]
        for i, txt in enumerate(row):
            xc = (xs[i] + xs[i + 1]) / 2
            _draw_centered_text(c, xc, texto_y, txt, font_name="Helvetica", font_size=10)

        texto_y -= linha_altura

    return box_y - 14


def desenhar_tabela_falhas_quebrada(c, left_margin, right_margin, y_top, titulo, subtrechos):
    box_width = right_margin - left_margin
    box_x = left_margin

    pad_left = 6
    pad_right = 6

    col_w = [0.22, 0.22, 0.56]
    inner_w = box_width - (pad_left + pad_right) * 2
    widths = [inner_w * w for w in col_w]

    x0 = box_x + pad_left
    xs = [x0]
    for w in widths:
        xs.append(xs[-1] + w)

    linha_altura = 15
    header_h = 18
    n_rows = max(1, len(subtrechos))
    box_height = 30 + (header_h if subtrechos else 0) + n_rows * linha_altura
    box_y = y_top - box_height

    c.setFillColorRGB(0.95, 0.95, 0.95)
    c.rect(box_x, box_y, box_width, box_height, fill=1, stroke=0)
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.8)
    c.rect(box_x, box_y, box_width, box_height, fill=0, stroke=1)

    c.line(box_x, box_y + box_height - 18, box_x + box_width, box_y + box_height - 18)
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(box_x + 10, box_y + box_height - 14, titulo)

    texto_y = box_y + box_height - 32

    if not subtrechos:
        c.setFont("Helvetica", 10)
        c.drawString(box_x + 12, texto_y, "Não foram identificados trechos para classificar o tipo de falha.")
        return box_y - 14

    headers = ["Prof. inic.", "Prof. final", "Tipo de falha"]
    c.setLineWidth(0.5)
    c.line(box_x + 8, texto_y - 4, box_x + box_width - 8, texto_y - 4)

    for i, htxt in enumerate(headers):
        xc = (xs[i] + xs[i + 1]) / 2
        _draw_centered_text(c, xc, texto_y, htxt, font_name="Helvetica-Bold", font_size=10)

    texto_y -= header_h

    for stc in subtrechos:
        row = [_fmt2(stc["pi"], " m"), _fmt2(stc["pf"], " m"), stc["tipo"]]
        for i, txt in enumerate(row):
            xc = (xs[i] + xs[i + 1]) / 2
            _draw_centered_text(c, xc, texto_y, txt, font_name="Helvetica", font_size=10)

        texto_y -= linha_altura

    return box_y - 14


def _unir_intervalos_mesmo_problema(subtrechos, tol=1e-9):
    """
    Une intervalos consecutivos/contíguos que possuem o mesmo tipo de problema.
    Retorna lista no formato:
    [
        {"pi": ..., "pf": ..., "tipo": ...},
        ...
    ]
    """
    if not subtrechos:
        return []

    # ordena por profundidade inicial
    itens = sorted(
        [
            {
                "pi": float(x["pi"]),
                "pf": float(x["pf"]),
                "tipo": str(x["tipo"])
            }
            for x in subtrechos
            if x.get("pi") is not None and x.get("pf") is not None
        ],
        key=lambda z: (z["pi"], z["pf"])
    )

    if not itens:
        return []

    unidos = [itens[0].copy()]

    for atual in itens[1:]:
        ultimo = unidos[-1]

        mesmo_tipo = atual["tipo"] == ultimo["tipo"]
        contiguo_ou_sobreposto = atual["pi"] <= (ultimo["pf"] + tol)

        if mesmo_tipo and contiguo_ou_sobreposto:
            ultimo["pf"] = max(ultimo["pf"], atual["pf"])
        else:
            unidos.append(atual.copy())

    return unidos


def desenhar_tabela_interpretacoes_syga(c, left_margin, right_margin, y_top, interpretacoes):
    """
    interpretacoes: lista de dicts com:
      - tipo
      - intervalos -> lista de strings
      - texto
    """
    if not interpretacoes:
        return y_top

    box_width = right_margin - left_margin
    box_x = left_margin

    pad_left = 6
    pad_right = 6

    col_w = [0.34, 0.28, 0.38]
    inner_w = box_width - (pad_left + pad_right) * 2
    widths = [inner_w * w for w in col_w]

    x0 = box_x + pad_left
    xs = [x0]
    for w in widths:
        xs.append(xs[-1] + w)

    linha_altura = 15
    header_h = 18

    n_rows = 0
    for item in interpretacoes:
        n_rows += max(1, len(item.get("intervalos", [])))

    box_height = 30 + header_h + n_rows * linha_altura
    box_y = y_top - box_height

    c.setFillColorRGB(0.95, 0.95, 0.95)
    c.rect(box_x, box_y, box_width, box_height, fill=1, stroke=0)
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.8)
    c.rect(box_x, box_y, box_width, box_height, fill=0, stroke=1)

    c.line(box_x, box_y + box_height - 18, box_x + box_width, box_y + box_height - 18)
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(box_x + 10, box_y + box_height - 14, "Interpretações SYGA")

    texto_y = box_y + box_height - 32

    headers = ["Problema", "Intervalo(s)", "Interpretação"]
    c.setLineWidth(0.5)
    c.line(box_x + 8, texto_y - 4, box_x + box_width - 8, texto_y - 4)

    for i, htxt in enumerate(headers):
        xc = (xs[i] + xs[i + 1]) / 2
        _draw_centered_text(c, xc, texto_y, htxt, font_name="Helvetica-Bold", font_size=10)

    texto_y -= header_h

    for item in interpretacoes:
        tipo = item.get("tipo", "")
        intervalos = item.get("intervalos", [])
        texto = item.get("texto", "")

        if not intervalos:
            intervalos = ["—"]

        for k, intervalo_txt in enumerate(intervalos):
            # Problema e interpretação só aparecem na primeira linha do grupo
            if k == 0:
                xc_tipo = (xs[0] + xs[1]) / 2
                xc_texto = (xs[2] + xs[3]) / 2
                _draw_centered_text(c, xc_tipo, texto_y, str(tipo), font_name="Helvetica", font_size=10)
                _draw_centered_text(c, xc_texto, texto_y, str(texto), font_name="Helvetica", font_size=10)

            xc_intervalo = (xs[1] + xs[2]) / 2
            _draw_centered_text(c, xc_intervalo, texto_y, str(intervalo_txt), font_name="Helvetica", font_size=10)

            texto_y -= linha_altura

    return box_y - 14


def _montar_interpretacoes_syga(sub_under, sub_over):
    """
    Agrupa os subtrechos por tipo de problema, mesmo que estejam em
    profundidades separadas.

    Retorna lista de dicts no formato:
    [
        {
            "tipo": "...",
            "intervalos": ["150.00 m–200.00 m", "350.00 m–400.00 m"],
            "texto": "..."
        },
        ...
    ]
    """
    todos = []

    for item in (sub_under or []):
        todos.append({
            "pi": float(item["pi"]),
            "pf": float(item["pf"]),
            "tipo": str(item["tipo"])
        })

    for item in (sub_over or []):
        todos.append({
            "pi": float(item["pi"]),
            "pf": float(item["pf"]),
            "tipo": str(item["tipo"])
        })

    if not todos:
        return []

    # ordena por profundidade
    todos = sorted(todos, key=lambda x: (x["pi"], x["pf"]))

    agrupados = {}
    ordem_tipos = []

    for item in todos:
        tipo = item["tipo"]
        intervalo_txt = f'{_fmt2(item["pi"], " m")}–{_fmt2(item["pf"], " m")}'

        if tipo not in agrupados:
            agrupados[tipo] = []
            ordem_tipos.append(tipo)

        agrupados[tipo].append(intervalo_txt)

    interpretacoes = []
    for tipo in ordem_tipos:
        interpretacoes.append({
            "tipo": tipo,
            "intervalos": agrupados[tipo],
            "texto": f"[Texto para: {tipo}]"
        })

    return interpretacoes


def _desenhar_pagina_fase(
    c, width, height, logo, footer_y,
    titulo, fase_broca, fase_revestimento,
    prof_ini, prof_fim,
    df_cmp_global,
    draw_header,
    incluir_fim=False
):
    # -------------------------------------------------
    # Validação ANTES de desenhar a página
    # -------------------------------------------------
    if prof_ini is None or prof_fim is None or pd.isna(prof_ini) or pd.isna(prof_fim) or prof_fim <= prof_ini:
        return

    if df_cmp_global is None or df_cmp_global.empty:
        return

    df_cmp_fase = _filtrar_intervalo(
        df_cmp_global,
        "Profundidade (m)",
        prof_ini,
        prof_fim,
        incluir_fim=incluir_fim
    )

    # Se não houver dados na fase, não gera página
    if df_cmp_fase.empty:
        return

    # -------------------------------------------------
    # Começa a desenhar a página
    # -------------------------------------------------
    y = draw_header(c, width, height, logo)

    left_margin = 40
    right_margin = width - 40

    c.setFont("Helvetica-Bold", 18)
    c.drawString(left_margin, y, titulo)
    y -= 10
    c.line(left_margin, y, right_margin, y)
    y -= 18

    # -------------------------------------------------
    # Tabela-resumo da fase (1 linha)
    # -------------------------------------------------
    def desenhar_tabela_resumo_fase(y_top):
        box_width = right_margin - left_margin
        box_x = left_margin

        pad_left = 6
        pad_right = 6

        col_w = [0.28, 0.30, 0.42]
        inner_w = box_width - (pad_left + pad_right) * 2
        widths = [inner_w * w for w in col_w]

        x0 = box_x + pad_left
        xs = [x0]
        for w in widths:
            xs.append(xs[-1] + w)

        linha_altura = 18
        header_h = 18
        box_height = 30 + header_h + linha_altura
        box_y = y_top - box_height

        c.setFillColorRGB(0.95, 0.95, 0.95)
        c.rect(box_x, box_y, box_width, box_height, fill=1, stroke=0)
        c.setStrokeColorRGB(0, 0, 0)
        c.setLineWidth(0.8)
        c.rect(box_x, box_y, box_width, box_height, fill=0, stroke=1)

        c.line(box_x, box_y + box_height - 18, box_x + box_width, box_y + box_height - 18)
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(box_x + 10, box_y + box_height - 14, "Dados da Fase")

        texto_y = box_y + box_height - 32

        headers = ["Diâmetro da fase", "Diâmetro do revestimento", "Intervalo analisado"]
        c.setLineWidth(0.5)
        c.line(box_x + 8, texto_y - 4, box_x + box_width - 8, texto_y - 4)

        for i, htxt in enumerate(headers):
            xc = (xs[i] + xs[i + 1]) / 2
            _draw_centered_text(c, xc, texto_y, htxt, font_name="Helvetica-Bold", font_size=10)

        texto_y -= header_h

        diam_fase_txt = _fmt_polegada(fase_broca)
        diam_revest_txt = (
            _fmt_polegada(fase_revestimento)
            if fase_revestimento is not None and pd.notna(fase_revestimento)
            else "Poço aberto"
        )
        intervalo_txt = f'{_fmt2(prof_ini, " m")} até {_fmt2(prof_fim, " m")}'

        row = [diam_fase_txt, diam_revest_txt, intervalo_txt]
        for i, txt in enumerate(row):
            xc = (xs[i] + xs[i + 1]) / 2
            _draw_centered_text(c, xc, texto_y, txt, font_name="Helvetica", font_size=10)

        return box_y - 14

    y = desenhar_tabela_resumo_fase(y)

    # -------------------------------------------------
    # Análise dos trechos
    # -------------------------------------------------
    segs_under = _seg_contiguos(
        df_base=df_cmp_fase,
        mask_bool=df_cmp_fase["is_under"],
        col_prof="Profundidade (m)",
        col_exec="Peso do Fluido (lb/gal)",
        col_req="Max Inferior"
    )

    segs_over = _seg_contiguos(
        df_base=df_cmp_fase,
        mask_bool=df_cmp_fase["is_over"],
        col_prof="Profundidade (m)",
        col_exec="Peso do Fluido (lb/gal)",
        col_req="Min Superior"
    )

    # Só desenha a tabela se houver trechos
    if segs_under:
        y = desenhar_tabela_segmentos(
            c=c,
            left_margin=left_margin,
            right_margin=right_margin,
            y_top=y,
            titulo="Trechos em Underbalance (Peso do Fluido < Limite Inferior)",
            segs=segs_under,
            col_balanco_label="Underbalance máx."
        )

    if segs_over:
        y = desenhar_tabela_segmentos(
            c=c,
            left_margin=left_margin,
            right_margin=right_margin,
            y_top=y,
            titulo="Trechos em Overbalance (Peso do Fluido > Limite Superior)",
            segs=segs_over,
            col_balanco_label="Overbalance máx."
        )

    sub_under = _subtrechos_por_falha_cmp(df_cmp_fase, segs_under, lado="inferior")
    sub_over = _subtrechos_por_falha_cmp(df_cmp_fase, segs_over, lado="superior")

    # Só desenha classificação se houver trechos
    if sub_under:
        y = desenhar_tabela_falhas_quebrada(
            c=c,
            left_margin=left_margin,
            right_margin=right_margin,
            y_top=y,
            titulo="Classificação do Tipo de Falha (Trechos em Underbalance)",
            subtrechos=sub_under
        )

    if sub_over:
        y = desenhar_tabela_falhas_quebrada(
            c=c,
            left_margin=left_margin,
            right_margin=right_margin,
            y_top=y,
            titulo="Classificação do Tipo de Falha (Trechos em Overbalance)",
            subtrechos=sub_over
        )

    # -------------------------------------------------
    # Interpretações SYGA
    # -------------------------------------------------
    interpretacoes_syga = _montar_interpretacoes_syga(sub_under, sub_over)

    if interpretacoes_syga:
        y = desenhar_tabela_interpretacoes_syga(
            c=c,
            left_margin=left_margin,
            right_margin=right_margin,
            y_top=y,
            interpretacoes=interpretacoes_syga
        )

    draw_footer(c, width, footer_y)
    c.showPage()


def _montar_df_cmp_global(df_mud, df_suav):
    if not isinstance(df_mud, pd.DataFrame) or df_mud.empty:
        return pd.DataFrame()

    if not isinstance(df_suav, pd.DataFrame) or df_suav.empty:
        return pd.DataFrame()

    df_mud2 = df_mud.copy()
    df_suav2 = df_suav.copy()

    if "Profundidade (m)" not in df_mud2.columns:
        for alt in ["Profundidade", "MD", "MD(m)"]:
            if alt in df_mud2.columns:
                df_mud2 = df_mud2.rename(columns={alt: "Profundidade (m)"})
                break

    col_exec = "Peso do Fluido Executado (lb/gal)"
    if "Profundidade (m)" not in df_mud2.columns or col_exec not in df_mud2.columns:
        return pd.DataFrame()

    df_mud2["Profundidade (m)"] = pd.to_numeric(df_mud2["Profundidade (m)"], errors="coerce")
    df_mud2[col_exec] = pd.to_numeric(df_mud2[col_exec], errors="coerce")

    df_mud2 = (
        df_mud2.dropna(subset=["Profundidade (m)", col_exec])
        .sort_values("Profundidade (m)")
        .reset_index(drop=True)
    )

    if df_mud2.empty:
        return pd.DataFrame()

    col_prof_suav = "Profundidade (m)" if "Profundidade (m)" in df_suav2.columns else "MD"
    obrig = [col_prof_suav, "Max Inferior", "Min Superior"]

    if col_prof_suav not in df_suav2.columns or any(c not in df_suav2.columns for c in obrig[1:]):
        return pd.DataFrame()

    cols_falha = [
        col_prof_suav,
        "Max Inferior",
        "Min Superior",
        "Gradiente de Pressão de Poros (lb/gal)",
        "Tração Inferior",
        "Comp Inferior σθA",
        "Comp Inferior σθB",
        "Tração Superior (σθA)",
        "Tração Superior (σθB)",
        "Comp Superior σθA",
        "Comp Superior σθB",
    ]
    cols_exist = [c for c in cols_falha if c in df_suav2.columns]
    df_suav2 = df_suav2[cols_exist].copy()

    for ccol in [col_prof_suav, "Max Inferior", "Min Superior"]:
        df_suav2[ccol] = pd.to_numeric(df_suav2[ccol], errors="coerce")

    df_suav2 = (
        df_suav2.dropna(subset=[col_prof_suav, "Max Inferior", "Min Superior"])
        .sort_values(col_prof_suav)
        .reset_index(drop=True)
    )

    if df_suav2.empty:
        return pd.DataFrame()

    prof_s = df_suav2[col_prof_suav].astype(float)

    rows_ref = []
    for d in df_mud2["Profundidade (m)"].astype(float).values:
        idx = (prof_s - float(d)).abs().idxmin()
        rows_ref.append(df_suav2.loc[idx].to_dict())

    df_cmp = df_mud2[["Profundidade (m)", col_exec]].copy()
    df_cmp = df_cmp.rename(columns={col_exec: "Peso do Fluido (lb/gal)"})

    df_cmp["Prof (ref df_suav)"] = [float(r.get(col_prof_suav, np.nan)) for r in rows_ref]
    df_cmp["Max Inferior"] = pd.to_numeric([r.get("Max Inferior", np.nan) for r in rows_ref], errors="coerce")
    df_cmp["Min Superior"] = pd.to_numeric([r.get("Min Superior", np.nan) for r in rows_ref], errors="coerce")

    extras = [
        "Gradiente de Pressão de Poros (lb/gal)",
        "Tração Inferior",
        "Comp Inferior σθA",
        "Comp Inferior σθB",
        "Tração Superior (σθA)",
        "Tração Superior (σθB)",
        "Comp Superior σθA",
        "Comp Superior σθB",
    ]
    for col in extras:
        df_cmp[col] = pd.to_numeric([r.get(col, np.nan) for r in rows_ref], errors="coerce")

    df_cmp["is_under"] = df_cmp["Peso do Fluido (lb/gal)"] < df_cmp["Max Inferior"]
    df_cmp["is_over"] = df_cmp["Peso do Fluido (lb/gal)"] > df_cmp["Min Superior"]

    return df_cmp


def _classificar_falha_row(row, lado, nd=6):
    if row is None:
        return "Não identificado"

    try:
        if lado == "inferior":
            lim = round(float(row["Max Inferior"]), nd)
            candidatos = [
                ("Gradiente de Pressão de Poros (lb/gal)", "Limitado pela pressão de poros"),
                ("Tração Inferior", "Falha por tração inf."),
                ("Comp Inferior σθA", "Falha por comp. inf. em σθA"),
                ("Comp Inferior σθB", "Falha por comp. inf. em σθB"),
            ]
        else:
            lim = round(float(row["Min Superior"]), nd)
            candidatos = [
                ("Tração Superior (σθA)", "Falha por tração sup. em σθA"),
                ("Tração Superior (σθB)", "Falha por tração sup. em σθB"),
                ("Comp Superior σθA", "Falha por comp. sup. em σθA"),
                ("Comp Superior σθB", "Falha por comp. sup. em σθB"),
            ]
    except Exception:
        return "Não identificado"

    for col, label in candidatos:
        try:
            v = row.get(col, np.nan)
            if pd.isna(v):
                continue
            if round(float(v), nd) == lim:
                return label
        except Exception:
            continue

    return "Não identificado"


def _subtrechos_por_falha_cmp(df_cmp_fase, segs, lado, tol_prof=1e-9):
    out = []
    if df_cmp_fase is None or df_cmp_fase.empty or not segs:
        return out

    prof_cmp = df_cmp_fase["Profundidade (m)"].astype(float)

    for s in segs:
        pi = float(s["pi"])
        pf = float(s["pf"])

        m = (prof_cmp >= (pi - tol_prof)) & (prof_cmp <= (pf + tol_prof))
        df_int = df_cmp_fase.loc[m].copy()

        if df_int.empty:
            continue

        df_int = df_int.sort_values("Profundidade (m)").reset_index(drop=True)
        df_int["tipo"] = df_int.apply(lambda r: _classificar_falha_row(r, lado=lado), axis=1)

        tipo_atual = None
        ini = None
        prof_prev = None

        for _, r in df_int.iterrows():
            p = float(r["Profundidade (m)"])
            t = str(r["tipo"])

            if tipo_atual is None:
                tipo_atual = t
                ini = p
            elif t != tipo_atual:
                out.append({"pi": ini, "pf": prof_prev, "tipo": tipo_atual})
                tipo_atual = t
                ini = p

            prof_prev = p

        if tipo_atual is not None and ini is not None and prof_prev is not None:
            out.append({"pi": ini, "pf": prof_prev, "tipo": tipo_atual})

    return out


def gerar_relatorio_pdf():
    hora_now = datetime.now() + timedelta(hours=0)
    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=letter)

    width, height = letter
    footer_y = 30

    # ==============================
    # Página 1
    # ==============================
    c.drawImage(logo, 230, height - 100, width=150, height=100)

    # --- Títulos principais ---
    c.setFont("Helvetica-Bold", 25)
    c.drawCentredString(width / 2, height - 260, "Syngular Geopressure Analysis - SYGA")

    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(width / 2, height - 295, "Relatório Final")

    # --- Informações do poço e responsável ---
    well_name = f"{st.session_state.poco}"
    user_name = f"{st.session_state.user_name}"

    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width / 2, height - 340, well_name)

    c.setFont("Helvetica", 15)
    c.drawCentredString(width / 2, height - 370, f"Responsável Técnico: {user_name}")

    # --- Rodapé ---
    c.setFont("Helvetica", 12)

    # Linha separadora
    c.line(30, 80, width - 30, 80)

    # Informações da empresa (lado esquerdo)
    c.drawString(40, 60, "Syngular Solutions")
    c.drawString(40, 45, "Houston, TX 77077")

    # Data e hora (lado direito)
    data_relatorio = datetime.today().strftime('%d/%m/%Y')
    hora_relatorio = hora_now.strftime('%H:%M')

    c.drawRightString(width - 40, 60, f"Data do Relatório: {data_relatorio} {hora_relatorio}")

    c.setFont("Helvetica-Bold", 18)
    c.drawRightString(75, 90,"V-1.0")

    c.showPage()

    footer_y = 20  # valor único para toda a página

    def draw_header(c, width, height, logo):
        logo_width = 150
        logo_height = 100
        logo_y = height - 90

        c.drawImage(
            logo,
            (width - logo_width) / 2,
            logo_y,
            width=logo_width,
            height=logo_height,
            preserveAspectRatio=True,
            mask='auto'
        )

        return logo_y - 10

    # ==============================
    # Página 2
    # ==============================
    y = draw_header(c, width, height, logo)

    left_margin = 40
    right_margin = width - 40

    # --- Título da seção ---
    c.setFont("Helvetica-Bold", 20)
    c.drawString(left_margin, y, "Dados do Poço")
    y -= 10
    c.line(left_margin, y, right_margin, y)

    # Espaço inicial
    y -= 14

    page_top = y
    page_bottom = footer_y + 20

    usable_h = page_top - page_bottom
    usable_w = right_margin - left_margin

    pad = 10
    mid_y = page_bottom + usable_h / 2

    def desenhar_caixa_padrao(titulo, x_left, x_right, y_top, y_bottom):
        box_x = x_left
        box_w = x_right - x_left
        box_y = y_bottom
        box_h = y_top - y_bottom

        if box_h < 40 or box_w < 80:
            return (x_left, x_right, y_top, y_bottom)

        # fundo
        c.setFillColorRGB(0.95, 0.95, 0.95)
        c.rect(box_x, box_y, box_w, box_h, fill=1, stroke=0)

        # borda
        c.setStrokeColorRGB(0, 0, 0)
        c.setLineWidth(0.8)
        c.rect(box_x, box_y, box_w, box_h, fill=0, stroke=1)

        # linha do título
        c.line(box_x, box_y + box_h - 18, box_x + box_w, box_y + box_h - 18)

        # título
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(box_x + box_w / 2, box_y + box_h - 14, titulo)

        # área útil
        inner_left = box_x + 10
        inner_right = box_x + box_w - 10
        inner_top = box_y + box_h - 24
        inner_bottom = box_y + 10

        return inner_left, inner_right, inner_top, inner_bottom

    box_sup_left = left_margin
    box_sup_right = right_margin
    box_sup_top = page_top
    box_sup_bottom = mid_y + 8

    sup_left, sup_right, sup_top, sup_bottom = desenhar_caixa_padrao(
        titulo="Referência de Elevação",
        x_left=box_sup_left,
        x_right=box_sup_right,
        y_top=box_sup_top,
        y_bottom=box_sup_bottom
    )

    if st.session_state.onshore:
        img_path = "rig.png"
    else:
        img_path = "rig_offshore.png"

    img_height = 240
    img_width = img_height * 0.65

    img_offset_x = 80
    img_offset_y = 0

    img_x = sup_left + img_offset_x
    img_y = sup_top - img_height + img_offset_y

    if img_y < sup_bottom:
        img_y = sup_bottom + 5

    c.drawImage(
        img_path,
        img_x,
        img_y,
        width=img_width,
        height=img_height,
        preserveAspectRatio=True,
        mask='auto'
    )

    texto_offset_x = 5
    texto_base_y = img_y + img_height
    texto_x = img_x + img_width + texto_offset_x

    pos_datum = -78
    pos_ele_datum = -108
    pos_airgap = -130
    pos_solo = -152
    pos_ele_solo = -190
    pos_nivel_mar = -228

    c.setFont("Helvetica-Bold", 12)

    c.drawString(
        texto_x,
        texto_base_y + pos_datum,
        f"DATUM: {st.session_state.datum}"
    )

    c.drawString(
        texto_x,
        texto_base_y + pos_airgap,
        f"Airgap: {st.session_state.rtkb:.2f} metros"
    )

    if st.session_state.onshore:
        c.drawString(
            texto_x,
            texto_base_y + pos_ele_datum,
            f"Elevação DATUM: {st.session_state.es:.2f} metros"
        )

        c.drawString(
            texto_x,
            texto_base_y + pos_solo,
            "Solo"
        )

        c.drawString(
            texto_x,
            texto_base_y + pos_ele_solo,
            f"Elevação do solo: {(st.session_state.es - st.session_state.rtkb):.2f} metros"
        )

        c.drawString(
            texto_x,
            texto_base_y + pos_nivel_mar,
            "Nível do mar"
        )
    else:
        pos_lamina = -160
        pos_leito = -230

        c.drawString(
            texto_x,
            texto_base_y + pos_lamina,
            f"Lâmina d'água: {st.session_state.lda:.2f} metros"
        )

        c.drawString(
            texto_x,
            texto_base_y + pos_leito,
            "Leito marinho"
        )

    box_map_left = left_margin
    box_map_right = right_margin
    box_map_top = mid_y - 10
    box_map_bottom = page_bottom

    nome_poco = st.session_state.get("poco", "Poço")

    map_left, map_right, map_top, map_bottom = desenhar_caixa_padrao(
        titulo=f'Localização do poço {nome_poco}',
        x_left=box_map_left,
        x_right=box_map_right,
        y_top=box_map_top,
        y_bottom=box_map_bottom
    )

    mapa_folium_pdf = st.session_state.get("mapa_folium_pdf", None)

    if mapa_folium_pdf is not None and map_top > map_bottom + 40:
        try:
            desenhar_mapa_folium_no_pdf(
                c=c,
                mapa_folium=mapa_folium_pdf,
                left=map_left,
                right=map_right,
                top=map_top,
                bottom=map_bottom
            )
        except Exception as e:
            c.setFont("Helvetica", 10)
            c.drawString(
                map_left,
                map_top - 15,
                f"Não foi possível inserir o mapa: {e}"
            )
    else:
        c.setFont("Helvetica", 10)
        c.drawString(
            map_left,
            map_top - 15,
            "Mapa: Não informado"
        )

    # Rodapé
    draw_footer(c, width, footer_y)
    c.showPage()

    # ==============================
    # Página 3
    # ==============================
    y = draw_header(c, width, height, logo)

    left_margin = 40
    right_margin = width - 40

    # --- Título da seção ---
    c.setFont("Helvetica-Bold", 20)
    c.drawString(left_margin, y, "Dados do Poço")
    y -= 10
    c.line(left_margin, y, right_margin, y)

    def _safe_str(v, default="Não informado"):
        if v is None or str(v).strip() == "":
            return default
        return str(v)

    def _fmt_coord(v, suf="", nd=2):
        try:
            if v is None or pd.isna(v):
                return "Não informado"
            return f"{float(v):.{nd}f}{suf}"
        except Exception:
            return "Não informado"

    def _fmt_float(v, suf="", nd=0):
        try:
            if v is None or pd.isna(v):
                return "—"
            return f"{float(v):.{nd}f}{suf}"
        except Exception:
            return "—"

    def desenhar_caixa_padrao(titulo, x_left, x_right, y_top, y_bottom):
        box_x = x_left
        box_w = x_right - x_left
        box_y = y_bottom
        box_h = y_top - y_bottom

        if box_h < 40 or box_w < 80:
            return (x_left, x_right, y_top, y_bottom)

        # fundo
        c.setFillColorRGB(0.95, 0.95, 0.95)
        c.rect(box_x, box_y, box_w, box_h, fill=1, stroke=0)

        # borda
        c.setStrokeColorRGB(0, 0, 0)
        c.setLineWidth(0.8)
        c.rect(box_x, box_y, box_w, box_h, fill=0, stroke=1)

        # linha do título
        c.line(box_x, box_y + box_h - 18, box_x + box_w, box_y + box_h - 18)

        # título centralizado
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(box_x + box_w / 2, box_y + box_h - 14, titulo)

        # área útil
        inner_left = box_x + 10
        inner_right = box_x + box_w - 10
        inner_top = box_y + box_h - 24
        inner_bottom = box_y + 10

        return inner_left, inner_right, inner_top, inner_bottom

    y -= 18

    comments = st.session_state.comments or ""
    comment_lines = wrap(comments, 95) if comments else ["Não informado"]

    info_linhas = [
        (
            "Usuário:", _safe_str(st.session_state.get("user_name", None)),
            "Companhia:", _safe_str(st.session_state.get("company_name", None))
        ),
        (
            "Poço:", _safe_str(st.session_state.get("poco", None)),
            "País:", _safe_str(st.session_state.get("country_name", None))
        ),
        (
            "Campo:", _safe_str(st.session_state.get("field_name", None)),
            "Zona UTM:", _safe_str(st.session_state.get("zona", None))
        ),
        (
            "Northing:", _fmt_coord(st.session_state.get("northing", None), " m", 2),
            "Easting:", _fmt_coord(st.session_state.get("easting", None), " m", 2)
        ),
        (
            "Hemisfério:", _safe_str(st.session_state.get("hem", None)),
            "", ""
        ),
    ]

    linha_altura = 16
    header_h = 28
    objetivo_h = 20 + len(comment_lines) * 14
    box_info_h = header_h + len(info_linhas) * linha_altura + 12 + objetivo_h + 8

    box_info_top = y
    box_info_bottom = y - box_info_h

    info_left, info_right, info_top, info_bottom = desenhar_caixa_padrao(
        titulo="Informações do Poço",
        x_left=left_margin,
        x_right=right_margin,
        y_top=box_info_top,
        y_bottom=box_info_bottom
    )

    col1_x = info_left + 2
    col2_x = info_left + (info_right - info_left) / 2 + 8
    texto_y = info_top - 8

    for label1, valor1, label2, valor2 in info_linhas:
        if label1:
            c.setFont("Helvetica-Bold", 10)
            c.drawString(col1_x, texto_y, label1)
            w1 = c.stringWidth(label1, "Helvetica-Bold", 10)

            c.setFont("Helvetica", 10)
            c.drawString(col1_x + w1 + 5, texto_y, str(valor1))

        if label2:
            c.setFont("Helvetica-Bold", 10)
            c.drawString(col2_x, texto_y, label2)
            w2 = c.stringWidth(label2, "Helvetica-Bold", 10)

            c.setFont("Helvetica", 10)
            txt_x = col2_x + w2 + 5
            c.drawString(txt_x, texto_y, str(valor2))

            if label2 == "País:":
                codigo_pais = paises.get(str(valor2))
                if codigo_pais:
                    flag_path = f"flag/{codigo_pais}.png"
                    try:
                        flag_x = txt_x + c.stringWidth(str(valor2), "Helvetica", 10) + 8
                        flag_y = texto_y - 2
                        c.drawImage(
                            flag_path,
                            flag_x,
                            flag_y,
                            width=18,
                            height=12,
                            preserveAspectRatio=True,
                            mask='auto'
                        )
                    except Exception:
                        pass

        texto_y -= linha_altura

    # separador
    c.line(info_left, texto_y + 4, info_right, texto_y + 4)

    # objetivo
    texto_y -= 12
    c.setFont("Helvetica-Bold", 10)
    c.drawString(col1_x, texto_y, "Objetivo:")
    texto_y -= 14

    c.setFont("Helvetica", 10)
    for line in comment_lines:
        c.drawString(col1_x, texto_y, line)
        texto_y -= 14

    pocos_dict = st.session_state.get("pocos", {})
    well_selected = st.session_state.get("well_selected", None)

    if well_selected not in pocos_dict and isinstance(pocos_dict, dict) and pocos_dict:
        well_selected = list(pocos_dict.keys())[0]

    poco_ativo = pocos_dict.get(well_selected, {}) if isinstance(pocos_dict, dict) else {}

    profs = poco_ativo.get("profundidade", [])
    fms = poco_ativo.get("formation", [])
    lits = poco_ativo.get("litologia", [])

    camadas = []
    if isinstance(profs, list) and isinstance(fms, list) and isinstance(lits, list):
        n = min(len(profs), len(fms), len(lits))
        for i in range(n):
            topo = profs[i]
            fm = fms[i]
            lit = lits[i]

            if (topo is None or (isinstance(topo, (int, float)) and float(topo) == 0.0)) and i != 0:
                continue

            camadas.append((topo, fm, lit))

    if not camadas:
        camadas = [(None, "Não informado", "—")]

    # usa TODAS as camadas
    camadas_visiveis = camadas

    y = box_info_bottom - 12

    linha_altura_lito = 11
    header_h_lito = 24
    margem_interna_lito = 8

    altura_tabela_lito = header_h_lito + (len(camadas_visiveis) + 1) * linha_altura_lito + margem_interna_lito

    box_lito_top = y
    box_lito_bottom = y - altura_tabela_lito

    lito_left, lito_right, lito_top, lito_bottom = desenhar_caixa_padrao(
        titulo="Litologia",
        x_left=left_margin,
        x_right=right_margin,
        y_top=box_lito_top,
        y_bottom=box_lito_bottom
    )

    x_topo = lito_left + 2
    x_fm = lito_left + (lito_right - lito_left) * 0.24
    x_lit = lito_left + (lito_right - lito_left) * 0.66

    yy = lito_top - 4
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x_topo, yy, "Topo (m)")
    c.drawString(x_fm, yy, "Formação")
    c.drawString(x_lit, yy, "Litologia")

    c.setLineWidth(0.5)
    c.line(lito_left, yy - 3, lito_right, yy - 3)

    yy -= 12
    c.setFont("Helvetica", 8.5)

    for topo, fm, lit in camadas_visiveis:
        c.drawString(x_topo, yy, _fmt_float(topo, " m", nd=2))
        c.drawString(x_fm, yy, (str(fm)[:32] if fm is not None else "—"))
        c.drawString(x_lit, yy, (str(lit)[:18] if lit is not None else "—"))
        yy -= linha_altura_lito

    fig_coluna = st.session_state.get("fig_coluna_lito", None)

    graph_top = box_lito_bottom - 8
    graph_bottom = footer_y + 18

    # só desenha se sobrar espaço suficiente
    if fig_coluna is not None and graph_top > graph_bottom + 60:
        desenhar_fig_plotly_no_pdf(
            c=c,
            fig_plotly=fig_coluna,
            left=left_margin,
            right=right_margin,
            top=graph_top,
            bottom=graph_bottom,
            titulo=None,
            scale=1.5
        )

    draw_footer(c, width, footer_y)
    c.showPage()

    # ==============================
    # Página 4 em diante - Análise por fase
    # ==============================
    if st.session_state.option == "Retroanálise":
        fases_df = st.session_state.get("fases_df", None)
        sapatas_df = st.session_state.get("sapatas_df", None)
        df_suav = st.session_state.get("df_suav", None)
        df_mud = st.session_state.get("df_mud", None)

        df_cmp_global = _montar_df_cmp_global(df_mud, df_suav)

        fase1 = fases_df.iloc[0]["Fase"] if isinstance(fases_df, pd.DataFrame) and len(fases_df) > 0 else None
        fase2 = fases_df.iloc[1]["Fase"] if isinstance(fases_df, pd.DataFrame) and len(fases_df) > 1 else None
        fase3 = fases_df.iloc[2]["Fase"] if isinstance(fases_df, pd.DataFrame) and len(fases_df) > 2 else None
        fase4 = fases_df.iloc[3]["Fase"] if isinstance(fases_df, pd.DataFrame) and len(fases_df) > 3 else None

        rev1 = sapatas_df.iloc[0]["Fase"] if isinstance(sapatas_df, pd.DataFrame) and len(sapatas_df) > 0 else None
        rev2 = sapatas_df.iloc[1]["Fase"] if isinstance(sapatas_df, pd.DataFrame) and len(sapatas_df) > 1 else None
        rev3 = sapatas_df.iloc[2]["Fase"] if isinstance(sapatas_df, pd.DataFrame) and len(sapatas_df) > 2 else None

        sap1 = sapatas_df.iloc[0]["Profundidade da sapata (m)"] if isinstance(sapatas_df, pd.DataFrame) and len(
            sapatas_df) > 0 else None
        sap2 = sapatas_df.iloc[1]["Profundidade da sapata (m)"] if isinstance(sapatas_df, pd.DataFrame) and len(
            sapatas_df) > 1 else None
        sap3 = sapatas_df.iloc[2]["Profundidade da sapata (m)"] if isinstance(sapatas_df, pd.DataFrame) and len(
            sapatas_df) > 2 else None

        td_final = None
        if isinstance(df_cmp_global, pd.DataFrame) and not df_cmp_global.empty:
            td_final = pd.to_numeric(df_cmp_global["Profundidade (m)"], errors="coerce").max()

        # -------- Fase 26" --------
        if fase1 is not None and sap1 is not None:
            _desenhar_pagina_fase(
                c, width, height, logo, footer_y,
                titulo=f'Análise da fase {_fmt_polegada(fase1)}, revestimento {_fmt_polegada(rev1)}',
                fase_broca=fase1,
                fase_revestimento=rev1,
                prof_ini=0.0,
                prof_fim=float(sap1),
                df_cmp_global=df_cmp_global,
                draw_header=draw_header,
                incluir_fim=False
            )

        # -------- Fase 17,5" --------
        if fase2 is not None and sap1 is not None and sap2 is not None:
            _desenhar_pagina_fase(
                c, width, height, logo, footer_y,
                titulo=f'Análise da fase {_fmt_polegada(fase2)}, revestimento {_fmt_polegada(rev2)}',
                fase_broca=fase2,
                fase_revestimento=rev2,
                prof_ini=float(sap1),
                prof_fim=float(sap2),
                df_cmp_global=df_cmp_global,
                draw_header=draw_header,
                incluir_fim=False
            )

        # -------- Fase 12,25" --------
        if fase3 is not None and sap2 is not None and sap3 is not None:
            _desenhar_pagina_fase(
                c, width, height, logo, footer_y,
                titulo=f'Análise da fase {_fmt_polegada(fase3)}, revestimento {_fmt_polegada(rev3)}',
                fase_broca=fase3,
                fase_revestimento=rev3,
                prof_ini=float(sap2),
                prof_fim=float(sap3),
                df_cmp_global=df_cmp_global,
                draw_header=draw_header,
                incluir_fim=False
            )

        # -------- Fase 8,5" --------
        if fase4 is not None and sap3 is not None and td_final is not None:
            _desenhar_pagina_fase(
                c, width, height, logo, footer_y,
                titulo=f'Análise da fase {_fmt_polegada(fase4)}, poço aberto',
                fase_broca=fase4,
                fase_revestimento=None,
                prof_ini=float(sap3),
                prof_fim=float(td_final),
                df_cmp_global=df_cmp_global,
                draw_header=draw_header,
                incluir_fim=True
            )

    # # ==============================
    # # Página 4
    # # ==============================
    #
    # pocos_dict = st.session_state.get("pocos", {})
    #
    # y = draw_header(c, width, height, logo)
    #
    # left_margin = 40
    # right_margin = width - 40
    #
    # c.setFillColorRGB(0, 0, 0)
    # c.setFont("Helvetica-Bold", 20)
    # c.drawString(left_margin, y, "Coluna litológica")
    # y -= 8
    # c.line(left_margin, y, right_margin, y)
    #
    # def _fmt_float(v, suf="", nd=0):
    #     try:
    #         if v is None or pd.isna(v):
    #             return "—"
    #         return f"{float(v):.{nd}f}{suf}"
    #     except Exception:
    #         return "—"
    #
    # well_selected = st.session_state.get("well_selected", None)
    # if well_selected not in pocos_dict and isinstance(pocos_dict, dict) and pocos_dict:
    #     well_selected = list(pocos_dict.keys())[0]
    #
    # poco_ativo = pocos_dict.get(well_selected, {}) if isinstance(pocos_dict, dict) else {}
    #
    # profs = poco_ativo.get("profundidade", [])
    # fms = poco_ativo.get("formation", [])
    # lits = poco_ativo.get("litologia", [])
    #
    # fig_coluna = st.session_state.get("fig_coluna_lito", None)
    #
    # box_width = right_margin - left_margin
    # box_x = left_margin
    #
    # y -= 14
    #
    # camadas = []
    # if isinstance(profs, list) and isinstance(fms, list) and isinstance(lits, list):
    #     n = min(len(profs), len(fms), len(lits))
    #     for i in range(n):
    #         topo = profs[i]
    #         fm = fms[i]
    #         lit = lits[i]
    #
    #         if (topo is None or (isinstance(topo, (int, float)) and float(topo) == 0.0)) and i != 0:
    #             continue
    #
    #         camadas.append((topo, fm, lit))
    #
    # if not camadas:
    #     camadas = [(None, "Não informado", "—")]
    #
    # linha_altura2 = 14
    # header_h = 28
    # box_height2 = header_h + (len(camadas) + 1) * linha_altura2
    #
    # box2_y = y - box_height2
    #
    # c.setFillColorRGB(0.95, 0.95, 0.95)
    # c.rect(box_x, box2_y, box_width, box_height2, fill=1, stroke=0)
    #
    # c.setStrokeColorRGB(0, 0, 0)
    # c.setLineWidth(0.8)
    # c.rect(box_x, box2_y, box_width, box_height2, fill=0, stroke=1)
    #
    # c.line(box_x, box2_y + box_height2 - 18, box_x + box_width, box2_y + box_height2 - 18)
    #
    # c.setFillColorRGB(0, 0, 0)
    # c.setFont("Helvetica-Bold", 12)
    # c.drawString(box_x + 10, box2_y + box_height2 - 14, "Camadas litológicas do poço ativo")
    #
    # x_topo = box_x + 12
    # x_fm = box_x + box_width * 0.30
    # x_lit = box_x + box_width * 0.70
    #
    # yy = box2_y + box_height2 - 34
    # c.setFont("Helvetica-Bold", 10)
    # c.drawString(x_topo, yy, "Topo (m)")
    # c.drawString(x_fm, yy, "Formação")
    # c.drawString(x_lit, yy, "Litologia")
    #
    # c.setLineWidth(0.5)
    # c.line(box_x + 10, yy - 4, box_x + box_width - 10, yy - 4)
    #
    # yy -= 14
    # c.setFont("Helvetica", 10)
    #
    # for topo, fm, lit in camadas:
    #     c.drawString(x_topo, yy, _fmt_float(topo, " m", nd=2))
    #     c.drawString(x_fm, yy, (str(fm)[:38] if fm is not None else "—"))
    #     c.drawString(x_lit, yy, (str(lit)[:18] if lit is not None else "—"))
    #     yy -= linha_altura2
    #
    # y = box2_y - 12
    #
    # if fig_coluna is not None and y > footer_y + 80:
    #     y = desenhar_fig_plotly_no_pdf(
    #         c=c,
    #         fig_plotly=fig_coluna,
    #         left=left_margin,
    #         right=right_margin,
    #         top=y,
    #         bottom=footer_y + 20,
    #         titulo=None,
    #         scale=3
    #     )
    #
    # draw_footer(c, width, footer_y)
    # c.showPage()

    # ==============================
    # Página 5
    # ==============================

    y = draw_header(c, width, height, logo)

    # Título da página
    c.setFont("Helvetica-Bold", 20)
    c.drawString(left_margin, y, "Gradiente de Sobrecarga")
    y -= 10
    c.line(left_margin, y, right_margin, y)

    gard_sel = st.session_state.get("gard", [])
    ex_sel = st.session_state.get("ex", "")
    ds_val = st.session_state.get("ds", None)
    rtkb_val = st.session_state.get("rtkb", None)
    es_val = st.session_state.get("es", None)
    lda_val = st.session_state.get("lda", None)
    onshore = st.session_state.get("onshore", False)

    if isinstance(gard_sel, list):
        gard_sel = ", ".join(gard_sel) if gard_sel else "Nenhum"

    if not ex_sel:
        ex_sel = "Não informado"

    linhas_esq = [
        ("Densidade:", gard_sel),
        ("Extrapolação:", ex_sel),
    ]

    linhas_dir = []

    if ex_sel == "Ativada":
        ds_txt = f"{ds_val:.3f} g/cm³" if ds_val is not None else "Não informado"
        linhas_dir.append(("Densidade média camadas sup.:", ds_txt))

    airgap_txt = f"{rtkb_val:.2f} m" if rtkb_val is not None else "Não informado"
    linhas_dir.append(("Air Gap:", airgap_txt))

    if onshore:
        nf_txt = f"{es_val:.2f} m" if es_val is not None else "Não informado"
        linhas_dir.append(("Elevação DATUM:", nf_txt))
    else:
        lda_txt = f"{lda_val:.2f} m" if lda_val is not None else "Não informado"
        linhas_dir.append(("Lâmina d'água:", lda_txt))

    num_linhas = max(len(linhas_esq), len(linhas_dir))
    linha_altura = 16
    box_height = 40 + num_linhas * linha_altura

    box_width = right_margin - left_margin
    box_x = left_margin

    y -= 25
    box_y = y - box_height

    c.setFillColorRGB(0.95, 0.95, 0.95)
    c.rect(box_x, box_y, box_width, box_height, fill=1, stroke=0)

    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.8)
    c.rect(box_x, box_y, box_width, box_height, fill=0, stroke=1)

    c.line(box_x, box_y + box_height - 22, box_x + box_width, box_y + box_height - 22)

    c.setFillColorRGB(0, 0, 0)

    # título da caixa
    c.setFont("Helvetica-Bold", 12)
    c.drawString(box_x + 10, box_y + box_height - 16, "Parâmetros Utilizados no Cálculo")

    col1_x = box_x + 12
    col2_x = box_x + box_width / 2
    texto_y = box_y + box_height - 36

    for i in range(num_linhas):

        if i < len(linhas_esq):
            label, valor = linhas_esq[i]

            c.setFont("Helvetica-Bold", 10)
            c.drawString(col1_x, texto_y, label)

            lw = c.stringWidth(label, "Helvetica-Bold", 10)

            c.setFont("Helvetica", 10)
            c.drawString(col1_x + lw + 4, texto_y, str(valor))

        if i < len(linhas_dir):
            label, valor = linhas_dir[i]

            c.setFont("Helvetica-Bold", 10)
            c.drawString(col2_x, texto_y, label)

            lw = c.stringWidth(label, "Helvetica-Bold", 10)

            c.setFont("Helvetica", 10)
            c.drawString(col2_x + lw + 4, texto_y, str(valor))

        texto_y -= linha_altura

    y = box_y - 20

    fig = st.session_state.fig_gs

    if fig is not None:
        y = desenhar_fig_mpl_no_pdf(
            c=c,
            fig=fig,
            left=left_margin,
            right=right_margin,
            top=y,
            bottom=footer_y + 20,
            titulo=None,
            dpi=200
        )

    # Rodapé
    draw_footer(c, width, footer_y)
    c.showPage()

    # ==============================
    # Página 6
    # ==============================
    y = draw_header(c, width, height, logo)

    left_margin = 40
    right_margin = width - 40

    # --- Título da seção ---
    c.setFont("Helvetica-Bold", 20)
    c.drawString(left_margin, y, "Gradiente de Pressão de Poros")

    y -= 10
    c.line(left_margin, y, right_margin, y)

    expoente = st.session_state.get("expoente", None)
    anormal = st.session_state.get("anormal", None)
    gn = st.session_state.get("gn", None)

    # suavizações
    spp = st.session_state.get("spp", False)
    s_gr = st.session_state.get("s_gr", False)
    suav_s = st.session_state.get("suav_s", False)

    # boyance / trending
    boyance_ativo = st.session_state.get("boyance", "Não")
    opcao_boyance = st.session_state.get("o_boyance", None)

    n_boyance = st.session_state.get("n_boyance", 0)
    n_trending = st.session_state.get("n_trending", 0)

    linhas_esq = []
    linhas_esq.append(("Expoente de Eaton:", f"{expoente}" if expoente else "Não informado"))
    linhas_esq.append(("Prof. zona anormal:", f"{anormal:.2f} m" if anormal else "Não informado"))
    linhas_esq.append(("Gradiente normal:", f"{gn:.2f} lb/gal" if gn else "Não informado"))
    linhas_esq.append(("Pressão de poros suavizada:", "Sim" if spp else "Não"))
    linhas_esq.append(("Raio gama suavizado:", "Sim" if s_gr else "Não"))
    linhas_esq.append(("Sônico suavizado:", "Sim" if suav_s else "Não"))

    linhas_dir = []

    # --- Boyance ---
    if boyance_ativo == "Sim":
        linhas_dir.append(("Boyance aplicado:", "Sim"))
        if opcao_boyance:
            # garante que é lista
            if isinstance(opcao_boyance, (list, tuple)) and len(opcao_boyance) > 1:
                # primeira linha com o label
                linhas_dir.append(("Opção:", str(opcao_boyance[0])))
                # demais linhas sem repetir o label (pra não estourar a largura)
                for extra in opcao_boyance[1:]:
                    linhas_dir.append(("", str(extra)))
            else:
                # caso venha 1 opção só (ou string)
                if isinstance(opcao_boyance, (list, tuple)):
                    linhas_dir.append(("Opção:", str(opcao_boyance[0]) if opcao_boyance else ""))
                else:
                    linhas_dir.append(("Opção:", str(opcao_boyance)))

        for i in range(n_boyance):
            prof_ini = st.session_state.get(f"prof_inicial_{i}", None)
            prof_fim = st.session_state.get(f"prof_final_{i}", None)
            fpr = st.session_state.get(f"fpr_{i}", None)

            texto = ""
            if prof_ini is not None and prof_fim is not None:
                texto += f"{prof_ini:.0f}–{prof_fim:.0f} m  "
            if fpr is not None:
                texto += f"Peso do fluido ={fpr} lb/gal"
            if texto:
                linhas_dir.append((f"Boyance {i + 1}:", texto))
    else:
        linhas_dir.append(("Boyance aplicado:", "Não"))

    # --- Trending (parâmetros) ---
    for i in range(n_trending):
        pp1 = st.session_state.get(f"pp1_{i}", None)
        pp2 = st.session_state.get(f"pp2_{i}", None)
        s1 = st.session_state.get(f"s1_{i}", None)
        s2 = st.session_state.get(f"s2_{i}", None)
        if pp1 and pp2 and s1 and s2:
            texto = f"{pp1:.0f}–{pp2:.0f} m  |  S1={s1}  S2={s2}"
            linhas_dir.append((f"Trending {i + 1}:", texto))

    # --- LBF (parâmetros) ---
    for i in range(n_trending):
        lbf = st.session_state.get(f"lbf_{i}", None)
        inclbf = st.session_state.get(f"inclbf_{i}", None)
        if lbf is not None and inclbf is not None:
            texto = f"Início={lbf}  Inclinação={inclbf}"
            linhas_dir.append((f"LBF {i + 1}:", texto))

    # ---- BOX ----
    num_linhas = max(len(linhas_esq), len(linhas_dir))
    linha_altura = 16
    box_height = 40 + num_linhas * linha_altura

    box_width = right_margin - left_margin
    box_x = left_margin

    y -= 25
    box_y = y - box_height

    c.setFillColorRGB(0.95, 0.95, 0.95)
    c.rect(box_x, box_y, box_width, box_height, fill=1, stroke=0)

    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.8)
    c.rect(box_x, box_y, box_width, box_height, fill=0, stroke=1)

    c.line(box_x, box_y + box_height - 22, box_x + box_width, box_y + box_height - 22)

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(box_x + 10, box_y + box_height - 16, "Dados utilizados no cálculo")

    col1_x = box_x + 12
    col2_x = box_x + box_width / 2
    texto_y = box_y + box_height - 36

    for i in range(num_linhas):
        if i < len(linhas_esq):
            label, valor = linhas_esq[i]
            c.setFont("Helvetica-Bold", 10)
            c.drawString(col1_x, texto_y, label)
            lw = c.stringWidth(label, "Helvetica-Bold", 10)
            c.setFont("Helvetica", 10)
            c.drawString(col1_x + lw + 4, texto_y, str(valor))

        if i < len(linhas_dir):
            label, valor = linhas_dir[i]
            c.setFont("Helvetica-Bold", 10)
            c.drawString(col2_x, texto_y, label)
            lw = c.stringWidth(label, "Helvetica-Bold", 10)
            c.setFont("Helvetica", 10)
            c.drawString(col2_x + lw + 4, texto_y, str(valor))

        texto_y -= linha_altura

    # ---- SUBTÍTULO LBF + LINHA (SEM NEGRITO) ----
    y = box_y - 18

    c.setFont("Helvetica", 14)  # sem negrito
    c.drawString(left_margin, y, "Linha Base de Folhelhos - LBF")
    y -= 8
    c.line(left_margin, y, right_margin, y)

    # espaço antes do gráfico
    y -= 12

    # ---- GRÁFICO (LBF = fig2) ----
    fig_lbf = st.session_state.get("fig2", None)
    if fig_lbf is not None and y > footer_y + 80:
        y = desenhar_fig_mpl_no_pdf(
            c=c,
            fig=fig_lbf,
            left=left_margin,
            right=right_margin,
            top=y,
            bottom=footer_y + 20,
            titulo=None
        )

    # Rodapé Página 4
    draw_footer(c, width, footer_y)
    c.showPage()

    # ==============================
    # Página 7
    # ==============================
    y = draw_header(c, width, height, logo)

    left_margin = 40
    right_margin = width - 40

    # --- Título principal ---
    c.setFont("Helvetica-Bold", 20)
    c.drawString(left_margin, y, "Gradiente de Pressão de Poros")

    y -= 10
    c.line(left_margin, y, right_margin, y)

    y -= 18
    c.setFont("Helvetica", 14)  # sem negrito
    c.drawString(left_margin, y, "Trending")
    y -= 8
    c.line(left_margin, y, right_margin, y)
    y -= 12

    fig_trending = st.session_state.get("fig1", None)
    if fig_trending is not None and y > footer_y + 80:
        y = desenhar_fig_mpl_no_pdf(
            c=c,
            fig=fig_trending,
            left=left_margin,
            right=right_margin,
            top=y,
            bottom=footer_y + 20,
            titulo=None
        )

    y -= 10

    min_space_for_next_plot = 220  # altura mínima segura
    if y < footer_y + min_space_for_next_plot:
        draw_footer(c, width, footer_y)
        c.showPage()

        y = draw_header(c, width, height, logo)
        left_margin = 40
        right_margin = width - 40

        c.setFont("Helvetica-Bold", 20)
        c.drawString(left_margin, y, "Gradiente de Pressão de Poros")
        y -= 10
        c.line(left_margin, y, right_margin, y)
        y -= 18

    c.setFont("Helvetica", 14)  # sem negrito
    c.drawString(left_margin, y, "Gradiente de Pressão de Poros")
    y -= 8
    c.line(left_margin, y, right_margin, y)
    y -= 12

    fig_pp = st.session_state.get("fig_pp", None)
    if fig_pp is not None and y > footer_y + 80:
        y = desenhar_fig_mpl_no_pdf(
            c=c,
            fig=fig_pp,
            left=left_margin,
            right=right_margin,
            top=y,
            bottom=footer_y + 20,
            titulo=None
        )

    draw_footer(c, width, footer_y)
    c.showPage()

    # ==============================
    # Página 8
    # ==============================
    y = draw_header(c, width, height, logo)

    left_margin = 40
    right_margin = width - 40

    # --- Título da seção ---
    c.setFont("Helvetica-Bold", 20)
    c.drawString(left_margin, y, "Estabilidade de Poço – Tensões em volta do poço")

    y -= 10
    c.line(left_margin, y, right_margin, y)

    phi = st.session_state.get("phi", None)  # (°)
    lft = st.session_state.get("lft", None)  # (psi)
    m_input = st.session_state.get("m", None)  # profundidade digitada
    ucs = st.session_state.get("ucs", None)  # método UCS

    sapatas_df = st.session_state.get("sapatas_df", None)
    zona2 = st.session_state.get("zona2", None)
    prisoes_df = st.session_state.get("prisoes_coluna_df", None)
    dados_lito = st.session_state.get("dados_lito", None)

    profundidade_proxima = None
    try:
        y_series = st.session_state.get("y", None)
        if y_series is not None and m_input is not None:
            profundidade_proxima = y_series.loc[(y_series - m_input).abs().idxmin()]
    except Exception:
        profundidade_proxima = None

    linhas_esq = []
    linhas_dir = []

    # --- ESQUERDA: parâmetros principais ---
    linhas_esq.append(("Ângulo de fricção:", f"{phi}°" if phi is not None else "Não informado"))
    linhas_esq.append(("Limite falha por tração:", f"{lft} psi" if lft is not None else "Não informado"))
    linhas_esq.append((
        "Profundidade analisada:",
        f"{profundidade_proxima:.2f} m" if isinstance(profundidade_proxima,
                                                      (int, float)) else "Não informado"
    ))
    linhas_esq.append(("Correlação para cálculo do UCS:", str(ucs) if ucs else "Não informado"))

    # --- ESQUERDA: Pontos de prisão de coluna ---
    if prisoes_df is not None and hasattr(prisoes_df, "empty") and not prisoes_df.empty:
        linhas_esq.append(("Pontos de prisão de coluna (TVD):", ""))

        col_prof_pr = "Profundidade da prisão (m)"
        try:
            pr_ord = prisoes_df.sort_values(by=col_prof_pr, ascending=True)
        except Exception:
            pr_ord = prisoes_df

        for _, row in pr_ord.iterrows():
            prof_p = row.get(col_prof_pr, "—")
            linhas_esq.append(("", f"{prof_p} m"))
    else:
        linhas_esq.append(("Pontos de prisão de coluna (TVD):", "Não informado"))

    # --- ESQUERDA: Litologias consideradas + lista topo/base/lito/So ---
    considera_lito = (
            dados_lito is not None
            and hasattr(dados_lito, "empty")
            and not dados_lito.empty
    )

    linhas_esq.append(("Litologias consideradas:", "Sim" if considera_lito else "Não"))

    if considera_lito:
        # sem linha em branco para reduzir o "buraco"
        linhas_esq.append(("", "Topo–Base | Litologia | So (psi)"))

        try:
            dados_ord = dados_lito.sort_values(by="Topo (m)", ascending=True)
        except Exception:
            dados_ord = dados_lito

        for _, row in dados_ord.iterrows():
            topo = row.get("Topo (m)", "—")
            base = row.get("Base (m)", "—")
            lit = row.get("Litologia", "—")
            so = row.get("So (psi)", "—")
            linhas_esq.append(("", f"{topo}–{base} m | {lit} | {so} psi"))

    # --- DIREITA: Sapatas ---
    if sapatas_df is not None and hasattr(sapatas_df, "empty") and not sapatas_df.empty:
        linhas_dir.append(("Profundidade das sapatas (TVD):", ""))

        col_fase = "Fase"
        col_prof = "Profundidade da sapata (m)"

        try:
            sapatas_ord = sapatas_df.sort_values(by=col_prof, ascending=True)
        except Exception:
            sapatas_ord = sapatas_df

        for _, row in sapatas_ord.iterrows():
            fase = row.get(col_fase, "—")
            prof = row.get(col_prof, "—")
            linhas_dir.append(("", f"Sapata {fase}: {prof} m"))
    else:
        linhas_dir.append(("Profundidade das sapatas (TVD):", "Não informado"))

    # --- DIREITA: Zonas de perda ---
    if zona2 and isinstance(zona2, list) and len(zona2) > 0:
        linhas_dir.append(("Zonas de perda:", ""))

        try:
            zona2_ord = sorted(zona2, key=lambda x: float(x[0]))
        except Exception:
            zona2_ord = zona2

        for z in zona2_ord:
            try:
                prof_z = z[0]
                ppg_z = z[1]
                linhas_dir.append(("", f"{prof_z} m | {ppg_z} lb/gal"))
            except Exception:
                linhas_dir.append(("", str(z)))
    else:
        linhas_dir.append(("Zonas de perda:", "Não informado"))

    num_linhas = max(len(linhas_esq), len(linhas_dir))
    linha_altura = 16
    box_height = 40 + num_linhas * linha_altura

    box_width = right_margin - left_margin
    box_x = left_margin

    y -= 25
    box_y = y - box_height

    c.setFillColorRGB(0.95, 0.95, 0.95)
    c.rect(box_x, box_y, box_width, box_height, fill=1, stroke=0)

    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.8)
    c.rect(box_x, box_y, box_width, box_height, fill=0, stroke=1)

    c.line(box_x, box_y + box_height - 22, box_x + box_width, box_y + box_height - 22)

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(box_x + 10, box_y + box_height - 16, "Dados utilizados no cálculo")

    col1_x = box_x + 12
    col2_x = box_x + box_width / 2
    texto_y = box_y + box_height - 36

    for i in range(num_linhas):
        # --- coluna esquerda ---
        if i < len(linhas_esq):
            label, valor = linhas_esq[i]
            if str(label).strip():
                c.setFont("Helvetica-Bold", 10)
                c.drawString(col1_x, texto_y, str(label))
                lw = c.stringWidth(str(label), "Helvetica-Bold", 10)
                c.setFont("Helvetica", 10)
                c.drawString(col1_x + lw + 4, texto_y, str(valor))
            else:
                c.setFont("Helvetica", 10)
                c.drawString(col1_x + 12, texto_y, str(valor))

        # --- coluna direita ---
        if i < len(linhas_dir):
            label, valor = linhas_dir[i]
            if str(label).strip():
                c.setFont("Helvetica-Bold", 10)
                c.drawString(col2_x, texto_y, str(label))
                lw = c.stringWidth(str(label), "Helvetica-Bold", 10)
                c.setFont("Helvetica", 10)
                c.drawString(col2_x + lw + 4, texto_y, str(valor))
            else:
                c.setFont("Helvetica", 10)
                c.drawString(col2_x + 12, texto_y, str(valor))

        texto_y -= linha_altura

    y = box_y - 18

    c.setFont("Helvetica", 14)  # sem negrito
    c.drawString(left_margin, y, "Trajetória do poço")
    y -= 8
    c.line(left_margin, y, right_margin, y)
    y -= 12

    fig_traj = st.session_state.get("fig_traj", None)
    if fig_traj is not None and y > footer_y + 60:
        try:
            y = desenhar_fig_plotly_no_pdf(
                c=c,
                fig_plotly=fig_traj,
                left=left_margin,
                right=right_margin,
                top=y,
                bottom=footer_y + 20,
                titulo=None,
                scale=1.5
            )
        except Exception as e:
            c.setFont("Helvetica", 10)
            c.drawString(left_margin, y - 10, f"Não foi possível exportar a trajetória para imagem: {e}")
    else:
        c.setFont("Helvetica", 10)
        c.drawString(left_margin, y - 10, "Trajetória do poço: Não informado")

    # Rodapé
    draw_footer(c, width, footer_y)
    c.showPage()

    # ==============================
    # Página 9
    # ==============================
    y = draw_header(c, width, height, logo)

    left_margin = 40
    right_margin = width - 40

    # --- Título da seção ---
    c.setFont("Helvetica-Bold", 20)
    c.drawString(left_margin, y, "Janela operacional")

    y -= 10
    c.line(left_margin, y, right_margin, y)

    # espaço antes do gráfico
    y -= 20

    # --- Gráfico (Janela Operacional) ---
    fig_jo = st.session_state.get("fig_jo", None)
    if fig_jo is not None and y > footer_y + 80:
        y = desenhar_fig_mpl_no_pdf(
            c=c,
            fig=fig_jo,
            left=left_margin,
            right=right_margin,
            top=y,
            bottom=footer_y + 20,
            titulo=None,
            dpi=200
        )
    else:
        c.setFont("Helvetica", 10)
        c.drawString(left_margin, y - 10, "Gráfico da janela operacional: Não informado")

    # Rodapé Página 7
    draw_footer(c, width, footer_y)
    c.showPage()

    # ==============================
    # Página 10
    # ==============================
    y = draw_header(c, width, height, logo)

    left_margin = 40
    right_margin = width - 40

    # --- Título da seção ---
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 20)
    c.drawString(left_margin, y, "Critério de falha de Mohr Coulomb")

    y -= 8
    c.line(left_margin, y, right_margin, y)

    def _fmt(v, suf=""):
        return f"{float(v):.2f}{suf}" if v is not None and pd.notna(v) else "Não informado"

    def _aprox(a, b, tol=1e-3):
        if a is None or b is None or (pd.isna(a) or pd.isna(b)):
            return False
        return abs(float(a) - float(b)) <= tol

    # >>>>>>>>>>>>>> TUDO AGORA USA df_suav <<<<<<<<<<<<<<
    df_suav = st.session_state.get("df_suav", None)

    ppg_val = st.session_state.get("ppg", None)
    ppg_txt = f"{ppg_val:.2f} lb/gal" if ppg_val is not None else "Não informado"

    prof_analisada = st.session_state.get("profundidade_proxima", None)

    if isinstance(prof_analisada, pd.Series):
        if not prof_analisada.empty:
            prof_analisada = float(prof_analisada.iloc[0])
        else:
            prof_analisada = None

    if prof_analisada is None:
        y_val = st.session_state.get("y", None)
        if isinstance(y_val, pd.Series):
            prof_analisada = float(y_val.iloc[0]) if not y_val.empty else None
        else:
            prof_analisada = y_val

    prof_analisada_txt = (
        f"{float(prof_analisada):.2f} m"
        if prof_analisada is not None and pd.notna(prof_analisada)
        else "Não informado"
    )

    # coluna de profundidade usada para localizar a linha analisada
    coluna_ref = "Profundidade (m)"
    if st.session_state.get("t_prof", "TVD") != "TVD":
        coluna_ref = "MD"  # se df_suav tiver MD

    # linha da profundidade analisada (para janela/status/falha)
    row_s = None
    if (
            isinstance(df_suav, pd.DataFrame)
            and not df_suav.empty
            and prof_analisada is not None
            and coluna_ref in df_suav.columns
    ):
        idx = (df_suav[coluna_ref].astype(float) - float(prof_analisada)).abs().idxmin()
        row_s = df_suav.loc[idx]

    prof_crit_txt = "Não informado"
    pp_txt = "Não informado"
    tr_inf_txt = "Não informado"
    comp_inf_a_txt = "Não informado"
    comp_inf_b_txt = "Não informado"

    if isinstance(df_suav, pd.DataFrame) and not df_suav.empty:
        col_max = "Max Inferior"
        col_prof = "Profundidade (m)" if "Profundidade (m)" in df_suav.columns else coluna_ref

        cols_needed_inf = [
            col_max,
            col_prof,
            "Gradiente de Pressão de Poros (lb/gal)",
            "Tração Inferior",
            "Comp Inferior σθA",
            "Comp Inferior σθB",
        ]

        if all(cn in df_suav.columns for cn in cols_needed_inf):
            s_max = df_suav[col_max]
            if s_max.notna().any():
                idx_crit = s_max.idxmax()
                linha_crit = df_suav.loc[idx_crit]

                prof_crit_txt = _fmt(linha_crit[col_prof], " m")
                pp_txt = _fmt(linha_crit["Gradiente de Pressão de Poros (lb/gal)"], " lb/gal")
                tr_inf_txt = _fmt(linha_crit["Tração Inferior"], " lb/gal")
                comp_inf_a_txt = _fmt(linha_crit["Comp Inferior σθA"], " lb/gal")
                comp_inf_b_txt = _fmt(linha_crit["Comp Inferior σθB"], " lb/gal")

    linhas_esq_1 = [
        ("Profundidade crítica:", prof_crit_txt),
        ("Peso do fluido atual:", ppg_txt),
        ("Pressão de poros:", pp_txt),
    ]
    linhas_dir_1 = [
        ("Limite de falha por tração inferior:", tr_inf_txt),
        ("Limite de falha compressão inferior (σθA):", comp_inf_a_txt),
        ("Limite de falha compressão inferior (σθB):", comp_inf_b_txt),
    ]

    num_linhas_1 = max(len(linhas_esq_1), len(linhas_dir_1))
    linha_altura_1 = 15
    box_height_1 = 30 + num_linhas_1 * linha_altura_1

    box_width = right_margin - left_margin
    box_x = left_margin

    y -= 14
    box1_y = y - box_height_1

    # fundo
    c.setFillColorRGB(0.95, 0.95, 0.95)
    c.rect(box_x, box1_y, box_width, box_height_1, fill=1, stroke=0)
    # borda
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.8)
    c.rect(box_x, box1_y, box_width, box_height_1, fill=0, stroke=1)
    # linha do título interno
    c.line(box_x, box1_y + box_height_1 - 18, box_x + box_width, box1_y + box_height_1 - 18)

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(box_x + 10, box1_y + box_height_1 - 14, "Resumo do limite inferior da janela operacional")

    col1_x = box_x + 12
    col2_x = box_x + box_width / 2
    texto_y = box1_y + box_height_1 - 32

    for i in range(num_linhas_1):
        if i < len(linhas_esq_1):
            label, valor = linhas_esq_1[i]
            c.setFont("Helvetica-Bold", 10)
            c.drawString(col1_x, texto_y, label)
            lw = c.stringWidth(label, "Helvetica-Bold", 10)
            c.setFont("Helvetica", 10)
            c.drawString(col1_x + lw + 4, texto_y, str(valor))

        if i < len(linhas_dir_1):
            label, valor = linhas_dir_1[i]
            c.setFont("Helvetica-Bold", 10)
            c.drawString(col2_x, texto_y, label)
            lw = c.stringWidth(label, "Helvetica-Bold", 10)
            c.setFont("Helvetica", 10)
            c.drawString(col2_x + lw + 4, texto_y, str(valor))

        texto_y -= linha_altura_1

    # cursor abaixo da tabela 1
    y = box1_y - 18

    prof_crit2_txt = "Não informado"
    gs_txt = "Não informado"
    ts_a_txt = "Não informado"
    ts_b_txt = "Não informado"
    cs_a_txt = "Não informado"
    cs_b_txt = "Não informado"

    if isinstance(df_suav, pd.DataFrame) and not df_suav.empty:
        col_min = "Min Superior"
        col_prof = "Profundidade (m)" if "Profundidade (m)" in df_suav.columns else coluna_ref

        required_cols = [
            col_min,
            col_prof,
            "Gradiente de Sobrecarga (lb/gal)",
            "Tração Superior (σθA)",
            "Tração Superior (σθB)",
            "Comp Superior σθA",
            "Comp Superior σθB",
        ]

        if all(cn in df_suav.columns for cn in required_cols):
            serie_min = df_suav[col_min]
            if serie_min.notna().any():
                idx_crit2 = serie_min.idxmin()
                linha = df_suav.loc[idx_crit2]

                prof_crit2_txt = _fmt(linha[col_prof], " m")
                gs_txt = _fmt(linha["Gradiente de Sobrecarga (lb/gal)"], " lb/gal")
                ts_a_txt = _fmt(linha["Tração Superior (σθA)"], " lb/gal")
                ts_b_txt = _fmt(linha["Tração Superior (σθB)"], " lb/gal")
                cs_a_txt = _fmt(linha["Comp Superior σθA"], " lb/gal")
                cs_b_txt = _fmt(linha["Comp Superior σθB"], " lb/gal")

    linhas_esq_2 = [
        ("Profundidade crítica:", prof_crit2_txt),
        ("Peso do fluido atual:", ppg_txt),
        ("Gradiente de sobrecarga:", gs_txt),
    ]
    linhas_dir_2 = [
        ("Falha por tração superior (σθA):", ts_a_txt),
        ("Falha por tração superior (σθB):", ts_b_txt),
        ("Falha por compressão superior (σθA):", cs_a_txt),
        ("Falha por compressão superior (σθB):", cs_b_txt),
    ]

    num_linhas_2 = max(len(linhas_esq_2), len(linhas_dir_2))
    linha_altura_2 = 15
    box_height_2 = 30 + num_linhas_2 * linha_altura_2

    box2_y = y - box_height_2

    # fundo
    c.setFillColorRGB(0.95, 0.95, 0.95)
    c.rect(box_x, box2_y, box_width, box_height_2, fill=1, stroke=0)
    # borda
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.8)
    c.rect(box_x, box2_y, box_width, box_height_2, fill=0, stroke=1)
    # linha do título interno
    c.line(box_x, box2_y + box_height_2 - 18, box_x + box_width, box2_y + box_height_2 - 18)

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(box_x + 10, box2_y + box_height_2 - 14, "Resumo do limite superior da janela operacional")

    col1_x = box_x + 12
    col2_x = box_x + box_width / 2
    texto_y = box2_y + box_height_2 - 32

    for i in range(num_linhas_2):
        if i < len(linhas_esq_2):
            label, valor = linhas_esq_2[i]
            c.setFont("Helvetica-Bold", 10)
            c.drawString(col1_x, texto_y, label)
            lw = c.stringWidth(label, "Helvetica-Bold", 10)
            c.setFont("Helvetica", 10)
            c.drawString(col1_x + lw + 4, texto_y, str(valor))

        if i < len(linhas_dir_2):
            label, valor = linhas_dir_2[i]
            c.setFont("Helvetica-Bold", 10)
            c.drawString(col2_x, texto_y, label)
            lw = c.stringWidth(label, "Helvetica-Bold", 10)
            c.setFont("Helvetica", 10)
            c.drawString(col2_x + lw + 4, texto_y, str(valor))

        texto_y -= linha_altura_2

    # cursor abaixo da tabela 2
    y = box2_y - 14

    max_inf = min_sup = None
    janela_txt = "Não informado"
    falha_txt = ""
    status_txt = "Poço instável"
    status_color = (1, 0, 0)  # vermelho

    if row_s is not None:
        max_inf = row_s.get("Max Inferior", None)
        min_sup = row_s.get("Min Superior", None)

        if pd.notna(max_inf) and pd.notna(min_sup):
            janela_txt = f"{float(max_inf):.2f} < ρ < {float(min_sup):.2f} lb/gal"

            if ppg_val is not None and pd.notna(ppg_val):
                if float(max_inf) < float(ppg_val) < float(min_sup):
                    status_txt = "Poço estável"
                    status_color = (0, 0.6, 0)  # verde
                elif _aprox(ppg_val, max_inf):
                    falha_txt = _classificar_falha_no_limite(row_s, ppg_val, "inferior")
                elif _aprox(ppg_val, min_sup):
                    falha_txt = _classificar_falha_no_limite(row_s, ppg_val, "superior")

    linhas_esq = [
        ("Profundidade analisada:", prof_analisada_txt),
        ("Janela operacional:", janela_txt),
    ]
    linhas_dir = [
        ("Estado do poço:", status_txt),
    ]
    if falha_txt:
        linhas_dir.append(("", falha_txt))

    num_linhas = max(len(linhas_esq), len(linhas_dir))
    linha_altura = 15
    box_height = 30 + num_linhas * linha_altura

    box_width = right_margin - left_margin
    box_x = left_margin

    y -= 6
    box_y = y - box_height

    # fundo
    c.setFillColorRGB(0.95, 0.95, 0.95)
    c.rect(box_x, box_y, box_width, box_height, fill=1, stroke=0)

    # borda
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.8)
    c.rect(box_x, box_y, box_width, box_height, fill=0, stroke=1)

    # linha título interno
    c.line(box_x, box_y + box_height - 18, box_x + box_width, box_y + box_height - 18)

    # título
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(box_x + 10, box_y + box_height - 14, "Avaliação na profundidade analisada")

    # conteúdo
    col1_x = box_x + 12
    col2_x = box_x + box_width / 2
    texto_y = box_y + box_height - 32

    for i in range(num_linhas):

        if i < len(linhas_esq):
            label, valor = linhas_esq[i]
            c.setFont("Helvetica-Bold", 10)
            c.drawString(col1_x, texto_y, label)
            lw = c.stringWidth(label, "Helvetica-Bold", 10)
            c.setFont("Helvetica", 10)
            c.drawString(col1_x + lw + 4, texto_y, str(valor))

        if i < len(linhas_dir):
            label, valor = linhas_dir[i]

            if label:
                c.setFont("Helvetica-Bold", 10)
                c.drawString(col2_x, texto_y, label)
                lw = c.stringWidth(label, "Helvetica-Bold", 10)

                if label == "Estado do poço:":
                    c.setFillColorRGB(*status_color)
                    c.setFont("Helvetica-Bold", 10)
                    c.drawString(col2_x + lw + 4, texto_y, str(valor))
                    c.setFillColorRGB(0, 0, 0)
                else:
                    c.setFont("Helvetica", 10)
                    c.drawString(col2_x + lw + 4, texto_y, str(valor))

            else:
                c.setFont("Helvetica-Bold", 10)
                c.drawString(col2_x, texto_y, str(valor))

        texto_y -= linha_altura

    # atualiza cursor para o gráfico
    y = box_y - 10

    fig_mohr = st.session_state.get("fig_mohr", None)

    if fig_mohr is not None and y > footer_y + 80:
        y = desenhar_fig_plotly_no_pdf(
            c=c,
            fig_plotly=fig_mohr,
            left=left_margin,
            right=right_margin,
            top=y,
            bottom=footer_y + 20,
            titulo=None,
            scale=1.2
        )

    draw_footer(c, width, footer_y)
    c.showPage()

    # # ==============================
    # # Página 11
    # # ==============================
    # df_suav = st.session_state.get("df_suav", None)
    # df_mud = st.session_state.get("df_mud", None)  # <<<< NOVO (vem do Excel)
    #
    # tem_mud = isinstance(df_mud, pd.DataFrame) and not df_mud.empty
    # tem_suav = isinstance(df_suav, pd.DataFrame) and not df_suav.empty
    #
    # if not (tem_mud and tem_suav):
    #     pass
    # else:
    #     # --------- Helpers ---------
    #     def _to_num(s):
    #         return pd.to_numeric(s, errors="coerce")
    #
    #     def _fmt2(v, suf=""):
    #         return f"{float(v):.2f}{suf}" if v is not None and pd.notna(v) else "—"
    #
    #     def _seg_contiguos(df_base, mask_bool, col_prof, col_exec, col_req):
    #         segs = []
    #         if df_base.empty:
    #             return segs
    #
    #         mask = mask_bool.fillna(False).astype(bool).values
    #         prof = df_base[col_prof].astype(float).values
    #         execv = df_base[col_exec].astype(float).values
    #         reqv = df_base[col_req].astype(float).values
    #
    #         i = 0
    #         n = len(df_base)
    #         while i < n:
    #             if not mask[i]:
    #                 i += 1
    #                 continue
    #
    #             j = i
    #             while j + 1 < n and mask[j + 1]:
    #                 j += 1
    #
    #             prof_ini = prof[i]
    #             prof_fim = prof[j]
    #
    #             exec_seg = execv[i:j + 1]
    #             req_seg = reqv[i:j + 1]
    #
    #             diffs = np.abs(req_seg - exec_seg)
    #             if np.isfinite(diffs).any():
    #                 k_rel = int(np.nanargmax(diffs))
    #                 k_abs = i + k_rel
    #                 diff_max = float(np.nanmax(diffs))
    #                 prof_pior = float(prof[k_abs])
    #             else:
    #                 diff_max = np.nan
    #                 prof_pior = np.nan
    #
    #             segs.append({
    #                 "pi": prof_ini,
    #                 "pf": prof_fim,
    #                 "prof_pior": prof_pior,
    #                 "exec_min": np.nanmin(exec_seg) if np.isfinite(exec_seg).any() else np.nan,
    #                 "exec_max": np.nanmax(exec_seg) if np.isfinite(exec_seg).any() else np.nan,
    #                 "req_min": np.nanmin(req_seg) if np.isfinite(req_seg).any() else np.nan,
    #                 "req_max": np.nanmax(req_seg) if np.isfinite(req_seg).any() else np.nan,
    #                 "diff_max": diff_max
    #             })
    #
    #             i = j + 1
    #
    #         return segs
    #
    #     def _draw_centered_text(c, x_center, y, text, font_name="Helvetica", font_size=9):
    #         c.setFont(font_name, font_size)
    #         tw = c.stringWidth(str(text), font_name, font_size)
    #         c.drawString(x_center - tw / 2, y, str(text))
    #
    #     def _classificar_falha(df_suav2, prof_ref, lado, nd=6):
    #         if df_suav2 is None or df_suav2.empty or prof_ref is None or pd.isna(prof_ref):
    #             return "Não identificado"
    #
    #         col_prof = "Profundidade (m)" if "Profundidade (m)" in df_suav2.columns else "MD"
    #         if col_prof not in df_suav2.columns:
    #             return "Não identificado"
    #
    #         try:
    #             idx = (df_suav2[col_prof].astype(float) - float(prof_ref)).abs().idxmin()
    #             row = df_suav2.loc[idx]
    #         except Exception:
    #             return "Não identificado"
    #
    #         if lado == "inferior":
    #             lim_col = "Max Inferior"
    #             candidatos = [
    #                 ("Gradiente de Pressão de Poros (lb/gal)", "Limitado pela pressão de poros"),
    #                 ("Tração Inferior", "Falha por tração inferior"),
    #                 ("Comp Inferior σθA", "Falha por compressão inferior em σθA"),
    #                 ("Comp Inferior σθB", "Falha por compressão inferior em σθB"),
    #             ]
    #         else:
    #             lim_col = "Min Superior"
    #             candidatos = [
    #                 ("Tração Superior (σθA)", "Falha por tração superior em σθA"),
    #                 ("Tração Superior (σθB)", "Falha por tração superior em σθB"),
    #                 ("Comp Superior σθA", "Falha por compressão superior em σθA"),
    #                 ("Comp Superior σθB", "Falha por compressão superior em σθB"),
    #             ]
    #
    #         if lim_col not in row or pd.isna(row[lim_col]):
    #             return "Não identificado"
    #
    #         lim = round(float(row[lim_col]), nd)
    #
    #         for col, label in candidatos:
    #             if col not in row or pd.isna(row[col]):
    #                 continue
    #             v = round(float(row[col]), nd)
    #             if lim == v:
    #                 return label
    #
    #         return "Não identificado"
    #
    #     # --------- Sanitiza df_mud (NOVO) ---------
    #     df_mud2 = df_mud.copy()
    #
    #     # garante coluna de profundidade
    #     if "Profundidade (m)" not in df_mud2.columns:
    #         # tenta fallback
    #         for alt in ["Profundidade", "MD", "MD(m)"]:
    #             if alt in df_mud2.columns:
    #                 df_mud2 = df_mud2.rename(columns={alt: "Profundidade (m)"})
    #                 break
    #
    #     col_plan = "Peso do Fluido Planejado (lb/gal)"
    #     col_exec = "Peso do Fluido Executado (lb/gal)"
    #
    #     for ccol in ["Profundidade (m)", col_plan, col_exec]:
    #         if ccol not in df_mud2.columns:
    #             df_mud2[ccol] = np.nan
    #
    #     df_mud2["Profundidade (m)"] = _to_num(df_mud2["Profundidade (m)"])
    #     df_mud2[col_plan] = _to_num(df_mud2[col_plan])
    #     df_mud2[col_exec] = _to_num(df_mud2[col_exec])
    #
    #     df_mud2 = df_mud2.dropna(subset=["Profundidade (m)"]).copy()
    #     df_mud2 = df_mud2.sort_values("Profundidade (m)").reset_index(drop=True)
    #
    #     col_mud_use = "Peso do Fluido Executado (lb/gal)"
    #
    #     if col_mud_use not in df_mud2.columns:
    #         df_mud2[col_mud_use] = np.nan
    #
    #     df_mud2[col_mud_use] = _to_num(df_mud2[col_mud_use])
    #
    #     # remove linhas sem peso executado
    #     df_mud2 = df_mud2.dropna(subset=[col_mud_use]).reset_index(drop=True)
    #
    #     if df_mud2.empty:
    #         pass
    #     else:
    #
    #         if "col_mud_use" not in locals():
    #             pass
    #         else:
    #             # --------- Sanitiza df_suav ---------
    #             col_prof_suav = "Profundidade (m)" if "Profundidade (m)" in df_suav.columns else "MD"
    #             col_maxinf = "Max Inferior"
    #             col_minsup = "Min Superior"
    #
    #             if (col_prof_suav not in df_suav.columns) or (col_maxinf not in df_suav.columns) or (
    #                     col_minsup not in df_suav.columns):
    #                 pass
    #             else:
    #                 cols_falha = [
    #                     col_prof_suav,
    #                     col_maxinf,
    #                     col_minsup,
    #                     "Gradiente de Pressão de Poros (lb/gal)",
    #                     "Tração Inferior",
    #                     "Comp Inferior σθA",
    #                     "Comp Inferior σθB",
    #                     "Tração Superior (σθA)",
    #                     "Tração Superior (σθB)",
    #                     "Comp Superior σθA",
    #                     "Comp Superior σθB",
    #                 ]
    #                 cols_exist = [c for c in cols_falha if c in df_suav.columns]
    #
    #                 df_suav2 = df_suav[cols_exist].copy()
    #                 df_suav2[col_prof_suav] = _to_num(df_suav2[col_prof_suav])
    #                 df_suav2[col_maxinf] = _to_num(df_suav2[col_maxinf])
    #                 df_suav2[col_minsup] = _to_num(df_suav2[col_minsup])
    #
    #                 df_suav2 = (
    #                     df_suav2
    #                     .dropna(subset=[col_prof_suav])
    #                     .sort_values(col_prof_suav)
    #                     .reset_index(drop=True)
    #                 )
    #
    #                 # --------- Match por profundidade mais próxima ---------
    #                 prof_s = df_suav2[col_prof_suav].astype(float)
    #
    #                 max_list, min_list, prof_ref_list = [], [], []
    #                 for d in df_mud2["Profundidade (m)"].astype(float).values:
    #                     idx = (prof_s - float(d)).abs().idxmin()
    #                     prof_ref_list.append(df_suav2.loc[idx, col_prof_suav])
    #                     max_list.append(df_suav2.loc[idx, col_maxinf])
    #                     min_list.append(df_suav2.loc[idx, col_minsup])
    #
    #                 df_cmp = df_mud2[["Profundidade (m)", col_mud_use]].copy()
    #                 df_cmp = df_cmp.rename(
    #                     columns={col_mud_use: "Peso do Fluido (lb/gal)"})  # reusa nomes do seu pipeline
    #                 df_cmp["Prof (ref df_suav)"] = prof_ref_list
    #                 df_cmp["Max Inferior"] = _to_num(pd.Series(max_list))
    #                 df_cmp["Min Superior"] = _to_num(pd.Series(min_list))
    #
    #                 # --------- Máscaras ---------
    #                 ppg = df_cmp["Peso do Fluido (lb/gal)"]
    #                 mx = df_cmp["Max Inferior"]
    #                 mn = df_cmp["Min Superior"]
    #
    #                 mask_le_max = (ppg < mx)  # underbalance
    #                 mask_gt_min = (ppg > mn)  # overbalance
    #
    #                 # --------- Segmentos contíguos ---------
    #                 segs_under = _seg_contiguos(
    #                     df_base=df_cmp,
    #                     mask_bool=mask_le_max,
    #                     col_prof="Profundidade (m)",
    #                     col_exec="Peso do Fluido (lb/gal)",
    #                     col_req="Max Inferior"
    #                 )
    #
    #                 segs_over = _seg_contiguos(
    #                     df_base=df_cmp,
    #                     mask_bool=mask_gt_min,
    #                     col_prof="Profundidade (m)",
    #                     col_exec="Peso do Fluido (lb/gal)",
    #                     col_req="Min Superior"
    #                 )
    #
    #                 # --------- PDF ---------
    #                 y = draw_header(c, width, height, logo)
    #                 left_margin = 40
    #                 right_margin = width - 40
    #
    #                 c.setFillColorRGB(0, 0, 0)
    #                 c.setFont("Helvetica-Bold", 20)
    #                 c.drawString(left_margin, y, f"Análise do Peso do Fluido na Janela Operacional")
    #                 y -= 8
    #                 c.line(left_margin, y, right_margin, y)
    #                 y -= 14
    #
    #                 def desenhar_tabela_segmentos(titulo, segs, y_top, col_balanco_label):
    #                     box_width = right_margin - left_margin
    #                     box_x = left_margin
    #
    #                     pad_left = 6
    #                     pad_right = 6
    #
    #                     col_w = [0.20, 0.20, 0.21, 0.21, 0.18]
    #                     inner_w = box_width - (pad_left + pad_right) * 2
    #                     widths = [inner_w * w for w in col_w]
    #
    #                     x0 = box_x + pad_left
    #                     xs = [x0]
    #                     for w in widths:
    #                         xs.append(xs[-1] + w)
    #
    #                     linha_altura = 15
    #                     header_h = 18
    #                     n_rows = max(1, len(segs))
    #                     box_height = 30 + header_h + n_rows * linha_altura
    #                     box_y = y_top - box_height
    #
    #                     c.setFillColorRGB(0.95, 0.95, 0.95)
    #                     c.rect(box_x, box_y, box_width, box_height, fill=1, stroke=0)
    #                     c.setStrokeColorRGB(0, 0, 0)
    #                     c.setLineWidth(0.8)
    #                     c.rect(box_x, box_y, box_width, box_height, fill=0, stroke=1)
    #
    #                     c.line(box_x, box_y + box_height - 18, box_x + box_width, box_y + box_height - 18)
    #                     c.setFillColorRGB(0, 0, 0)
    #                     c.setFont("Helvetica-Bold", 12)
    #                     c.drawString(box_x + 10, box_y + box_height - 14, titulo)
    #
    #                     texto_y = box_y + box_height - 32
    #
    #                     if not segs:
    #                         c.setFont("Helvetica", 10)
    #                         c.drawString(box_x + 12, texto_y,
    #                                      "Não foram identificados trechos para esta condição.")
    #                         return box_y - 14
    #
    #                     headers = ["Prof. inic.", "Prof. final", "Requerido", "Executado", col_balanco_label]
    #                     c.setLineWidth(0.5)
    #                     c.line(box_x + 8, texto_y - 4, box_x + box_width - 8, texto_y - 4)
    #
    #                     for i, htxt in enumerate(headers):
    #                         xc = (xs[i] + xs[i + 1]) / 2
    #                         _draw_centered_text(c, xc, texto_y, htxt, font_name="Helvetica-Bold", font_size=10)
    #
    #                     texto_y -= header_h
    #
    #                     for s in segs:
    #                         pi, pf = s["pi"], s["pf"]
    #                         req_min, req_max = s["req_min"], s["req_max"]
    #                         exec_min, exec_max = s["exec_min"], s["exec_max"]
    #                         diff_max = s.get("diff_max", np.nan)
    #
    #                         cell = [
    #                             _fmt2(pi, " m"),
    #                             _fmt2(pf, " m"),
    #                             f"{_fmt2(req_min)}–{_fmt2(req_max)} lb/gal",
    #                             f"{_fmt2(exec_min)}–{_fmt2(exec_max)} lb/gal",
    #                             f"{_fmt2(diff_max)} lb/gal"
    #                         ]
    #
    #                         for i, txt in enumerate(cell):
    #                             xc = (xs[i] + xs[i + 1]) / 2
    #                             _draw_centered_text(c, xc, texto_y, txt, font_name="Helvetica", font_size=10)
    #
    #                         texto_y -= linha_altura
    #
    #                     return box_y - 14
    #
    #                 y = desenhar_tabela_segmentos(
    #                     titulo="Trechos em Underbalance (Peso do Fluido < Limite Inferior)",
    #                     segs=segs_under,
    #                     y_top=y,
    #                     col_balanco_label="Underbalance máx."
    #                 )
    #
    #                 y = desenhar_tabela_segmentos(
    #                     titulo="Trechos em Overbalance (Peso do Fluido > Limite Superior)",
    #                     segs=segs_over,
    #                     y_top=y,
    #                     col_balanco_label="Overbalance máx."
    #                 )
    #
    #                 # --------- Subtrechos por tipo de falha (mantém sua lógica) ---------
    #                 def _subtrechos_por_falha(df_cmp, df_suav2, segs, lado, tol_prof=1e-9):
    #                     out = []
    #                     if df_cmp is None or df_cmp.empty or not segs:
    #                         return out
    #
    #                     col_prof_cmp = "Profundidade (m)"
    #                     col_prof_ref = "Prof (ref df_suav)"
    #                     if col_prof_cmp not in df_cmp.columns or col_prof_ref not in df_cmp.columns:
    #                         return out
    #
    #                     prof_cmp = df_cmp[col_prof_cmp].astype(float)
    #                     pref_cmp = df_cmp[col_prof_ref].astype(float)
    #
    #                     for s in segs:
    #                         pi = float(s["pi"])
    #                         pf = float(s["pf"])
    #
    #                         m = (prof_cmp >= (pi - tol_prof)) & (prof_cmp <= (pf + tol_prof))
    #                         df_int = df_cmp.loc[m, [col_prof_cmp, col_prof_ref]].copy()
    #
    #                         if df_int.empty:
    #                             prof_ref = s.get("prof_pior", np.nan)
    #                             if pd.isna(prof_ref):
    #                                 prof_ref = (pi + pf) / 2.0
    #                             tipo = _classificar_falha(df_suav2=df_suav2, prof_ref=prof_ref, lado=lado, nd=6)
    #                             out.append({"pi": pi, "pf": pf, "tipo": tipo})
    #                             continue
    #
    #                         df_int = df_int.sort_values(col_prof_cmp).reset_index(drop=True)
    #
    #                         tipos = []
    #                         for _, r in df_int.iterrows():
    #                             tipo_i = _classificar_falha(
    #                                 df_suav2=df_suav2,
    #                                 prof_ref=float(r[col_prof_ref]),
    #                                 lado=lado,
    #                                 nd=6
    #                             )
    #                             tipos.append(tipo_i)
    #
    #                         df_int["tipo"] = tipos
    #
    #                         tipo_atual = None
    #                         ini = None
    #                         prof_prev = None
    #
    #                         for k in range(len(df_int)):
    #                             p = float(df_int.loc[k, col_prof_cmp])
    #                             t = str(df_int.loc[k, "tipo"])
    #
    #                             if tipo_atual is None:
    #                                 tipo_atual = t
    #                                 ini = p
    #                             elif t != tipo_atual:
    #                                 out.append({"pi": ini, "pf": prof_prev, "tipo": tipo_atual})
    #                                 tipo_atual = t
    #                                 ini = p
    #
    #                             prof_prev = p
    #
    #                         if tipo_atual is not None and ini is not None and prof_prev is not None:
    #                             out.append({"pi": ini, "pf": prof_prev, "tipo": tipo_atual})
    #
    #                     return out
    #
    #                 def desenhar_tabela_falhas_quebrada(titulo, subtrechos, y_top):
    #                     box_width = right_margin - left_margin
    #                     box_x = left_margin
    #
    #                     pad_left = 6
    #                     pad_right = 6
    #
    #                     col_w = [0.22, 0.22, 0.56]
    #                     inner_w = box_width - (pad_left + pad_right) * 2
    #                     widths = [inner_w * w for w in col_w]
    #
    #                     x0 = box_x + pad_left
    #                     xs = [x0]
    #                     for w in widths:
    #                         xs.append(xs[-1] + w)
    #
    #                     linha_altura = 15
    #                     header_h = 18
    #                     n_rows = max(1, len(subtrechos))
    #                     box_height = 30 + (header_h if subtrechos else 0) + n_rows * linha_altura
    #                     box_y = y_top - box_height
    #
    #                     c.setFillColorRGB(0.95, 0.95, 0.95)
    #                     c.rect(box_x, box_y, box_width, box_height, fill=1, stroke=0)
    #                     c.setStrokeColorRGB(0, 0, 0)
    #                     c.setLineWidth(0.8)
    #                     c.rect(box_x, box_y, box_width, box_height, fill=0, stroke=1)
    #
    #                     c.line(box_x, box_y + box_height - 18, box_x + box_width, box_y + box_height - 18)
    #                     c.setFillColorRGB(0, 0, 0)
    #                     c.setFont("Helvetica-Bold", 12)
    #                     c.drawString(box_x + 10, box_y + box_height - 14, titulo)
    #
    #                     texto_y = box_y + box_height - 32
    #
    #                     if not subtrechos:
    #                         c.setFont("Helvetica", 10)
    #                         c.drawString(box_x + 12, texto_y,
    #                                      "Não foram identificados trechos para classificar o tipo de falha.")
    #                         return box_y - 14
    #
    #                     headers = ["Prof. inic.", "Prof. final", "Tipo de falha"]
    #                     c.setLineWidth(0.5)
    #                     c.line(box_x + 8, texto_y - 4, box_x + box_width - 8, texto_y - 4)
    #
    #                     for i, htxt in enumerate(headers):
    #                         xc = (xs[i] + xs[i + 1]) / 2
    #                         _draw_centered_text(c, xc, texto_y, htxt, font_name="Helvetica-Bold", font_size=10)
    #
    #                     texto_y -= header_h
    #
    #                     for stc in subtrechos:
    #                         pi = stc["pi"]
    #                         pf = stc["pf"]
    #                         tipo = stc["tipo"]
    #
    #                         row = [_fmt2(pi, " m"), _fmt2(pf, " m"), tipo]
    #                         for i, txt in enumerate(row):
    #                             xc = (xs[i] + xs[i + 1]) / 2
    #                             _draw_centered_text(c, xc, texto_y, txt, font_name="Helvetica", font_size=10)
    #
    #                         texto_y -= linha_altura
    #
    #                     return box_y - 14
    #
    #                 sub_under = _subtrechos_por_falha(df_cmp=df_cmp, df_suav2=df_suav2, segs=segs_under,
    #                                                   lado="inferior")
    #                 sub_over = _subtrechos_por_falha(df_cmp=df_cmp, df_suav2=df_suav2, segs=segs_over,
    #                                                  lado="superior")
    #
    #                 y = desenhar_tabela_falhas_quebrada(
    #                     titulo="Classificação do Tipo de Falha (Trechos em Underbalance)",
    #                     subtrechos=sub_under,
    #                     y_top=y
    #                 )
    #
    #                 y = desenhar_tabela_falhas_quebrada(
    #                     titulo="Classificação do Tipo de Falha (Trechos em Overbalance)",
    #                     subtrechos=sub_over,
    #                     y_top=y
    #                 )
    #
    #                 draw_footer(c, width, footer_y)
    #                 c.showPage()

    # ==============================
    # Página 12
    # ==============================

    left_margin = 40
    right_margin = width - 40

    # --- coleta do session_state ---
    fig_fratura = st.session_state.get("fig_fratura", None)
    df_f = st.session_state.get("df_f", None)

    tt = st.session_state.get("tt", [])  # ["LOT"/"FIT"...]
    pp = st.session_state.get("pp", [])  # profundidades
    lt = st.session_state.get("lt", [])  # pesos equivalentes (lb/gal)

    # ✅ condição: só gera a página se houver FIT/LOT informado
    tem_lot_fit = any(str(x).upper() in ["LOT", "FIT"] for x in tt)

    if tem_lot_fit:

        y = draw_header(c, width, height, logo)

        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 20)
        c.drawString(left_margin, y, "Gradiente de Fratura - Método das tensões mínimas")

        y -= 8
        c.line(left_margin, y, right_margin, y)

        origem_lots = "Inseridos pelo usuário" if st.session_state.get("lot", False) else "Base de dados"
        aux_flag = "Sim" if st.session_state.get("auxiliar", False) else "Não"

        # contagens
        n_lot = sum(1 for x in tt if str(x).upper() == "LOT")
        n_fit = sum(1 for x in tt if str(x).upper() == "FIT")

        linhas_pontos = []
        if tt and pp and lt and len(tt) == len(pp) == len(lt):
            linhas_pontos.append(("Origem dos LOT/FIT:", origem_lots))
            linhas_pontos.append(("Ponto auxiliar K:", aux_flag))

            for i, (tipo, prof, peso) in enumerate(zip(tt, pp, lt), start=1):
                try:
                    prof_txt = f"{float(prof):.0f} m" if prof is not None and pd.notna(prof) else "—"
                    peso_txt = f"{float(peso):.2f} lb/gal" if peso is not None and pd.notna(peso) else "—"
                except Exception:
                    prof_txt, peso_txt = "—", "—"

                linhas_pontos.append((f"{str(tipo).upper()} {i}:", f"{prof_txt} | {peso_txt}"))
        else:
            linhas_pontos.append(("LOT/FIT:", "Não informado"))

        gf_min_txt = "Não informado"
        gf_max_txt = "Não informado"

        if isinstance(df_f, pd.DataFrame) and not df_f.empty:
            if "Gradiente de Fratura (lb/gal)" in df_f.columns:
                try:
                    gf_min_txt = f"{float(df_f['Gradiente de Fratura (lb/gal)'].min()):.2f} lb/gal"
                    gf_max_txt = f"{float(df_f['Gradiente de Fratura (lb/gal)'].max()):.2f} lb/gal"
                except Exception:
                    pass

        linhas_esq = [
            ("Qtd. LOT:", str(n_lot)),
            ("Qtd. FIT:", str(n_fit)),
            ("G. Fratura (mín):", gf_min_txt),
            ("G. Fratura (máx):", gf_max_txt),
        ]

        # tudo que é “pontos” vai para o lado direito, um abaixo do outro
        linhas_dir = linhas_pontos

        num_linhas = max(len(linhas_esq), len(linhas_dir))
        linha_altura = 15
        box_height = 30 + num_linhas * linha_altura

        box_width = right_margin - left_margin
        box_x = left_margin

        y -= 14
        box_y = y - box_height

        # fundo
        c.setFillColorRGB(0.95, 0.95, 0.95)
        c.rect(box_x, box_y, box_width, box_height, fill=1, stroke=0)

        # borda
        c.setStrokeColorRGB(0, 0, 0)
        c.setLineWidth(0.8)
        c.rect(box_x, box_y, box_width, box_height, fill=0, stroke=1)

        # linha do título interno
        c.line(box_x, box_y + box_height - 18, box_x + box_width, box_y + box_height - 18)

        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(box_x + 10, box_y + box_height - 14, "Resumo dos dados utilizados")

        col1_x = box_x + 12
        col2_x = box_x + box_width / 2
        texto_y = box_y + box_height - 32

        for i in range(num_linhas):

            if i < len(linhas_esq):
                label, valor = linhas_esq[i]
                c.setFont("Helvetica-Bold", 10)
                c.drawString(col1_x, texto_y, label)
                lw = c.stringWidth(label, "Helvetica-Bold", 10)
                c.setFont("Helvetica", 10)
                c.drawString(col1_x + lw + 4, texto_y, str(valor))

            if i < len(linhas_dir):
                label, valor = linhas_dir[i]
                c.setFont("Helvetica-Bold", 10)
                c.drawString(col2_x, texto_y, label)
                lw = c.stringWidth(label, "Helvetica-Bold", 10)
                c.setFont("Helvetica", 10)
                c.drawString(col2_x + lw + 4, texto_y, str(valor))

            texto_y -= linha_altura

        # cursor abaixo da tabela
        y = box_y - 12

        if fig_fratura is not None and y > footer_y + 80:
            y = desenhar_fig_mpl_no_pdf(
                c=c,
                fig=fig_fratura,
                left=left_margin,
                right=right_margin,
                top=y,
                bottom=footer_y + 20,
                titulo=None,
                dpi=180
            )

        draw_footer(c, width, footer_y)
        c.showPage()

    else:
        pass

    # texto das anotações
    anotacoes_txt = st.session_state.get("anotacoes", "")
    anotacoes_txt = anotacoes_txt.strip()

    if anotacoes_txt:
        # ==============================
        # Página 13
        # ==============================
        y = draw_header(c, width, height, logo)

        left_margin = 40
        right_margin = width - 40

        # --- Título da seção ---
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 20)
        c.drawString(left_margin, y, "Anotações")

        y -= 8
        c.line(left_margin, y, right_margin, y)

        y -= 18

        # Caixa padrão (cinza + borda)
        box_width = right_margin - left_margin
        box_x = left_margin

        # altura “máxima” disponível (até o rodapé)
        top = y
        bottom = footer_y + 20
        box_height = top - bottom

        c.setFillColorRGB(0.95, 0.95, 0.95)
        c.rect(box_x, bottom, box_width, box_height, fill=1, stroke=0)

        c.setStrokeColorRGB(0, 0, 0)
        c.setLineWidth(0.8)
        c.rect(box_x, bottom, box_width, box_height, fill=0, stroke=1)

        # linha do título interno
        c.line(box_x, top - 22, box_x + box_width, top - 22)

        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(box_x + 10, top - 16, "Registro de anotações")

        # conteúdo
        text_x = box_x + 12
        text_y = top - 36
        max_width = box_width - 24
        line_height = 12

        c.setFillColorRGB(0, 0, 0)
        text_y = draw_wrapped_text(
            c=c,
            text=anotacoes_txt,
            x=text_x,
            y=text_y,
            max_width=max_width,
            line_height=line_height,
            font_name="Helvetica",
            font_size=10
        )

        c.showPage()

    c.save()
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue()


def geo_page():
    st.title('Syngular Geopressure Analysis - SYGA')

    # tabs = st.tabs(['Entrada de Dados', 'Coluna litológica', 'Gradiente de Sobrecarga',
    #                 'Gradiente de Pressão de Poros', 'Estabilidade de Poço', 'Anotações', 'RTP', 'Relatório', 'Informações Sobre o SYGA' ])
    tabs = st.tabs(['Entrada de Dados', 'Coluna litológica', 'Gradiente de Sobrecarga',
                    'Gradiente de Pressão de Poros', 'Estabilidade de Poço', 'Anotações', 'Relatório', 'Informações Sobre o SYGA'])

    # Carregar Dados
    with tabs[0]:
        c1, c2, c3 = st.columns((1, 1, 1))
        with c3:
            container = st.container(border=True)
            with container:
                st.markdown("### Upload de Arquivo Excel")

                col1, col2 = st.columns(2)
                with col1:
                    # Input para escolher o intervalo de linhas (step)
                    step = st.number_input("Intervalo entre linhas para leitura", min_value=1, value=1, step=1)
                traj_modo = st.selectbox(
                    "Trajetória utilizada para os cálculos",
                    ["Planejada", "Executada"],
                    index=0,
                    key="traj_modo"
                )

                uploaded_file = st.file_uploader("***Envie o seu arquivo Excel***", type=["xlsx", "xls", "xlsm"])

                if uploaded_file:
                    file_bytes = uploaded_file.getvalue()

                    st.session_state.main_xlsm = file_bytes
                    st.session_state.wb = carregar_workbook(file_bytes)
                    try:
                        # Selecionar a aba desejada
                        sheet_name = "Perfilagens"

                        # Carregar os dados da aba selecionada
                        df_full = pd.read_excel(uploaded_file, sheet_name=sheet_name)

                        # Aplicar o intervalo (step)
                        df = df_full.iloc[::step].reset_index(drop=True)

                        # Garantir que o último registro original está no df
                        if not df_full.iloc[-1].equals(df.iloc[-1]):
                            # Adiciona o último registro ao dataframe
                            df = pd.concat([df, df_full.iloc[[-1]]], ignore_index=True)

                        # Colunas obrigatórias
                        colunas_obrigatorias = [
                            "Profundidade",
                            "MD",
                            "Perfil de densidade",
                            "Perfil sônico",
                            "Perfil Raio Gama"
                        ]

                        # Verificação
                        colunas_faltantes = [c for c in colunas_obrigatorias if c not in df.columns]

                        if colunas_faltantes:
                            st.error(
                                "❌ O arquivo enviado não contém todas as colunas obrigatórias.\n\n"
                                f"**Colunas ausentes:** {', '.join(colunas_faltantes)}"
                            )
                            st.stop()
                        else:
                            st.session_state.df1 = df

                        # Atualiza a segunda coluna com a contagem de linhas
                        with col2:
                            st.write("")
                            st.write("")
                            st.markdown(f"**Total de linhas carregadas:** {len(df)}")

                        # Mostrar a tabela resultante
                        st.write("Dados Importados:")
                        st.dataframe(df, use_container_width=True, hide_index=True)

                    except Exception as e:
                        st.error(f"Erro ao processar o arquivo: {e}")

                    # lê a aba Início e preenche session_state
                    try:
                        info_ini = _ler_inicio_do_xlsm(st.session_state.wb)

                        if info_ini["poco"] is not None:
                            st.session_state.poco = info_ini["poco"]

                        # Objetivo -> comments
                        st.session_state.comments = info_ini["comments"]

                        # Coordenadas
                        if info_ini["easting"] is not None:
                            st.session_state.easting = info_ini["easting"]

                        if info_ini["northing"] is not None:
                            st.session_state.northing = info_ini["northing"]

                    except Exception as e:
                        st.warning(f"Não foi possível ler a aba 'Início' para preencher dados do poço: {e}")

                    try:
                        df_mud = _ler_peso_fluido_do_xlsm(st.session_state.wb)
                        st.session_state.df_mud = df_mud

                    except Exception as e:
                        st.warning(f"Não foi possível ler pesos do fluido na aba 'Geopressões': {e}")

                    try:
                        df_eventos = pd.read_excel(uploaded_file, sheet_name="Eventos")

                        # limpa nomes de colunas (resolve "MD Final " com espaço)
                        df_eventos.columns = [str(c).strip() for c in df_eventos.columns]

                        # garante as colunas esperadas
                        col_req_evt = ["MD Inicial", "MD Final", "Evento"]
                        falt = [c for c in col_req_evt if c not in df_eventos.columns]
                        if falt:
                            st.warning(f"A aba 'Eventos' existe, mas faltam colunas: {', '.join(falt)}")
                            st.session_state.df_eventos = pd.DataFrame(columns=col_req_evt)
                        else:
                            # normaliza tipos e strings
                            df_eventos = df_eventos[col_req_evt].copy()
                            df_eventos["MD Inicial"] = pd.to_numeric(df_eventos["MD Inicial"], errors="coerce")
                            df_eventos["MD Final"] = pd.to_numeric(df_eventos["MD Final"], errors="coerce")
                            df_eventos["Evento"] = (
                                df_eventos["Evento"]
                                .astype("string")
                                .fillna("")
                                .str.strip()
                                .str.replace(r"[,\.;:]+$", "", regex=True)  # remove vírgula no final (ex: "Overpull,")
                            )

                            # remove linhas inválidas
                            df_eventos = df_eventos.dropna(subset=["MD Inicial"])
                            df_eventos = df_eventos[df_eventos["Evento"] != ""].reset_index(drop=True)

                            st.session_state.df_eventos = df_eventos

                    except Exception:
                        # se não tiver a aba, só deixa vazio
                        st.session_state.df_eventos = pd.DataFrame(columns=["MD Inicial", "MD Final", "Evento"])

                    try:
                        # 1) lê trajetória do próprio XLSM obrigatório
                        df2 = _ler_trajetoria_do_xlsm(st.session_state.wb, st.session_state.traj_modo)
                        st.session_state.df2 = df2

                        # 2) gera df_interp imediatamente
                        st.session_state.df_interp = _gerar_df_interp_a_partir_df1_df2(st.session_state.df1,
                                                                                       st.session_state.df2)

                    except Exception as e:
                        st.warning(f"Não foi possível carregar/interpolar trajetória pela aba 'Trajetória': {e}")

                    try:
                        sapatas_df = _ler_sapatas_do_xlsm(st.session_state.wb)
                        st.session_state.sapatas_df = sapatas_df
                    except Exception as e:
                        st.warning(f"Não foi possível ler as sapatas do Excel: {e}")

                    try:
                        fases_df = _ler_fases_do_xlsm(st.session_state.wb)
                        st.session_state.fases_df = fases_df
                    except Exception as e:
                        st.warning(f"Não foi possível ler as fases do Excel: {e}")

        with c1:
            container = st.container(border=True)  # Criando um container com borda
            with container:
                st.markdown('#### Informações básicas do poço')
                st.selectbox("Objetivo do estudo", ["Retroanálise", "Previsão de Geopressões"], key="option", index=1)
                st.text_input('Nome do Usuário', max_chars=None, key='user_name', type="default")
                # st.text_input('País', max_chars=None, key='country_name', type="default")
                lista_paises = list(paises.keys())
                st.selectbox(
                    "País",
                    options=lista_paises,
                    index=lista_paises.index("Brasil"),
                    key="country_name"
                )
                codigo_pais = paises.get(st.session_state.country_name)
                flag_path = f"Flag/{codigo_pais}.png" if codigo_pais else None

                st.text_input('Nome da Companhia', max_chars=None, key='company_name', type="default", value="Petroreconcavo")
                st.text_input('Nome do Campo', max_chars=None, key='field_name', type="default")

                col1, col2 = st.columns((1, 0.5))
                with col1:
                    st.text_input('Nome do Poço', max_chars=None, key='poco', type="default")
                with col2:
                    st.write('')
                    st.write('')
                    st.checkbox('Poço Onshore', key="onshore", value=True)
                st.text_input('Datum', key='datum', value='RTKB')
                st.text_area('Objetivo do Poço', max_chars=None, key='comments')

        with c2:
            with st.container(border=True):
                # Função para calcular distância geodésica (em metros) entre dois pontos lat/lon
                def haversine(lat1, lon1, lat2, lon2):
                    R = 6371000
                    phi1 = math.radians(lat1)
                    phi2 = math.radians(lat2)
                    dphi = math.radians(lat2 - lat1)
                    dlambda = math.radians(lon2 - lon1)
                    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
                    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
                    return R * c

                # --- CARREGAR POÇOS DO YAML ---
                with open("pocos.yaml", "r", encoding="utf-8") as f:
                    dados_yaml = yaml.safe_load(f)

                pocos = dados_yaml['pocos']

                # --- INPUT MANUAL DO POÇO BASE ---
                st.markdown("### Coordenadas do Poço")
                st.number_input("Zona UTM", min_value=1, max_value=60, value=24, key='zona')
                st.radio("Hemisfério", ("Norte", "Sul"), index=1, key='hem')
                st.number_input("Coordenada Leste (Easting)", min_value=100000.0, max_value=900000.0, value=569886.5,
                                format="%.2f", key='easting')
                st.number_input("Coordenada Norte (Northing)", min_value=0.0, max_value=10000000.0, value=8571669.07,
                                format="%.2f", key='northing')
                st.number_input("Raio de busca (km)", min_value=0.1, value=0.1, format="%.2f", key='raio')

                lat_base, lon_base = utm.to_latlon(
                    st.session_state.easting,
                    st.session_state.northing,
                    st.session_state.zona,
                    northern=(st.session_state.hem == "Norte")
                )

                # --- CRIA MAPA COM POÇO BASE E CÍRCULO ---
                m = folium.Map(
                    location=[lat_base, lon_base],
                    zoom_start=17,
                    zoom_control=False,
                    attributionControl=False,
                    tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
                    attr='Esri'
                )

                folium.Marker(
                    [lat_base, lon_base],
                    popup=st.session_state.poco if st.session_state.poco else "Poço",
                    icon=folium.CustomIcon('poço.png', icon_size=(30, 30))
                ).add_to(m)

                folium.Circle(
                    location=[lat_base, lon_base],
                    radius=st.session_state.raio * 1000,
                    color='green',
                    fill=True,
                    fill_opacity=0.2,
                    popup=f"Raio: {st.session_state.raio:.2f} km"
                ).add_to(m)

                if uploaded_file:
                    st.session_state.profundidade_maxima = max(df['Profundidade'])

                profundidade_maxima = st.session_state.get('profundidade_maxima', None)

                if profundidade_maxima is not None:
                    pocos_filtrados = [
                        poco for poco in pocos
                        if poco.get("profundidade_vertical_m") is not None and
                           poco["profundidade_vertical_m"] <= profundidade_maxima
                    ]
                else:
                    pocos_filtrados = pocos

                # --- PLOTAR POÇOS E CALCULAR DISTÂNCIAS ---
                dados_pontos = []
                for poco in pocos_filtrados:
                    e = poco['coordenadas']['easting']
                    n = poco['coordenadas']['northing']
                    lat_p, lon_p = utm.to_latlon(e, n, poco['zona_utm'], northern=(st.session_state.hem == "Norte"))
                    dist = haversine(lat_base, lon_base, lat_p, lon_p)
                    dentro_do_raio = dist <= (st.session_state.raio * 1000)

                    cor = 'green' if dentro_do_raio else 'red'
                    popup_text = f"{poco['nome']}<br>Distância: {dist / 1000:.2f} km"
                    # popup_text = f"{poco['nome']}<br>E: {e:.2f}, N: {n:.2f}<br>Distância: {dist / 1000:.2f} km"

                    folium.Marker(
                        location=[lat_p, lon_p],
                        popup=folium.Popup(popup_text, max_width=300),
                        icon=folium.Icon(color=cor, icon='map-marker')
                    ).add_to(m)

                    dados_pontos.append({
                        "Nome": poco['nome'],
                        "Easting": e,
                        "Northing": n,
                        "Distância (km)": round(dist / 1000, 2),
                        "Dentro do Raio": "Sim" if dentro_do_raio else "Não",
                        "Profundidade Medida (m)": poco.get("profundidade_medida_m", None),
                        "Profundidade Vertical (m)": poco.get("profundidade_vertical_m", None),
                        "Peso Eq. (lb/gal)": poco.get("peso_eq_lb_gal", None)
                    })

                # --- MOSTRAR MAPA ---
                st.session_state["mapa_folium_pdf"] = m
                st_folium(m, use_container_width=True, height=400)
                m.save('filename.png')
                # --- TABELA DE PONTOS DENTRO DO RAIO ---
                df_resultado = pd.DataFrame(dados_pontos)
                df_dentro = df_resultado[df_resultado["Dentro do Raio"] == "Sim"].sort_values(
                    by="Distância (km)").reset_index(drop=True)
                df_dentro_exibir = df_dentro.drop(columns=["Dentro do Raio"])

    # Coluna Litológica
    with tabs[1]:

        st.title("Construção da coluna litológica")

        profundidades = []
        formacoes = []
        litologias = []
        st.session_state.well_name = st.session_state.poco

        with st.container(border=True):

            st.subheader("Idades geológicas")

            st.selectbox(
                "Inserir Idade Geológica",
                ['Não', 'Sim'],
                key="idg"
            )

            if st.session_state.idg == 'Sim':

                st.number_input(
                    "Quantidade de idades geológicas",
                    min_value=1,
                    step=1,
                    key="n_id"
                )

                prof_ini_id = []
                prof_fim_id = []
                idade_geo = []

                # profundidade máxima do poço
                try:
                    prof_max = df["Profundidade"].max()
                except:
                    prof_max = 0

                for i in range(int(st.session_state.n_id)):

                    col1, col2, col3 = st.columns(3)

                    # =============================
                    # PROFUNDIDADE INICIAL
                    # =============================

                    with col1:

                        if i == 0:

                            p_ini = st.number_input(
                                f"Profundidade inicial **Intervalo {i + 1}**",
                                step=1.0,
                                format="%f",
                                min_value=0.0,
                                key=f'prof_inicial_2_{i}'
                            )

                        else:

                            p_ini = st.number_input(
                                f"Profundidade inicial **Intervalo {i + 1}**",
                                value=st.session_state.get(f'prof_final_2_{i - 1}', 0.0),
                                disabled=True,
                                key=f'prof_inicial_2_{i}'
                            )

                    # =============================
                    # PROFUNDIDADE FINAL
                    # =============================

                    with col2:

                        if i == int(st.session_state.n_id) - 1:

                            p_fim = st.number_input(
                                f"Profundidade final **Intervalo {i + 1}**",
                                value=float(prof_max + 100),
                                disabled=True,
                                key=f'prof_final_2_{i}'
                            )

                        else:

                            p_fim = st.number_input(
                                f"Profundidade final **Intervalo {i + 1}**",
                                step=1.0,
                                format="%f",
                                min_value=0.0,
                                key=f'prof_final_2_{i}'
                            )

                    # =============================
                    # IDADE
                    # =============================

                    with col3:

                        idade = st.text_input(
                            f"Idade Geológica {i + 1}",
                            key=f'idg_{i}'
                        )

                    prof_ini_id.append(p_ini)
                    prof_fim_id.append(p_fim)
                    idade_geo.append(idade)

                # dataframe final (mesmo formato usado pela função idade_formacao)
                st.session_state.df_idade = pd.DataFrame({
                    "Topo (m)": prof_ini_id,
                    "Base (m)": prof_fim_id,
                    "Idade": idade_geo
                })

        with st.container(border=True):

            st.subheader("Descrição das camadas litológicas")

            if "pocos" not in st.session_state:
                st.session_state.pocos = {"Poço": {}}

            if "well_selected" not in st.session_state:
                st.session_state.well_selected = "Poço"

            pocos = st.session_state.pocos.keys()

            selected = st.session_state.well_selected

            # ===== AUTO-IMPORT LITOLOGIA DO XLSM (ABA "Litologia") =====
            try:
                if "main_xlsm" in st.session_state and st.session_state.main_xlsm is not None:

                    if st.session_state.get("lito_import_ok", False) is False:
                        df_lito_excel = _ler_litologia_do_xlsm(st.session_state.wb)

                        st.session_state.df_lito_excel = df_lito_excel

                        _aplicar_litologia_no_state(selected, df_lito_excel)

                        st.session_state.lito_import_ok = True

            except Exception as e:
                st.warning(f"Falha ao importar litologia do XLSM: {e}")

            try:
                if 'formation' in st.session_state.pocos[selected]:
                    n_fm = len(st.session_state.pocos[selected]['formation'])
                else:
                    n_fm = 5
            except (KeyError, IndexError):
                n_fm = 5

            excel_carregado = st.session_state.get("lito_import_ok", False)

            num_formacoes = st.number_input(
                "Número de formações:",
                min_value=1,
                max_value=20,
                key='n_fm',
                value=int(st.session_state.get("n_fm", 5)),
                disabled=excel_carregado
            )

            for i in range(st.session_state.n_fm):

                try:
                    if 'profundidade' in st.session_state.pocos[selected]:
                        p = st.session_state.pocos[selected]['profundidade'][i]
                    else:
                        p = 0.0

                    if 'formation' in st.session_state.pocos[selected]:
                        f = st.session_state.pocos[selected]['formation'][i]
                    else:
                        f = ''

                    if 'litologia' in st.session_state.pocos[selected]:
                        y = st.session_state.pocos[selected]['litologia'][i]
                    else:
                        y = 'Arenito'

                except (KeyError, IndexError):
                    p = 0.0
                    f = ''
                    y = 'Arenito'

                col1, col2, col3 = st.columns(3)

                with col1:
                    prof = st.number_input(
                        f"Topo da {i + 1}ª formação (TVD)",
                        key=f"prof_{i}",
                        value=p
                    )

                with col2:
                    fm = st.text_input(
                        f"Nome da {i + 1}ª formação",
                        key=f"fm_{i}",
                        value=f
                    )

                with col3:

                    lithology = [
                        "Arenito",
                        "Folhelho",
                        "Calcário",
                        "Carbonato",
                        "Siltito",
                        "Conglomerado",
                        "Halita",
                        "Basalto"
                    ]

                    lit = st.selectbox(
                        f"Litologia da {i + 1}ª formação",
                        lithology,
                        key=f"lit_{i}",
                        index=lithology.index(y)
                    )

                profundidades.append(prof)
                formacoes.append(fm)
                litologias.append(lit)

            st.session_state.pocos[selected]["profundidade"] = profundidades
            st.session_state.pocos[selected]["formation"] = formacoes
            st.session_state.pocos[selected]["litologia"] = litologias

            if "perfil" not in st.session_state.pocos[selected]:
                st.session_state.pocos[selected]["perfil"] = {}

            if "df_idade" in st.session_state and not st.session_state.df_idade.empty:

                base_final = st.session_state.df_idade["Base (m)"].max()

            else:

                base_final = max(profundidades) if profundidades else 0

            # base da última litologia +100
            st.session_state.pocos[selected]["tvd"] = base_final + 100

        with st.container(border=True):

            st.markdown("### Visualização")

            fig = plot_correlacao_com_logs(
                st.session_state.pocos,
                [False, False, False],
                True,
                True,
                list(st.session_state.pocos.keys()),
                False,
                escala=(30, 300)
            )

            st.session_state.fig_coluna_lito = fig

            st.plotly_chart(fig)

    # Gradiente de Sobrecarga
    with tabs[2]:
        tb = st.tabs(['Gradiente de Sobrecarga', 'Tabela de Dados Calculados'])
        with tb[0]:
            if uploaded_file:
                col1, col2, col3 = st.columns((0.9, 0.9, 1))
                with col1:
                    # if st.session_state.s_gr:
                    df.insert(
                        loc=5,
                        column='Raio Gama Suavizado',
                        value=suavizar(
                            df['Profundidade'],
                            df['Perfil Raio Gama']
                        )
                    )
                    # ENTRADA DE DADOS
                    with st.container(border=True):
                        if st.session_state.onshore:
                            x = False
                        else:
                            x = True
                        st.segmented_control("***Correlação para estimativa da densidade da formação***",
                                             ['Perfil de Densidade', 'Gardner'],
                                             selection_mode="multi",
                                             default='Perfil de Densidade',
                                             key='gard',
                                             width="stretch")
                        # st.radio('***Correlação para estimativa da densidade da formação***',
                        #          ['Perfil de Densidade', 'Gardner'], key="gard", index=0,
                        #          disabled=False)
                        st.segmented_control("***Correlação para estimativa da densidade da formação***",
                                             ['Miller', 'Bourgoyne'],
                                             selection_mode="multi",
                                             default='Miller',
                                             key='gard_2',
                                             width="stretch",
                                             disabled=True)
                        # st.radio('',
                        #          ['Miller', 'Bourgoyne'], key="gard2", index=0,
                        #          disabled=True)
                        st.segmented_control("***Extrapolação***",
                                             ['Desativada', 'Ativada'],
                                             selection_mode="single",
                                             default='Desativada',
                                             key='ex',
                                             width="stretch",
                                             disabled=x)

                        # MÉDIA DAS DENSIDADES CASO FOR FAZER A EXTRAPOLAÇÃO
                        if st.session_state.ex == 'Ativada':
                            if st.session_state.onshore:
                                if not df['Perfil de densidade'].isnull().all():
                                    st.session_state.md = round(df['Perfil de densidade'].dropna().head(10).mean(), 2)
                                else:
                                    st.session_state.md = 2.0
                            else:
                                st.session_state.md = 2.0
                            st.number_input('Insira o ***Valor médio da densidade das camadas superiores***',
                                            step=0.1, format='%f', key='ds', min_value=0.0, value=st.session_state.md)

                        # CÁLCULOS
                        with st.form("gs_form", border=False):
                            # if st.session_state.gard == 'Miller':
                            #     parametros_miller()
                            # INSERIR DADOS
                            st.markdown('***Dados de elevação***')
                            st.number_input('***Air Gap***', step=1.0, format='%f', key='rtkb', min_value=0.0, value=5.0)
                            if st.session_state.onshore:
                                st.number_input('***Elevação do DATUM***', step=1.0, format='%f', key='es', min_value=0.0, value=140.0)
                                st.session_state.nf = st.session_state.rtkb + st.session_state.es

                            else:
                                st.number_input("Insira o valor da ***Lâmina D'água***", step=1.0, format='%f',
                                                key='lda', min_value=0.0)
                            float_sonic = df['Perfil sônico'].apply(lambda x: isinstance(x, float)).any()
                            float_dens = df['Perfil de densidade'].apply(lambda x: isinstance(x, float)).any()

                            if float_dens or float_sonic:
                                # OFFSHORE
                                if not st.session_state.onshore:
                                    if st.session_state.ex == 'Desativada':
                                        df_sup = pd.DataFrame({
                                            'Profundidade': [st.session_state.rtkb, st.session_state.lda],
                                            'Perfil de densidade': [0, 1.03],
                                            'Perfil sônico': [None, None]
                                        })
                                    else:
                                        prof_start = (st.session_state.rtkb + st.session_state.lda) + 1
                                        prof_end = df['Profundidade'].iloc[0]

                                        if prof_end > prof_start:
                                            prof = list(range(int(prof_start), int(prof_end) + 1))
                                        else:
                                            prof = []

                                        df_sup = pd.DataFrame({
                                            'Profundidade': [st.session_state.rtkb,
                                                             st.session_state.rtkb + st.session_state.lda] + prof,
                                            'Perfil de densidade': [None, 1.03] + [None] * len(prof),
                                            'Perfil sônico': [None, None] + [None] * len(prof)
                                        })
                                    # Concatena dataframe superior com o dataframe principal
                                    df = pd.concat([df_sup, df], ignore_index=True)

                                    dt = [0]
                                    dd, dm = [], []

                                    if 'Gardner' in st.session_state.gard:
                                        mask = df['Perfil de densidade'].isna()

                                        sonico = df.loc[mask, 'Perfil sônico']

                                        df.loc[mask, 'Perfil de densidade'] = (
                                                0.23 * ((1_000_000 / sonico) ** 0.25)
                                        )
                                    # # Cálculo da densidade total linha a linha
                                    # for index in range(1, len(df)):
                                    #     dens = df['Perfil de densidade'].iloc[index]
                                    #     sonic = df['Perfil sônico'].iloc[index]
                                    #
                                    #     if not pd.isnull(dens) and dens != 0:
                                    #         dt.append(dens)
                                    #     elif pd.isnull(dens) and not pd.isnull(sonic) and sonic != 0:
                                    #         dt.append(0.23 * ((10 ** 6 / sonic) ** 0.25))
                                    #     else:
                                    #         dt.append(st.session_state.md)
                                    #
                                    # df['Perfil de densidade'] = dt

                                    # Cálculo de ΔD (m)
                                    dd.append(df['Profundidade'].iloc[0])
                                    for i in range(1, len(df)):
                                        dd.append(
                                            df['Profundidade'].iloc[i] - df['Profundidade'].iloc[i - 1])
                                    df['ΔD (m)'] = dd

                                    # Cálculo de Densidade x ΔD x 1,422 (psi)
                                    for i in range(len(df)):
                                        dm.append(df['Perfil de densidade'].iloc[i] * df['ΔD (m)'].iloc[i] * 1.422)

                                    df['Densidade x ΔD x 1,422 (psi)'] = dm
                                    df['Pressão de Sobrecarga (psi)'] = df['Densidade x ΔD x 1,422 (psi)'].cumsum()

                                    gs = []
                                    for i in range(len(df)):
                                        if df['Profundidade'].iloc[i] <= st.session_state.rtkb:
                                            gs.append(None)
                                        else:
                                            gs.append(df['Pressão de Sobrecarga (psi)'].iloc[i] / (
                                                        0.1704 * df['Profundidade'].iloc[i]))

                                    df['Gradiente de Sobrecarga (lb/gal)'] = gs

                                # ONSHORE
                                if st.session_state.onshore:
                                    # SE EXTRAPOLAR E NÃO TIVER DADOS DE DENSIDADE, DENSIDADE DAS CAMADAS EXTRAPOLADAS IGUAL A 2
                                    # SE NÃO EXTRAPOLAR, NÃO FAZ NADA
                                    if st.session_state.ex == 'Ativada':
                                        prof_min = df['Profundidade'][0]
                                        ext = np.arange(0, prof_min, 1)
                                        ext_list = list(ext)
                                        prof_list = df['Profundidade'].tolist()
                                        dens_list = df['Perfil de densidade'].tolist()
                                        son_list = df['Perfil sônico'].tolist()
                                        gr_list = df['Perfil Raio Gama'].tolist()
                                        gr_list_s = df['Raio Gama Suavizado'].tolist()
                                        profundidade = ext_list + prof_list
                                        densidade = [
                                                        st.session_state.ds if prof > st.session_state.rtkb else 0
                                                        for prof in ext_list
                                                    ] + dens_list

                                        sonico = [None] * len(ext_list) + son_list
                                        gr = [None] * len(ext_list) + gr_list
                                        gr_s = [None] * len(ext_list) + gr_list_s

                                    else:
                                        profundidade = df['Profundidade'].tolist()
                                        densidade = df['Perfil de densidade'].tolist()
                                        sonico = df['Perfil sônico'].tolist()
                                        gr = df['Perfil Raio Gama'].tolist()
                                        gr_s = df['Raio Gama Suavizado'].tolist()

                                    profundidade_solo = []
                                    for i in range(len(profundidade)):
                                        if st.session_state.ex == 'Ativada':
                                            if profundidade[i] <= st.session_state.rtkb:
                                                profundidade_solo.append(0)
                                            else:
                                                profundidade_solo.append(profundidade[i] - st.session_state.rtkb)
                                        else:
                                            profundidade_solo.append(profundidade[i])

                                    st.session_state.ext_df = pd.DataFrame({
                                        'Profundidade em relação a mesa rotativa (m)': profundidade,
                                        'Profundidade em relação ao solo (m)': profundidade_solo,
                                        'Densidade (g/cm³)': densidade,
                                        'Sônico (µs/pé)': sonico,
                                        'Perfil Raio Gama': gr,
                                        'Raio Gama Suavizado': gr_s
                                    })

                                    if st.session_state.ex == 'Ativada':
                                        st.session_state.ext_df.loc[0, 'Densidade (g/cm³)'] = 0

                                    if 'Gardner' in st.session_state.gard:
                                        mask = st.session_state.ext_df['Densidade (g/cm³)'].isna()

                                        sonico = st.session_state.ext_df.loc[mask, 'Sônico (µs/pé)']

                                        st.session_state.ext_df.loc[mask, 'Densidade (g/cm³)'] = (
                                                0.23 * ((1_000_000 / sonico) ** 0.25)
                                        )

                                    dd = []

                                    for i in range(len(st.session_state.ext_df)):
                                        if i == 0:
                                            if st.session_state.ex == 'Desativada':
                                                dd.append(st.session_state.ext_df[
                                                              'Profundidade em relação a mesa rotativa (m)'].iloc[0])
                                            else:
                                                dd.append(0)
                                        else:
                                            if st.session_state.ex == 'Ativada':
                                                if st.session_state.ext_df[
                                                    'Profundidade em relação a mesa rotativa (m)'].iloc[
                                                    i] <= st.session_state.rtkb:
                                                    dd.append(0)
                                                else:
                                                    dd.append(st.session_state.ext_df[
                                                                  'Profundidade em relação a mesa rotativa (m)'].iloc[
                                                                  i] - st.session_state.ext_df[
                                                                  'Profundidade em relação a mesa rotativa (m)'].iloc[
                                                                  i - 1])
                                            else:
                                                dd.append(st.session_state.ext_df[
                                                              'Profundidade em relação a mesa rotativa (m)'].iloc[i] -
                                                          st.session_state.ext_df[
                                                              'Profundidade em relação a mesa rotativa (m)'].iloc[
                                                              i - 1])
                                    st.session_state.ext_df['ΔD (m)'] = dd

                                    dm = []
                                    for i in range(len(st.session_state.ext_df)):
                                        if st.session_state.ext_df['Profundidade em relação a mesa rotativa (m)'].iloc[
                                            i] <= st.session_state.rtkb:
                                            dm.append(0)
                                        else:
                                            dm.append(st.session_state.ext_df['Densidade (g/cm³)'].iloc[i] *
                                                      st.session_state.ext_df['ΔD (m)'].iloc[i] * 1.422)

                                    st.session_state.ext_df['Densidade (g/cm³) x ΔD x 1,422 (psi)'] = dm

                                    st.session_state.ext_df['Pressão de Sobrecarga (psi)'] = st.session_state.ext_df[
                                        'Densidade (g/cm³) x ΔD x 1,422 (psi)'].cumsum()

                                    gs = []
                                    for i in range(len(st.session_state.ext_df)):
                                        if st.session_state.ext_df['Profundidade em relação a mesa rotativa (m)'].iloc[
                                            i] < st.session_state.rtkb:
                                            gs.append(0)
                                        else:
                                            gs.append(
                                                st.session_state.ext_df['Pressão de Sobrecarga (psi)'].iloc[i] /
                                                (0.1704 * st.session_state.ext_df[
                                                    'Profundidade em relação a mesa rotativa (m)'].iloc[i]))

                                    st.session_state.ext_df['Gradiente de Sobrecarga (lb/gal)'] = gs

                                # FALTA IMPLEMENTAR
                                # MILLER
                                # BOURGOYNE

                            else:
                                st.warning('Não foram encontrados dados válidos de perfilagem')

                            st.form_submit_button("Calcular Gradiente de Sobrecarga", use_container_width=True,
                                                  type='primary')

                # Gráfico do Gradiente de Sobrecarga
                with col3:
                    # Gráfico de Gradiente de Sobrecarga
                    container = st.container(border=True)  # Criando um container com borda
                    with container:
                        if st.button('🔄 Limpar Dados - Gradiente de Sobrecarga', use_container_width=True,
                                     type='primary'):
                            keys_to_clear = [
                                'gard', 'ex', 'bop', 'ds', 'rtkb', 'lda', 'es', 'md',
                                'nf', 'datum', 'ext_df'
                            ]
                            for key in keys_to_clear:
                                if key in st.session_state:
                                    del st.session_state[key]
                            st.rerun()

                        def reset_config():
                            if st.session_state.ogs == "Gradiente (lb/gal)":
                                st.session_state.x_min_s = 7
                                st.session_state.x_max_s = 23
                                st.session_state.x_step_s = 1

                            else:
                                st.session_state.x_min_s = 0
                                st.session_state.x_max_s = int(
                                    st.session_state.ext_df['Pressão de Sobrecarga (psi)'].max()) + 200
                                st.session_state.x_step_s = 500

                            st.session_state.y_min_s = 0
                            st.session_state.y_max_s = int(df['Profundidade'].max()) + 100
                            st.session_state.y_step_s = 200

                        with st.expander("Configurações do Gráfico", expanded=False):
                            st.segmented_control("***Opção de Gráfico***", ['Gradiente (lb/gal)',
                                                                            'Pressão (psi)'],
                                                 selection_mode="single",
                                                 default='Gradiente (lb/gal)', key='ogs', width="stretch")
                            st.selectbox("Profundidade", ['TVD', 'MD'], key="t_prof_s")
                            st.number_input("Eixo X - mínimo", value=7, step=1, key="x_min_s")
                            st.number_input("Eixo X - máximo", value=23, step=1, key="x_max_s")
                            st.number_input("Passo do eixo X", value=1, step=1, key="x_step_s")

                            st.number_input("Eixo Y - mínimo", value=0,
                                            step=100, key="y_min_s")
                            st.number_input("Eixo Y - máximo", value=int(df['Profundidade'].max()) + 100,
                                            step=100, key="y_max_s")
                            st.number_input("Passo do eixo Y", value=200, step=50, key="y_step_s")
                            # Botão de reset com callback
                            st.button("Resetar Configurações Gráficas - Gradiente de Sobrecarga", on_click=reset_config,
                                      type="primary", use_container_width=True)
                        if st.session_state.ogs == "Gradiente (lb/gal)":
                            if st.session_state.onshore:
                                st.session_state.oes = st.session_state.ext_df['Gradiente de Sobrecarga (lb/gal)']
                                st.session_state.profs = st.session_state.ext_df['Profundidade em relação a mesa rotativa (m)']
                            else:
                                st.session_state.oes = df['Gradiente de Sobrecarga (lb/gal)']
                                st.session_state.profs = df['Profundidade']
                            st.session_state.oesl = "G. de Sobrecarga"

                        else:
                            st.session_state.oes = st.session_state.ext_df['Pressão de Sobrecarga (psi)']
                            st.session_state.oesl = "P. de Sobrecarga"

                        # Ajuste da figura
                        st.session_state.fig_gs, ax = plt.subplots(figsize=(8, 10))
                        if st.session_state.onshore:
                            if st.session_state.rtkb and st.session_state.es != 0:
                                if st.session_state.ex == 'Desativada':
                                    ax.plot(st.session_state.oes, st.session_state.profs,
                                            color='black', linestyle='-', linewidth=2, label=st.session_state.oesl)
                                else:
                                    ax.plot(st.session_state.oes,
                                            st.session_state.ext_df['Profundidade em relação a mesa rotativa (m)'],
                                            color='black', linestyle='-', linewidth=2, label=st.session_state.oesl)

                        else:
                            if st.session_state.rtkb and st.session_state.lda != 0:
                                ax.plot(st.session_state.oes, st.session_state.profs,
                                        color='black', linestyle='-', linewidth=2, label=st.session_state.oesl)

                        if st.session_state.ogs == "Gradiente (lb/gal)":
                            ax.set_title('Gradiente de Sobrecarga (lb/gal)', fontsize=14, fontweight='bold')
                            ax.set_xlabel('Gradiente (ppg)', fontsize=12)
                            ax.set_ylabel('Profundidade em relação à mesa rotativa', fontsize=12)
                        else:
                            ax.set_title('Pressão de Sobrecarga (psi)', fontsize=14, fontweight='bold')
                            ax.set_xlabel('Pressão (psi)', fontsize=12)
                            ax.set_ylabel('Profundidade em relação à mesa rotativa', fontsize=12)
                        ax.set_ylabel('Profundidade em relação à mesa rotativa', fontsize=12)
                        ax.invert_yaxis()
                        max_depth = int(df['Profundidade'].max()) + 100
                        margin = 20
                        ax.set_yticks(
                            range(st.session_state.y_min_s, st.session_state.y_max_s, st.session_state.y_step_s))
                        ax.set_ylim(st.session_state.y_max_s, st.session_state.y_min_s)
                        ax.set_xticks(
                            range(st.session_state.x_min_s, st.session_state.x_max_s, st.session_state.x_step_s))
                        ax.set_xlim(st.session_state.x_min_s, st.session_state.x_max_s)
                        ax.grid(True, linestyle='--', alpha=0.5)
                        ax.legend(
                            loc='upper right',
                            fontsize=8,
                            frameon=True,
                            shadow=True,
                            fancybox=True,
                            framealpha=1,
                            facecolor='white',
                            edgecolor='gray'
                        )
                        add_watermark(
                            ax,
                            logo_path="logo.png",
                            xy=(0.50, 0.5),
                            zoom=0.2,
                            alpha=0.2,
                            zorder=0
                        )

                        st.pyplot(st.session_state.fig_gs)

                # Imagem Datum
                with col2:
                    container = st.container(border=True)  # Criando um container com borda
                    with container:
                        if st.session_state.onshore:
                            image = Image.open("rig.png")
                            st.image(image, width=200)
                            st.markdown(
                                f"""
                                <div style='position: absolute; top: -240px; left: 200px;
                                            padding: 10px; border-radius: 5px;'>
                                    <strong>Datum: </strong> {st.session_state.datum} <br> 
                                </div>
                                <div style='position: absolute; top: -197px; left: 200px;
                                            padding: 10px; border-radius: 5px;'>
                                    <strong>Elevação DATUM: </strong> {st.session_state.es:.2f} m <br>
                                </div>
                                <div style='position: absolute; top: -170px; left: 200px;
                                            padding: 10px; border-radius: 5px;'>
                                    <strong>Air gap: </strong> {st.session_state.rtkb:.2f} m <br>
                                </div>
                                <div style='position: absolute; top: -140px; left: 200px;
                                            padding: 10px; border-radius: 5px;'>
                                    <strong>Solo <br>
                                </div>
                                <div style='position: absolute; top: -90px; left: 200px;
                                            padding: 10px; border-radius: 5px;'>
                                    <strong>Elevação do solo: </strong> {st.session_state.es - st.session_state.rtkb:.2f} m <br>
                                </div>
                                <div style='position: absolute; top: -43px; left: 200px;
                                            padding: 10px; border-radius: 5px;'>
                                    <strong>Nível do Mar </strong>
                                </div>
                                """,
                                unsafe_allow_html=True
                            )

                        else:
                            # if not st.session_state.bop:
                            image = Image.open("rig_offshore.png")
                            st.image(image, width=200)
                            st.markdown(
                                f"""
                                <div style='position: absolute; top: -240px; left: 200px;
                                            padding: 10px; border-radius: 5px;'>
                                    <strong>Datum: </strong> {st.session_state.datum} <br> 
                                </div>
                                <div style='position: absolute; top: -170px; left: 200px;
                                            padding: 10px; border-radius: 5px;'>
                                    <strong>Air gap: </strong> {st.session_state.rtkb:.2f} m <br>
                                </div>
                                <div style='position: absolute; top: -90px; left: 200px;
                                            padding: 10px; border-radius: 5px;'>
                                    <strong>Lâmina D'água: </strong> {st.session_state.lda:.2f} m <br>
                                </div>
                                <div style='position: absolute; top: -43px; left: 200px;
                                            padding: 10px; border-radius: 5px;'>
                                    <strong>Leito Marinho </strong>
                                </div>
                                """,
                                unsafe_allow_html=True
                            )

            else:
                st.error('Por favor, insira um documento!', icon="🚨")

        # Ver Dataframes
        with tb[1]:
            if uploaded_file:
                if not st.session_state.onshore:
                    st.dataframe(df, use_container_width=True, hide_index=True)
                else:
                    if st.session_state.rtkb and st.session_state.es != 0:
                        st.dataframe(st.session_state.ext_df, use_container_width=True, hide_index=True)

    # Gradiente de Pressão de Poros
    with tabs[3]:
        tb = st.tabs(['Gradiente de Pressão de Poros', 'Tabela de Dados Calculados'])
        with tb[0]:
            if uploaded_file:
                if st.session_state.rtkb != 0:
                    coluna1, coluna2, coluna3 = st.columns((1, 1, 1))
                    # ENTRADA DE DADOS
                    with coluna1:
                        if "suav_s" not in st.session_state:
                            st.session_state.suav_s = False
                        if "fpp" not in st.session_state:
                            st.session_state.fpp = 0.01
                        with st.container(border=True):
                            st.segmented_control("***Mecanismo Gerador de Pressão de Poros***", ['Subcompactação','Transferência Lateral'],
                                                 selection_mode="single",default='Subcompactação', key='mgpp', width="stretch", disabled=True)
                            if st.session_state.onshore:
                                profundidade = st.session_state.ext_df['Profundidade em relação a mesa rotativa (m)']
                                densidade = st.session_state.ext_df['Densidade (g/cm³)']
                                sonico = st.session_state.ext_df['Sônico (µs/pé)']
                                raio_gama = st.session_state.ext_df['Perfil Raio Gama']
                                raio_gama_s = st.session_state.ext_df['Raio Gama Suavizado']
                            else:
                                profundidade = df['Profundidade']
                                densidade = df['Perfil de densidade']
                                sonico = df['Perfil sônico']
                                raio_gama = df['Perfil Raio Gama']
                                raio_gama_s = df['Raio Gama Suavizado']

                            # Criação do DataFrame principal
                            df_pp = pd.DataFrame({
                                'Profundidade (m)': profundidade,
                                'Perfil de densidade (g/cm³)': densidade,
                                'Perfil sônico (µs/pé)': sonico,
                                'Perfil Raio Gama': raio_gama,
                                'Raio Gama Suavizado': raio_gama_s
                            })

                            try:
                                if st.session_state.onshore:
                                    # normal(st.session_state.anormal, st.session_state.gn)
                                    normal(df)

                                m = (np.log10(st.session_state.s2 / st.session_state.s1)) / (
                                        st.session_state.pp2 - st.session_state.pp1)
                                rn = []
                                if not st.session_state.suav_s:
                                    s = df_pp['Perfil sônico (µs/pé)']
                                else:
                                    s = df_pp['Perfil sônico suavizado (µs/pé)']
                                for i in range(len(df_pp)):
                                    if pd.isnull(s.iloc[i]) and not pd.isnull(
                                            df_pp['Perfil de densidade (g/cm³)'].iloc[i]):
                                        # rn.append(None)
                                        rn.append(st.session_state.s1 * 10 ** (m * (
                                                df_pp['Profundidade (m)'].iloc[i] - st.session_state.pp1)))
                                    elif not pd.isnull(s.iloc[i]) and s.iloc[i] != 0:
                                        rn.append(st.session_state.s1 * 10 ** (m * (
                                                df_pp['Profundidade (m)'].iloc[i] - st.session_state.pp1)))
                                    elif pd.isnull(s.iloc[i]) and pd.isnull(
                                            df_pp['Perfil de densidade (g/cm³)'].iloc[i]):
                                        rn.append(None)
                                df_pp.insert(4, 'Perfil sônico (µs/pé) Reta Normal', pd.Series(rn))
                            except Exception as e:
                                pass

                            # Aplica a suavização condicionalmente
                            if st.session_state.suav_s:
                                perfil_sonico_suav = suavizar(profundidade, sonico)
                                df_pp.insert(3, 'Perfil sônico suavizado (µs/pé)', perfil_sonico_suav)

                            if 'boyance' not in st.session_state:
                                st.session_state.boyance = 'Não'

                            if st.session_state.boyance == 'Sim':
                                with st.expander("Boyance", expanded=True):
                                    if 'n_boyance' not in st.session_state:
                                        st.session_state.n_boyance = 1
                                    col_add_2, col_rem_2 = st.columns(2)
                                    with col_add_2:
                                        if st.button(
                                                "Adicionar Intervalo Boyance",
                                                type="primary",
                                                use_container_width=True,
                                                key="b_add_boyance"
                                        ):
                                            st.session_state.n_boyance += 1

                                    with col_rem_2:
                                        if st.button(
                                                "Remover Intervalo Boyance",
                                                type="primary",
                                                use_container_width=True,
                                                key="b_rem_boyance"
                                        ):
                                            if st.session_state.n_boyance > 1:
                                                st.session_state.n_boyance -= 1

                                    st.segmented_control("***Pressão***",
                                                         ['Base do arenito = Topo do Folhelho',
                                                          'Topo do arenito = Base do Folhelho'],
                                                         selection_mode="multi",
                                                         default='Base do arenito = Topo do Folhelho',
                                                         key='o_boyance',
                                                         width="stretch")

                                    for i in range(st.session_state.n_boyance):
                                        with st.expander(f"### Boyance - Intervalo {i + 1}", expanded=True):
                                            st.markdown(f"### Boyance - Intervalo {i + 1}")

                                            # ===== INTERVALO DE PROFUNDIDADE (SÓ SE > 1 BLOCO) =====
                                            if st.session_state.n_boyance > 1:
                                                colun1, colun2 = st.columns(2)
                                                with colun1:
                                                    if i == 0:
                                                        st.number_input(
                                                            "Profundidade inicial",
                                                            step=1.0,
                                                            format="%f",
                                                            min_value=0.0,
                                                            key=f'prof_inicial_{i}'
                                                        )
                                                    else:
                                                        st.number_input(
                                                            "Profundidade inicial",
                                                            value=st.session_state.get(f'prof_final_{i - 1}', 0.0),
                                                            disabled=True,
                                                            key=f'prof_inicial_{i}'
                                                        )

                                                with colun2:
                                                    if i == st.session_state.n_boyance - 1:
                                                        st.number_input(
                                                            "Profundidade final",
                                                            value=df_pp['Profundidade (m)'].max()+100,
                                                            disabled=True,
                                                            key=f'prof_final_{i}'
                                                        )
                                                    else:
                                                        st.number_input(
                                                            "Profundidade final",
                                                            step=1.0,
                                                            format="%f",
                                                            min_value=0.0,
                                                            key=f'prof_final_{i}'
                                                        )

                                            # Inicializa o FPR apenas uma vez
                                            if f"fpr_{i}" not in st.session_state:
                                                st.session_state[f"fpr_{i}"] = 8.5

                                            st.number_input(
                                                "Peso do Fluido Contido nos Poros das Rochas",
                                                min_value=1.,
                                                format='%1f',
                                                key=f'fpr_{i}',
                                                step=0.5,
                                                value=8.5
                                            )

                                    boyances = []

                                    for i in range(st.session_state.n_boyance):
                                        if f'fpr_{i}' in st.session_state:
                                            boyances.append({
                                                'fpr': st.session_state.get(f'fpr_{i}'),
                                                'prof_inicial': st.session_state.get(f'prof_ini_{i}', None),
                                                'prof_final': st.session_state.get(f'prof_fim_{i}', None)
                                            })

                            if 'n_trending' not in st.session_state:
                                st.session_state.n_trending = 1

                            col_add, col_rem = st.columns(2)

                            with col_add:
                                if st.button(
                                        "Adicionar Trending/LBF",
                                        type="primary",
                                        use_container_width=True,
                                        key="b_add_trending"
                                ):
                                    st.session_state.n_trending += 1

                            with col_rem:
                                if st.button(
                                        "Remover Trending/LBF",
                                        type="primary",
                                        use_container_width=True,
                                        key="b_rem_trending"
                                ):
                                    if st.session_state.n_trending > 1:
                                        idx = st.session_state.n_trending - 1

                                        keys_to_remove = [
                                            f'pp1_{idx}', f'pp2_{idx}',
                                            f's1_{idx}', f's2_{idx}',
                                            f'lbf_{idx}', f'inclbf_{idx}',
                                            f'prof_ini_{idx}', f'prof_fim_{idx}'
                                        ]

                                        for k in keys_to_remove:
                                            if k in st.session_state:
                                                del st.session_state[k]

                                        st.session_state.n_trending -= 1
                            with st.form("p_poros", border=False):
                                if st.session_state.mgpp == 'Subcompactação':
                                    with st.expander("Informações Gerais", expanded=True):
                                        st.number_input('Expoente de Eaton', step=1.0, format='%f', key='expoente', value = 3.0)
                                        st.number_input('Profundidade de início da zona anormal', step=100.0,format='%f', key='anormal', value=400.0)
                                        st.number_input('Gradiente Normal', step=1.0, format='%f', key='gn', value=8.5)

                                    for i in range(st.session_state.n_trending):
                                        with st.expander(f"### Trending / LBF {i + 1}", expanded=True):
                                            # ===== INTERVALO DE PROFUNDIDADE (SÓ SE > 1 BLOCO) =====
                                            if st.session_state.n_trending > 1:
                                                colun1, colun2 = st.columns(2)

                                                with colun1:
                                                    st.number_input(
                                                        "Profundidade inicial",
                                                        step=1.0,
                                                        format="%f",
                                                        min_value=0.0,
                                                        key=f'prof_ini_{i}'
                                                    )

                                                with colun2:
                                                    st.number_input(
                                                        "Profundidade final",
                                                        step=1.0,
                                                        format="%f",
                                                        min_value=0.0,
                                                        key=f'prof_fim_{i}'
                                                    )

                                            # -------- TRENDING --------
                                            with st.expander("Trending", expanded=False):
                                                col1, col2 = st.columns((1, 1))

                                                with col1:
                                                    st.number_input(
                                                        'Profundidade 1',
                                                        step=1.0,
                                                        format='%f',
                                                        min_value=0.0,
                                                        value=400.,
                                                        key=f'pp1_{i}'
                                                    )
                                                    st.number_input(
                                                        'Profundidade 2',
                                                        step=1.0,
                                                        format='%f',
                                                        min_value=0.0,
                                                        value=1000.,
                                                        key=f'pp2_{i}'
                                                    )

                                                with col2:
                                                    st.number_input(
                                                        'Leitura 1 do Sônico',
                                                        step=1.0,
                                                        format='%f',
                                                        min_value=0.0,
                                                        value=110.,
                                                        key=f's1_{i}'
                                                    )
                                                    st.number_input(
                                                        'Leitura 2 do Sônico',
                                                        step=1.0,
                                                        format='%f',
                                                        min_value=0.0,
                                                        value=87.,
                                                        key=f's2_{i}'
                                                    )

                                            # -------- LINHA BASE DE FOLHELHOS --------
                                            with st.expander("Linha Base de Folhelhos", expanded=False):
                                                st.number_input(
                                                    'Ponto inicial da LBF',
                                                    step=10.0,
                                                    format='%f',
                                                    min_value=1.0,
                                                    value=111.0,
                                                    key=f'lbf_{i}',
                                                    help=(
                                                        "**Linha Base de Folhelhos (LBF)** 📉\n\n"
                                                        "- Representa o comportamento esperado dos **folhelhos normalmente compactados**.\n"
                                                        "- Traçada no registro **raio gama (GAPI) × profundidade (m)**."
                                                    )
                                                )

                                                st.number_input(
                                                    'Inclinação da LBF',
                                                    step=0.1,
                                                    format='%f',
                                                    value=0.0,
                                                    key=f'inclbf_{i}'
                                                )

                                            trendings = []

                                            for i in range(st.session_state.n_trending):
                                                if f'pp1_{i}' in st.session_state:
                                                    trendings.append({
                                                        'pp1': st.session_state.get(f'pp1_{i}'),
                                                        'pp2': st.session_state.get(f'pp2_{i}'),
                                                        's1': st.session_state.get(f's1_{i}'),
                                                        's2': st.session_state.get(f's2_{i}'),
                                                        'lbf': st.session_state.get(f'lbf_{i}'),
                                                        'inclbf': st.session_state.get(f'inclbf_{i}'),
                                                        'prof_ini': st.session_state.get(f'prof_ini_{i}', None),
                                                        'prof_fim': st.session_state.get(f'prof_fim_{i}', None)
                                                    })

                                    # Garante que as colunas existam, mesmo sem trending válido
                                    if 'Perfil sônico (µs/pé) Reta Normal' not in df_pp.columns:
                                        df_pp['Perfil sônico (µs/pé) Reta Normal'] = np.nan

                                    if 'LBF_calc' not in df_pp.columns:
                                        df_pp['LBF_calc'] = np.nan

                                    for tr in trendings:
                                        try:
                                            prof = df_pp['Profundidade (m)']

                                            # -------- VALIDAÇÃO DOS PARÂMETROS --------
                                            if (
                                                    tr['s1'] in (None, 0) or
                                                    tr['s2'] in (None, 0) or
                                                    tr['pp1'] is None or
                                                    tr['pp2'] is None or
                                                    tr['s2'] == tr['s1'] or
                                                    tr['pp2'] == tr['pp1']
                                            ):
                                                continue

                                            # ---------- TRENDING ----------
                                            if st.session_state.onshore:
                                                # Caso ONshore (modelo logarítmico)
                                                m = (np.log10(tr['s2'] / tr['s1'])) / (tr['pp2'] - tr['pp1'])
                                                s_normal = tr['s1'] * 10 ** (m * (prof - tr['pp1']))
                                            else:
                                                # Caso OFFshore (modelo linear invertido)
                                                m = (tr['pp2'] - tr['pp1']) / (tr['s2'] - tr['s1'])
                                                b = -(m * tr['s2']) + tr['pp2']
                                                s_normal = (prof - b) / m

                                            # ---------- INTERVALO DO TRENDING ----------
                                            if tr['prof_ini'] is not None and tr['prof_fim'] is not None:
                                                mask = (prof >= tr['prof_ini']) & (prof <= tr['prof_fim'])
                                            else:
                                                mask = np.ones(len(df_pp), dtype=bool)

                                            # ---------- NOVA VERIFICAÇÃO (OFFSHORE) ----------
                                            if not st.session_state.onshore:
                                                mask = mask & (prof > st.session_state.lda)

                                            df_pp.loc[mask, 'Perfil sônico (µs/pé) Reta Normal'] = s_normal[mask]

                                            # ---------- LBF ----------
                                            prof_ref = prof.min()
                                            lbf_line = tr['inclbf'] * (prof - prof_ref) + tr['lbf']
                                            df_pp.loc[mask, 'LBF_calc'] = lbf_line[mask]

                                        except Exception:
                                            pass

                                    if st.session_state.onshore:
                                        df_pp['Gradiente de Sobrecarga (lb/gal)'] = st.session_state.ext_df[
                                            'Gradiente de Sobrecarga (lb/gal)']
                                    else:
                                        df_pp['Gradiente de Sobrecarga (lb/gal)'] = df[
                                            'Gradiente de Sobrecarga (lb/gal)']

                                    gp = []
                                    prof_min = df["Profundidade"].min()
                                    base_grad = st.session_state.gn
                                    if not st.session_state.suav_s:
                                        s = df_pp['Perfil sônico (µs/pé)']
                                    else:
                                        s = df_pp['Perfil sônico suavizado (µs/pé)']
                                    for i in range(len(df_pp)):
                                        perfil_sonico = s.iloc[i]
                                        if pd.isna(perfil_sonico):
                                            perfil_sonico = df_pp['Perfil sônico (µs/pé) Reta Normal'].iloc[i]
                                        perfil = perfil_sonico
                                        if st.session_state.onshore:
                                            x = (
                                                    df_pp['Gradiente de Sobrecarga (lb/gal)'].iloc[i]
                                                    - (
                                                            (df_pp['Gradiente de Sobrecarga (lb/gal)'].iloc[
                                                                 i] - base_grad)
                                                            * ((perfil /
                                                                df_pp['Perfil sônico (µs/pé) Reta Normal'].iloc[
                                                                    i]) ** (-st.session_state.expoente))
                                                    )
                                            )
                                            profundidade_atual = df_pp['Profundidade (m)'].iloc[i]
                                            if profundidade_atual < prof_min:
                                                if profundidade_atual < st.session_state.anormal:
                                                    idx_mais_proximo = (
                                                        (st.session_state.df_gfs[
                                                             'Profundidade (m)'] - profundidade_atual)
                                                        .abs()
                                                        .idxmin()
                                                    )
                                                    valor_gradiente = st.session_state.df_gfs.loc[
                                                        idx_mais_proximo, 'Gradiente de Pressão de Poros (lb/gal)'
                                                    ]
                                                    gp.append(valor_gradiente)
                                                else:
                                                    if x < st.session_state.gn:
                                                        gp.append(base_grad)
                                                    else:
                                                        gp.append(x)
                                            else:
                                                if "s_gr" not in st.session_state:
                                                    curva = df_pp['Perfil Raio Gama']
                                                else:
                                                    if not st.session_state.s_gr:
                                                        curva = df_pp['Perfil Raio Gama']
                                                    else:
                                                        curva = df_pp['Raio Gama Suavizado']
                                                if curva.iloc[i] >= df_pp['LBF_calc'].iloc[i]:

                                                    if profundidade_atual < st.session_state.anormal:
                                                        valor_gradiente = st.session_state.gn
                                                        gp.append(valor_gradiente)

                                                    else:
                                                        if x < st.session_state.gn:
                                                            gp.append(base_grad)
                                                        else:
                                                            gp.append(x)
                                                else:
                                                    if gp:
                                                        gp.append(gp[-1])
                                                    else:
                                                        gp.append(base_grad)
                                        else:
                                            x = (
                                                    df_pp['Gradiente de Sobrecarga (lb/gal)'].iloc[i]
                                                    - (
                                                            (df_pp['Gradiente de Sobrecarga (lb/gal)'].iloc[
                                                                 i] - st.session_state.gn)
                                                            * (
                                                                    (perfil / df_pp[
                                                                        'Perfil sônico (µs/pé) Reta Normal'].iloc[
                                                                        i])
                                                                    ** (-st.session_state.expoente)
                                                            )
                                                    )
                                            )

                                            if df_pp['Profundidade (m)'].iloc[i] <= st.session_state.lda:
                                                gp.append(None)
                                            elif df_pp['Profundidade (m)'].iloc[i] < st.session_state.anormal or x < st.session_state.gn:
                                                gp.append(st.session_state.gn)
                                            else:
                                                gp.append(x)

                                    # Resultado final
                                    df_pp['Gradiente de Pressão de Poros (lb/gal)'] = gp

                                    df_pp['Gradiente de Pressão de Poros Médio (lb/gal)'] = (
                                        df_pp['Gradiente de Pressão de Poros (lb/gal)']
                                        .rolling(window=20, min_periods=1)
                                        .mean()
                                    )

                                    # Ajusta valores se ultrapassarem variação máxima fpp
                                    for i in range(1, len(df_pp)):
                                        if df_pp.loc[i, 'Profundidade (m)'] >= st.session_state.anormal:
                                            anterior = df_pp.loc[
                                                i - 1, 'Gradiente de Pressão de Poros Médio (lb/gal)']
                                            atual = df_pp.loc[i, 'Gradiente de Pressão de Poros Médio (lb/gal)']
                                            if atual > anterior + st.session_state.fpp:
                                                df_pp.loc[
                                                    i, 'Gradiente de Pressão de Poros Médio (lb/gal)'] = anterior + st.session_state.fpp

                                    if "spp" not in st.session_state:
                                        st.session_state.spp = True
                                    try:
                                        if st.session_state.spp:
                                            df_pp['Gradiente de Pressão de Poros Suavizado (lb/gal)'] = suavizar_2(df_pp
                                                                       [
                                                                           'Profundidade (m)'],
                                                                       df_pp[
                                                                           'Gradiente de Pressão de Poros Médio (lb/gal)'],
                                                                       df_pp[
                                                                           'Gradiente de Pressão de Poros Médio (lb/gal)'])
                                    except Exception as e:
                                        pass

                                    df_pp.insert(
                                        loc=9,
                                        column='Pressão de Poros (psi)',
                                        value=0.1704 * df_pp['Gradiente de Pressão de Poros Médio (lb/gal)'] *
                                              df_pp['Profundidade (m)']
                                    )

                                    if st.session_state.boyance == 'Sim':
                                        df_pp["FPR_efetivo"] = np.nan
                                        if st.session_state.n_boyance == 1:
                                            df_pp["FPR_efetivo"] = st.session_state.get("fpr_0")
                                        else:
                                            for i in range(st.session_state.n_boyance):

                                                fpr = st.session_state.get(f"fpr_{i}")
                                                prof_ini = st.session_state.get(f"prof_inicial_{i}")
                                                prof_fim = st.session_state.get(f"prof_final_{i}")

                                                # Validações básicas
                                                if fpr is None:
                                                    continue
                                                if prof_ini is None or prof_fim is None:
                                                    continue

                                                # Máscara de profundidade
                                                mask = (
                                                        (df_pp["Profundidade (m)"] >= prof_ini) &
                                                        (df_pp["Profundidade (m)"] <= prof_fim)
                                                )

                                                df_pp.loc[mask, "FPR_efetivo"] = fpr

                                        # Preenche eventuais lacunas
                                        df_pp["FPR_efetivo"] = (
                                            df_pp["FPR_efetivo"]
                                            .ffill()
                                            .fillna(st.session_state.get("fpr_0"))
                                        )

                                        incremento = (
                                                0.1704
                                                * df_pp["FPR_efetivo"]
                                                * (df_pp["Profundidade (m)"] - df_pp["Profundidade (m)"].shift(1))
                                        )
                                        if not st.session_state.s_gr:
                                            x = df_pp["Perfil Raio Gama"]
                                        else:
                                            x = df_pp["Raio Gama Suavizado"]
                                        df_pp.insert(loc=12,
                                                     column='Formação',
                                                     value=np.where(x < df_pp["LBF_calc"],
                                                                    "Formação Permeável",
                                                                    "Formação Impermeável")
                                                     )

                                        topo_permeavel = (
                                                (df_pp["Formação"] == "Formação Permeável") &
                                                (df_pp["Formação"].shift(1) != "Formação Permeável")
                                        )

                                        # Insere a coluna na posição correta
                                        df_pp.insert(loc=13, column='Pressão Boyance (TA = BF)', value=np.nan)
                                        # Identifica o topo das camadas permeáveis
                                        topo_permeavel = (
                                                (df_pp["Formação"] == "Formação Permeável") &
                                                (df_pp["Formação"].shift(1) != "Formação Permeável")
                                        )

                                        if not st.session_state.spp:
                                            y = df_pp['Gradiente de Pressão de Poros Médio (lb/gal)']
                                        else:
                                            y = df_pp['Gradiente de Pressão de Poros Suavizado (lb/gal)']

                                        # Valor inicial no topo da camada
                                        df_pp.loc[topo_permeavel, 'Pressão Boyance (TA = BF)'] = (
                                                y.shift(1)
                                                * 0.1704
                                                * df_pp["Profundidade (m)"]
                                        )

                                        # Cria ID de camada permeável
                                        id_camada = topo_permeavel.cumsum()

                                        # Soma acumulada linha a linha dentro de cada camada permeável
                                        mask_perm = df_pp["Formação"] == "Formação Permeável"

                                        serie_ta_bf = (
                                            df_pp.loc[mask_perm]
                                            .groupby(id_camada[mask_perm], group_keys=False)
                                            .apply(
                                                lambda g: g['Pressão Boyance (TA = BF)'].iloc[0] + incremento.loc[
                                                    g.index].cumsum()
                                            )
                                        )

                                        serie_ta_bf = serie_ta_bf.sort_index()
                                        df_pp.loc[serie_ta_bf.index, 'Pressão Boyance (TA = BF)'] = serie_ta_bf

                                        # Calcula a boyance normalmente
                                        boyance_calc = np.where(
                                            df_pp["Formação"] == "Formação Impermeável",
                                            y,
                                            df_pp["Pressão Boyance (TA = BF)"] / (0.1704 * df_pp["Profundidade (m)"])
                                        )

                                        # Converte para Series para facilitar o tratamento
                                        boyance_calc = pd.Series(boyance_calc, index=df_pp.index)

                                        # Onde for NaN ou None, substitui por y
                                        boyance_calc = boyance_calc.fillna(y)

                                        # Insere no DataFrame
                                        df_pp.insert(
                                            loc=14,
                                            column='Boyance (lb/gal) (TA = BF)',
                                            value=boyance_calc
                                        )

                                        df_pp.insert(loc=15, column='Pressão Boyance (BA = TF)', value=np.nan)

                                        # Identifica a BASE das camadas permeáveis
                                        base_permeavel = (
                                                (df_pp["Formação"] == "Formação Permeável") &
                                                (df_pp["Formação"].shift(-1) != "Formação Permeável")
                                        )

                                        # Valor inicial na base (igual ao valor da formação impermeável abaixo)
                                        idx = df_pp.index[base_permeavel]

                                        df_pp.loc[idx, 'Pressão Boyance (BA = TF)'] = np.where(
                                            df_pp.loc[idx, "Formação"].shift(-1) == "Formação Impermeável",
                                            df_pp.loc[idx, 'Pressão de Poros (psi)'].shift(-1),
                                            df_pp.loc[idx, 'Pressão de Poros (psi)']
                                        )

                                        # Cria ID de camada permeável (de cima para baixo)
                                        id_camada = (
                                                (df_pp["Formação"] == "Formação Permeável") &
                                                (df_pp["Formação"].shift(1) != "Formação Permeável")
                                        ).cumsum()

                                        # Calcula a pressão de baixo para cima dentro de cada camada permeável
                                        mask_perm = df_pp["Formação"] == "Formação Permeável"

                                        serie_ba_tf = (
                                            df_pp.loc[mask_perm]
                                            .groupby(id_camada[mask_perm], group_keys=False)
                                            .apply(
                                                lambda g: (
                                                        g['Pressão Boyance (BA = TF)'].iloc[-1]
                                                        - incremento.loc[g.index][::-1].cumsum()[::-1]
                                                )
                                            )
                                        )

                                        serie_ba_tf = serie_ba_tf.sort_index()
                                        df_pp.loc[serie_ba_tf.index, 'Pressão Boyance (BA = TF)'] = serie_ba_tf

                                        # Garante ordenação por profundidade
                                        df_pp = df_pp.sort_values("Profundidade (m)").reset_index(drop=True)

                                        # Flag da formação impermeável
                                        impermeavel = df_pp["Formação"] == "Formação Impermeável"

                                        # Verifica se existe impermeável acima (excluindo a linha atual)
                                        impermeavel_acima = impermeavel.shift(fill_value=False).cumsum() > 0

                                        df_pp.insert(
                                            loc=16,
                                            column='Boyance (lb/gal) (BA = TF)',
                                            value=np.where(
                                                (~impermeavel) & (impermeavel_acima),
                                                df_pp["Pressão Boyance (BA = TF)"] / (
                                                            0.1704 * df_pp["Profundidade (m)"]),
                                                y
                                            )
                                        )


                                else:
                                    st.warning('Funcionalidade ainda não disponível', icon="⚠️")

                                st.form_submit_button("Calcular Gradiente de Pessão de Poros", use_container_width=True,type='primary')

                    # GRÁFICO DO GRADIENTE DE PRESSÃO DE POROS
                    with coluna3:
                        with st.container(border=True):
                            if "idg" not in st.session_state:
                                st.session_state.idg = 'Não'

                            st.session_state.fig_pp = plt.figure(figsize=(8, 10))

                            if st.session_state.idg == 'Sim':
                                # === COM coluna de idade ===
                                gs = gridspec.GridSpec(
                                    1, 4,
                                    width_ratios=[0.1, 0.2, 0.21, 1],
                                    wspace=0
                                )

                                ax_idade = st.session_state.fig_pp.add_subplot(gs[0])
                                ax1 = st.session_state.fig_pp.add_subplot(gs[1], sharey=ax_idade)

                                ax_gap = st.session_state.fig_pp.add_subplot(gs[2])
                                ax_gap.axis('off')

                                ax = st.session_state.fig_pp.add_subplot(gs[3], sharey=ax_idade)

                                idade_formacao(ax_idade, st.session_state.df_idade, df_pp['Profundidade (m)'].max()+100)

                                # evita duplicar rótulos de profundidade
                                plt.setp(ax1.get_yticklabels(), visible=False)
                                plt.setp(ax.get_yticklabels(), visible=False)

                            else:
                                # === SEM coluna de idade ===
                                gs = gridspec.GridSpec(
                                    1, 3,
                                    width_ratios=[0.2, 0.21, 1],
                                    wspace=0
                                )

                                ax1 = st.session_state.fig_pp.add_subplot(gs[0])
                                ax_gap = st.session_state.fig_pp.add_subplot(gs[1])
                                ax_gap.axis('off')

                                ax = st.session_state.fig_pp.add_subplot(gs[2], sharey=ax1)

                                plt.setp(ax.get_yticklabels(), visible=False)

                            x = len(df_pp['Profundidade (m)']) - len(df['Profundidade'])
                            st.write("")

                            def reset_config():
                                if st.session_state.ogp == "Gradiente (lb/gal)":
                                    st.session_state.x_min_pp = 7
                                    st.session_state.x_max_pp = 23
                                    st.session_state.x_step = 1
                                else:
                                    st.session_state.x_min_pp = 0
                                    st.session_state.x_max_pp = int(
                                        st.session_state.ext_df['Pressão de Sobrecarga (psi)'].max()) + 200
                                    st.session_state.x_step = 500
                                st.session_state.y_min_pp = 0
                                st.session_state.y_max_pp = int(df_pp['Profundidade (m)'].max()) + 100

                                st.session_state.y_step = 200

                            with st.expander("Configurações do Gráfico", expanded=False):
                                st.segmented_control("***Opção de Gráfico***",
                                                     ['Gradiente (lb/gal)', 'Pressão (psi)'],
                                                     selection_mode="single",
                                                     default='Gradiente (lb/gal)', key='ogp',
                                                     width="stretch")
                                colu1, colu2 = st.columns(2)
                                with colu1:
                                    st.checkbox('Suavizar Pressão de Poros', key="spp", value=True)
                                    st.checkbox('Suavizar Sônico', key="suav_s", value=False)
                                with colu2:
                                    st.checkbox('Sobrecarga', key="grafpp", value=True)
                                    st.checkbox('Suavizar Raio Gama', key="s_gr", value=False)
                                    st.session_state.ss = True

                                st.selectbox("Visualizar peso do fluido planejado", ['Não', 'Sim'], key='fpl')
                                st.selectbox("Visualizar peso do fluido executado", ['Não', 'Sim'], key='fex', index=1)
                                st.number_input('Alfa', step=1., format='%f',
                                                key='alfa_pp',
                                                value=4., min_value=1., max_value=20.)
                                st.selectbox("Boyance Ativado", ['Não', 'Sim'], key='boyance')
                                st.number_input('Filtro da Pressão de Poros', step=0.01, format='%f',
                                                key='fpp',
                                                value=0.01, min_value=0.001, max_value=1.0)
                                st.number_input('Nível de suavização', step=0.1, format='%f', key='frac', value=0.1)
                                st.number_input('Gaussiano', step=1., format='%f', key='gauss', value=50.)
                                st.number_input("Eixo X - mínimo", value=7, step=1, key="x_min_pp")
                                st.number_input("Eixo X - máximo", value=23, step=1, key="x_max_pp")
                                st.number_input("Passo do eixo X", value=1, step=1, key="x_step")

                                st.number_input("Eixo Y - mínimo", value=0,
                                                step=100, key="y_min_pp")
                                st.number_input("Eixo Y - máximo",
                                                value=int(df_pp['Profundidade (m)'].max()) + 100,
                                                step=100, key="y_max_pp")
                                st.number_input("Passo do eixo Y", value=200, step=50, key="y_step")
                                # Botão de reset com callback
                                st.button("Resetar Eixos - Gradiente de Pressão de Poros",
                                          on_click=reset_config,
                                          type="primary", use_container_width=True)
                            try:
                                if st.session_state.ogp == "Gradiente (lb/gal)":
                                    st.session_state.oep = df_pp[
                                        'Gradiente de Pressão de Poros Médio (lb/gal)']
                                    st.session_state.oepl = 'G. de Pressão de Poros'
                                else:
                                    st.session_state.oep = df_pp['Pressão de Poros (psi)']
                                    st.session_state.oepl = 'P. de Poros'
                            except Exception as e:
                                pass

                            # idx = df_pp.index[df_pp["Profundidade (m)"] > st.session_state.es][0]
                            if not st.session_state.spp:
                                try:
                                    ax.plot(st.session_state.oep, df_pp['Profundidade (m)'], color='orange',
                                            linestyle='-',
                                            linewidth=2, label=st.session_state.oepl)

                                except Exception as e:
                                    pass

                            else:
                                if st.session_state.ogp == "Gradiente (lb/gal)":
                                    ax.plot(df_pp['Gradiente de Pressão de Poros Suavizado (lb/gal)'],
                                            df_pp['Profundidade (m)'], color='blue', linestyle='-',
                                            linewidth=2,
                                            label=st.session_state.oepl)
                                else:
                                    ax.plot(df_pp['Pressão de Poros Suavizado (psi)'],
                                            df_pp['Profundidade (m)'],
                                            color='blue',
                                            linestyle='-',
                                            linewidth=2, label=st.session_state.oepl)

                            if st.session_state.boyance == 'Sim':
                                try:
                                    selecoes = st.session_state.o_boyance

                                    if 'Topo do arenito = Base do Folhelho' in selecoes:
                                        ax.plot(
                                            df_pp['Boyance (lb/gal) (TA = BF)'],
                                            df_pp['Profundidade (m)'],
                                            color='red',
                                            linestyle='-',
                                            linewidth=2,
                                            label='Boyance (lb/gal) (TA = BF)'
                                        )

                                    if 'Base do arenito = Topo do Folhelho' in selecoes:
                                        ax.plot(
                                            df_pp['Boyance (lb/gal) (BA = TF)'],
                                            df_pp['Profundidade (m)'],
                                            color='green',
                                            linestyle='-',
                                            linewidth=2,
                                            label='Boyance (lb/gal) (BA = TF)'
                                        )
                                except Exception as e:
                                    st.error(f"Erro: {e}")
                            if st.session_state.grafpp:
                                if st.session_state.ogp == "Gradiente (lb/gal)":
                                    ax.plot(df_pp['Gradiente de Sobrecarga (lb/gal)'][1:],
                                            df_pp['Profundidade (m)'][1:], color='black', linestyle='-',
                                            linewidth=2,
                                            label=st.session_state.oesl)
                                else:
                                    ax.plot(st.session_state.ext_df['Pressão de Sobrecarga (psi)'][1:],
                                            st.session_state.ext_df[
                                                'Profundidade em relação a mesa rotativa (m)'][1:],
                                            color='black', linestyle='-', linewidth=2,
                                            label=st.session_state.oesl)

                            if "df_mud" in st.session_state and isinstance(st.session_state["df_mud"], pd.DataFrame):
                                df_mud = st.session_state["df_mud"].copy()

                                mostrar_planejado = st.session_state.get("fpl", "Não") == "Sim"
                                mostrar_executado = st.session_state.get("fex", "Não") == "Sim"

                                # Planejado
                                if mostrar_planejado and df_mud["Peso do Fluido Planejado (lb/gal)"].notna().any():
                                    ax.plot(
                                        df_mud["Peso do Fluido Planejado (lb/gal)"],
                                        df_mud["Profundidade (m)"],
                                        linestyle="-",
                                        color="green",
                                        linewidth=2,
                                        label="Peso do Fluido (Planejado)",
                                        zorder=5
                                    )

                                # Executado
                                if mostrar_executado and df_mud["Peso do Fluido Executado (lb/gal)"].notna().any():
                                    ax.plot(
                                        df_mud["Peso do Fluido Executado (lb/gal)"],
                                        df_mud["Profundidade (m)"],
                                        linestyle="-",
                                        color="mediumvioletred",
                                        linewidth=2,
                                        label="Peso do Fluido (Executado)",
                                        zorder=5
                                    )

                            lito(
                                ax1,
                                df_pp,
                                profundidades,
                                litologias,
                                st.session_state.y_max_pp
                            )

                            # Configurações do gráfico
                            if st.session_state.ogp == "Gradiente (lb/gal)":
                                ax.set_title('Gradiente de Pressão de Poros (lb/gal)', fontsize=14,
                                             fontweight='bold')
                                ax.set_xlabel('Gradiente (ppg)', fontsize=12)
                                ax.set_ylabel('Profundidade (m)', fontsize=12)
                            else:
                                ax.set_title('Pressão de Poros (psi)', fontsize=14, fontweight='bold')
                                ax.set_xlabel('Pressão (psi)', fontsize=12)
                                ax.set_ylabel('Profundidade (m)', fontsize=12)
                            ax.invert_yaxis()
                            ax.tick_params(axis='y', which='both', left=True, labelleft=True)
                            ax.set_yticks(range(st.session_state.y_min_pp, st.session_state.y_max_pp,
                                                st.session_state.y_step))
                            ax.set_ylim(st.session_state.y_max_pp, st.session_state.y_min_pp)
                            ax.set_xticks(range(st.session_state.x_min_pp, st.session_state.x_max_pp,
                                                st.session_state.x_step))
                            ax.set_xlim(st.session_state.x_min_pp, st.session_state.x_max_pp)
                            ax.grid(True, linestyle='--', alpha=0.5)
                            ax.legend(
                                loc='upper right',
                                fontsize=8,
                                frameon=True,
                                shadow=True,
                                fancybox=True,
                                framealpha=1,
                                facecolor='white',
                                edgecolor='gray'
                            )
                            add_watermark(
                                ax,
                                logo_path="logo.png",
                                xy=(0.50, 0.5),
                                zoom=0.2,
                                alpha=0.2,
                                zorder=0
                            )

                            # plt.subplots_adjust(wspace=0.45)
                            # Exibe o gráfico no Streamlit
                            st.pyplot(st.session_state.fig_pp)

                    # TRENDING E LBF
                    with coluna2:
                        with st.container(border=True):
                            st.segmented_control("Gráficos", ['LBF', 'Trending'],selection_mode="single",default='LBF',
                                                 key='graf', width="stretch")

                            TRENDING_COLORS = [
                                "#0072B2",  # azul forte
                                "#009E73",  # verde-azulado
                                "#56B4E9",  # azul claro
                                "#00BFC4",  # ciano
                                "#1F78B4",  # azul médio
                            ]

                            if st.session_state.graf == 'Trending':
                                fig = plt.figure(figsize=(8, 10))

                                if st.session_state.idg == 'Sim':
                                    # ===== COM coluna de idade =====
                                    gs = gridspec.GridSpec(
                                        1, 4,
                                        width_ratios=[0.1, 0.2, 0.21, 1],
                                        wspace=0
                                    )

                                    ax_idade = fig.add_subplot(gs[0])
                                    ax1 = fig.add_subplot(gs[1], sharey=ax_idade)

                                    ax_gap = fig.add_subplot(gs[2])
                                    ax_gap.axis('off')

                                    ax = fig.add_subplot(gs[3], sharey=ax_idade)

                                    idade_formacao(ax_idade, st.session_state.df_idade, st.session_state.y_max_pp)

                                    # Evita duplicar rótulos de profundidade
                                    plt.setp(ax1.get_yticklabels(), visible=False)
                                    plt.setp(ax.get_yticklabels(), visible=False)

                                else:
                                    # ===== SEM coluna de idade =====
                                    gs = gridspec.GridSpec(
                                        1, 3,
                                        width_ratios=[0.18, 0.21, 1],
                                        wspace=0
                                    )

                                    ax1 = fig.add_subplot(gs[0])
                                    ax_gap = fig.add_subplot(gs[1])
                                    ax_gap.axis('off')

                                    ax = fig.add_subplot(gs[2], sharey=ax1)

                                    plt.setp(ax.get_yticklabels(), visible=False)

                                # salva a figura no session_state
                                st.session_state.fig1 = fig

                                lito(
                                    ax1,
                                    df_pp,
                                    profundidades,
                                    litologias,
                                    st.session_state.y_max_pp
                                )

                                # Perfil sônico (vermelho)
                                ax.semilogx(
                                    df_pp['Perfil sônico (µs/pé)'],
                                    df_pp['Profundidade (m)'],
                                    color='red',
                                    linewidth=2,
                                    label="Sônico"
                                )

                                if st.session_state.suav_s:
                                    ax.semilogx(df_pp['Perfil sônico suavizado (µs/pé)'],
                                                df_pp['Profundidade (m)'],
                                                color='blue', linestyle='-', linewidth=2,
                                                label="Sônico Suavizado")

                                if not st.session_state.s_gr:
                                    mask = df_pp['Perfil Raio Gama'] >= df_pp['LBF_calc']
                                    df_rg = df_pp[mask]
                                    ax.semilogx(df_rg['Perfil sônico (µs/pé)'],
                                                df_rg['Profundidade (m)'],
                                                marker='o', linestyle='None',
                                                color='cyan', markersize=1,
                                                label="Sônico ≥ LBF")
                                else:
                                    mask = df_pp['Raio Gama Suavizado'] >= df_pp['LBF_calc']
                                    df_rg = df_pp[mask]
                                    ax.semilogx(df_rg['Perfil sônico (µs/pé)'],
                                                df_rg['Profundidade (m)'],
                                                marker='o', linestyle='None',
                                                color='cyan', markersize=1,
                                                label="Sônico ≥ LBF")

                                # Loop em todos os Trendings
                                for idx, tr in enumerate(trendings):

                                    try:
                                        color = TRENDING_COLORS[idx % len(TRENDING_COLORS)]

                                        m = (np.log10(tr['s2'] / tr['s1'])) / (tr['pp2'] - tr['pp1'])
                                        prof = df_pp['Profundidade (m)']
                                        s_normal = tr['s1'] * 10 ** (m * (prof - tr['pp1']))

                                        if st.session_state.onshore:
                                            mask_base = prof.notna()  # todas as profundidades válidas
                                        else:
                                            mask_base = prof > st.session_state.lda

                                        if tr['prof_ini'] is not None and tr['prof_fim'] is not None:
                                            mask_tr = (prof >= tr['prof_ini']) & (prof <= tr['prof_fim'])
                                            mask_final = mask_base & mask_tr
                                        else:
                                            mask_final = mask_base

                                        ax.semilogx(
                                            s_normal[mask_final],
                                            prof[mask_final],
                                            linestyle='--',
                                            linewidth=3,
                                            color=color,
                                            label=f'Trending {idx + 1}'
                                        )

                                    except Exception:
                                        pass

                                ax.set_title('Sônico x Profundidade (m)', fontsize=14, fontweight='bold')
                                ax.set_xlabel('Perfil sônico (µs/pé)', fontsize=12)
                                ax.set_ylabel('Profundidade (m)', fontsize=12)
                                # Ticks do eixo X
                                xticks = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100,
                                          200, 300, 400, 500, 600, 700, 800, 900, 1000, 1100]
                                ax.set_xticks(xticks)
                                from matplotlib.ticker import LogLocator, NullFormatter, ScalarFormatter

                                # Major ticks → apenas potências de 10
                                ax.xaxis.set_major_locator(LogLocator(base=10.0))
                                ax.xaxis.set_major_formatter(ScalarFormatter())

                                # Minor ticks → subdivisões (20–90, 200–900)
                                ax.xaxis.set_minor_locator(
                                    LogLocator(base=10.0, subs=np.arange(2, 10) * 0.1)
                                )
                                ax.xaxis.set_minor_formatter(NullFormatter())  # sem números nos minors

                                # Grid
                                ax.grid(True, which='major', linestyle='--', alpha=0.6)
                                ax.grid(True, which='minor', linestyle=':', alpha=0.4)

                                # Limites do eixo Y
                                if st.session_state.onshore:
                                    x = len(df_pp) - len(df)
                                    if 0 <= x < len(df_pp):
                                        min_depth = df_pp['Profundidade (m)'].iloc[x] - 50
                                    else:
                                        min_depth = df_pp['Profundidade (m)'].iloc[0] - 50
                                else:
                                    if len(df_pp) > 2:
                                        min_depth = df_pp['Profundidade (m)'].iloc[2] - 50
                                    else:
                                        min_depth = df_pp['Profundidade (m)'].iloc[0] - 50

                                ax.invert_yaxis()
                                ax.tick_params(axis='y', which='both', left=True, labelleft=True)
                                ax.set_yticks(range(st.session_state.y_min_pp, st.session_state.y_max_pp,
                                                    st.session_state.y_step))
                                ax.set_ylim(st.session_state.y_max_pp, st.session_state.y_min_pp)
                                ax.grid(True, linestyle='--', alpha=0.5)
                                ax.legend(
                                    loc='upper right',
                                    fontsize=8,
                                    frameon=True,
                                    shadow=True,
                                    fancybox=True,
                                    framealpha=1,
                                    facecolor='white',
                                    edgecolor='gray'
                                )
                                add_watermark(
                                    ax,
                                    logo_path="logo.png",
                                    xy=(0.50, 0.5),
                                    zoom=0.2,
                                    alpha=0.2,
                                    zorder=0
                                )

                                plt.subplots_adjust(wspace=0.3)
                                st.pyplot(st.session_state.fig1)

                            else:
                                # Figura e eixos
                                fig = plt.figure(figsize=(8, 10))

                                if st.session_state.idg == 'Sim':
                                    # ===== COM coluna de idade =====
                                    gs = gridspec.GridSpec(
                                        1, 4,
                                        width_ratios=[0.1, 0.2, 0.21, 1],
                                        wspace=0
                                    )

                                    ax_idade = fig.add_subplot(gs[0])
                                    ax1 = fig.add_subplot(gs[1], sharey=ax_idade)

                                    ax_gap = fig.add_subplot(gs[2])
                                    ax_gap.axis('off')

                                    ax = fig.add_subplot(gs[3], sharey=ax_idade)

                                    idade_formacao(ax_idade, st.session_state.df_idade, st.session_state.y_max_pp)

                                    plt.setp(ax1.get_yticklabels(), visible=False)
                                    plt.setp(ax.get_yticklabels(), visible=False)

                                else:
                                    # ===== SEM coluna de idade =====
                                    gs = gridspec.GridSpec(
                                        1, 3,
                                        width_ratios=[0.18, 0.21, 1],
                                        wspace=0
                                    )

                                    ax1 = fig.add_subplot(gs[0])
                                    ax_gap = fig.add_subplot(gs[1])
                                    ax_gap.axis('off')

                                    ax = fig.add_subplot(gs[2], sharey=ax1)

                                    plt.setp(ax.get_yticklabels(), visible=False)

                                # salva no session_state
                                st.session_state.fig2 = fig

                                lito(
                                    ax1,
                                    df_pp,
                                    profundidades,
                                    litologias,
                                    st.session_state.y_max_pp
                                )

                                # Perfil Raio Gama (vermelho)
                                ax.plot(
                                    df['Perfil Raio Gama'],
                                    df['Profundidade'],
                                    color='red',
                                    linewidth=2,
                                    label="Perfil Raio Gama"
                                )

                                for idx, tr in enumerate(trendings):

                                    try:
                                        color = TRENDING_COLORS[idx % len(TRENDING_COLORS)]

                                        prof = df_pp['Profundidade (m)']

                                        prof_ref = (
                                            tr['prof_ini']
                                            if tr['prof_ini'] is not None
                                            else prof.min()
                                        )

                                        # 🔹 LBF ANCORADA NO PONTO INICIAL
                                        lbf_line = tr['inclbf'] * (prof - prof_ref) + tr['lbf']

                                        if st.session_state.onshore:
                                            mask_base = prof.notna()
                                        else:
                                            mask_base = prof > st.session_state.lda

                                        if tr['prof_ini'] is not None and tr['prof_fim'] is not None:
                                            mask_tr = (prof >= tr['prof_ini']) & (prof <= tr['prof_fim'])
                                            mask_final = mask_base & mask_tr
                                        else:
                                            mask_final = mask_base

                                        ax.plot(
                                            lbf_line[mask_final],
                                            prof[mask_final],
                                            linewidth=3,
                                            color=color,
                                            label=f'LBF {idx + 1}'
                                        )

                                    except Exception:
                                        pass

                                x = len(df_pp) - len(df)
                                if st.session_state.s_gr:
                                    if st.session_state.ex == 'Desativada':
                                        ax.plot(df_pp['Raio Gama Suavizado'], df_pp['Profundidade (m)'], color='blue',
                                                linestyle='-', linewidth=2, label="Perfil Raio Gama")
                                    else:
                                        ax.plot(df_pp['Raio Gama Suavizado'][x:], df_pp['Profundidade (m)'][x:], color='blue',
                                                linestyle='-', linewidth=2, label="Perfil Raio Gama")
                                ax.set_title('GR x Profundidade (m)', fontsize=14, fontweight='bold')
                                ax.set_xlabel('Perfil Raio Gama (GAPI)', fontsize=12)
                                ax.set_ylabel('Profundidade (m)', fontsize=12)


                                ax.set_xlim(0, df_pp["Perfil Raio Gama"].max() + 20)

                                ax.invert_yaxis()
                                ax.tick_params(axis='y', which='both', left=True, labelleft=True)
                                ax.set_yticks(range(st.session_state.y_min_pp, st.session_state.y_max_pp,
                                                    st.session_state.y_step))
                                ax.set_ylim(st.session_state.y_max_pp, st.session_state.y_min_pp)
                                ax.grid(True, linestyle='--', alpha=0.5)
                                ax.legend(
                                    loc='upper right',
                                    fontsize=8,
                                    frameon=True,
                                    shadow=True,
                                    fancybox=True,
                                    framealpha=1,
                                    facecolor='white',
                                    edgecolor='gray'
                                )
                                add_watermark(
                                    ax,
                                    logo_path="logo.png",
                                    xy=(0.50, 0.5),
                                    zoom=0.2,
                                    alpha=0.2,
                                    zorder=0
                                )

                                st.pyplot(st.session_state.fig2)

                else:
                    st.error('Preencha corretamente a aba "Gradiente de Sobrecarga"', icon="🚨")
            else:
                st.error('Por favor, insira um documento!', icon="🚨")

        # Ver Dataframes
        with tb[1]:
            if uploaded_file:
                try:
                    st.dataframe(df_pp, use_container_width=True, hide_index=True)
                except Exception as e:
                    pass

    # Estabilidade de Poço
    with tabs[4]:
        x = 0
        if x == 0:
            tabss = st.tabs(['Tensões em Volta do Poço', 'Gradiente de Fratura'])
            # Gradiente de Fratura
            with tabss[1]:
                tb = st.tabs(['Gradiente de Fratura', 'Tabela de Dados Calculados'])
                with tb[0]:
                    if uploaded_file:
                        if "tvp" not in st.session_state:
                            st.session_state.tvp = False
                        # try:
                        if all(value != 0 for value in [st.session_state.rtkb]) and all(value != 0 for value in [st.session_state.gn,
                                                                 st.session_state.anormal, st.session_state.expoente]):
                            colu1, colu2, colu3 = st.columns((1, 1, 1))
                            with colu1:
                                with st.container(border=True):
                                    with st.expander('Leak Off Test', expanded=True):
                                        colun1, colun2 = st.columns(2)
                                        with colun1:
                                            st.checkbox('Inserir LOTs', key='lot', value=False)
                                        with colun2:
                                            st.checkbox(
                                                'K auxiliar',
                                                key='auxiliar',
                                                value=False,
                                                help=(
                                                    "**Ponto Auxiliar (K=0.01)** \n\n"
                                                    "- Adiciona um ponto inicial em **1 m de profundidade** com um valor de K muito baixo (0.01).\n"
                                                    "- Útil quando há apenas **um dado de LOT**, garantindo que os cálculos de gradiente de fratura e interpolação comecem desde o topo do poço.\n"
                                                    "- Serve como **referência inicial**, evitando lacunas nos gráficos e melhorando a consistência dos resultados.\n"
                                                    "- Não altera os dados reais; é apenas um ponto de apoio para os cálculos."
                                                )
                                            )

                                        if st.session_state.lot:
                                            if 'add' not in st.session_state:
                                                st.session_state.add = []
                                            c1, c2 = st.columns(2)
                                            with c1:
                                                if st.button('Add :heavy_plus_sign:', key='add_bt',
                                                             use_container_width=True):
                                                    st.session_state.add.append(1)
                                            with c2:
                                                if st.button('Delete :x:', key='delete_bt', use_container_width=True):
                                                    if st.session_state.add:
                                                        st.session_state.add.pop(-1)
                                                    else:
                                                        st.toast('There are no action to delete!')

                                            col1, col2, col3 = st.columns((1, 1, 1))
                                            with col1:
                                                st.selectbox('LOT ou FIT 1', ('LOT', 'FIT'), key='r1_tipo')
                                                st.selectbox('LOT ou FIT 2', ('LOT', 'FIT'), key='r2_tipo')
                                                for i in range(len(st.session_state.add)):
                                                    st.selectbox(f'LOT ou FIT {i + 3}', ('LOT', 'FIT'),
                                                                 key=f'r_{i + 3}_tipo')
                                            with col2:
                                                st.number_input('Profundidade 1', step=1.0, format='%f', key='p1_proff',
                                                                min_value=0.0)
                                                st.number_input('Profundidade 2', step=1.0, format='%f', key='p2_proff',
                                                                min_value=0.0)
                                                for i in range(len(st.session_state.add)):
                                                    st.number_input(f'Profundidade {i + 3}', step=1.0, format='%f',
                                                                    min_value=0.0, key=f'p{i + 3}_proff')
                                            with col3:
                                                st.number_input('Peso Equivalente 1', step=1.0, format='%f', key='l1_leak',
                                                                min_value=0.0)
                                                st.number_input('Peso Equivalente 2', step=1.0, format='%f', key='l2_leak',
                                                                min_value=0.0)
                                                for i in range(len(st.session_state.add)):
                                                    st.number_input(f'Peso Equivalente {i + 3}', step=1.0, format='%f',
                                                                    min_value=0.0,
                                                                    key=f'l{i + 3}_leak')

                                            # ---- Captura e ordena as chaves corretamente ----

                                            tipo_keys = [k for k in st.session_state.keys() if k.endswith('_tipo')]
                                            tipo_keys = sorted(tipo_keys, key=lambda x: int(re.search(r'\d+', x).group()))
                                            tt = [st.session_state[k] for k in tipo_keys]

                                            prof_keys = [k for k in st.session_state.keys() if k.endswith('_proff')]
                                            prof_keys = sorted(prof_keys, key=lambda x: int(re.search(r'\d+', x).group()))
                                            pp = [st.session_state[k] for k in prof_keys]

                                            leak_keys = [k for k in st.session_state.keys() if k.endswith('_leak')]
                                            leak_keys = sorted(leak_keys, key=lambda x: int(re.search(r'\d+', x).group()))
                                            lt = [st.session_state[k] for k in leak_keys]

                                            st.session_state.tt = tt
                                            st.session_state.pp = pp
                                            st.session_state.lt = lt

                                        else:
                                            colunas_para_exibir = ["Nome", "Distância (km)", "Profundidade Vertical (m)",
                                                                   "Peso Eq. (lb/gal)"]
                                            df_filtrado = df_dentro_exibir[colunas_para_exibir]
                                            st.markdown("### LOTs de poços próximos")
                                            st.dataframe(df_filtrado, use_container_width=True, hide_index=True)
                                            pp = df_filtrado['Profundidade Vertical (m)'].tolist()
                                            lt = df_filtrado['Peso Eq. (lb/gal)'].tolist()
                                            st.session_state.pp = pp
                                            st.session_state.lt = lt
                                            tt = ["LOT"] * len(st.session_state.pp)
                                            st.session_state.tt = tt

                                        # ---- Funções auxiliares ----
                                        def get_closest_depth(df, target_depth):
                                            closest_index = (df['Profundidade (m)'] - target_depth).abs().idxmin()
                                            return df.iloc[closest_index]['Gradiente de Pressão de Poros Médio (lb/gal)']

                                        def get_closest_depth_value(target_depth):
                                            if st.session_state.onshore:
                                                df_ext = st.session_state.ext_df
                                                closest_index = (df_ext[
                                                                     'Profundidade em relação a mesa rotativa (m)'] - target_depth).abs().idxmin()
                                            else:
                                                df_ext = df
                                                closest_index = (df_ext['Profundidade'] - target_depth).abs().idxmin()
                                            return df_ext.iloc[closest_index]['Pressão de Sobrecarga (psi)']

                                        # --- Botão para gerar o DataFrame ---
                                        if st.button('Calibrar curva de K', use_container_width=True,
                                                     type='primary'):
                                            # Lista de profundidades e valores de LOT/FIT ajustados
                                            pp_all = st.session_state.pp
                                            lt_all_adjusted = [
                                                st.session_state.lt[i] + 0.5 if st.session_state.tt[i] == "FIT" else
                                                st.session_state.lt[i]
                                                for i in range(len(st.session_state.tt))
                                            ]

                                            # Monta DataFrame com todos os pontos (LOT e FIT ajustado)
                                            gf = pd.DataFrame({'Profundidade (m)': pp_all})
                                            gf['Gradiente de P. de poros (psi)'] = [get_closest_depth(df_pp, d) for d in
                                                                                    pp_all]
                                            gf['P. Poros (psi)'] = 0.1704 * gf['Profundidade (m)'] * gf[
                                                'Gradiente de P. de poros (psi)']
                                            gf['P. Absorção (psi)'] = 0.1704 * gf['Profundidade (m)'] * lt_all_adjusted
                                            gf['P. Sobrecarga (psi)'] = [get_closest_depth_value(d) for d in pp_all]
                                            gf['K'] = (gf['P. Absorção (psi)'] - gf['P. Poros (psi)']) / (
                                                    gf['P. Sobrecarga (psi)'] - gf['P. Poros (psi)'])

                                            # Se o usuário marcou o ponto auxiliar
                                            if st.session_state.auxiliar:
                                                nova_linha = pd.DataFrame([{
                                                    'Profundidade (m)': 1,
                                                    'Gradiente de P. de poros (psi)': None,
                                                    'P. Poros (psi)': None,
                                                    'P. Absorção (psi)': None,
                                                    'P. Sobrecarga (psi)': None,
                                                    'K': 0.01
                                                }])
                                                gf = pd.concat([nova_linha, gf], ignore_index=True)

                                            # Ordena e salva no session_state
                                            gf = gf.sort_values(by='Profundidade (m)').reset_index(drop=True)
                                            st.session_state.gf = gf
                                            st.session_state.edited_gf = gf.copy()

                                    with st.expander('Relação das tensões', expanded=True):
                                        st.markdown("### Variação de K com a Profundidade")

                                        # Inicializa edited_gf se não existir
                                        if 'edited_gf' not in st.session_state:
                                            # Se gf existir, copia, senão cria um DataFrame vazio com as colunas necessárias
                                            if 'gf' in st.session_state:
                                                st.session_state.edited_gf = st.session_state.gf.copy()
                                            else:
                                                st.session_state.edited_gf = pd.DataFrame(columns=['Profundidade (m)', 'K'])

                                        # --- Atualização automática do DataFrame de edição ---
                                        if 'gf' in st.session_state:
                                            if not st.session_state.edited_gf.equals(st.session_state.gf):
                                                st.session_state.edited_gf = st.session_state.gf.copy()

                                        # Editor de tabela
                                        st.session_state.edited_gf = st.data_editor(
                                            st.session_state.edited_gf[['Profundidade (m)', 'K']],
                                            use_container_width=True,
                                            hide_index=True
                                        )

                                        # Atualiza o DataFrame base temporariamente para os cálculos
                                        gf = st.session_state.edited_gf.copy()

                                        # Cálculo com tratamento de exceção
                                        try:
                                            b, log_a = np.polyfit(gf['K'], np.log(gf['Profundidade (m)']), 1)
                                            a = np.exp(log_a)
                                        except Exception as e:
                                            b, a = 1, 1
                                        st.session_state['a'] = a
                                        st.session_state['b'] = b

                                    with st.expander('Zonas de perda', expanded=True):
                                        with st.form('loss_form', border=False):
                                            st.markdown("### Inserir Zonas de Perda de Circulação")
                                            z = pd.DataFrame({
                                                'Profundidade da zona de perda (m)': [0.0],
                                                'Peso do fluido (lb/gal)': [0.0]
                                            })
                                            if 'zonas' not in st.session_state:
                                                st.session_state.zonas = []
                                            cols_to_check = ["Profundidade da zona de perda (m)", "Peso do fluido (lb/gal)"]
                                            st.session_state.edited_z = st.data_editor(z, hide_index=True,
                                                                                       num_rows='dynamic',
                                                                                       key='edited')
                                            if st.form_submit_button('Inserir Zonas de Perda', use_container_width=True,
                                                                     type='primary'):
                                                st.session_state.zonas = []
                                                if (st.session_state.edited_z[cols_to_check] != 0).all(axis=1).any():
                                                    for i, value in enumerate(
                                                            st.session_state.edited_z["Profundidade da zona de perda (m)"]):
                                                        st.session_state.zonas.append([value, st.session_state.edited_z[
                                                            "Peso do fluido (lb/gal)"][i]])

                            if st.session_state.spp:
                                df_f = pd.DataFrame({
                                    'Profundidade (m)': df_pp['Profundidade (m)'],
                                    'MD': df['MD'],
                                    'Gradiente de Sobrecarga (lb/gal)': df_pp['Gradiente de Sobrecarga (lb/gal)'],
                                    'Gradiente de Pressão de Poros (lb/gal)': df_pp[
                                        'Gradiente de Pressão de Poros Suavizado (lb/gal)']
                                })

                            else:
                                df_f = pd.DataFrame({
                                    'Profundidade (m)': df_pp['Profundidade (m)'],
                                    'MD': df['MD'],
                                    'Gradiente de Sobrecarga (lb/gal)': df_pp['Gradiente de Sobrecarga (lb/gal)'],
                                    'Gradiente de Pressão de Poros (lb/gal)': df_pp['Gradiente de Pressão de Poros Médio (lb/gal)']
                                })

                            # Cálculo de K
                            df_f['K'] = (np.log(df_f['Profundidade (m)']) - np.log(a)) / b

                            df_f['Gradiente de Fratura (lb/gal)'] = df_f.apply(
                                lambda row: row['Gradiente de Pressão de Poros (lb/gal)']
                                            + row['K'] * (row['Gradiente de Sobrecarga (lb/gal)'] - row[
                                    'Gradiente de Pressão de Poros (lb/gal)']),
                                axis=1
                            )

                            # Substitui valores negativos por 0
                            df_f['Gradiente de Fratura (lb/gal)'] = np.where(df_f['Gradiente de Fratura (lb/gal)'] < 0, 0,
                                                                             df_f['Gradiente de Fratura (lb/gal)'])

                            df_f.insert(
                                loc=4,
                                column='Pressão de Fratura (psi)',
                                value=0.1704 * df_f['Gradiente de Fratura (lb/gal)'] * df_f['Profundidade (m)']
                            )

                            for i in st.session_state.zonas:
                                index = (df_f['Profundidade (m)'] - i[0]).abs().idxmin()
                                df_f.loc[index, 'Gradiente de Fratura (lb/gal)'] = i[1]

                            with colu3:
                                with st.container(border=True):
                                    fig1 = plt.figure(figsize=(8, 9.95))

                                    if st.session_state.idg == 'Sim':
                                        # ===== COM coluna de idade =====
                                        gs = gridspec.GridSpec(
                                            1, 4,
                                            width_ratios=[0.1, 0.2, 0.21, 1],
                                            wspace=0
                                        )

                                        ax_idade = fig1.add_subplot(gs[0])
                                        ax1 = fig1.add_subplot(gs[1], sharey=ax_idade)

                                        ax_gap = fig1.add_subplot(gs[2])
                                        ax_gap.axis('off')

                                        ax = fig1.add_subplot(gs[3], sharey=ax_idade)

                                        idade_formacao(ax_idade, st.session_state.df_idade, st.session_state.y_max_pp)

                                        # evita duplicar rótulos de profundidade
                                        plt.setp(ax1.get_yticklabels(), visible=False)
                                        plt.setp(ax.get_yticklabels(), visible=False)

                                    else:
                                        # ===== SEM coluna de idade =====
                                        gs = gridspec.GridSpec(
                                            1, 3,
                                            width_ratios=[0.18, 0.21, 1],
                                            wspace=0
                                        )

                                        ax1 = fig1.add_subplot(gs[0])
                                        ax_gap = fig1.add_subplot(gs[1])
                                        ax_gap.axis('off')

                                        ax = fig1.add_subplot(gs[2], sharey=ax1)

                                        plt.setp(ax.get_yticklabels(), visible=False)

                                    with st.form("form_gradiente_fratura", border=False):
                                        submitted = st.form_submit_button("Calcular Gradiente de Fratura",
                                                                          use_container_width=True, type='primary')
                                        if submitted:
                                            st.session_state.fratura_calculada = True
                                            st.session_state.tvp = True

                                    def reset_config():
                                        if st.session_state.ogf == "Gradiente (lb/gal)":
                                            st.session_state.x_min_f = 7
                                            st.session_state.x_max_f = 23
                                            st.session_state.x_step_f = 1
                                        else:
                                            st.session_state.x_min_f = 0
                                            st.session_state.x_max_f = int(
                                                st.session_state.ext_df['Pressão de Sobrecarga (psi)'].max()) + 200
                                            st.session_state.x_step_f = 500
                                        st.session_state.y_min_f = 0
                                        st.session_state.y_max_f = int(df_f['Profundidade (m)'].max()) + 100
                                        st.session_state.y_step_f = 200

                                    lito(
                                        ax1,
                                        df_pp,
                                        profundidades,
                                        litologias,
                                        st.session_state.y_max_pp
                                    )
                                    # profundidades=st.session_state.get("profundidades", None),
                                    # litologias=st.session_state.get("litologias", None)

                                    with st.expander("Configurações do Gráfico", expanded=False):
                                        st.segmented_control("***Opção de Gráfico***", ['Gradiente (lb/gal)',
                                                                                        'Pressão (psi)'],
                                                             selection_mode="single",
                                                             default='Gradiente (lb/gal)', key='ogf', width="stretch")
                                        col1, col2 = st.columns((1, 1))
                                        with col1:
                                            st.checkbox('Sobrecarga', key="gras", value=False)
                                        with col2:
                                            st.checkbox('Pressão de Poros', key="grap", value=False)
                                        st.selectbox('Profundidade', ['MD', 'TVD'], key="t_prof_f")
                                        st.number_input("Eixo X - mínimo", value=7, step=1, key="x_min_f")
                                        st.number_input("Eixo X - máximo", value=23, step=1, key="x_max_f")
                                        st.number_input("Passo do eixo X", value=1, step=1, key="x_step_f")
                                        st.number_input("Eixo Y - mínimo", value=0,
                                                        step=100, key="y_min_f")
                                        st.number_input("Eixo Y - máximo", value=int(df_f['Profundidade (m)'].max()) + 100,
                                                        step=100, key="y_max_f")
                                        st.number_input("Passo do eixo Y", value=200, step=50, key="y_step_f")
                                        # Botão de reset com callback
                                        st.button("Resetar Eixos - Gradiente de Fratura", on_click=reset_config,
                                                  type="primary",
                                                  use_container_width=True)
                                    try:
                                        if st.session_state.ogf == "Gradiente (lb/gal)":
                                            st.session_state.oef = df_f['Gradiente de Fratura (lb/gal)']
                                            st.session_state.oefl = 'G. de Pressão de Poros'
                                        else:
                                            st.session_state.oef = df_f['Pressão de Fratura (psi)']
                                            st.session_state.oefl = 'P. de Fratura'
                                    except Exception as e:
                                        pass

                                    if st.session_state.get("fratura_calculada", False):
                                        # if st.session_state.lot:
                                        if st.session_state.onshore:
                                            if st.session_state.ogf == "Gradiente (lb/gal)":
                                                ax.plot(st.session_state.oef, df_f['Profundidade (m)'],
                                                        color='brown', linestyle='-', linewidth=2,
                                                        label=st.session_state.oefl)
                                            else:
                                                ax.plot(st.session_state.oef, df_f['Profundidade (m)'],
                                                        color='brown', linestyle='-', linewidth=2,
                                                        label=st.session_state.oefl)

                                        else:
                                            if st.session_state.ogf == "Gradiente (lb/gal)":
                                                ax.plot(st.session_state.oef[x:],
                                                        df_f['Profundidade (m)'][x:], color='brown', linestyle='-',
                                                        linewidth=2, label=st.session_state.oefl)
                                            else:
                                                ax.plot(st.session_state.oef[x:],
                                                        df_f['Profundidade (m)'][x:], color='brown', linestyle='-',
                                                        linewidth=2, label=st.session_state.oefl)

                                        if st.session_state.gras:
                                            if st.session_state.ogf == "Gradiente (lb/gal)":
                                                if st.session_state.onshore:
                                                    ax.plot(df_f['Gradiente de Sobrecarga (lb/gal)'],
                                                            df_f['Profundidade (m)'],
                                                            color='black', linestyle='-', linewidth=2,
                                                            label="G. Sobrecarga")
                                                else:
                                                    ax.plot(df_f['Gradiente de Sobrecarga (lb/gal)'][2:],
                                                            df_f['Profundidade (m)'][2:], color='black', linestyle='-',
                                                            linewidth=2, label="G. Sobrecarga")
                                            else:
                                                if st.session_state.onshore:
                                                    ax.plot(st.session_state.ext_df['Pressão de Sobrecarga (psi)'],
                                                            st.session_state.ext_df[
                                                                'Profundidade em relação a mesa rotativa (m)'],
                                                            color='black', linestyle='-', linewidth=2,
                                                            label="P. Sobrecarga")
                                                else:
                                                    ax.plot(st.session_state.ext_df['Pressão Sobrecarga (lb/gal)'][2:],
                                                            st.session_state.ext_df['Profundidade (m)'][2:], color='black',
                                                            linestyle='-',
                                                            linewidth=2, label="P. Sobrecarga")

                                        if st.session_state.grap:
                                            if st.session_state.ogf == "Gradiente (lb/gal)":
                                                if st.session_state.onshore:
                                                    if not st.session_state.spp:
                                                        ax.plot(df_f['Gradiente de Pressão de Poros (lb/gal)'],
                                                                df_f['Profundidade (m)'], color='orange', linestyle='-',
                                                                linewidth=2, label="G. Pressão de Poros")
                                                    else:
                                                        ax.plot(df_f['Gradiente de Pressão de Poros (lb/gal)'],
                                                                df_f['Profundidade (m)'], color='orange', linestyle='-',
                                                                linewidth=2, label="G. Pressão de Poros")
                                                else:
                                                    ax.plot(df_f['Gradiente de Pressão de Poros (lb/gal)'][x:],
                                                            df_f['Profundidade (m)'][x:], color='orange', linestyle='-',
                                                            linewidth=2, label="G. Pressão de Poros")
                                            else:
                                                if st.session_state.onshore:
                                                    if not st.session_state.spp:
                                                        ax.plot(df_pp['Pressão de Poros (psi)'],
                                                                df_pp['Profundidade (m)'], color='orange', linestyle='-',
                                                                linewidth=2, label="P. de Poros")
                                                    else:
                                                        ax.plot(df_pp['Pressão de Poros Suavizado (psi)'],
                                                                df_pp['Profundidade (m)'], color='orange', linestyle='-',
                                                                linewidth=2, label="P. de Poros")
                                                else:
                                                    ax.plot(df_pp['Pressão de Poros (lb/gal)'][x:],
                                                            df_pp['Profundidade (m)'][x:], color='orange', linestyle='-',
                                                            linewidth=2, label="P. de Poros")

                                            if "df_mud" in st.session_state and isinstance(st.session_state["df_mud"],
                                                                                           pd.DataFrame):
                                                df_mud = st.session_state["df_mud"].copy()

                                                # Executado
                                                if mostrar_executado and df_mud[
                                                    "Peso do Fluido Executado (lb/gal)"].notna().any():
                                                    ax.plot(
                                                        df_mud["Peso do Fluido Executado (lb/gal)"],
                                                        df_mud["Profundidade (m)"],
                                                        linestyle="-",
                                                        color="mediumvioletred",
                                                        linewidth=2,
                                                        label="Peso do Fluido (Executado)",
                                                        zorder=5
                                                    )

                                        if st.session_state.ogf == "Gradiente (lb/gal)":
                                            # Mantém os valores originais
                                            lot_x = [lt[i] for i in range(len(tt)) if tt[i] == "LOT"]
                                            lot_y = [pp[i] for i in range(len(tt)) if tt[i] == "LOT"]

                                        else:
                                            # Faz o cálculo com multiplicação
                                            lot_x = [lt[i] * 0.1704 * pp[i] for i in range(len(tt)) if tt[i] == "LOT"]
                                            lot_y = [pp[i] for i in range(len(tt)) if tt[i] == "LOT"]

                                        fit_x = [lt[i] for i in range(len(tt)) if tt[i] == "FIT"]
                                        fit_y = [pp[i] for i in range(len(tt)) if tt[i] == "FIT"]

                                        if lot_x and lot_y:
                                            ax.scatter(lot_x, lot_y, color='red', label="LOT's", zorder=5, marker='D', s=50)

                                        if fit_x and fit_y:
                                            ax.scatter(fit_x, fit_y, color='blue', label="FIT's", zorder=5, marker='^',
                                                       s=50)

                                        if 'edited_z' in st.session_state and not st.session_state.edited_z.empty:
                                            # pega os valores das colunas
                                            x_vals = st.session_state.edited_z["Peso do fluido (lb/gal)"]
                                            y_vals = st.session_state.edited_z["Profundidade da zona de perda (m)"]

                                            mask = (x_vals != 0) & (y_vals != 0)
                                            if mask.any():
                                                ax.scatter(
                                                    x_vals[mask],
                                                    y_vals[mask],
                                                    facecolors='brown',
                                                    edgecolors='black',
                                                    marker='o',
                                                    s=80,
                                                    label="Perdas de Circulação",
                                                    zorder=6
                                                )

                                        # Janela Operacional: Preenchimento entre pressão de poros e fratura
                                        if st.session_state.ogf == "Gradiente (lb/gal)":
                                            if st.session_state.grap:
                                                if st.session_state.onshore:
                                                    x1 = df_f['Gradiente de Pressão de Poros (lb/gal)']
                                                    x2 = df_f['Gradiente de Fratura (lb/gal)']
                                                    y = df_f['Profundidade (m)']
                                                else:
                                                    x1 = df_f['Gradiente de Pressão de Poros (lb/gal)'][x:]
                                                    x2 = df_f['Gradiente de Fratura (lb/gal)'][x:]
                                                    y = df_f['Profundidade (m)'][x:]

                                                ax.fill_betweenx(
                                                    y,
                                                    x1,
                                                    x2,
                                                    where=(x2 > x1),
                                                    color='lightgreen',
                                                    alpha=0.2,
                                                    label='Janela Operacional',
                                                    interpolate=True
                                                )

                                        else:
                                            # Usa curvas de pressão (psi)
                                            if st.session_state.spp:
                                                x1 = df_pp['Pressão de Poros Suavizado (psi)']
                                            else:
                                                x1 = df_pp['Pressão de Poros (psi)']
                                            x2 = df_f['Pressão de Fratura (psi)']
                                            y = df_f['Profundidade (m)']

                                            ax.fill_betweenx(
                                                y,
                                                x1,
                                                x2,
                                                where=(x2 > x1),
                                                color='lightgreen',
                                                alpha=0.2,
                                                label='Janela Operacional',
                                                interpolate=True
                                            )

                                    ax.set_title('Geopressões (lb/gal)', fontsize=14, fontweight='bold')
                                    ax.set_xlabel('Gradiente (ppg)', fontsize=12)
                                    ax.set_ylabel('Profundidade (m)', fontsize=12)
                                    ax.invert_yaxis()
                                    ax.set_ylim(st.session_state.y_max_f, st.session_state.y_min_f)
                                    ax.tick_params(axis='y', which='both', left=True, labelleft=True)
                                    ax.set_yticks(range(st.session_state.y_min_f, st.session_state.y_max_f,
                                                        st.session_state.y_step_f))
                                    ax.set_xticks(range(st.session_state.x_min_f, st.session_state.x_max_f,
                                                        st.session_state.x_step_f))
                                    ax.set_xlim(st.session_state.x_min_f, st.session_state.x_max_f)
                                    ax.grid(True, linestyle='--', alpha=0.5)
                                    ax.legend(
                                        loc='upper right',
                                        fontsize=8,
                                        frameon=True,
                                        shadow=True,
                                        fancybox=True,
                                        framealpha=1,
                                        facecolor='white',
                                        edgecolor='gray'
                                    )
                                    add_watermark(
                                        ax,
                                        logo_path="logo.png",
                                        xy=(0.50, 0.5),
                                        zoom=0.2,
                                        alpha=0.2,
                                        zorder=0
                                    )

                                    st.pyplot(fig1)

                            with colu2:
                                with st.container(border=True):

                                    # Botão de limpar dados e restaurar o DataFrame original
                                    if st.button('🔄 Limpar Dados - Gradiente de Fratura', use_container_width=True,
                                                 type='primary'):
                                        # Limpa variáveis específicas
                                        for key in list(st.session_state.keys()):
                                            if key.startswith(('r', 'p', 'l')) and key.endswith(
                                                    ('_prof', '_leak')) or key in ['add', 'gras', 'grap',
                                                                                   'fratura_calculada', 'tvp']:
                                                del st.session_state[key]

                                        # Restaura o DataFrame editável para os dados originais
                                        st.session_state.edited_gf = st.session_state.gf.copy()

                                        # Recarrega a página para refletir a limpeza
                                        st.rerun()

                                    st.markdown("### Método das Tensões Mínimas")

                                    fig2 = plt.figure(figsize=(8, 10))

                                    if st.session_state.idg == 'Sim':
                                        # ===== COM coluna de idade =====
                                        gs = gridspec.GridSpec(
                                            1, 4,
                                            width_ratios=[0.1, 0.2, 0.21, 1],
                                            wspace=0
                                        )

                                        ax_idade = fig2.add_subplot(gs[0])
                                        ax1 = fig2.add_subplot(gs[1], sharey=ax_idade)

                                        ax_gap = fig2.add_subplot(gs[2])
                                        ax_gap.axis('off')

                                        ax = fig2.add_subplot(gs[3], sharey=ax_idade)

                                        idade_formacao(ax_idade, st.session_state.df_idade, st.session_state.y_max_pp)

                                        # evita duplicar rótulos de profundidade
                                        plt.setp(ax1.get_yticklabels(), visible=False)
                                        plt.setp(ax.get_yticklabels(), visible=False)

                                    else:
                                        # ===== SEM coluna de idade =====
                                        gs = gridspec.GridSpec(
                                            1, 3,
                                            width_ratios=[0.18, 0.21, 1],
                                            wspace=0
                                        )

                                        ax1 = fig2.add_subplot(gs[0])
                                        ax_gap = fig2.add_subplot(gs[1])
                                        ax_gap.axis('off')

                                        ax = fig2.add_subplot(gs[2], sharey=ax1)

                                        plt.setp(ax.get_yticklabels(), visible=False)

                                    lito(
                                        ax1,
                                        df_pp,
                                        profundidades,
                                        litologias,
                                        st.session_state.y_max_pp

                                    )

                                    ax.plot(
                                        gf['K'], gf['Profundidade (m)'],
                                        color='blue', linestyle='None', linewidth=2,
                                        marker='o', markersize=10,
                                        markerfacecolor='black', markeredgecolor='red',
                                        label="K"
                                    )
                                    k_values = np.linspace(gf['K'].min(), gf['K'].max(), 200)
                                    depth_trend = a * np.exp(b * k_values)

                                    ax.plot(k_values, depth_trend, color='red', linestyle='--', linewidth=2,
                                            label='Linha de Tendência Exponencial de K')

                                    ax.set_title('K x Profundidade', fontsize=14, fontweight='bold')
                                    ax.set_xlabel('K', fontsize=12)
                                    ax.set_ylabel('Profundidade (m)', fontsize=12)
                                    ax.invert_yaxis()

                                    max_depth = int(st.session_state.df1['Profundidade'].max())
                                    ax.tick_params(axis='y', which='both', left=True, labelleft=True)
                                    ax.set_ylim(st.session_state.y_max_f, st.session_state.y_min_f)
                                    ax.set_yticks(range(st.session_state.y_min_f, st.session_state.y_max_f,
                                                        st.session_state.y_step_f))
                                    ax.grid(True, linestyle='--', alpha=0.5)
                                    ax.legend(
                                        loc='upper right',
                                        fontsize=8,
                                        frameon=True,
                                        shadow=True,
                                        fancybox=True,
                                        framealpha=1,
                                        facecolor='white',
                                        edgecolor='gray'
                                    )
                                    add_watermark(
                                        ax,
                                        logo_path="logo.png",
                                        xy=(0.50, 0.5),
                                        zoom=0.2,
                                        alpha=0.2,
                                        zorder=0
                                    )

                                    st.pyplot(fig2)

                            # depois de st.pyplot(fig1)
                            st.session_state.fig_fratura = fig1
                            st.session_state.df_f = df_f.copy()

                            # guarde também os parâmetros e flags importantes
                            st.session_state.gf = gf.copy()  # tabela K x profundidade (editada)
                            st.session_state.a_k = float(a) if a is not None else None
                            st.session_state.b_k = float(b) if b is not None else None
                            st.session_state.tt = tt
                            st.session_state.pp = pp
                            st.session_state.lt = lt
                            st.session_state.zonas_perda = st.session_state.get("zonas", [])

                            # depois de st.pyplot(fig2)
                            st.session_state.fig_k_prof = fig2

                            def salvar_matplotlib_para_pdf(fig, path, dpi=200):
                                fig.savefig(
                                    path,
                                    dpi=dpi,
                                    bbox_inches="tight",
                                    facecolor="white"
                                )



                        # except Exception as e:
                        else:
                            st.error('Preencha corretamente a aba "Gradiente de Pressão de Poros"', icon="🚨")

                    else:
                        st.error('Por favor, insira um documento!', icon="🚨")

                # Ver Dataframes
                with tb[1]:
                    if uploaded_file:
                        if all(value != 0 for value in [st.session_state.rtkb]):
                            if all(value != 0 for value in
                                   [st.session_state.gn, st.session_state.anormal, st.session_state.expoente]):
                                st.dataframe(df_f, use_container_width=True, hide_index=True)

            # Tensões em Volta do Poço
            with tabss[0]:
                if uploaded_file:
                    if st.session_state.rtkb != 0:
                        if "fs" not in st.session_state:
                            st.session_state.fs = 0.5

                        if all(value != 0 for value in [st.session_state.rtkb]) and all(
                                value != 0 for value in [st.session_state.gn,
                                                         st.session_state.anormal, st.session_state.expoente]):
                            tb = st.tabs(
                                ['Critério de Falha - Mohr Coulomb', 'Visualização 3D das Tensões', 'Configurações',
                                 'Dados Calculados'])
                            if st.session_state.onshore:
                                profundidade = df_pp['Profundidade (m)']
                                md = df['MD']
                                densidade = df_pp['Perfil de densidade (g/cm³)']
                                sonico = df_pp['Perfil sônico (µs/pé)']
                                grad_sobrecarga = df_pp['Gradiente de Sobrecarga (lb/gal)']
                            else:
                                profundidade = df['Profundidade']
                                md = df['MD']
                                densidade = df['Perfil de densidade']
                                sonico = df['Perfil sônico']
                                grad_sobrecarga = df['Gradiente de Sobrecarga (lb/gal)']

                            if "suavi_s" not in st.session_state:
                                st.session_state.suavi_s = False
                            if st.session_state.suavi_s:
                                sonico = suavizar(profundidade, sonico)

                            df_tvp = pd.DataFrame({
                                'Profundidade (m)': profundidade,
                                'MD': md,
                                'Perfil de densidade (g/cm³)': densidade,
                                'Perfil sônico (µs/pé)': sonico,
                                'Gradiente de Sobrecarga (lb/gal)': grad_sobrecarga
                            })

                            # ABA DE CONFIGURAÇÕES
                            with tb[2]:
                                # if uploaded_file:
                                c1, c2 = st.columns(2)
                                with c1:
                                    with st.container(border=True):
                                        coluna1, coluna2 = st.columns(2)
                                        st.markdown('### Configurações dos Cálculos')
                                        with coluna1:
                                            st.button("Inserir direções das tensões in situ", key="tis",
                                                      use_container_width=True, type='primary', on_click=tensoes)
                                        with coluna2:
                                            if st.button("Resetar direções das tensões in situ", key="resd",
                                                         use_container_width=True, type='primary'):
                                                st.session_state.direct = False
                                                dir_H = 51
                                                dir_h = 141
                                        if "direct" not in st.session_state:
                                            st.session_state.direct = False
                                        if st.session_state.direct:
                                            df_pp = pd.merge_asof(
                                                df_pp,
                                                st.session_state.tise,
                                                on="Profundidade (m)",
                                                direction="backward"  # pega o valor anterior ou igual
                                            )
                                            dir_H = df_pp['Direção SH']
                                            dir_h = df_pp['Direção SH'] + 90

                                        else:
                                            dir_H = 51
                                            dir_h = 141

                                        with coluna1:
                                            st.button(
                                                "Inserir relação das tensões horizontais",
                                                key="rel_tens_btn",
                                                use_container_width=True,
                                                type="primary",
                                                on_click=rel_hor
                                            )

                                        with coluna2:
                                            if st.button(
                                                    "Resetar relação das tensões horizontais",
                                                    key="reset_rel_tens",
                                                    use_container_width=True,
                                                    type="primary"
                                            ):
                                                st.session_state.rel_hor = False
                                                st.session_state.SH = 0.61
                                                st.session_state.Sh = 0.6

                                                if "rel_hor_df" in st.session_state:
                                                    del st.session_state.rel_hor_df

                                                st.rerun()

                                        st.write("")
                                        st.number_input("Raio do poço", value=1, step=1, key='rw')
                                        st.write("")
                                        st.number_input("Raio de investigação", value=1, step=1, key='r')
                                # =========================
                                # Configuração dos grupos
                                # =========================
                                PLOT_DEFAULTS_ORIGINAIS = {
                                    "jo": True,
                                    "suav_max_inf": True,
                                    "suav_min_sup": True,
                                    "ijo": True,
                                    "sjo": True,
                                    "li": False,
                                    "ls": False,
                                    "show_pp": False,
                                    "gs": False,
                                    "ti": False,
                                    "cia": False,
                                    "cib": False,
                                    "tsa": False,
                                    "tsb": False,
                                    "csa": False,
                                    "csb": False,
                                    "suav_cia": False,
                                    "suav_cib": False,
                                    "suav_tsa": False,
                                    "suav_tsb": False,
                                    "suav_csa": False,
                                    "suav_csb": False,
                                }

                                PLOT_ALL_TRUE = {
                                    "jo": True,
                                    "li": False,
                                    "ls": False,
                                    "show_pp": True,
                                    "gs": True,
                                    "ti": True,
                                    "cia": True,
                                    "cib": True,
                                    "tsa": True,
                                    "tsb": True,
                                    "csa": True,
                                    "csb": True,
                                    "ijo": False,
                                    "sjo": False,
                                }

                                SMOOTH_ALL_TRUE = {
                                    "suav_max_inf": True,
                                    "suav_min_sup": True,
                                    "suav_cia": True,
                                    "suav_cib": True,
                                    "suav_tsa": True,
                                    "suav_tsb": True,
                                    "suav_csa": True,
                                    "suav_csb": True,
                                }

                                def set_session_defaults(defaults: dict):
                                    for key, value in defaults.items():
                                        st.session_state[key] = value

                                def init_session_defaults(defaults: dict):
                                    for key, value in defaults.items():
                                        if key not in st.session_state:
                                            st.session_state[key] = value

                                CHECKBOX_GROUPS_COL1 = [
                                    (
                                        "##### Janela Operacional",
                                        [
                                            ("Janela Operacional", "jo", True),
                                            ("Limite Inferior da Janela Operacional", "li", False),
                                            ("Limite Superior da Janela Operacional", "ls", False),
                                            ("FS Inferior da Janela Operacional", "ijo", True),
                                            ("FS Superior da Janela Operacional", "sjo", True),
                                        ],
                                    ),
                                    (
                                        "##### Limites Inferiores",
                                        [
                                            ("Pressão de Poros", "show_pp", False),
                                            ("Gradiente de Sobrecarga", "gs", False),
                                            ("Tração Inferior", "ti", False),
                                            ("Comp Inferior σθA", "cia", False),
                                            ("Comp Inferior σθB", "cib", False),
                                        ],
                                    ),
                                    (
                                        "##### Limites Superiores",
                                        [
                                            ("Tração Superior (σθA)", "tsa", False),
                                            ("Tração Superior (σθB)", "tsb", False),
                                            ("Comp Superior σθA", "csa", False),
                                            ("Comp Superior σθB", "csb", False),
                                        ],
                                    ),
                                    (
                                        "##### Suavizar Perfil Sônico",
                                        [
                                            ("Suavizar Sônico", "suavi_s", False),
                                        ],
                                    ),
                                ]

                                CHECKBOX_GROUPS_COL2 = [
                                    (
                                        "##### Janela Operacional Suavizada",
                                        [
                                            ("Suavizar Limite Inferior da Janela Operacional", "suav_max_inf",
                                             True),
                                            ("Suavizar Limite Superior da Janela Operacional", "suav_min_sup",
                                             True),
                                        ],
                                    ),
                                    (
                                        "##### Limites Inferiores Suavizados",
                                        [
                                            ("Suavizar Comp Inferior σθA", "suav_cia", False),
                                            ("Suavizar Comp Inferior σθB", "suav_cib", False),
                                        ],
                                    ),
                                    (
                                        "##### Limites Superiores Suavizados",
                                        [
                                            ("Suavizar Tração Superior σθA", "suav_tsa", False),
                                            ("Suavizar Tração Superior σθB", "suav_tsb", False),
                                            ("Suavizar Comp Superior σθA", "suav_csa", False),
                                            ("Suavizar Comp Superior σθB", "suav_csb", False),
                                        ],
                                    ),
                                    (
                                        "##### Testes de Absorção (FIT/LOT)",
                                        [
                                            ("Testes de Absorção", "tab", True),
                                        ],
                                    ),
                                ]

                                def render_checkbox_groups(groups):
                                    for titulo, items in groups:
                                        with st.expander(titulo, expanded=False):
                                            for label, key, default in items:
                                                if key not in st.session_state:
                                                    st.session_state[key] = default
                                                st.checkbox(label, key=key)

                                # =========================
                                # Bloco c2 reescrito
                                # =========================
                                with c2:
                                    with st.container(border=True):
                                        st.markdown("### Curvas Plotadas no Gráfico")

                                        # inicialização segura dos estados
                                        init_session_defaults(PLOT_DEFAULTS_ORIGINAIS)
                                        init_session_defaults({
                                            "suav_max_inf": True,
                                            "suav_min_sup": True,
                                            "tab": True,
                                            "suavi_s": False,
                                        })

                                        if st.button(
                                                "Restaurar Padrões Originais",
                                                key="plot_all_o",
                                                use_container_width=True,
                                                type="primary"
                                        ):
                                            set_session_defaults(PLOT_DEFAULTS_ORIGINAIS)

                                        col1, col2 = st.columns(2)

                                        with col1:
                                            if st.button(
                                                    "Plotar todas as curvas",
                                                    key="plot_all",
                                                    use_container_width=True,
                                                    type="primary"
                                            ):
                                                set_session_defaults(PLOT_ALL_TRUE)

                                            render_checkbox_groups(CHECKBOX_GROUPS_COL1)

                                        with col2:
                                            if st.button(
                                                    "Suavizar todas as curvas",
                                                    key="smooth_all",
                                                    use_container_width=True,
                                                    type="primary"
                                            ):
                                                set_session_defaults(SMOOTH_ALL_TRUE)

                                            render_checkbox_groups(CHECKBOX_GROUPS_COL2)

                            @st.dialog("Adicionar Dados Direcionais")
                            def direcional():
                                st.markdown("## Dados Direcionais (automáticos do XLSM)")

                                if "df2" in st.session_state:
                                    st.markdown(f"**Trajetória usada:** {st.session_state.get('traj_modo', 'Planejada')}")
                                    st.dataframe(st.session_state.df2, use_container_width=True, hide_index=True)
                                else:
                                    st.warning("Trajetória não carregada ainda (aba 'Trajetória' do XLSM).")

                                if "df_interp" in st.session_state:
                                    st.success("df_interp já foi gerado automaticamente.")
                                else:
                                    st.warning("df_interp ainda não existe (verifique a aba 'Trajetória').")

                            def carregar_coesao():
                                with open("coesao.yaml", "r") as f:
                                    return yaml.safe_load(f)["Coesao"]

                            @st.dialog("Litologias")
                            def coe_lito():

                                # Carrega o YAML
                                coesao_dict = carregar_coesao()

                                # Inicializa o dataframe no session_state se não existir
                                if "dados_lito" not in st.session_state:
                                    st.session_state.dados_lito = pd.DataFrame({
                                        "Topo (m)": [],
                                        "Base (m)": [],
                                        "Litologia": [],
                                        "So (psi)": []
                                    })

                                st.markdown("### Adicionar Intervalos de Litologia")

                                col1, col2 = st.columns(2)

                                with col1:
                                    topo = st.number_input(
                                        "Profundidade Topo (m)",
                                        min_value=0.0,
                                        step=0.5,
                                        format="%.2f"
                                    )

                                with col2:
                                    base = st.number_input(
                                        "Profundidade Base (m)",
                                        min_value=0.0,
                                        step=0.5,
                                        format="%.2f"
                                    )

                                litologia = st.selectbox(
                                    "Litologia",
                                    ["Arenito", "Folhelho", "Calcário", "Siltito", "Conglomerado", "Halita"]
                                )

                                # Converter acentuados → YAML keys
                                chave_yaml = (
                                    litologia
                                    .replace("á", "a").replace("Á", "A")
                                    .replace("í", "i").replace("Í", "I")
                                    .replace("ó", "o").replace("Ó", "O")
                                    .replace("ç", "c").replace("Ç", "C")
                                )

                                so_lito = coesao_dict[chave_yaml]

                                # Botão para adicionar intervalo
                                if st.button("Adicionar Intervalo", key="add_lito", use_container_width=True,
                                             type="primary"):
                                    if base <= topo:
                                        st.warning("⚠ A Base deve ser maior que o Topo!")
                                    else:
                                        nova_linha = pd.DataFrame({
                                            "Topo (m)": [topo],
                                            "Base (m)": [base],
                                            "Litologia": [litologia],
                                            "So (psi)": [so_lito]
                                        })

                                        st.session_state.dados_lito = pd.concat(
                                            [st.session_state.dados_lito, nova_linha],
                                            ignore_index=True
                                        )

                                st.markdown("---")
                                st.markdown("### Intervalos Inseridos")

                                # Botão limpar tudo
                                if st.button("Limpar tudo", key="clear_lito", use_container_width=True, type="primary"):
                                    st.session_state.dados_lito = pd.DataFrame({
                                        "Topo (m)": [],
                                        "Base (m)": [],
                                        "Litologia": [],
                                        "So (psi)": []
                                    })

                                # Exibe tabela
                                st.dataframe(
                                    st.session_state.dados_lito,
                                    use_container_width=True,
                                    hide_index=True
                                )

                                # Botão para finalizar inserção
                                if st.button("Inserir valores", type="primary", use_container_width=True):
                                    st.rerun()

                            # ABA CRITÉRIO DE FALHA
                            with tb[0]:
                                c1, c2, c3 = st.columns((0.8, 1, 1))
                                #ENTRADA DE DADOS
                                with c1:
                                    with ((st.container(border=True))):
                                        with st.expander("Entrada de Dados", expanded=True):
                                            st.markdown('### Entrada de Dados')
                                            if "ts" not in st.session_state:
                                                st.session_state.ts = False

                                            if st.button("Considerar Litologia das Formações", use_container_width=True,
                                                         type="primary", key="lito_button"):
                                                coe_lito()

                                            with st.form("jop", border=False):
                                                col1, col2 = st.columns((0.8, 1))
                                                with col1:
                                                    st.number_input('Ângulo de fricção (Φ)', key='phi', value=30)
                                                    st.number_input('Limite de falha por tração', key='lft', value=0, disabled=True)
                                                    st.number_input('Profundidade (m)', key='m', value=700.00, format="%.2f")
                                                    st.number_input('Peso do fluido (lb/gal)', key='ppg', value=9.,
                                                                    format="%.2f", step=0.5)
                                                st.selectbox('Método de cálculo do UCS', ['Lacy', 'Mechpro'], key='ucs', index=1)
                                                st.selectbox("Profundidade", ['TVD', 'MD'], key="t_prof")

                                                if st.session_state.t_prof == "TVD":
                                                    st.session_state.y = df_tvp['Profundidade (m)']
                                                    prof_final = st.session_state.y_max_pp
                                                else:
                                                    st.session_state.y = df_tvp['MD']
                                                    prof_final = df_tvp['MD'].max() + 100

                                                profundidade_proxima = st.session_state.y.loc[
                                                    (st.session_state.y - st.session_state.m).abs().idxmin()
                                                ]

                                                if not st.session_state.spp:
                                                    gpo = df_pp['Gradiente de Pressão de Poros Médio (lb/gal)']
                                                else:
                                                    gpo = df_pp['Gradiente de Pressão de Poros Suavizado (lb/gal)']

                                                df_tvp.insert(
                                                    loc=4,
                                                    column='Gradiente de Pressão de Poros (lb/gal)',
                                                    value=gpo
                                                )

                                                if "df_interp" in st.session_state:
                                                    df_tvp["Inc"] = np.interp(st.session_state.y,
                                                                              st.session_state.df_interp["Profundidade"],
                                                                              st.session_state.df_interp["Inc (°)"])
                                                    df_tvp["Azi"] = np.interp(st.session_state.y,
                                                                              st.session_state.df_interp["Profundidade"],
                                                                              st.session_state.df_interp["Azi (°)"])

                                                else:
                                                    if "Inc" not in df_tvp.columns:
                                                        df_tvp.insert(loc=5, column="Inc", value=0)
                                                    if "Azi" not in df_tvp.columns:
                                                        df_tvp.insert(loc=6, column="Azi", value=0)

                                                if st.form_submit_button("Gerar Janela Operacional", use_container_width=True,
                                                                         type='primary'):
                                                    st.session_state.ts = True

                                        # ZONAS DE PERDA
                                        with st.expander('Zonas de perda', expanded=False):
                                            with st.form('los_form', border=False):
                                                st.markdown("### Perdas de Circulação")
                                                z = pd.DataFrame({
                                                    'Profundidade da zona de perda (m)': [0.0],
                                                    'Peso do fluido (lb/gal)': [0.0]
                                                })
                                                if 'zona2' not in st.session_state:
                                                    st.session_state.zona2 = []
                                                cols_to_check = ["Profundidade da zona de perda (m)",
                                                                 "Peso do fluido (lb/gal)"]
                                                st.session_state.edited_z2 = st.data_editor(z, hide_index=True,
                                                                                            num_rows='dynamic',
                                                                                            key='edited2')
                                                if st.form_submit_button('Inserir Zonas de Perda', use_container_width=True,
                                                                         type='primary'):
                                                    st.session_state.zona2.clear()
                                                    if (st.session_state.edited_z2[cols_to_check] != 0).all(axis=1).any():
                                                        for i, value in enumerate(
                                                                st.session_state.edited_z2[
                                                                    "Profundidade da zona de perda (m)"]):
                                                            st.session_state.zona2.append(
                                                                [value, st.session_state.edited_z2[
                                                                    "Peso do fluido (lb/gal)"][i]])

                                        # PROFUNDIDADE DAS PRISÕES DE COLUNA
                                        with st.expander('Prisões de Coluna', expanded=False):
                                            with st.form('prisoes_coluna_form', clear_on_submit=False, border=False):
                                                st.markdown("### Prisões de Coluna (TVD)")

                                                # DataFrame inicial
                                                z_prisao = pd.DataFrame({
                                                    'Profundidade da prisão (m)': [0.0]
                                                })

                                                # Editor de dados dinâmico
                                                edited_prisoes = st.data_editor(
                                                    z_prisao,
                                                    hide_index=True,
                                                    num_rows='dynamic',
                                                    key='edited_prisoes_coluna'
                                                )

                                                # Botão de envio
                                                submitted_prisoes = st.form_submit_button(
                                                    'Inserir Prisões de Coluna',
                                                    use_container_width=True,
                                                    type='primary'
                                                )

                                                if submitted_prisoes:
                                                    # Garante que a coluna exista
                                                    if 'Profundidade da prisão (m)' not in edited_prisoes.columns:
                                                        edited_prisoes['Profundidade da prisão (m)'] = 0.0

                                                    # Filtra apenas profundidades válidas
                                                    prisoes_df = edited_prisoes[
                                                        edited_prisoes['Profundidade da prisão (m)'] > 0
                                                        ].copy()

                                                    # Converte tipo
                                                    prisoes_df['Profundidade da prisão (m)'] = prisoes_df[
                                                        'Profundidade da prisão (m)'
                                                    ].astype(float)

                                                    # Salva no session_state
                                                    st.session_state.prisoes_coluna_df = prisoes_df

                                        # ===== Cálculos iniciais =====
                                        if st.session_state.ts:
                                            df_tvp.insert(
                                                loc=7,
                                                column='DTS',
                                                value=((1 / (((0.8042 * (
                                                        ((1000000 / df_tvp['Perfil sônico (µs/pé)']) / 3.281) / 1000)) -
                                                              0.8559) * 1000)) * 1000000) / 3.281
                                            )

                                            df_tvp.insert(
                                                loc=8,
                                                column='Poisson',
                                                value=(0.5 * (df_tvp['DTS'] / df_tvp['Perfil sônico (µs/pé)']) ** 2 - 1) / (
                                                        (df_tvp['DTS'] / df_tvp['Perfil sônico (µs/pé)']) ** 2 - 1)
                                            )
                                            if st.session_state.ucs == 'Lacy':
                                                df_tvp.insert(
                                                    loc=9,
                                                    column='G dinam (MMpsi)',
                                                    value=(1.34 * 10 ** 10 * df_tvp['Perfil de densidade (g/cm³)'] / (
                                                            df_tvp['DTS'] ** 2)) / 10 ** 6
                                                )
                                                df_tvp.insert(
                                                    loc=10,
                                                    column='E dinâmico (MMpsi)',
                                                    value=2 * df_tvp['G dinam (MMpsi)'] * (1 + df_tvp['Poisson'])
                                                )
                                                df_tvp.insert(
                                                    loc=11,
                                                    column='E estático (MMpsi)',
                                                    value=0.018 * (df_tvp['E dinâmico (MMpsi)'] ** 2) + 0.422 * df_tvp[
                                                        'E dinâmico (MMpsi)']
                                                )
                                                df_tvp.insert(
                                                    loc=12,
                                                    column='UCS (psi)',
                                                    value=(0.2787 * df_tvp['E estático (MMpsi)'] ** 2 + 2.458 * df_tvp[
                                                        'E estático (MMpsi)']) * 1000
                                                )
                                            else:
                                                df_tvp.insert(
                                                    loc=9,
                                                    column='Vsh',
                                                    value=0.5
                                                )
                                                df_tvp.insert(
                                                    loc=10,
                                                    column='UCS (psi)',
                                                    value=145.0377 * 1.9e-20 * (
                                                            1000 * df_tvp['Perfil de densidade (g/cm³)']) ** 2 * (
                                                                  304800 / df_tvp['Perfil sônico (µs/pé)']) ** 4 *
                                                          ((1 + df_tvp['Poisson']) / (1 - df_tvp['Poisson'])) ** 2 * (
                                                                  1 - 2 * df_tvp['Poisson']) * (
                                                                  1 + 0.79 * df_tvp['Vsh'])
                                                )
                                            # Calcula So original
                                            df_tvp["So (psi)"] = (df_tvp['UCS (psi)'] *
                                                                  (1 - np.sin(np.radians(st.session_state.phi)))) / \
                                                                 (2 * np.cos(np.radians(st.session_state.phi)))

                                            # Se existir dados de litologia, substitui intervalos
                                            if "dados_lito" in st.session_state and not st.session_state.dados_lito.empty:

                                                lito_df = st.session_state.dados_lito.copy()

                                                # Copia coluna original
                                                so = df_tvp["So (psi)"].copy()

                                                # Aplica intervalos topo–base
                                                for _, row in lito_df.iterrows():
                                                    topo = row["Topo (m)"]
                                                    base = row["Base (m)"]
                                                    so_val = row["So (psi)"]

                                                    mascara = (df_tvp["Profundidade (m)"] >= topo) & (
                                                                df_tvp["Profundidade (m)"] < base)
                                                    so.loc[mascara] = so_val

                                                df_tvp["So (psi)"] = so


                                            if "rel_hor_df" in st.session_state and not st.session_state.rel_hor_df.empty:
                                                df_tvp = pd.merge_asof(
                                                    df_tvp.sort_values("Profundidade (m)"),
                                                    st.session_state.rel_hor_df.sort_values("Profundidade (m)"),
                                                    on="Profundidade (m)",
                                                    direction="backward"
                                                )

                                                rel_sh = df_tvp["SH% Sobrecarga"].fillna(0.7)
                                                rel_shmin = df_tvp["Sh% Sobrecarga"].fillna(0.7)

                                                r1 = rel_sh * df_tvp['Gradiente de Sobrecarga (lb/gal)']
                                                r2 = rel_shmin * df_tvp['Gradiente de Sobrecarga (lb/gal)']
                                            else:
                                                r1 = 0.61 * df_tvp['Gradiente de Sobrecarga (lb/gal)']
                                                r2 = 0.6 * df_tvp['Gradiente de Sobrecarga (lb/gal)']

                                            # Inserindo a coluna no DataFrame
                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('So (psi)') + 1,
                                                column='SH (lb/gal)',
                                                value=r1
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('SH (lb/gal)') + 1,
                                                column='Direção SH',
                                                value=dir_H
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('Direção SH') + 1,
                                                column='Sh (lb/gal)',
                                                value=r2
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('Sh (lb/gal)') + 1,
                                                column='Direção Sh',
                                                value=dir_h
                                            )

                                            #Eixos invertidos, por isso foi somado 90°

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('Direção Sh') + 1,
                                                column='τxy',
                                                value=round(((np.cos(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) * np.cos(np.radians(df_tvp['Inc'])) *
                                                        -np.sin(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90))) * df_tvp['SH (lb/gal)'] +
                                                       (np.sin(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) * np.cos(np.radians(df_tvp['Inc'])) *
                                                        np.cos(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90))) * df_tvp['Sh (lb/gal)']),0)
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('τxy') + 1,
                                                column='τyz',
                                                value=round(((-np.sin(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) * np.cos(np.radians(df_tvp['Azi'])) *
                                                        np.sin(np.radians(df_tvp['Inc'])) * df_tvp['SH (lb/gal)']) +
                                                       (np.cos(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) * np.sin(np.radians(df_tvp['Azi'])) *
                                                        np.sin(np.radians(df_tvp['Inc'])) * df_tvp['Sh (lb/gal)'])),2)
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('τyz') + 1,
                                                column='τzx',
                                                value=round(((np.cos(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) * np.sin(np.radians(df_tvp['Inc'])) *
                                                        np.cos(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) * np.cos(np.radians(df_tvp['Inc'])) *
                                                        df_tvp['SH (lb/gal)']) + (np.sin(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) *
                                                        np.sin(np.radians(df_tvp['Inc'] )) * np.sin(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) *
                                                        np.cos(np.radians(df_tvp['Inc'])) * df_tvp['Sh (lb/gal)']) +
                                                        (np.cos(np.radians(df_tvp['Inc'])) * -np.sin(np.radians(df_tvp['Inc'])) *
                                                         df_tvp['Gradiente de Sobrecarga (lb/gal)'])),2)
                                            )

                                            thetaA = np.degrees(np.arctan(df_tvp['τyz'] / df_tvp['τzx']))
                                            thetaA[df_tvp['τzx'] == 0] = np.nan

                                            df_tvp.insert(
                                                df_tvp.columns.get_loc('τzx') + 1,
                                                'θA (°)',
                                                thetaA.ffill().fillna(0)
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('θA (°)') + 1,
                                                column='θB (°)',
                                                value=df_tvp['θA (°)'] + 90
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('θB (°)') + 1,
                                                column='τθa',
                                                value=round(2 * (df_tvp['τyz'] * np.cos(np.radians(df_tvp['θA (°)'])) - df_tvp[
                                                    'τzx'] * np.sin(np.radians(df_tvp['θA (°)']))),0)
                                            )

                                            # ==== Inicialmente Pw = 8.8 ====
                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('τθa') + 1,
                                                column='Pw',
                                                value=st.session_state.ppg
                                            )
                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('Pw') + 1,
                                                column='rw',
                                                value=st.session_state.rw
                                            )
                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('rw') + 1,
                                                column='r',
                                                value=st.session_state.r
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('r') + 1,
                                                column='σx',
                                                value=((np.cos(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) * np.cos(np.radians(df_tvp['Inc'])))**2 *
                                                       df_tvp['SH (lb/gal)'] + (np.sin(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) *
                                                       np.cos(np.radians(df_tvp['Inc'])))**2 * df_tvp['Sh (lb/gal)'] +
                                                       (np.sin(np.radians(df_tvp['Inc'])))**2 * df_tvp['Gradiente de Sobrecarga (lb/gal)'])
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('σx') + 1,
                                                column='σy',
                                                value=((-np.sin(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)))**2 * df_tvp['SH (lb/gal)'] +
                                                       (np.cos(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)))**2 * df_tvp['Sh (lb/gal)'])
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('σy') + 1,
                                                column='σr',
                                                value=(0.5 * (df_tvp['σx'] + df_tvp['σy']) * (
                                                        1 - ((df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)))) +
                                                      (0.5 * (df_tvp['σx'] - df_tvp['σy']) * (
                                                              1 - (4 * (df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)) +
                                                              (3 * (df_tvp['rw'] ** 4) / (df_tvp['r'] ** 4))) * np.cos(
                                                          np.radians((2 * df_tvp['θA (°)'])))) +
                                                      (df_tvp['τxy'] * (
                                                              1 - (4 * (df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)) + (
                                                              3 * (df_tvp['rw'] ** 4) /
                                                              (df_tvp['r'] ** 4)) * np.sin(
                                                          np.radians(2 * df_tvp['θA (°)']))) + (
                                                               (df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)) * df_tvp['Pw'])
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('σr') + 1,
                                                column='σθA',
                                                value=(0.5 * (df_tvp['σx'] + df_tvp['σy']) * (
                                                        1 + (df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2))) +
                                                      (-0.5 * (df_tvp['σx'] - df_tvp['σy']) * (
                                                              1 + 3 * ((df_tvp['rw'] ** 4) / (df_tvp['r'] ** 4))) *
                                                       np.cos(np.radians(2 * df_tvp['θA (°)']))) -
                                                      (df_tvp['τxy'] * (
                                                              1 + (3 * (df_tvp['rw'] ** 4) / (df_tvp['r'] ** 4))) * np.sin(
                                                          np.radians(2 * df_tvp['θA (°)']))) -
                                                      ((df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)) * df_tvp['Pw']
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('σθA') + 1,
                                                column='σθB',
                                                value=(0.5 * (df_tvp['σx'] + df_tvp['σy']) * (
                                                        1 + (df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2))) +
                                                      (-0.5 * (df_tvp['σx'] - df_tvp['σy']) * (
                                                              1 + 3 * ((df_tvp['rw'] ** 4) / (df_tvp['r'] ** 4))) *
                                                       np.cos(np.radians(2 * df_tvp['θB (°)']))) -
                                                      (df_tvp['τxy'] * (
                                                              1 + (3 * (df_tvp['rw'] ** 4) / (df_tvp['r'] ** 4))) * np.sin(
                                                          np.radians(2 * df_tvp['θB (°)']))) -
                                                      ((df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)) * df_tvp['Pw']
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('σθB') + 1,
                                                column='σa',
                                                value=((np.cos(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) * np.sin(np.radians(df_tvp['Inc'])))**2 *
                                                       df_tvp['SH (lb/gal)'] + (np.sin(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) *
                                                       np.sin(np.radians(df_tvp['Inc'])))**2 * df_tvp['Sh (lb/gal)'] +
                                                       (np.cos(np.radians(df_tvp['Inc'])))**2 * df_tvp['Gradiente de Sobrecarga (lb/gal)'])
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('σa') + 1,
                                                column='σr efetivo (psi)',
                                                value=(df_tvp['σr'] - df_tvp[
                                                    'Gradiente de Pressão de Poros (lb/gal)']) * 0.1704 *
                                                      df_tvp['Profundidade (m)']
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('σr efetivo (psi)') + 1,
                                                column='σθA efetivo (psi)',
                                                value=(df_tvp['σθA'] - df_tvp[
                                                    'Gradiente de Pressão de Poros (lb/gal)']) * 0.1704 *
                                                      df_tvp['Profundidade (m)']
                                            )

                                            df_tvp.insert(
                                                loc=df_tvp.columns.get_loc('σθA efetivo (psi)') + 1,
                                                column='σθB efetivo (psi)',
                                                value=(df_tvp['σθB'] - df_tvp[
                                                    'Gradiente de Pressão de Poros (lb/gal)']) * 0.1704 *
                                                      df_tvp['Profundidade (m)']
                                            )

                                            # ==== Pw que zera cada tensão efetiva ====
                                            coef_r = (df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)
                                            coef_t = -(df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)

                                            sigma_r_sem_pw = df_tvp['σr'] - coef_r * df_tvp['Pw']
                                            sigma_ta_sem_pw = df_tvp['σθA'] - coef_t * df_tvp['Pw']
                                            sigma_tb_sem_pw = df_tvp['σθB'] - coef_t * df_tvp['Pw']

                                            df_tvp['Tração Inferior'] = (df_tvp[
                                                                             'Gradiente de Pressão de Poros (lb/gal)'])

                                            df_tvp['Tração Superior (σθA)'] = (df_tvp[
                                                                                   'Gradiente de Pressão de Poros (lb/gal)'] - sigma_ta_sem_pw) / coef_t
                                            df_tvp['Tração Superior (σθB)'] = (df_tvp[
                                                                                   'Gradiente de Pressão de Poros (lb/gal)'] - sigma_tb_sem_pw) / coef_t

                                            # ================== Pw na falha por compressão ==================
                                            phi_rad = np.radians(st.session_state.phi)
                                            t = np.tan(phi_rad)

                                            coef_base = (df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)  # (rw^2/r^2)
                                            Kdepth = 0.1704 * df_tvp['Profundidade (m)']  # psi por (lb/gal)
                                            coef_psi = Kdepth * coef_base  # coeficiente do termo linear de Pw já em psi

                                            def sigma_r_sem_pw(theta_deg):
                                                th = np.radians(2 * theta_deg)
                                                return (0.5 * (df_tvp['σx'] + df_tvp['σy']) * (1 - coef_base)
                                                        + 0.5 * (df_tvp['σx'] - df_tvp['σy']) * (
                                                                1 - 4 * coef_base + 3 * (coef_base ** 2)) * np.cos(th)
                                                        + df_tvp['τxy'] * (
                                                                1 - 4 * coef_base + 3 * (coef_base ** 2)) * np.sin(th))

                                            def sigma_t_sem_pw(theta_deg):
                                                th = np.radians(2 * theta_deg)
                                                return (0.5 * (df_tvp['σx'] + df_tvp['σy']) * (1 + coef_base)
                                                        - 0.5 * (df_tvp['σx'] - df_tvp['σy']) * (
                                                                1 + 3 * (coef_base ** 2)) * np.cos(
                                                            th)
                                                        + df_tvp['τxy'] * (1 + 3 * (coef_base ** 2)) * np.sin(th))

                                            Ar_A = sigma_r_sem_pw(df_tvp['θA (°)'])
                                            At_A = sigma_t_sem_pw(df_tvp['θA (°)'])
                                            Ar_B = sigma_r_sem_pw(df_tvp['θB (°)'])
                                            At_B = sigma_t_sem_pw(df_tvp['θB (°)'])
                                            Pp = df_tvp['Gradiente de Pressão de Poros (lb/gal)']

                                            C_A_psi = Kdepth * (0.5 * (At_A + Ar_A) - Pp)
                                            R0_A_psi = Kdepth * (0.5 * (At_A - Ar_A))

                                            C_B_psi = Kdepth * (0.5 * (At_B + Ar_B) - Pp)
                                            R0_B_psi = Kdepth * (0.5 * (At_B - Ar_B))

                                            S0 = df_tvp['So (psi)']

                                            K_A = (t * C_A_psi + S0) ** 2 / (1 + t ** 2)
                                            K_B = (t * C_B_psi + S0) ** 2 / (1 + t ** 2)

                                            sqrtK_A = np.sqrt(np.maximum(K_A, 0))
                                            sqrtK_B = np.sqrt(np.maximum(K_B, 0))

                                            den_ok = np.where(coef_psi != 0, coef_psi, np.nan)

                                            Pw_A_1 = (R0_A_psi - sqrtK_A) / den_ok
                                            Pw_A_2 = (R0_A_psi + sqrtK_A) / den_ok
                                            Pw_B_1 = (R0_B_psi - sqrtK_B) / den_ok
                                            Pw_B_2 = (R0_B_psi + sqrtK_B) / den_ok

                                            Pw_A_inf = np.minimum(Pw_A_1, Pw_A_2)
                                            Pw_A_sup = np.maximum(Pw_A_1, Pw_A_2)
                                            Pw_B_inf = np.minimum(Pw_B_1, Pw_B_2)
                                            Pw_B_sup = np.maximum(Pw_B_1, Pw_B_2)

                                            # Adiciona as colunas ao DataFrame
                                            df_tvp.insert(loc=df_tvp.columns.get_loc('Tração Superior (σθB)') + 1,
                                                          column='Comp Inferior σθA', value=Pw_A_inf)
                                            df_tvp.insert(loc=df_tvp.columns.get_loc('Comp Inferior σθA') + 1,
                                                          column='Comp Superior σθA', value=Pw_A_sup)
                                            df_tvp.insert(loc=df_tvp.columns.get_loc('Comp Superior σθA') + 1,
                                                          column='Comp Inferior σθB', value=Pw_B_inf)
                                            df_tvp.insert(loc=df_tvp.columns.get_loc('Comp Inferior σθB') + 1,
                                                          column='Comp Superior σθB', value=Pw_B_sup)
                                            # ==================================================================
                                            with col2:
                                                opcao_tracao = st.radio(
                                                    "Selecione o tipo de falha:",
                                                    (
                                                        "Peso de Fluido Escolhido",
                                                        "Tração Inferior",
                                                        "Tração Superior σθA",
                                                        "Tração Superior σθB",
                                                        "Comp Inferior σθA",
                                                        "Comp Superior σθA",
                                                        "Comp Inferior σθB",
                                                        "Comp Superior σθB"
                                                    ), key='op'
                                                )

                                                if opcao_tracao == "Peso de Fluido Escolhido":
                                                    peso_fluido = st.session_state.ppg
                                                else:
                                                    coluna_ref = 'Profundidade (m)' if st.session_state.t_prof == "TVD" else 'MD'
                                                    linha = df_tvp.loc[df_tvp[coluna_ref] == profundidade_proxima].iloc[0]
                                                    if opcao_tracao == "Tração Inferior":
                                                        peso_fluido = linha['Tração Inferior']
                                                    elif opcao_tracao == "Tração Superior σθA":
                                                        peso_fluido = linha['Tração Superior (σθA)']
                                                    elif opcao_tracao == "Tração Superior σθB":
                                                        peso_fluido = linha['Tração Superior (σθB)']
                                                    elif opcao_tracao == "Comp Inferior σθA":
                                                        peso_fluido = linha['Comp Inferior σθA']
                                                    elif opcao_tracao == "Comp Superior σθA":
                                                        peso_fluido = linha['Comp Superior σθA']
                                                    elif opcao_tracao == "Comp Inferior σθB":
                                                        peso_fluido = linha['Comp Inferior σθB']
                                                    elif opcao_tracao == "Comp Superior σθB":
                                                        peso_fluido = linha['Comp Superior σθB']

                                                    # Atualiza Pw no dataframe
                                                    df_tvp['Pw'] = peso_fluido

                                                    # ==== Recalcular todas as tensões com o novo Pw ====
                                                    # σr
                                                    df_tvp['σr'] = ((0.5 * (df_tvp['σx'] + df_tvp['σy']) * (
                                                        1 - ((df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)))) +
                                                      (0.5 * (df_tvp['σx'] - df_tvp['σy']) * (
                                                              1 - (4 * (df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)) +
                                                              (3 * (df_tvp['rw'] ** 4) / (df_tvp['r'] ** 4))) * np.cos(
                                                          np.radians((2 * df_tvp['θA (°)'])))) +
                                                      (df_tvp['τxy'] * (
                                                              1 - (4 * (df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)) + (
                                                              3 * (df_tvp['rw'] ** 4) /
                                                              (df_tvp['r'] ** 4)) * np.sin(
                                                          np.radians(2 * df_tvp['θA (°)']))) + (
                                                               (df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)) * df_tvp['Pw']))

                                                    # σθA
                                                    df_tvp['σθA'] = ((0.5 * (df_tvp['σx'] + df_tvp['σy']) * (
                                                        1 + (df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2))) +
                                                      (-0.5 * (df_tvp['σx'] - df_tvp['σy']) * (
                                                              1 + 3 * ((df_tvp['rw'] ** 4) / (df_tvp['r'] ** 4))) *
                                                       np.cos(np.radians(2 * df_tvp['θA (°)']))) -
                                                      (df_tvp['τxy'] * (
                                                              1 + (3 * (df_tvp['rw'] ** 4) / (df_tvp['r'] ** 4))) * np.sin(
                                                          np.radians(2 * df_tvp['θA (°)']))) -
                                                      ((df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)) * df_tvp['Pw'])

                                                    # σθB
                                                    df_tvp['σθB'] = ((0.5 * (df_tvp['σx'] + df_tvp['σy']) * (
                                                        1 + (df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2))) +
                                                      (-0.5 * (df_tvp['σx'] - df_tvp['σy']) * (
                                                              1 + 3 * ((df_tvp['rw'] ** 4) / (df_tvp['r'] ** 4))) *
                                                       np.cos(np.radians(2 * df_tvp['θB (°)']))) -
                                                      (df_tvp['τxy'] * (
                                                              1 + (3 * (df_tvp['rw'] ** 4) / (df_tvp['r'] ** 4))) * np.sin(
                                                          np.radians(2 * df_tvp['θB (°)']))) -
                                                      ((df_tvp['rw'] ** 2) / (df_tvp['r'] ** 2)) * df_tvp['Pw'])

                                                    # σa
                                                    df_tvp['σa'] = ((np.cos(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) * np.sin(np.radians(df_tvp['Inc'])))**2 *
                                                       df_tvp['SH (lb/gal)'] + (np.sin(np.radians(df_tvp['Azi'] - df_tvp['Direção SH'] + 90)) *
                                                       np.sin(np.radians(df_tvp['Inc'])))**2 * df_tvp['Sh (lb/gal)'] +
                                                       (np.cos(np.radians(df_tvp['Inc'])))**2 * df_tvp['Gradiente de Sobrecarga (lb/gal)'])

                                                    # Recalcular tensões efetivas
                                                    df_tvp['σr efetivo (psi)'] = (df_tvp['σr'] - df_tvp[
                                                        'Gradiente de Pressão de Poros (lb/gal)']) * 0.1704 * df_tvp[
                                                                                     'Profundidade (m)']
                                                    df_tvp['σθA efetivo (psi)'] = (df_tvp['σθA'] - df_tvp[
                                                        'Gradiente de Pressão de Poros (lb/gal)']) * 0.1704 * df_tvp[
                                                                                      'Profundidade (m)']
                                                    df_tvp['σθB efetivo (psi)'] = (df_tvp['σθB'] - df_tvp[
                                                        'Gradiente de Pressão de Poros (lb/gal)']) * 0.1704 * df_tvp[
                                                                                      'Profundidade (m)']

                                                # Atualiza as variáveis para uso posterior
                                                linha = df_tvp.loc[st.session_state.y == profundidade_proxima].iloc[0]
                                                sr_ef = linha['σr efetivo (psi)']
                                                sta_ef = linha['σθA efetivo (psi)']
                                                stb_ef = linha['σθB efetivo (psi)']
                                                S0 = linha['So (psi)']

                                                phi_rad = np.radians(st.session_state.phi)
                                                t = np.tan(phi_rad)

                                                # Pontos A
                                                ra = abs(sr_ef - sta_ef) / 2
                                                Ca = (sr_ef + sta_ef) / 2
                                                Ra = abs(-np.radians(st.session_state.phi) * Ca - S0) / (
                                                    np.sqrt(np.radians(st.session_state.phi) ** 2 + 1 ** 2))

                                                # Pontos B
                                                Rb = abs(sr_ef - stb_ef) / 2
                                                Cb = (sr_ef + stb_ef) / 2
                                                Ra = abs(-np.radians(st.session_state.phi) * Cb - S0) / (
                                                    np.sqrt(np.radians(st.session_state.phi) ** 2 + 1 ** 2))

                                                df_tvp['Max Inferior'] = df_tvp[
                                                    ['Gradiente de Pressão de Poros (lb/gal)', 'Tração Inferior',
                                                     'Comp Inferior σθA', 'Comp Inferior σθB']].max(axis=1)

                                                # Coluna com o menor valor entre Tração Superior σθA, Tração Superior σθB, Comp Superior σθA e Comp Superior σθB
                                                df_tvp['Min Superior'] = df_tvp[
                                                    ['Tração Superior (σθA)', 'Tração Superior (σθB)', 'Comp Superior σθA',
                                                     'Comp Superior σθB']].min(axis=1)

                                                st.session_state.df_tvp = df_tvp

                                def criterio_disponivel(df):
                                    colunas = [
                                        'Tração Inferior',
                                        'Comp Inferior σθA',
                                        'Comp Inferior σθB',
                                        'Tração Superior (σθA)',
                                        'Tração Superior (σθB)',
                                        'Comp Superior σθA',
                                        'Comp Superior σθB'
                                    ]
                                    return all(c in df.columns for c in colunas)

                                # GRÁFICO - MOHR
                                with c2:
                                    with st.container(border=True):
                                        with st.expander('Critério de Falha - Mohr Coulomb', expanded=True):
                                            st.markdown('### Critério de Falha - Mohr Coulomb')
                                            if not criterio_disponivel(df_tvp):
                                                st.warning(
                                                    "⚠️ Calcule as tensões antes de avaliar o critério de falha.")

                                            if criterio_disponivel(df_tvp):
                                                df_suav = pd.DataFrame()
                                                df_suav['Profundidade (m)'] = df_tvp['Profundidade (m)']
                                                df_suav['MD'] = df_tvp['MD']
                                                df_suav['Gradiente de Sobrecarga (lb/gal)'] = df_tvp[
                                                    'Gradiente de Sobrecarga (lb/gal)']
                                                df_suav['Gradiente de Pressão de Poros (lb/gal)'] = df_tvp[
                                                    'Gradiente de Pressão de Poros (lb/gal)']
                                                df_suav['Tração Inferior'] = suavizar(df_tvp['Profundidade (m)'],
                                                                                      df_tvp['Tração Inferior'])
                                                df_suav['Comp Inferior σθA'] = suavizar(df_tvp['Profundidade (m)'],
                                                                                        df_tvp['Comp Inferior σθA'])
                                                df_suav['Comp Inferior σθB'] = suavizar(df_tvp['Profundidade (m)'],
                                                                                        df_tvp['Comp Inferior σθB'])
                                                df_suav['Max Inferior'] = df_suav[
                                                    ['Gradiente de Pressão de Poros (lb/gal)', 'Tração Inferior',
                                                     'Comp Inferior σθA', 'Comp Inferior σθB']].max(axis=1)
                                                df_suav['Tração Superior (σθA)'] = suavizar(
                                                    df_tvp['Profundidade (m)'],
                                                    df_tvp['Tração Superior (σθA)'])
                                                df_suav['Tração Superior (σθB)'] = suavizar(
                                                    df_tvp['Profundidade (m)'],
                                                    df_tvp['Tração Superior (σθB)'])
                                                df_suav['Comp Superior σθA'] = suavizar(df_tvp['Profundidade (m)'],
                                                                                        df_tvp['Comp Superior σθA'])
                                                df_suav['Comp Superior σθB'] = suavizar(df_tvp['Profundidade (m)'],
                                                                                        df_tvp['Comp Superior σθB'])
                                                df_suav['Min Superior'] = df_suav[
                                                    ['Tração Superior (σθA)', 'Tração Superior (σθB)',
                                                     'Comp Superior σθA',
                                                     'Comp Superior σθB']].min(axis=1)
                                                st.session_state.df_suav = df_suav

                                                colu1, colu2, colu3 = st.columns(3)
                                                with colu1:
                                                    if opcao_tracao == "Peso de Fluido Escolhido":
                                                        linha2 = \
                                                            linha = df_tvp.loc[st.session_state.y == profundidade_proxima].iloc[0]
                                                        max_inferior = max(linha2['Tração Inferior'],
                                                                           linha2['Comp Inferior σθA'],
                                                                           linha2['Comp Inferior σθB'])
                                                        min_superior = min(linha2['Tração Superior (σθA)'],
                                                                           linha2['Tração Superior (σθB)'],
                                                                           linha2['Comp Superior σθA'],
                                                                           linha2['Comp Superior σθB'])

                                                        if max_inferior < peso_fluido < min_superior:
                                                            st.markdown(
                                                                """
                                                                <div style="
                                                                    display: flex;
                                                                    justify-content: center;
                                                                    margin-top: 0px;
                                                                ">
                                                                    <div style="
                                                                        color: green;
                                                                        font-weight: bold;
                                                                        border: 2px solid black;
                                                                        border-radius: 10px;
                                                                        padding: 10px 10px;
                                                                    ">
                                                                        Poço Estável
                                                                    </div>
                                                                </div>
                                                                """,
                                                                unsafe_allow_html=True
                                                            )
                                                        else:
                                                            st.markdown(
                                                                """
                                                                <div style="
                                                                    display: flex;
                                                                    justify-content: center;
                                                                    margin-top: 3px;
                                                                ">
                                                                    <div style="
                                                                        color: red;
                                                                        font-weight: bold;
                                                                        border: 2px solid black;
                                                                        border-radius: 10px;
                                                                        padding: 10px 10px;
                                                                    ">
                                                                        Poço Instável
                                                                    </div>
                                                                </div>
                                                                """,
                                                                unsafe_allow_html=True
                                                            )
                                                    # else:
                                                    #     coluna_ref = 'Profundidade (m)' if st.session_state.t_prof == "TVD" else 'MD'
                                                    #
                                                    #     linha2 = df_tvp.loc[df_tvp[coluna_ref] == profundidade_proxima].iloc[0]
                                                    #
                                                    #     max_inferior = max(linha2['Tração Inferior'],
                                                    #                        linha2['Comp Inferior σθA'],
                                                    #                        linha2['Comp Inferior σθB'])
                                                    #
                                                    #     min_superior = min(linha2['Tração Superior (σθA)'],
                                                    #                        linha2['Tração Superior (σθB)'],
                                                    #                        linha2['Comp Superior σθA'],
                                                    #                        linha2['Comp Superior σθB'])
                                                    #
                                                    #     if peso_fluido <= max_inferior:
                                                    #         tipo_falha = "Colapso"
                                                    #         falha = max_inferior
                                                    #     elif peso_fluido >= min_superior:
                                                    #         tipo_falha = "Fratura"
                                                    #         falha = min_superior
                                                    #
                                                    #     else:
                                                    #         tipo_falha = None
                                                    #         falha = None
                                                    #
                                                    #     if falha is not None:
                                                    #         st.markdown(
                                                    #             f"""
                                                    #             <div style="
                                                    #                 display: flex;
                                                    #                 justify-content: center;
                                                    #                 margin-top: 3px;
                                                    #             ">
                                                    #                 <div style="
                                                    #                     color: red;
                                                    #                     font-weight: bold;
                                                    #                     border: 2px solid black;
                                                    #                     border-radius: 10px;
                                                    #                     padding: 12px 20px;
                                                    #                 ">
                                                    #                     {tipo_falha} em: {falha:.2f} ppg
                                                    #                 </div>
                                                    #             </div>
                                                    #             """,
                                                    #             unsafe_allow_html=True
                                                    #         )
                                                    #
                                                    #     else:
                                                    #         st.markdown(
                                                    #             """
                                                    #             <div style="
                                                    #                 display: flex;
                                                    #                 justify-content: center;
                                                    #                 margin-top: 3px;
                                                    #             ">
                                                    #                 <div style="
                                                    #                     color: green;
                                                    #                     font-weight: bold;
                                                    #                     border: 2px solid black;
                                                    #                     border-radius: 10px;
                                                    #                     padding: 12px 20px;
                                                    #                 ">
                                                    #                     Poço Estável
                                                    #                 </div>
                                                    #             </div>
                                                    #             """,
                                                    #             unsafe_allow_html=True
                                                    #         )

                                                with colu2:
                                                    st.markdown(
                                                        f"""
                                                        <div style="
                                                            display: flex;
                                                            justify-content: center;
                                                            margin-top: 0px;
                                                        ">
                                                            <div style="
                                                                color: black;
                                                                font-weight: bold;
                                                                border: 2px solid black;
                                                                border-radius: 10px;
                                                                padding: 0px 0px;
                                                                text-align: center;
                                                            ">
                                                                Peso do Fluido<br>
                                                                <span style="color: red;">{st.session_state.ppg:.2f}</span> lb/gal
                                                            </div>
                                                        </div>
                                                        """,
                                                        unsafe_allow_html=True
                                                    )

                                                centro_A = (sr_ef + sta_ef) / 2
                                                raio_A = abs(sta_ef - sr_ef) / 2
                                                centro_B = (sr_ef + stb_ef) / 2
                                                raio_B = abs(stb_ef - sr_ef) / 2

                                                ang = np.linspace(0, np.pi, 200)
                                                x_A = centro_A + raio_A * np.cos(ang)
                                                y_A = raio_A * np.sin(ang)
                                                x_B = centro_B + raio_B * np.cos(ang)
                                                y_B = raio_B * np.sin(ang)

                                                # Figura
                                                fig = go.Figure()
                                                if raio_A >= raio_B:
                                                    fig.add_trace(
                                                        go.Scatter(x=x_A, y=y_A, fill='toself',
                                                                   fillcolor='lightblue',
                                                                   line=dict(color='black', width=3),
                                                                   name='Círculo A'))
                                                    fig.add_trace(
                                                        go.Scatter(x=x_B, y=y_B, fill='toself', fillcolor='gold',
                                                                   line=dict(color='black', width=3),
                                                                   name='Círculo B'))
                                                else:
                                                    fig.add_trace(
                                                        go.Scatter(x=x_B, y=y_B, fill='toself', fillcolor='gold',
                                                                   line=dict(color='black', width=3),
                                                                   name='Círculo B'))
                                                    fig.add_trace(
                                                        go.Scatter(x=x_A, y=y_A, fill='toself',
                                                                   fillcolor='lightblue',
                                                                   line=dict(color='black', width=3),
                                                                   name='Círculo A'))

                                                # Pontos de tensões
                                                fig.add_trace(
                                                    go.Scatter(x=[sr_ef], y=[0], mode='markers+text', text=["σ'r"],
                                                               textposition='bottom center',
                                                               marker=dict(color='black', size=8),
                                                               showlegend=False))
                                                fig.add_trace(
                                                    go.Scatter(x=[sta_ef], y=[0], mode='markers+text',
                                                               text=["σ'θ A"],
                                                               textposition='bottom center',
                                                               marker=dict(color='black', size=8),
                                                               showlegend=False))
                                                fig.add_trace(
                                                    go.Scatter(x=[stb_ef], y=[0], mode='markers+text',
                                                               text=["σ'θ B"],
                                                               textposition='bottom center',
                                                               marker=dict(color='black', size=8),
                                                               showlegend=False))

                                                # Rótulos
                                                fig.add_trace(
                                                    go.Scatter(x=[centro_A], y=[raio_A / 2], mode='text',
                                                               text=['A'],
                                                               textposition="middle center",
                                                               textfont=dict(size=14, color='black'),
                                                               showlegend=False))
                                                fig.add_trace(
                                                    go.Scatter(x=[centro_B], y=[raio_B / 2], mode='text',
                                                               text=['B'],
                                                               textposition="middle center",
                                                               textfont=dict(size=14, color='black'),
                                                               showlegend=False))

                                                # Linha de critério de falha por tração
                                                y_tracao_inicio = 0
                                                y_tracao_fim = df_tvp.loc[
                                                    st.session_state.y == profundidade_proxima, 'So (psi)'
                                                ].values[0]
                                                # Limites do gráfico
                                                x_min = min(sr_ef, sta_ef, stb_ef) - 0.1 * max(raio_A, raio_B)
                                                x_max = max(sr_ef, sta_ef, stb_ef) + 0.1 * max(raio_A, raio_B)

                                                fig.add_shape(
                                                    type="line", x0=0, y0=y_tracao_inicio,
                                                    x1=0, y1=y_tracao_fim,
                                                    line=dict(color="red", width=3),
                                                    name='Critério de Falha por Tração',
                                                    showlegend=True
                                                )

                                                # Linha de critério de falha por compressão, começa onde termina a linha vermelha
                                                x_compressao_inicio = 0
                                                y_compressao_inicio = y_tracao_fim  # toca a linha de tração
                                                phi_rad = np.radians(st.session_state.phi)
                                                m = np.tan(phi_rad)

                                                x_reta = np.linspace(x_compressao_inicio, x_max, 100)
                                                y_reta = m * (x_reta - x_compressao_inicio) + y_compressao_inicio

                                                fig.add_trace(go.Scatter(
                                                    x=x_reta,
                                                    y=y_reta,
                                                    mode='lines',
                                                    line=dict(color='green', width=3),
                                                    name='Critério de Falha por Compressão',
                                                    showlegend=True
                                                ))

                                                def intersecao_circulo_reta(centro, raio, m, x0, y0):
                                                    a = 1 + m ** 2
                                                    b = -2 * centro + 2 * m * (y0 - m * x0)
                                                    c = centro ** 2 + (y0 - m * x0) ** 2 - raio ** 2
                                                    delta = b ** 2 - 4 * a * c
                                                    if delta < 0:
                                                        return None
                                                    raiz_delta = np.sqrt(delta)
                                                    x1 = (-b + raiz_delta) / (2 * a)
                                                    x2 = (-b - raiz_delta) / (2 * a)
                                                    return [(x1, m * (x1 - x0) + y0), (x2, m * (x2 - x0) + y0)]

                                                # Parâmetros da reta
                                                m = np.tan(np.radians(st.session_state.phi))
                                                x0, y0 = x_compressao_inicio, y_compressao_inicio

                                                # Interseções com A e B (pega o ponto de maior x se existir)
                                                for nome, centro, raio in [('A', centro_A, raio_A),
                                                                           ('B', centro_B, raio_B)]:
                                                    pontos = intersecao_circulo_reta(centro, raio, m, x0, y0)
                                                    if pontos:
                                                        x, y = max(pontos, key=lambda p: p[0])
                                                        st.session_state[f'x{nome.lower()}'] = x
                                                        st.session_state[f'y{nome.lower()}'] = y

                                                # Adiciona pontos no gráfico, se aplicável
                                                if st.session_state.op in ['Comp Inferior σθA',
                                                                           'Comp Superior σθA'] and 'xa' in st.session_state:
                                                    fig.add_trace(go.Scatter(
                                                        x=[st.session_state.xa], y=[st.session_state.ya],
                                                        mode='markers',
                                                        marker=dict(color='red', size=10, symbol='circle'),
                                                        name='Falha por Compressão em A'
                                                    ))

                                                if st.session_state.op in ['Comp Inferior σθB',
                                                                           'Comp Superior σθB'] and 'xb' in st.session_state:
                                                    fig.add_trace(go.Scatter(
                                                        x=[st.session_state.xb], y=[st.session_state.yb],
                                                        mode='markers',
                                                        marker=dict(color='red', size=10, symbol='circle'),
                                                        name='Falha por Compressão em B'
                                                    ))

                                                fig.update_layout(
                                                    xaxis_title="Tensão normal σ (psi)",
                                                    yaxis_title="Tensão cisalhante τ (psi)",
                                                    xaxis=dict(autorange=True),
                                                    yaxis=dict(autorange=True, scaleanchor="x", scaleratio=1),
                                                    margin=dict(l=50, r=50, t=50, b=100),
                                                    legend=dict(orientation='h', x=0.5, xanchor='center', y=-0.25,
                                                                yanchor='top')
                                                )

                                                st.session_state.fig_mohr = fig
                                                st.plotly_chart(st.session_state.fig_mohr, use_container_width=True)

                                        with st.expander('Trajetória do Poço', expanded=False):
                                            st.markdown('### Trajetória do Poço')
                                            try:
                                                def find_directional_columns(df):
                                                    md_col = inc_col = azi_col = None
                                                    for c in df.columns:
                                                        cl = c.lower()
                                                        if md_col is None and any(k in cl for k in
                                                                                  ["md", "measured depth",
                                                                                   "profund",
                                                                                   "profundidade"]):
                                                            md_col = c
                                                        if inc_col is None and any(k in cl for k in
                                                                                   ["inc", "incl", "inclinação",
                                                                                    "inclination"]):
                                                            inc_col = c
                                                        if azi_col is None and any(
                                                                k in cl for k in
                                                                ["azi", "azim", "azimuth", "azimute"]):
                                                            azi_col = c
                                                    return md_col, inc_col, azi_col

                                                def standardize_directional(df):
                                                    md_col, inc_col, azi_col = find_directional_columns(df)
                                                    if md_col is None or inc_col is None or azi_col is None:
                                                        raise ValueError(
                                                            "Não foram encontradas colunas MD / Inc / Azi na tabela.")
                                                    df2 = df[[md_col, inc_col, azi_col]].rename(
                                                        columns={md_col: "MD", inc_col: "Inc", azi_col: "Azi"}
                                                    )
                                                    df2 = df2.dropna().astype(float).sort_values(
                                                        "MD").reset_index(
                                                        drop=True)
                                                    if any(df2["MD"].diff().fillna(1) <= 0):
                                                        raise ValueError(
                                                            "Coluna MD deve ser estritamente crescente.")
                                                    return df2

                                                def minimum_curvature(df, x0=0.0, y0=0.0, tvd0=0.0,
                                                                      vsec_azimute_ref_graus=None):
                                                    MD = df["MD"].to_numpy(dtype=float)
                                                    Inc = np.radians(df["Inc"].to_numpy(dtype=float))
                                                    Azi = np.radians(df["Azi"].to_numpy(dtype=float))

                                                    Easting = [float(x0)]
                                                    Northing = [float(y0)]
                                                    TVD = [float(tvd0)]
                                                    DLS_list = []

                                                    for i in range(1, len(MD)):
                                                        dMD = MD[i] - MD[i - 1]
                                                        cosDL = (
                                                                np.sin(Inc[i - 1]) * np.sin(Inc[i]) * np.cos(
                                                            Azi[i] - Azi[i - 1])
                                                                + np.cos(Inc[i - 1]) * np.cos(Inc[i])
                                                        )
                                                        cosDL = np.clip(cosDL, -1.0, 1.0)
                                                        DL = np.arccos(cosDL)

                                                        DLS = np.degrees(DL) * 30.0 / dMD
                                                        DLS_list.append(float(DLS))

                                                        RF = 1.0 if DL < 1e-8 else (2.0 / DL) * np.tan(DL / 2.0)

                                                        dN = 0.5 * dMD * (
                                                                np.sin(Inc[i - 1]) * np.cos(Azi[i - 1]) +
                                                                np.sin(Inc[i]) * np.cos(Azi[i])
                                                        ) * RF
                                                        dE = 0.5 * dMD * (
                                                                np.sin(Inc[i - 1]) * np.sin(Azi[i - 1]) +
                                                                np.sin(Inc[i]) * np.sin(Azi[i])
                                                        ) * RF
                                                        dTVD = 0.5 * dMD * (np.cos(Inc[i - 1]) + np.cos(Inc[i])) * RF

                                                        Easting.append(Easting[-1] + dE)
                                                        Northing.append(Northing[-1] + dN)
                                                        TVD.append(TVD[-1] + dTVD)

                                                    # --- Afastamento horizontal absoluto (Departure / HD) ---
                                                    East_arr = np.asarray(Easting, dtype=float)
                                                    North_arr = np.asarray(Northing, dtype=float)
                                                    afast_h = np.sqrt(East_arr ** 2 + North_arr ** 2)

                                                    # --- VSEC: projeção numa direção de referência ---
                                                    # Se você não passar referência, uso o azimute do 1º ponto (padrão simples)
                                                    if vsec_azimute_ref_graus is None:
                                                        az_ref = float(np.degrees(Azi[0])) if len(Azi) else 0.0
                                                    else:
                                                        az_ref = float(vsec_azimute_ref_graus)

                                                    az_ref_rad = np.radians(az_ref)
                                                    vsec = East_arr * np.sin(az_ref_rad) + North_arr * np.cos(
                                                        az_ref_rad)

                                                    return pd.DataFrame({
                                                        "MD": MD,
                                                        "Inclinação (°)": np.degrees(Inc),
                                                        "Azimute (°)": np.degrees(Azi),
                                                        "Easting": Easting,
                                                        "Northing": Northing,
                                                        "TVD": TVD,
                                                        "Dogleg Severity (°/30m)": [0.0] + DLS_list,
                                                        "Afastamento Horizontal (m)": afast_h,
                                                        "VSEC (m)": vsec,
                                                    })

                                                df_to_use = None

                                                # 1) Prioridade: df2 (trajetória escolhida no tabs[0]) -> é o "dado real"
                                                if "df2" in st.session_state and isinstance(st.session_state["df2"],
                                                                                            pd.DataFrame):
                                                    df_to_use = st.session_state["df2"].copy()

                                                # 2) Se você quiser permitir usar df_interp (interpolada para coincidir com df1),
                                                # mantenha como alternativa — mas para minimum curvature, df2 já é suficiente.
                                                elif "df_interp" in st.session_state and isinstance(
                                                        st.session_state["df_interp"], pd.DataFrame):
                                                    df_to_use = st.session_state["df_interp"].copy()

                                                if df_to_use is None or df_to_use.empty:
                                                    st.info(
                                                        "Trajetória não encontrada. Verifique se o XLSM possui a aba 'Trajetória' e se foi lida no upload.")
                                                    st.stop()

                                                if df_to_use is None:
                                                    st.info(
                                                        "Nenhum arquivo direcional disponível. Carregue ou insira uma planilha.")
                                                    st.stop()

                                                # Ajustar nomes de colunas comuns
                                                rename_map = {"Profundidade": "MD", "Inc (°)": "Inc",
                                                              "Azi (°)": "Azi"}
                                                df_to_use = df_to_use.rename(
                                                    columns={k: v for k, v in rename_map.items() if
                                                             k in df_to_use.columns})

                                                # Padronizar e validar
                                                df_proc = standardize_directional(df_to_use)

                                                # Interpolar até a superfície com base nos dados reais de df2, se disponível
                                                # Se o usuário quiser expandir até o zero (ou até o menor MD), faça aqui
                                                if st.session_state.get("ex", "Desativada") == "Ativada":
                                                    first_md = float(df_proc["MD"].iloc[0])
                                                    if first_md > 0:
                                                        md_extra = np.arange(0.0, first_md, 1.0)
                                                        df_extra = pd.DataFrame({
                                                            "MD(m)": md_extra,
                                                            "Incl": np.interp(md_extra, df_proc["MD"], df_proc["Inc"]),
                                                            "Azimute": np.interp(md_extra, df_proc["MD"], df_proc["Azi"]),
                                                        })
                                                        df_proc = pd.concat([df_extra, df_proc],
                                                                            ignore_index=True).drop_duplicates(
                                                            "MD").sort_values("MD")

                                                df_out = minimum_curvature(df_proc, x0=0.0, y0=0.0, tvd0=0.0)
                                                st.session_state.df_out_traj = df_out.copy()

                                                fig = go.Figure()

                                                nx, ny = 80, 80
                                                x_min, x_max = df_out["Easting"].min() - 50, df_out[
                                                    "Easting"].max() + 50
                                                y_min, y_max = df_out["Northing"].min() - 50, df_out[
                                                    "Northing"].max() + 50
                                                xg, yg = np.meshgrid(np.linspace(x_min, x_max, nx),
                                                                     np.linspace(y_min, y_max, ny))
                                                z_superficie = np.zeros_like(xg)

                                                fig.add_trace(go.Surface(
                                                    x=xg, y=yg, z=z_superficie,
                                                    colorscale=[[0, "#aaaaaa"], [1, "#aaaaaa"]],
                                                    showscale=False, opacity=0.6, name="Solo", showlegend=True
                                                ))

                                                z_plot = df_out["TVD"]
                                                fig.add_trace(go.Scatter3d(
                                                    x=df_out["Easting"], y=df_out["Northing"], z=z_plot,
                                                    mode="lines", line=dict(color="red", width=5),
                                                    name="Trajetória"
                                                ))

                                                fig.add_trace(go.Scatter3d(
                                                    x=[df_out["Easting"].iloc[0]],
                                                    y=[df_out["Northing"].iloc[0]],
                                                    z=[z_plot.iloc[0]],
                                                    mode="markers",
                                                    marker=dict(size=8, color="blue", symbol="circle"),
                                                    name="Cabeça do poço"
                                                ))

                                                fig.add_trace(go.Scatter3d(
                                                    x=[df_out["Easting"].iloc[-1]],
                                                    y=[df_out["Northing"].iloc[-1]],
                                                    z=[z_plot.iloc[-1]],
                                                    mode="markers+text",
                                                    marker=dict(size=8, color="green", symbol="circle"),
                                                    textposition="top center",
                                                    name="Alvo"
                                                ))

                                                ponto_idx = (
                                                            df_out["TVD"] - profundidade_proxima).abs().idxmin()
                                                ponto_x = df_out.loc[ponto_idx, "Easting"]
                                                ponto_y = df_out.loc[ponto_idx, "Northing"]
                                                ponto_z = df_out.loc[ponto_idx, "TVD"]

                                                fig.add_trace(go.Scatter3d(
                                                    x=[ponto_x],
                                                    y=[ponto_y],
                                                    z=[ponto_z],
                                                    mode="markers+text",
                                                    marker=dict(size=8, color="orange", symbol="diamond"),
                                                    text=f"Profundidade: {profundidade_proxima:.1f} m",
                                                    textposition="top center",
                                                    name="Profundidade Analisada"
                                                ))

                                                fig.update_layout(
                                                    height=800,
                                                    width=600,
                                                    scene=dict(
                                                        xaxis=dict(title="Easting"),
                                                        yaxis=dict(title="Northing"),
                                                        zaxis=dict(title="TVD (m)", autorange="reversed"),
                                                        camera=dict(
                                                            eye=dict(x=2., y=2., z=1.6)
                                                        )
                                                    ),
                                                    legend=dict(
                                                        font=dict(size=10, color="black"),
                                                        bgcolor="rgba(255,255,255,0.7)",
                                                        bordercolor="gray",
                                                        borderwidth=1,
                                                        x=1,
                                                        y=-0.15,
                                                        xanchor="right",
                                                        yanchor="top",
                                                    )
                                                )

                                                st.session_state.fig_traj = fig

                                                st.plotly_chart(fig, use_container_width=True)

                                            except Exception as e:
                                                st.warning(f"Erro ao calcular/plotar trajetória: {e}")
                                                pass

                                        with st.expander('Tabela de Dados da Trajetória', expanded=False):
                                            st.markdown('### Tabela de Dados da Trajetória')
                                            try:
                                                st.dataframe(df_out, use_container_width=True)
                                            except Exception as e:
                                                st.warning(f"Erro ao calcular/plotar trajetória: {e}")
                                                pass

                                # GRÁFICO JANELA OPERACIONAL
                                with c3:
                                    with st.container(border=True):
                                        st.markdown('### Janela Operacional')
                                        if not criterio_disponivel(df_tvp):
                                            st.warning(
                                                "⚠️ Calcule as tensões antes de avaliar o critério de falha.")
                                        if criterio_disponivel(df_tvp):
                                            with colu3:
                                                df_tvp['Max Inferior'] = df_tvp[
                                                    ['Tração Inferior', 'Comp Inferior σθA', 'Comp Inferior σθB']
                                                ].max(axis=1)

                                                df_tvp['Min Superior'] = df_tvp[
                                                    ['Tração Superior (σθA)', 'Tração Superior (σθB)',
                                                     'Comp Superior σθA', 'Comp Superior σθB']
                                                ].min(axis=1)

                                                max_inferior = df_tvp['Max Inferior'].max()
                                                min_superior = df_tvp['Min Superior'].min()

                                                st.markdown(
                                                    f"""
                                                        <div style="
                                                            display: flex;
                                                            justify-content: center;
                                                            margin-top: 0px;
                                                        ">
                                                            <div style="
                                                                color: black;
                                                                font-weight: bold;
                                                                border: 2px solid black;
                                                                border-radius: 10px;
                                                                padding: 0px 0px;
                                                                text-align: center;
                                                            ">
                                                                Janela Op.<br>
                                                                <span style="color: red;">{max_inferior + st.session_state.fs:.2f}</span> &lt; ρ &lt; <span style="color: red;">{min_superior - st.session_state.fs:.2f}</span>
                                                            </div>
                                                        </div>
                                                        """,
                                                    unsafe_allow_html=True
                                                )
                                            st.session_state.fig_jo = plt.figure(figsize=(8, 10))

                                            if st.session_state.idg == 'Sim':
                                                # ===== COM coluna de idade =====
                                                gs = gridspec.GridSpec(
                                                    1, 4,
                                                    width_ratios=[0.1, 0.2, 0.21, 1],
                                                    wspace=0
                                                )

                                                ax_idade = st.session_state.fig_jo.add_subplot(gs[0])
                                                ax1 = st.session_state.fig_jo.add_subplot(gs[1], sharey=ax_idade)

                                                ax_gap = st.session_state.fig_jo.add_subplot(gs[2])
                                                ax_gap.axis('off')

                                                ax = st.session_state.fig_jo.add_subplot(gs[3], sharey=ax_idade)

                                                idade_formacao(ax_idade, st.session_state.df_idade, st.session_state.y_max_pp)

                                                # evita duplicar rótulos de profundidade
                                                plt.setp(ax1.get_yticklabels(), visible=False)
                                                plt.setp(ax.get_yticklabels(), visible=False)

                                            else:
                                                # ===== SEM coluna de idade =====
                                                gs = gridspec.GridSpec(
                                                    1, 3,
                                                    width_ratios=[0.18, 0.21, 1],
                                                    wspace=0
                                                )

                                                ax1 = st.session_state.fig_jo.add_subplot(gs[0])
                                                ax_gap = st.session_state.fig_jo.add_subplot(gs[1])
                                                ax_gap.axis('off')

                                                ax = st.session_state.fig_jo.add_subplot(gs[2], sharey=ax1)

                                                plt.setp(ax.get_yticklabels(), visible=False)

                                            # Limites da Janela Operacional
                                            x_max_inf = df_suav[
                                                'Max Inferior'] if st.session_state.suav_max_inf else \
                                                df_tvp['Max Inferior']
                                            x_min_sup = df_suav[
                                                'Min Superior'] if st.session_state.suav_min_sup else \
                                                df_tvp['Min Superior']

                                            # ZONAS DE PERDA DIMINUEM LIMITE SUPERIOR
                                            for i in st.session_state.zona2:
                                                profundidade_zona = i[0]
                                                peso_zona = i[1]

                                                if st.session_state.suav_min_sup:

                                                    if st.session_state.t_prof == "TVD":
                                                        coluna_profundidade = df_suav['Profundidade (m)']
                                                    else:
                                                        coluna_profundidade = df_suav['MD']

                                                    index = (
                                                                coluna_profundidade - profundidade_zona).abs().idxmin()
                                                    df_suav.loc[index, 'Min Superior'] = peso_zona

                                                else:

                                                    if st.session_state.t_prof == "MD":
                                                        coluna_profundidade = df_tvp['Profundidade (m)']
                                                    else:
                                                        coluna_profundidade = df_tvp['MD']

                                                    index = (
                                                                coluna_profundidade - profundidade_zona).abs().idxmin()
                                                    df_tvp.loc[index, 'Min Superior'] = peso_zona

                                            if "ponto_a"not in st.session_state:
                                                st.session_state.ponto_a = 'Não'

                                            if st.session_state.ponto_a == 'Sim':
                                                if st.session_state.ppg is not None and st.session_state.m is not None:
                                                    ax.scatter(
                                                        peso_fluido,
                                                        profundidade_proxima,
                                                        color='red',
                                                        edgecolor='black',
                                                        s=80,
                                                        zorder=10,
                                                        label='Peso do Fluido Selecionado'
                                                    )

                                            # Janela Operacional completa
                                            if st.session_state.jo:
                                                ax.plot(x_max_inf, st.session_state.y, color='blue',
                                                        linestyle='-', linewidth=2,
                                                        label="Limite Inferior da Janela Operacional")
                                                ax.plot(x_min_sup, st.session_state.y, color='red',
                                                        linestyle='-', linewidth=2,
                                                        label="Limite Superior da Janela Operacional")

                                            colunas_texto = [
                                                'Profundidade (m)',
                                                'Gradiente de Sobrecarga (lb/gal)',
                                                'SH (lb/gal)',
                                                'Sh (lb/gal)',
                                                'Direção SH',
                                                'Direção Sh'
                                            ]

                                            def ajustar_angulo_360(ang):
                                                if pd.isna(ang):
                                                    return ang
                                                while ang > 360:
                                                    ang -= 360
                                                return ang

                                            if all(col in df_tvp.columns for col in colunas_texto):
                                                try:
                                                    # Sempre usa a profundidade_proxima
                                                    idx_tensoes = (df_tvp[
                                                                       'Profundidade (m)'] - profundidade_proxima).abs().idxmin()
                                                    linha_t = df_tvp.loc[idx_tensoes]

                                                    sv = linha_t['Gradiente de Sobrecarga (lb/gal)']
                                                    dir_shmax = ajustar_angulo_360(linha_t['Direção SH'])
                                                    dir_shmin = ajustar_angulo_360(linha_t['Direção Sh'])

                                                    # Captura azimute final do poço, se existir
                                                    azimute_poco = np.nan
                                                    if 'Azi' in df_tvp.columns and pd.notna(linha_t.get('Azi', np.nan)):
                                                        azimute_poco = ajustar_angulo_360(linha_t['Azi'])
                                                    elif 'Azimute' in df_tvp.columns and pd.notna(
                                                            linha_t.get('Azimute', np.nan)):
                                                        azimute_poco = ajustar_angulo_360(linha_t['Azimute'])
                                                    elif 'Azimuth' in df_tvp.columns and pd.notna(
                                                            linha_t.get('Azimuth', np.nan)):
                                                        azimute_poco = ajustar_angulo_360(linha_t['Azimuth'])

                                                    if pd.notna(sv) and sv != 0:
                                                        if 'SH% Sobrecarga' in df_tvp.columns and pd.notna(
                                                                linha_t.get('SH% Sobrecarga', np.nan)):
                                                            rel_shmax = linha_t['SH% Sobrecarga']
                                                        else:
                                                            rel_shmax = linha_t['SH (lb/gal)'] / sv

                                                        if 'Sh% Sobrecarga' in df_tvp.columns and pd.notna(
                                                                linha_t.get('Sh% Sobrecarga', np.nan)):
                                                            rel_shmin = linha_t['Sh% Sobrecarga']
                                                        else:
                                                            rel_shmin = linha_t['Sh (lb/gal)'] / sv

                                                        texto_tensoes = (
                                                            f"SH = {rel_shmax:.2f}·σv | Dir. SH = {dir_shmax:.1f}°\n"
                                                            f"Sh = {rel_shmin:.2f}·σv | Dir. Sh = {dir_shmin:.1f}°"
                                                        )

                                                        if pd.notna(azimute_poco):
                                                            texto_tensoes += f"\nAzimute final do poço = {azimute_poco:.1f}°"

                                                        ax.text(
                                                            0.98, 0.99,
                                                            texto_tensoes,
                                                            transform=ax.transAxes,
                                                            fontsize=9,
                                                            verticalalignment='top',
                                                            horizontalalignment='right',
                                                            bbox=dict(
                                                                boxstyle='round',
                                                                facecolor='white',
                                                                alpha=0.85,
                                                                edgecolor='black'
                                                            ),
                                                            zorder=20
                                                        )
                                                except Exception:
                                                    pass

                                            # Limites da Janela Operacional
                                            x_max_inf = np.asarray(
                                                df_suav['Max Inferior'] if st.session_state.suav_max_inf else df_tvp[
                                                    'Max Inferior'],
                                                dtype=float
                                            )
                                            x_min_sup = np.asarray(
                                                df_suav['Min Superior'] if st.session_state.suav_min_sup else df_tvp[
                                                    'Min Superior'],
                                                dtype=float
                                            )

                                            y_vals = np.asarray(st.session_state.y, dtype=float)

                                            # FS inferior com sua lógica:
                                            if st.session_state.ijo:
                                                x_fs_base_inf = np.asarray(x_max_inf, dtype=float) + float(
                                                    st.session_state.fs)

                                                # Se o próximo for menor que o atual, repete o atual
                                                x_fs_inf = np.maximum.accumulate(x_fs_base_inf)

                                                ax.plot(
                                                    x_fs_inf,
                                                    y_vals,
                                                    color='gold',
                                                    linestyle='--',
                                                    linewidth=2,
                                                    label="FS Inferior da Janela Operacional"
                                                )
                                            else:
                                                x_fs_inf = np.asarray(x_max_inf, dtype=float).copy()

                                            # FS superior
                                            if st.session_state.sjo:
                                                x_fs_sup = np.asarray(x_min_sup, dtype=float) - float(
                                                    st.session_state.fs)

                                                ax.plot(
                                                    x_fs_sup,
                                                    y_vals,
                                                    color='tomato',
                                                    linestyle='--',
                                                    linewidth=2,
                                                    label="FS Superior da Janela Operacional"
                                                )
                                            else:
                                                x_fs_sup = np.asarray(x_min_sup, dtype=float).copy()

                                            # Janela útil (verde)
                                            mascara_janela = x_fs_sup > x_fs_inf
                                            if np.any(mascara_janela):
                                                ax.fill_betweenx(
                                                    y_vals,
                                                    x_fs_inf,
                                                    x_fs_sup,
                                                    where=mascara_janela,
                                                    interpolate=True,
                                                    color='lightgreen',
                                                    alpha=0.25,
                                                    label='Janela Operacional'
                                                )

                                            # Faixa consumida pelo FS inferior (vermelho)
                                            if st.session_state.ijo:
                                                mascara_inf = x_fs_inf > x_max_inf
                                                if np.any(mascara_inf):
                                                    ax.fill_betweenx(
                                                        y_vals,
                                                        x_max_inf,
                                                        x_fs_inf,
                                                        where=mascara_inf,
                                                        interpolate=True,
                                                        color='lightcoral',
                                                        alpha=0.25,
                                                    )

                                            # Faixa consumida pelo FS superior (vermelho)
                                            if st.session_state.sjo:
                                                mascara_sup = x_min_sup > x_fs_sup
                                                if np.any(mascara_sup):
                                                    ax.fill_betweenx(
                                                        y_vals,
                                                        x_fs_sup,
                                                        x_min_sup,
                                                        where=mascara_sup,
                                                        interpolate=True,
                                                        color='lightcoral',
                                                        alpha=0.25,
                                                    )

                                            # Apenas limite inferior
                                            if st.session_state.li and not st.session_state.jo:
                                                ax.plot(x_max_inf, st.session_state.y, color='blue',
                                                        linestyle='-', linewidth=2,
                                                        label="Limite Inferior da Janela Operacional")

                                            # Apenas limite superior
                                            if st.session_state.ls and not st.session_state.jo:
                                                ax.plot(x_min_sup, st.session_state.y, color='red',
                                                        linestyle='-', linewidth=2,
                                                        label="Limite Superior da Janela Operacional")

                                            # Tração e Compressão com lógica de suavização
                                            if st.session_state.tsa:
                                                if st.session_state.suav_tsa:
                                                    ax.plot(df_suav['Tração Superior (σθA)'],
                                                            st.session_state.y, color='green',
                                                            linestyle='-',
                                                            linewidth=2,
                                                            label="Tração Superior σθA")
                                                else:
                                                    ax.plot(df_tvp['Tração Superior (σθA)'], st.session_state.y,
                                                            color='green',
                                                            linestyle='-', linewidth=2,
                                                            label="Tração Superior σθA")

                                            if st.session_state.tsb:
                                                if st.session_state.suav_tsb:
                                                    ax.plot(df_suav['Tração Superior (σθB)'],
                                                            st.session_state.y, color='purple',
                                                            linestyle='-',
                                                            linewidth=2,
                                                            label="Tração Superior σθB")
                                                else:
                                                    ax.plot(df_tvp['Tração Superior (σθB)'], st.session_state.y,
                                                            color='purple',
                                                            linestyle='-', linewidth=2,
                                                            label="Tração Superior σθB")

                                            if st.session_state.cia:
                                                if st.session_state.suav_cia:
                                                    ax.plot(df_suav['Comp Inferior σθA'], st.session_state.y,
                                                            color='turquoise',
                                                            linestyle='-',
                                                            linewidth=2,
                                                            label="Comp Inferior σθA")
                                                else:
                                                    ax.plot(df_tvp['Comp Inferior σθA'], st.session_state.y,
                                                            color='turquoise',
                                                            linestyle='-', linewidth=2,
                                                            label="Comp Inferior σθA")

                                            if st.session_state.csa:
                                                if st.session_state.suav_csa:
                                                    ax.plot(df_suav['Comp Superior σθA'], st.session_state.y,
                                                            color='lime',
                                                            linestyle='-',
                                                            linewidth=2,
                                                            label="Comp Superior σθA")
                                                else:
                                                    ax.plot(df_tvp['Comp Superior σθA'], st.session_state.y,
                                                            color='lime',
                                                            linestyle='-', linewidth=2,
                                                            label="Comp Superior σθA")

                                            if st.session_state.cib:
                                                if st.session_state.suav_cib:
                                                    ax.plot(df_suav['Comp Inferior σθB'], st.session_state.y,
                                                            color='brown',
                                                            linestyle='-',
                                                            linewidth=2,
                                                            label="Comp Inferior σθB")
                                                else:
                                                    ax.plot(df_tvp['Comp Inferior σθB'], st.session_state.y,
                                                            color='brown',
                                                            linestyle='-', linewidth=2,
                                                            label="Comp Inferior σθB")

                                            if st.session_state.csb:
                                                if st.session_state.suav_csb:
                                                    ax.plot(df_suav['Comp Superior σθB'], st.session_state.y,
                                                            color='gray',
                                                            linestyle='-',
                                                            linewidth=2,
                                                            label="Comp Superior σθB")
                                                else:
                                                    ax.plot(df_tvp['Comp Superior σθB'], st.session_state.y,
                                                            color='gray',
                                                            linestyle='-', linewidth=2,
                                                            label="Comp Superior σθB")

                                            # Pressão de poros e Tração Inferior (sempre originais)
                                            if st.session_state.gs:
                                                ax.plot(df_tvp['Gradiente de Sobrecarga (lb/gal)'],
                                                        st.session_state.y,
                                                        color='black', linestyle='-', linewidth=2,
                                                        label="Gradiente de Sobrecarga")
                                            if st.session_state.show_pp:
                                                ax.plot(df_tvp['Gradiente de Pressão de Poros (lb/gal)'],
                                                        st.session_state.y,
                                                        color='orange', linestyle='-', linewidth=2,
                                                        label="Pressão de Poros")
                                            if st.session_state.ti:
                                                ax.plot(df_tvp['Tração Inferior'], st.session_state.y,
                                                        color='teal', linestyle='-',
                                                        linewidth=2,
                                                        label="Tração Inferior")

                                            if st.session_state.ponto_a == 'Sim':
                                                # Linha horizontal indicando a profundidade analisada
                                                ax.axhline(
                                                    y=profundidade_proxima,
                                                    color='dodgerblue',
                                                    linestyle='--',
                                                    linewidth=1,
                                                )

                                                if "x_max" not in st.session_state:
                                                    st.session_state.x_max = 23
                                                if "x_min" not in st.session_state:
                                                    st.session_state.x_min = 7

                                                # Texto acima e à direita da linha
                                                ax.text(
                                                    st.session_state.x_max - (
                                                            st.session_state.x_max - st.session_state.x_min) * 0.01,
                                                    profundidade_proxima - 3,
                                                    f"Profundidade Analisada ({profundidade_proxima:.1f} m)",
                                                    color='black',
                                                    fontsize=10,
                                                    verticalalignment='bottom',
                                                    horizontalalignment='right',
                                                    zorder=8
                                                )

                                            if st.session_state.get("spt", "Não") == "Sim":
                                                if 'sapatas_df' in st.session_state and not st.session_state.sapatas_df.empty:
                                                    plotted_label = False

                                                    # Proteções básicas
                                                    df_tvp_local = None
                                                    if 'df_tvp' in locals():
                                                        df_tvp_local = df_tvp
                                                    elif 'df_tvp' in globals():
                                                        df_tvp_local = globals().get('df_tvp')
                                                    elif 'df_tvp' in st.session_state:
                                                        df_tvp_local = st.session_state.df_tvp

                                                    for _, row in st.session_state.sapatas_df.iterrows():
                                                        fase = float(row['Fase'])
                                                        tvd_informada = float(row['Profundidade da sapata (m)'])

                                                        # Decide profundidade a plotar (y_plot)
                                                        y_plot = tvd_informada
                                                        tvd_display = tvd_informada
                                                        md_display = None

                                                        try:
                                                            if st.session_state.t_prof == "TVD":
                                                                y_plot = tvd_informada
                                                            else:
                                                                # visualizar por MD → converter TVD informado para MD buscando o registro mais próximo em df_tvp
                                                                if (
                                                                        df_tvp_local is not None
                                                                        and 'Profundidade (m)' in df_tvp_local.columns
                                                                        and 'MD' in df_tvp_local.columns
                                                                ):
                                                                    idx_closest = (df_tvp_local[
                                                                                       'Profundidade (m)'] - tvd_informada).abs().idxmin()
                                                                    md_val = float(df_tvp_local.loc[idx_closest, 'MD'])
                                                                    y_plot = md_val
                                                                    md_display = md_val
                                                                else:
                                                                    y_plot = tvd_informada
                                                        except Exception:
                                                            y_plot = tvd_informada

                                                        # Triângulo / marcador na posição calculada
                                                        ax.scatter(
                                                            st.session_state.x_min + 0.5,
                                                            y_plot - 12,
                                                            color='black',
                                                            marker='v',
                                                            s=100,
                                                            label='Sapatas' if not plotted_label else "",
                                                            zorder=7
                                                        )

                                                        # Linha horizontal na posição calculada
                                                        ax.hlines(
                                                            y=y_plot,
                                                            xmin=st.session_state.x_min,
                                                            xmax=st.session_state.x_max,
                                                            colors='black',
                                                            linestyles='--',
                                                            linewidth=1,
                                                            zorder=5
                                                        )

                                                        # Texto à direita
                                                        if md_display is None:
                                                            text_str = f"Sapata {fase:.3f}, {tvd_display:.2f} m"
                                                        else:
                                                            text_str = f"Sapata {fase:.3f}, TVD {tvd_display:.2f} m (MD {md_display:.2f} m)"

                                                        ax.text(
                                                            st.session_state.x_max - (
                                                                        st.session_state.x_max - st.session_state.x_min) * 0.01,
                                                            y_plot - 3,
                                                            text_str,
                                                            color='black',
                                                            fontsize=10,
                                                            verticalalignment='bottom',
                                                            horizontalalignment='right',
                                                            zorder=8
                                                        )

                                                        plotted_label = True

                                            # ===== PLOT DAS PRISÕES DE COLUNA (SOMENTE MARCADOR + LINHA) =====
                                            if 'prisoes_coluna_df' in st.session_state and not st.session_state.prisoes_coluna_df.empty:
                                                plotted_label = False

                                                # Proteção para df_tvp
                                                df_tvp_local = None
                                                if 'df_tvp' in locals():
                                                    df_tvp_local = df_tvp
                                                elif 'df_tvp' in globals():
                                                    df_tvp_local = globals().get('df_tvp')
                                                elif 'df_tvp' in st.session_state:
                                                    df_tvp_local = st.session_state.df_tvp

                                                for _, row in st.session_state.prisoes_coluna_df.iterrows():
                                                    tvd_informada = float(row['Profundidade da prisão (m)'])

                                                    y_plot = tvd_informada

                                                    try:
                                                        if st.session_state.t_prof == "TVD":
                                                            y_plot = tvd_informada
                                                        else:
                                                            if (
                                                                    df_tvp_local is not None
                                                                    and 'Profundidade (m)' in df_tvp_local.columns
                                                                    and 'MD' in df_tvp_local.columns
                                                            ):
                                                                idx_closest = (
                                                                        df_tvp_local[
                                                                            'Profundidade (m)'] - tvd_informada
                                                                ).abs().idxmin()

                                                                y_plot = float(
                                                                    df_tvp_local.loc[idx_closest, 'MD'])
                                                    except Exception:
                                                        y_plot = tvd_informada

                                                    # Marcador
                                                    ax.scatter(
                                                        st.session_state.x_min + 0.5,
                                                        y_plot,
                                                        color='red',
                                                        marker='X',
                                                        s=90,
                                                        label='Prisão de Coluna' if not plotted_label else "",
                                                        zorder=7
                                                    )

                                                    # Linha horizontal
                                                    ax.hlines(
                                                        y=y_plot,
                                                        xmin=st.session_state.x_min,
                                                        xmax=st.session_state.x_max,
                                                        colors='red',
                                                        linestyles='--',
                                                        linewidth=1,
                                                        zorder=5
                                                    )

                                                    plotted_label = True

                                            if 'edited_z2' in st.session_state and not st.session_state.edited_z.empty:
                                                # pega os valores das colunas
                                                x_vals = st.session_state.edited_z2["Peso do fluido (lb/gal)"]
                                                y_vals = st.session_state.edited_z2[
                                                    "Profundidade da zona de perda (m)"]

                                                mask = (x_vals != 0) & (y_vals != 0)
                                                if mask.any():
                                                    ax.scatter(
                                                        x_vals[mask],
                                                        y_vals[mask],
                                                        facecolors='brown',
                                                        edgecolors='black',
                                                        marker='o',
                                                        s=80,
                                                        label="Perdas de Circulação",
                                                        zorder=6
                                                    )

                                            # Separa LOT e FIT de acordo com tt
                                            lot_x = [lt[i] for i in range(len(tt)) if tt[i] == "LOT"]
                                            lot_y = [pp[i] for i in range(len(tt)) if tt[i] == "LOT"]

                                            fit_x = [lt[i] for i in range(len(tt)) if tt[i] == "FIT"]
                                            fit_y = [pp[i] for i in range(len(tt)) if tt[i] == "FIT"]

                                            if st.session_state.tab:
                                                if lot_x and lot_y:
                                                    ax.scatter(lot_x, lot_y, color='red', label="LOT's",
                                                               zorder=5,
                                                               marker='D', s=50)
                                                else:
                                                    pass

                                                if fit_x and fit_y:
                                                    ax.scatter(fit_x, fit_y, color='blue', label="FIT's",
                                                               zorder=5,
                                                               marker='^', s=50)
                                                else:
                                                    pass
                                            if "df_mud" in st.session_state and isinstance(st.session_state["df_mud"],
                                                                                           pd.DataFrame):
                                                df_mud = st.session_state["df_mud"].copy()

                                                # Executado
                                                if mostrar_executado and df_mud[
                                                    "Peso do Fluido Executado (lb/gal)"].notna().any():
                                                    ax.plot(
                                                        df_mud["Peso do Fluido Executado (lb/gal)"],
                                                        df_mud["Profundidade (m)"],
                                                        linestyle="-",
                                                        color="mediumvioletred",
                                                        linewidth=2,
                                                        label="Peso do Fluido (Executado)",
                                                        zorder=5
                                                    )

                                            def reset_config():
                                                st.session_state.x_min = 7
                                                st.session_state.x_max = 23
                                                st.session_state.x_step_tvp = 1
                                                st.session_state.y_min = 0
                                                st.session_state.y_max = int(st.session_state.y.max()) + 100
                                                st.session_state.y_step_tvp = 200
                                                st.session_state.fs = 0.5

                                            lito(
                                                ax1,
                                                df_pp,
                                                profundidades,
                                                litologias,
                                                prof_final
                                            )

                                            with st.expander("Configurações do Gráfico", expanded=False):
                                                with st.expander("Configurações dos Eixos", expanded=False):
                                                    st.selectbox("Inserir ponto analisado", ['Não', 'Sim'], key="ponto_a",index=0)
                                                    st.selectbox("Inserir Sapatas", ['Não', 'Sim'], key="spt", index=0)
                                                    st.number_input("Eixo X - mínimo", value=7, step=1,
                                                                    key="x_min")
                                                    st.number_input("Eixo X - máximo", value=23, step=1,
                                                                    key="x_max")
                                                    st.number_input("Passo do eixo X", value=1, step=1,
                                                                    key="x_step_tvp")

                                                    st.number_input("Eixo Y - mínimo", value=0, step=100,
                                                                    key="y_min")
                                                    st.number_input(
                                                        "Eixo Y - máximo",
                                                        value=int(st.session_state.y.max()) + 100,
                                                        step=100,
                                                        key="y_max"
                                                    )
                                                    st.number_input(
                                                        "Passo do eixo Y",
                                                        value=50,
                                                        step=50,
                                                        key="y_step_tvp"
                                                    )

                                                    st.number_input(
                                                        "Fator de Segurança da Janela Operacional",
                                                        value=0.5,
                                                        step=0.1,
                                                        key='fs',
                                                        help=(
                                                            "**Fator de Segurança (FS)** ⚠️\n\n"
                                                            "- Coeficiente que reduz a **janela operacional**.\n"
                                                            "- Introduz uma margem de segurança contra **kick** ou **fratura**.\n"
                                                            "- Essencial para o planejamento seguro da perfuração."
                                                        )
                                                    )

                                                    st.button(
                                                        "Resetar Eixos - Tensões em Volta do Poço",
                                                        on_click=reset_config,
                                                        type="primary",
                                                        use_container_width=True
                                                    )

                                                with st.expander("Configurações da Legenda", expanded=False):
                                                    st.checkbox('Exibir Legendas', key='leg', value=True)
                                                    if st.session_state.leg:
                                                        legendas_pt = {
                                                            "Inferior direito": "lower right",
                                                            "Melhor posição": "best",
                                                            "Superior direito": "upper right",
                                                            "Superior esquerdo": "upper left",
                                                            "Inferior esquerdo": "lower left",
                                                            "Direita": "right",
                                                            "Centro esquerdo": "center left",
                                                            "Centro direito": "center right",
                                                            "Inferior central": "lower center",
                                                            "Superior central": "upper center",
                                                            "Central": "center",
                                                        }

                                                        escolha_legenda = st.selectbox(
                                                            "Posição da legenda",
                                                            list(legendas_pt.keys()),
                                                            index=0
                                                        )

                                                        fontsize_legenda = st.number_input(
                                                            "Tamanho da fonte da legenda",
                                                            min_value=4,
                                                            max_value=40,
                                                            step=1,
                                                            value=6
                                                        )

                                            # Configurações do gráfico
                                            ax.set_title('Janela Operacional (lb/gal)', fontsize=14,
                                                         fontweight='bold')
                                            ax.set_xlabel('Gradiente (ppg)', fontsize=12)
                                            ax.set_ylabel('Profundidade (m)', fontsize=12)
                                            ax.invert_yaxis()
                                            ax.tick_params(axis='y', which='both', left=True, labelleft=True)
                                            ax.set_ylim(st.session_state.y_max, st.session_state.y_min)
                                            ax.set_yticks(range(st.session_state.y_min, st.session_state.y_max,
                                                                st.session_state.y_step_tvp))
                                            ax.set_xticks(
                                                range(int(st.session_state.x_min), int(st.session_state.x_max),
                                                      int(st.session_state.x_step_tvp)))
                                            ax.set_xlim(st.session_state.x_min, st.session_state.x_max)
                                            ax.grid(True, linestyle='--', alpha=0.5)

                                            if st.session_state.leg:
                                                loc_legenda = legendas_pt[escolha_legenda]
                                                ax.legend(
                                                    loc=loc_legenda,
                                                    fontsize=fontsize_legenda,
                                                    frameon=True,
                                                    shadow=True,
                                                    fancybox=True,
                                                    framealpha=1,
                                                    facecolor='white',
                                                    edgecolor='gray'
                                                )

                                            add_watermark(
                                                ax,
                                                logo_path="logo.png",
                                                xy=(0.50, 0.5),
                                                zoom=0.2,
                                                alpha=0.2,
                                                zorder=0
                                            )
                                            plt.subplots_adjust(wspace=0.3)
                                            st.pyplot(st.session_state.fig_jo)

                            # VIZUALIZAÇÃO 3D DAS TENSÕES
                            with tb[1]:
                                if criterio_disponivel(df_tvp):
                                    colun1, colun2, colun3 = st.columns(3)
                                    with colun1:
                                        todos_parametros = ["σr", "σθ", "σa", "Dir. Tensões principais",
                                                            "Tensões Horizontais", "Coordenadas geográficas",
                                                            "Profundidade", "Direção do poço",
                                                            "Sistemas de coordenadas do poço"]
                                        with st.container(border=True, height=536):
                                            st.markdown('#### Opções de visualização')
                                            vistas = ["Selecione a vista", "Vista de planta",
                                                      "Vista de seção N/S",
                                                      "Vista de seção E/W", "Vista axial do poço",
                                                      "Dentro do Poço"]
                                            view = st.selectbox("Selecione o ponto de visualização", vistas)
                                            st.number_input('Profundidade (m)', key='m2', value=0.00,
                                                            format="%.2f")
                                            mostrar_todos = st.checkbox("Exibir todas as tensões",
                                                                        value=True)
                                            if not mostrar_todos:
                                                selecionados = st.multiselect("Escolha parâmetros a exibir:",
                                                                              todos_parametros)
                                            else:
                                                selecionados = todos_parametros
                                            show_t = st.checkbox('Exibir descrição', value=False,
                                                                 key='show_t')
                                            failure = st.checkbox('Mapa de falha na parede do poço',
                                                                  value=False, key='fail')
                                            around = st.checkbox('Influência ao redor do poço', value=False,
                                                                 key='around')

                                    if show_t:
                                        st.session_state.show = True
                                    else:
                                        st.session_state.show = False

                                    if failure:
                                        st.session_state.ff = True
                                    else:
                                        st.session_state.ff = False

                                    if around:
                                        st.session_state.arr = True
                                    else:
                                        st.session_state.arr = False

                                    with colun3:
                                        with st.container(border=True, height=536):
                                            st.markdown('### Configurações')
                                            ra = st.number_input('Raio de investigação', min_value=2)
                                            options = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]
                                            op = st.selectbox('Opacidade do mapa de falha', options,
                                                              key='opacity', index=7)
                                            lg = st.selectbox('Exibir linhas de grade', ["Sim", "Não"])

                                    # Pega o índice da profundidade mais próxima de m2
                                    idx = (df_tvp['Profundidade (m)'] - st.session_state.m2).abs().idxmin()

                                    # Seleciona a linha correspondente
                                    linha = df_tvp.loc[idx]

                                    # Extrai os valores
                                    sr = linha['σr efetivo (psi)']
                                    sig_t = linha['σθA efetivo (psi)']
                                    sa = linha['σa']
                                    tvd = linha['Profundidade (m)']
                                    pressure = linha['Gradiente de Pressão de Poros (lb/gal)']
                                    sig_a = sa * 0.1704 * tvd - pressure * 0.1704 * tvd
                                    sig_h = linha['Sh (lb/gal)'] * 0.1704 * tvd
                                    sig_H = linha['SH (lb/gal)'] * 0.1704 * tvd
                                    ang_theta = linha['θA (°)']
                                    ang_azi = linha['Azi']
                                    ang_inc = linha['Inc']
                                    ang_horizontal = linha['Direção SH']
                                    # Criação do gráfico
                                    figura = criar_grafico(selecionados, sr, sig_t, sig_a, sig_h, sig_H,
                                                           ang_theta, tvd, ang_azi, ang_inc, ra, op,
                                                           ang_horizontal, view, lg)

                                    with colun2:
                                        with st.container(border=True):
                                            st.markdown('#### Visualização 3D das Tensões')
                                            st.plotly_chart(figura, use_container_width=True)

                            # DATAFRAMES
                            with tb[3]:
                                if criterio_disponivel(df_tvp):
                                    if uploaded_file:
                                        with st.container(border=True):
                                            try:
                                                st.dataframe(
                                                    st.session_state.dados_lito,
                                                    use_container_width=True,
                                                    hide_index=True
                                                )
                                            except Exception:
                                                pass
                                        with st.container(border=True):
                                            st.dataframe(df_tvp, use_container_width=True, hide_index=True)
                                        with st.container(border=True):
                                            st.dataframe(df_suav, use_container_width=True, hide_index=True)

                        else:
                            st.error('Preencha corretamente a aba "Gradiente de Pressão de Poros"', icon="🚨")
                    else:
                        st.error('Preencha corretamente a aba "Gradiente de Sobrecarga"', icon="🚨")
                else:
                    st.error('Por favor, insira um documento!', icon="🚨")

    #Anotações
    with tabs[5]:
        if "anotacoes" not in st.session_state:
            st.session_state.anotacoes = ""

        with st.container(border=True):
            def _interp_em_tvd(df_traj: pd.DataFrame, tvd_alvo: float):
                tvd = df_traj["TVD"].to_numpy(dtype=float)
                md = df_traj["MD"].to_numpy(dtype=float)

                if "Afastamento Horizontal (m)" in df_traj.columns:
                    hd = df_traj["Afastamento Horizontal (m)"].to_numpy(dtype=float)
                else:
                    e = df_traj["Easting"].to_numpy(dtype=float)
                    n = df_traj["Northing"].to_numpy(dtype=float)
                    hd = np.sqrt(e ** 2 + n ** 2)

                # ordena por TVD para garantir interp correta
                order = np.argsort(tvd)
                tvd = tvd[order]
                md = md[order]
                hd = hd[order]

                if tvd_alvo < tvd.min() or tvd_alvo > tvd.max():
                    raise ValueError(f"TVD fora do intervalo da trajetória ({tvd.min():.2f} a {tvd.max():.2f}).")

                md_i = float(np.interp(tvd_alvo, tvd, md))
                hd_i = float(np.interp(tvd_alvo, tvd, hd))
                return md_i, hd_i

            def _segmento_traj_por_tvd(df_traj: pd.DataFrame, tvd_ini: float, tvd_fim: float):
                if tvd_fim < tvd_ini:
                    tvd_ini, tvd_fim = tvd_fim, tvd_ini

                md_all = df_traj["MD"].to_numpy(dtype=float)
                tvd_all = df_traj["TVD"].to_numpy(dtype=float)

                if "Afastamento Horizontal (m)" in df_traj.columns:
                    hd_all = df_traj["Afastamento Horizontal (m)"].to_numpy(dtype=float)
                else:
                    e = df_traj["Easting"].to_numpy(dtype=float)
                    n = df_traj["Northing"].to_numpy(dtype=float)
                    hd_all = np.sqrt(e ** 2 + n ** 2)

                # ordena por TVD
                order = np.argsort(tvd_all)
                md_all = md_all[order]
                tvd_all = tvd_all[order]
                hd_all = hd_all[order]

                if tvd_ini < tvd_all.min() or tvd_fim > tvd_all.max():
                    raise ValueError(
                        f"Trecho fora do intervalo da trajetória em TVD ({tvd_all.min():.2f} a {tvd_all.max():.2f})."
                    )

                mask = (tvd_all >= tvd_ini) & (tvd_all <= tvd_fim)
                md_seg = md_all[mask]
                tvd_seg = tvd_all[mask]
                hd_seg = hd_all[mask]

                md_i = float(np.interp(tvd_ini, tvd_all, md_all))
                hd_i = float(np.interp(tvd_ini, tvd_all, hd_all))

                md_f = float(np.interp(tvd_fim, tvd_all, md_all))
                hd_f = float(np.interp(tvd_fim, tvd_all, hd_all))

                if md_seg.size == 0 or tvd_seg[0] != tvd_ini:
                    md_seg = np.insert(md_seg, 0, md_i)
                    tvd_seg = np.insert(tvd_seg, 0, tvd_ini)
                    hd_seg = np.insert(hd_seg, 0, hd_i)

                if tvd_seg[-1] != tvd_fim:
                    md_seg = np.append(md_seg, md_f)
                    tvd_seg = np.append(tvd_seg, tvd_fim)
                    hd_seg = np.append(hd_seg, hd_f)

                return hd_seg, tvd_seg, md_seg

            st.markdown("### Trajetória 2D (TVD x Afastamento Horizontal)")

            if "df_out_traj" not in st.session_state or not isinstance(st.session_state.df_out_traj, pd.DataFrame):
                st.info(
                    "A trajetória ainda não foi calculada.")
            else:
                df_traj = st.session_state.df_out_traj.copy()

                # =========================
                # EVENTOS (opções)
                # =========================
                EVENTOS = [
                    "Drag",
                    "Overpull",
                    "Repasse",
                    "Coluna presa",
                    "Ameaça de prisão",
                    "Coluna topou",
                    "Ferramenta de perfilagem topou",
                    "Perda parcial de circulação",
                    "Perda severa de circulação",
                ]

                # =========================
                # Estilo por evento (cor + símbolo)
                # =========================
                MAPA_EVENTO_ESTILO = {
                    "Drag": {"cor": "orange", "simbolo": "triangle-up"},
                    "Overpull": {"cor": "black", "simbolo": "square"},
                    "Repasse": {"cor": "blue", "simbolo": "circle"},
                    "Coluna presa": {"cor": "red", "simbolo": "x"},
                    "Ameaça de prisão": {"cor": "purple", "simbolo": "star-square"},
                    "Coluna topou": {"cor": "gray", "simbolo": "diamond"},
                    "Ferramenta de perfilagem topou": {"cor": "black", "simbolo": "diamond-open-dot"},
                    "Perda parcial de circulação": {"cor": "orange", "simbolo": "triangle-down"},
                    "Perda severa de circulação": {"cor": "red", "simbolo": "triangle-down"},
                }

                # =========================
                # Estados (somente o calculado)
                # =========================
                if "traj_marks_calc" not in st.session_state:
                    st.session_state.traj_marks_calc = pd.DataFrame(columns=[
                        "Tipo", "MD Inicial", "MD Final",
                        "TVD Inicial", "TVD Final",
                        "HD Inicial", "HD Final",
                        "Evento", "_cor_plot", "_simbolo"
                    ])

                # Layout
                col_left, col_right = st.columns((1, 0.75))

                # =========================
                # ESQUERDA — eventos vindos do Excel
                # =========================
                with col_left:
                    df_eventos = st.session_state.get("df_eventos", None)

                    if not isinstance(df_eventos, pd.DataFrame) or df_eventos.empty:
                        st.info("Nenhum evento encontrado na aba **Eventos** do Excel (df_eventos vazio).")
                        st.session_state.traj_marks_calc = pd.DataFrame(columns=[
                            "Tipo", "MD Inicial", "MD Final",
                            "TVD Inicial", "TVD Final",
                            "HD Inicial", "HD Final",
                            "Evento", "_cor_plot", "_simbolo"
                        ])
                    else:
                        with st.expander("Ver eventos lidos do Excel", expanded=True):
                            st.dataframe(df_eventos, use_container_width=True, hide_index=True)

                        df_in = df_eventos.copy()

                        # normaliza colunas esperadas
                        if "MD Inicial" in df_in.columns:
                            df_in["MD Inicial"] = pd.to_numeric(df_in["MD Inicial"], errors="coerce")
                        else:
                            df_in["MD Inicial"] = np.nan

                        if "MD Final" in df_in.columns:
                            df_in["MD Final"] = pd.to_numeric(df_in["MD Final"], errors="coerce")
                        else:
                            df_in["MD Final"] = np.nan

                        if "Evento" not in df_in.columns:
                            df_in["Evento"] = ""

                        df_in["Evento"] = df_in["Evento"].astype("string").fillna("").str.strip()

                        df_in = df_in.dropna(subset=["MD Inicial"])
                        df_in = df_in[df_in["Evento"] != ""]

                        rows = []
                        linhas_com_erro = []

                        for _, r in df_in.iterrows():
                            try:
                                md_i = float(r["MD Inicial"])
                                md_f = r["MD Final"]
                                evento = str(r["Evento"]).strip()

                                estilo = MAPA_EVENTO_ESTILO.get(evento, {"cor": "red", "simbolo": "diamond"})
                                cor_plot = estilo["cor"]
                                simbolo = estilo["simbolo"]

                                # PONTO
                                if pd.isna(md_f):
                                    tvd_i, hd_i = _interp_em_tvd(df_traj, md_i)
                                    rows.append({
                                        "Tipo": "Ponto",
                                        "MD Inicial": md_i,
                                        "MD Final": np.nan,
                                        "TVD Inicial": tvd_i,
                                        "TVD Final": np.nan,
                                        "HD Inicial": hd_i,
                                        "HD Final": np.nan,
                                        "Evento": evento,
                                        "_cor_plot": cor_plot,
                                        "_simbolo": simbolo
                                    })

                                # TRECHO
                                else:
                                    md_f = float(md_f)
                                    md_a, md_b = (md_i, md_f) if md_f >= md_i else (md_f, md_i)

                                    tvd_a, hd_a = _interp_em_tvd(df_traj, md_a)
                                    tvd_b, hd_b = _interp_em_tvd(df_traj, md_b)

                                    rows.append({
                                        "Tipo": "Trecho",
                                        "MD Inicial": md_a,
                                        "MD Final": md_b,
                                        "TVD Inicial": tvd_a,
                                        "TVD Final": tvd_b,
                                        "HD Inicial": hd_a,
                                        "HD Final": hd_b,
                                        "Evento": evento,
                                        "_cor_plot": cor_plot,
                                        "_simbolo": simbolo
                                    })


                            except Exception as e:
                                linhas_com_erro.append({"MD Inicial": r.get("MD Inicial"),"MD Final": r.get("MD Final"),
                                                        "Evento": r.get("Evento"),"Erro": str(e)})

                        df_calc = pd.DataFrame(rows)
                        if not df_calc.empty:
                            df_calc = df_calc.sort_values(["Tipo", "MD Inicial"]).reset_index(drop=True)

                        st.session_state.traj_marks_calc = df_calc

                        if linhas_com_erro:
                            st.warning("Alguns eventos foram ignorados.")

                            df_erros = pd.DataFrame(linhas_com_erro)

                            st.dataframe(
                                df_erros,
                                use_container_width=True,
                                hide_index=True
                            )

                # =========================
                # DIREITA — gráfico
                # =========================
                with col_right:
                    with st.container(border=True):

                        if "Afastamento Horizontal (m)" in df_traj.columns:
                            hd_traj = df_traj["Afastamento Horizontal (m)"].to_numpy(dtype=float)
                        else:
                            hd_traj = np.sqrt(
                                df_traj["Easting"].to_numpy(float) ** 2 + df_traj["Northing"].to_numpy(float) ** 2
                            )

                        fig2d = go.Figure()

                        hd_min = float(np.nanmin(hd_traj)) if len(hd_traj) else 0.0
                        hd_max = float(np.nanmax(hd_traj)) if len(hd_traj) else 0.0
                        hd_span = max(hd_max - hd_min, 1.0)

                        # Faixa automática da litologia
                        largura_lito = max(170.0, 0.14 * hd_span)
                        margem_lito = max(10.0, 0.015 * hd_span)

                        x1_lito = hd_min - margem_lito
                        x0_lito = x1_lito - largura_lito

                        x_left = x0_lito - max(10.0, 0.02 * hd_span)
                        x_right = hd_max + max(10.0, 0.03 * hd_span)

                        # Trajetória (sempre vermelha), sem legenda
                        fig2d.add_trace(go.Scatter(
                            x=hd_traj,
                            y=df_traj["TVD"],
                            mode="lines",
                            line=dict(color="red", width=4),
                            name="Trajetória",
                            showlegend=False
                        ))

                        df_notes = st.session_state.traj_marks_calc

                        if isinstance(df_notes, pd.DataFrame) and not df_notes.empty:

                            # ========= TRECHOS: linhas (sem legenda)
                            df_trechos = df_notes[df_notes["Tipo"] == "Trecho"].copy()
                            for _, r in df_trechos.iterrows():
                                cor_plot = r["_cor_plot"]
                                md_a = float(r["MD Inicial"])
                                md_b = float(r["MD Final"])
                                try:
                                    hd_seg, tvd_seg, md_seg = _segmento_traj_por_md(df_traj, md_a, md_b)

                                    fig2d.add_trace(go.Scatter(
                                        x=hd_seg,
                                        y=tvd_seg,
                                        mode="lines",
                                        line=dict(width=6, color=cor_plot),
                                        showlegend=False,
                                        hovertemplate=(
                                            "Trecho<br>"
                                            f"Evento: {r['Evento']}<br>"
                                            "MD: %{customdata[0]:.1f} m<br>"
                                            "TVD: %{y:.1f} m<br>"
                                            "HD: %{x:.1f} m<extra></extra>"
                                        ),
                                        customdata=np.column_stack([md_seg]),
                                    ))
                                except Exception:
                                    pass

                            # ========= TRECHOS: marcador NO INÍCIO (sem legenda)
                            tmarkers = []
                            for _, r in df_trechos.iterrows():
                                try:
                                    md_ini = float(r["MD Inicial"])
                                    tvd_ini, hd_ini = _interp_em_tvd(df_traj, md_ini)

                                    tmarkers.append({
                                        "Evento": r["Evento"],
                                        "MD_ini": md_ini,
                                        "MD_fim": float(r["MD Final"]),
                                        "TVD": tvd_ini,
                                        "HD": hd_ini,
                                        "_cor_plot": r["_cor_plot"],
                                        "_simbolo": r["_simbolo"]
                                    })
                                except Exception:
                                    pass

                            if tmarkers:
                                df_tmark = pd.DataFrame(tmarkers).sort_values(["Evento", "MD_ini"]).reset_index(
                                    drop=True)

                                for evento, grp in df_tmark.groupby("Evento"):
                                    estilo = MAPA_EVENTO_ESTILO.get(evento, {"cor": "red", "simbolo": "diamond"})

                                    fig2d.add_trace(go.Scatter(
                                        x=grp["HD"],
                                        y=grp["TVD"],
                                        mode="markers",
                                        name=str(evento),
                                        showlegend=False,
                                        marker=dict(
                                            size=12,
                                            symbol=estilo["simbolo"],
                                            color=estilo["cor"],
                                            line=dict(color="black", width=1.5)
                                        ),
                                        customdata=np.column_stack([
                                            grp["MD_ini"].to_numpy(float),
                                            grp["MD_fim"].to_numpy(float),
                                        ]),
                                        hovertemplate=(
                                            "Trecho<br>"
                                            f"Evento: {evento}<br>"
                                            "Início: %{customdata[0]:.1f} m<br>"
                                            "Fim: %{customdata[1]:.1f} m<br>"
                                            "TVD: %{y:.1f} m<br>"
                                            "HD: %{x:.1f} m<extra></extra>"
                                        ),
                                    ))

                            # ========= PONTOS: marcadores no gráfico (sem legenda)
                            df_pontos = df_notes[df_notes["Tipo"] == "Ponto"].copy()

                            for evento, grp in df_pontos.groupby("Evento"):
                                estilo = MAPA_EVENTO_ESTILO.get(evento, {"cor": "red", "simbolo": "diamond"})

                                fig2d.add_trace(go.Scatter(
                                    x=grp["HD Inicial"],
                                    y=grp["TVD Inicial"],
                                    mode="markers",
                                    name=str(evento),
                                    showlegend=False,
                                    marker=dict(
                                        size=12,
                                        symbol=estilo["simbolo"],
                                        color=estilo["cor"],
                                        line=dict(color="black", width=1.5)
                                    ),
                                    customdata=np.column_stack([grp["MD Inicial"].to_numpy(float)]),
                                    hovertemplate=(
                                        "Ponto<br>"
                                        f"Evento: {evento}<br>"
                                        "MD: %{customdata[0]:.2f} m<br>"
                                        "TVD: %{y:.2f} m<br>"
                                        "HD: %{x:.2f} m<extra></extra>"
                                    ),
                                ))

                            # =========================
                            # LEGENDA ÚNICA POR EVENTO
                            # =========================
                            pontos_por_evento = {}
                            trechos_por_evento = {}

                            df_p = df_notes[df_notes["Tipo"] == "Ponto"].copy()
                            for evento, grp in df_p.groupby("Evento"):
                                mds = sorted(grp["MD Inicial"].astype(float).tolist())
                                pontos_por_evento[str(evento)] = mds

                            df_t = df_notes[df_notes["Tipo"] == "Trecho"].copy()
                            for evento, grp in df_t.groupby("Evento"):
                                pares = list(zip(
                                    grp["MD Inicial"].astype(float).tolist(),
                                    grp["MD Final"].astype(float).tolist()
                                ))
                                pares = sorted((min(a, b), max(a, b)) for a, b in pares)
                                trechos_por_evento[str(evento)] = pares

                            eventos_all = sorted(set(pontos_por_evento.keys()) | set(trechos_por_evento.keys()))

                            for evento in eventos_all:
                                estilo = MAPA_EVENTO_ESTILO.get(evento, {"cor": "red", "simbolo": "diamond"})

                                linhas = [evento]

                                if evento in pontos_por_evento and pontos_por_evento[evento]:
                                    for md in pontos_por_evento[evento]:
                                        linhas.append(f"{md:.2f} m")

                                if evento in trechos_por_evento and trechos_por_evento[evento]:
                                    for a, b in trechos_por_evento[evento]:
                                        linhas.append(f"{a:.2f}–{b:.2f} m")

                                nome_legenda = "<br>".join(linhas)

                                fig2d.add_trace(go.Scatter(
                                    x=[None],
                                    y=[None],
                                    mode="markers",
                                    name=nome_legenda,
                                    showlegend=True,
                                    marker=dict(
                                        size=12,
                                        symbol=estilo["simbolo"],
                                        color=estilo["cor"],
                                        line=dict(color="black", width=1.5)
                                    ),
                                    hoverinfo="skip"
                                ))

                        rig_sizex = max(40.0, 0.06 * hd_span)
                        rig_sizey = max(40.0, 0.06 * hd_span)

                        rig_img = Image.open("rig_t.png")
                        fig2d.add_layout_image(
                            dict(
                                source=rig_img,
                                xref="x",
                                yref="y",
                                x=0,
                                y=0,
                                sizex=rig_sizex,
                                sizey=rig_sizey,
                                xanchor="center",
                                yanchor="bottom",
                                layer="above"
                            )
                        )

                        def get_paleta_lito():
                            return {
                                "Argilito": {"bg": "#9ACD32", "simbol": "|"},
                                "Folhelho": {"bg": "#9ACD32", "simbol": "-"},
                                "Siltito": {"bg": "#A67B5B", "simbol": "-"},
                                "Arenito": {"bg": "#FFD580", "simbol": "."},
                                "Diamictito": {"bg": "#E97451", "simbol": "."},
                                "Conglomerado": {"bg": "#CD853F", "simbol": "."},
                                "Anidrita / Gipsita": {"bg": "#E6E6FA", "simbol": "/"},
                                "Halita": {"bg": "#FFFFFF", "simbol": "."},
                                "Marga": {"bg": "#EEE8AA", "simbol": "-"},
                                "Calcário": {"bg": "#B0C4DE", "simbol": "."},
                                "Basalto": {"bg": "#2F4F4F", "simbol": "+"},
                                "SDR": {"bg": "#FF7F50", "simbol": "."},
                                "Crosta Oceânica": {"bg": "#000080", "simbol": "."},
                            }

                        def add_lito_track_plotly(
                                fig,
                                profundidades,
                                litologias,
                                base_final,
                                x0, x1,
                                xaxis="x", yaxis="y",
                                show_labels=False
                        ):
                            paleta = get_paleta_lito()

                            for i in range(len(profundidades)):
                                z_top = float(profundidades[i])
                                z_base = float(profundidades[i + 1]) if i < len(profundidades) - 1 else float(
                                    base_final)

                                lit = litologias[i]
                                estilo = paleta.get(lit, {"bg": "#CCCCCC", "simbol": "."})

                                fig.add_trace(go.Scatter(
                                    x=[x0, x1, x1, x0, x0],
                                    y=[z_top, z_top, z_base, z_base, z_top],
                                    fill="toself",
                                    mode="lines",
                                    line=dict(color="black", width=1),
                                    fillpattern=dict(
                                        shape=estilo["simbol"],
                                        fgcolor="black",
                                        bgcolor=estilo["bg"],
                                        size=4,
                                        solidity=0.05
                                    ),
                                    showlegend=False,
                                    xaxis=xaxis,
                                    yaxis=yaxis,
                                    hoverinfo="skip"
                                ))

                                if show_labels:
                                    altura_intervalo = abs(z_base - z_top)

                                    if altura_intervalo >= 30:
                                        font_size = 11
                                    elif altura_intervalo >= 15:
                                        font_size = 10
                                    elif altura_intervalo >= 8:
                                        font_size = 9
                                    else:
                                        font_size = 8

                                    if altura_intervalo >= 6:
                                        fig.add_annotation(
                                            x=(x0 + x1) / 2,
                                            y=(z_top + z_base) / 2,
                                            text=f"<b>{lit}</b>",
                                            showarrow=False,
                                            xanchor="center",
                                            yanchor="middle",
                                            font=dict(size=font_size, color="black"),
                                            xref=xaxis,
                                            yref=yaxis,
                                            align="center"
                                        )

                        add_lito_track_plotly(
                            fig2d,
                            profundidades=profundidades,
                            litologias=litologias,
                            base_final=st.session_state.y_max_pp,
                            x0=x0_lito, x1=x1_lito,
                            xaxis="x", yaxis="y",
                            show_labels=True
                        )

                        sapatas_df = st.session_state.get("sapatas_df", None)

                        if sapatas_df is not None and not sapatas_df.empty:

                            tvd_traj = df_traj["TVD"].to_numpy(dtype=float)

                            if "Afastamento Horizontal (m)" in df_traj.columns:
                                afast_traj = df_traj["Afastamento Horizontal (m)"].to_numpy(dtype=float)
                            else:
                                afast_traj = np.sqrt(
                                    df_traj["Easting"].to_numpy(float) ** 2 + df_traj["Northing"].to_numpy(float) ** 2
                                )

                            # garante interpolação correta em TVD
                            ordem = np.argsort(tvd_traj)
                            tvd_traj = tvd_traj[ordem]
                            afast_traj = afast_traj[ordem]

                            sapatas_plot = []
                            for _, row in sapatas_df.iterrows():
                                prof_sapata = row.get("Profundidade da sapata (m)", None)
                                fase = row.get("Fase", "")

                                if prof_sapata is None or pd.isna(prof_sapata):
                                    continue

                                prof_sapata = float(prof_sapata)

                                if prof_sapata < tvd_traj.min() or prof_sapata > tvd_traj.max():
                                    continue

                                x_sapata = float(np.interp(prof_sapata, tvd_traj, afast_traj))

                                sapatas_plot.append({
                                    "prof": prof_sapata,
                                    "fase": fase,
                                    "x": x_sapata
                                })

                            sapatas_plot = sorted(sapatas_plot, key=lambda s: s["prof"])

                            linha_half = max(30.0, 0.015 * hd_span)

                            cores_sapatas = [
                                "#000000",
                                "#1f77b4",
                                "#2ca02c",
                                "#ff7f0e",
                                "#d62728",
                                "#9467bd",
                                "#8c564b",
                                "#e377c2",
                            ]

                            for i, s in enumerate(sapatas_plot):
                                prof_sapata = s["prof"]
                                fase = s["fase"]
                                x_sapata = s["x"]
                                cor_sapata = cores_sapatas[i % len(cores_sapatas)]

                                # linha horizontal da sapata na trajetória
                                fig2d.add_trace(
                                    go.Scatter(
                                        x=[x_sapata - linha_half, x_sapata + linha_half],
                                        y=[prof_sapata, prof_sapata],
                                        mode="lines",
                                        line=dict(
                                            color=cor_sapata,
                                            width=3,
                                            dash="dot"
                                        ),
                                        showlegend=False,
                                        hovertemplate=(
                                            f"Sapata {fase}<br>"
                                            f"Profundidade: {prof_sapata:.0f} m<extra></extra>"
                                        )
                                    )
                                )

                                # item de legenda da sapata
                                fig2d.add_trace(
                                    go.Scatter(
                                        x=[None],
                                        y=[None],
                                        mode="lines",
                                        line=dict(
                                            color=cor_sapata,
                                            width=3,
                                            dash="dot"
                                        ),
                                        name=f"Sapata {fase} - {prof_sapata:.0f} m",
                                        showlegend=True,
                                        hoverinfo="skip"
                                    )
                                )

                        fig2d.update_layout(
                            title=dict(
                                text="Vista lateral + Eventos",
                                x=0.5,
                                xanchor="center"
                            ),
                            height=700,
                            margin=dict(l=20, r=20, t=40, b=20),
                            xaxis=dict(
                                range=[x_left, x_right]
                            ),
                            yaxis=dict(
                                title="TVD (m)",
                                autorange="reversed"
                            ),
                            showlegend=True,
                            legend=dict(
                                orientation="v",
                                yanchor="top",
                                y=1,
                                xanchor="left",
                                x=1.02
                            )
                        )

                        st.session_state.fig2d = fig2d
                        st.plotly_chart(fig2d, use_container_width=True)

        with st.container(border=True):
            st.markdown("## Anotações")
            with st.expander("Anotações do usuário", expanded=True):
                st.session_state.anotacoes = st.text_area(
                    "Registre aqui observações, decisões e pendências do estudo:",
                    value=st.session_state.anotacoes,
                    height=260,
                    placeholder="Ex.: Ajustar coesão na formação X; revisar LOT do poço Y; confirmar densidade do fluido na fase Z..."
                )

                col1, col2 = st.columns(2)
                with col1:
                    if st.button("💾 Salvar anotações", use_container_width=True, type="primary"):
                        st.toast("Anotações salvas.")
                with col2:
                    if st.button("🗑️ Limpar", use_container_width=True):
                        st.session_state.anotacoes = ""
                        st.rerun()

    # #RTP
    # with tabs[6]:
    #     st.write("Em andamento")

    # Relatório
    with tabs[6]:

        col_pdf1, col_pdf2, col_pdf3, col_pdf4, col_pdf5 = st.columns((0.4, 0.2, 1, 0.2, 0.3))

        with col_pdf1:

            if st.session_state.well_name == '':
                report_name = 'Relatorio_Final.pdf'
            else:
                report_name = f'{st.session_state.well_name}.pdf'

            view = st.button(':bookmark_tabs: Ver Relatório', key='pdf_view_bt', use_container_width=True)

            if view:
                with st.spinner("Gerando relatório..."):
                    st.session_state.pdf_bytes = gerar_relatorio_pdf()
                    st.session_state.pdf_ready = True
                    st.session_state.pdf_view_open = True

            if st.session_state.pdf_ready and st.session_state.pdf_bytes is not None:
                st.download_button(
                    label="⬇️ Baixar Relatório",
                    data=st.session_state.pdf_bytes,
                    file_name=report_name,
                    mime="application/pdf",
                    use_container_width=True
                )

        with col_pdf3:
            if st.session_state.pdf_view_open and st.session_state.pdf_bytes is not None:
                container_pdf = st.container(border=True, height=900)
                with container_pdf:
                    pdf_viewer(
                        input=st.session_state.pdf_bytes,
                        width=700,
                        pages_vertical_spacing=20
                    )

    # Informações Gerais
    with tabs[7]:
        st.title("📘 Informações Gerais do SYGA")

        # --- Dicionário com informações dos perfis ---
        parameters = {
            "Curvas de geopressões": {
                "Gradiente": {
                    "Sigla": ["Gradiente"],
                    "Nome": [
                        "Variação de pressão ao longo de um intervalo de profundidade, medido com base em um determinado referencial (Datum)"],
                    "Unidade": "massa/volume"
                },
                "Sobrecarga": {
                    "Sigla": ["Sobrecarga"],
                    "Nome": ["Pressão gerada em uma certa porfundidade devido ao"
                             " peso das camadass de rochas sobrepostas (pode ser expressa em gradiente)"],
                    "Unidade": "pressão ou massa/volume"
                },

                "Pressão de poros": {
                    "Sigla": ["Pressão de poros"],
                    "Nome": ["Pressão contida nos poros da rocha (normalmente expressada em gradiente)"],
                    "Unidade": "pressão ou massa/volume"
                },

                "Pressão de fratura": {
                    "Sigla": ["Pressão de fratura"],
                    "Nome": ["Pressão necessária para levar a falha da formação por tração,"
                             " podendo ser fratura superior ou inferior (normalmente expressada em gradiente)"],
                    "Unidade": "pressão ou massa/volume"
                },

                "Pressão de colapso": {
                    "Sigla": ["Pressão de colpaso"],
                    "Nome": ["Pressão necessária para levar a falha da formação por cisalhamento"
                             " podendo ser colapso superior ou inferior (pode ser expressa em gradiente)"],
                    "Unidade": "pressão ou massa/volume"
                },
            },
            "Perfil": {
                "Sônico compressional": {
                    "Sigla": ["DT", "DTCO"],
                    "Nome": ["Delta-T Compressional"],
                    "Unidade": "µs/ft"
                },
                "Sônico cisalhante": {
                    "Sigla": ["DTS", "DTSM"],
                    "Nome": ["Compressional Wave Transit Time", "Sonic Transit Time"],
                    "Unidade": "µs/ft"
                },
                "Densidade": {
                    "Sigla": ["RHOB", "DEN"],
                    "Nome": ["Standard Resolution Formation Density", "Bulk density"],
                    "Unidade": "g/cm³"
                },
                "Gamma-Ray": {
                    "Sigla": ["GR"],
                    "Nome": ["Gamma Ray"],
                    "Unidade": "gAPI"
                }
            }
        }

        # --- Layout com colunas ---
        col1, col2 = st.columns([1, 2])

        with col1:
            cat = st.selectbox(
                "Selecione a categoria:",
                options=list(parameters.keys())
            )
            perfil_selecionado = st.selectbox(
                "Selecione um perfil:",
                options=list(parameters[cat].keys())
            )

        with col2:
            dados = parameters[cat][perfil_selecionado]

            # Converte listas em string separada por vírgula

            if cat == "Curvas de geopressões":
                op1 = 'Nome'
                op2 = 'Definição'
                t = 'Curvas de Geopressões'
            else:
                op1 = 'Sigla(s)'
                op2 = 'Nome(s)'
                t = 'Perfil'
            siglas = ", ".join(dados["Sigla"])
            nomes = ", ".join(dados["Nome"])

            st.markdown(
                f"""
                ### ℹ️ {t} - **{perfil_selecionado}**

                - **{op1}:** `{siglas}`
                - **{op2}:** {nomes}
                - **Unidade:** *{dados['Unidade']}*
                """,
                unsafe_allow_html=True
            )

        # Um toque estético extra (barra de separação)
        st.markdown("---")
        st.info("💡 Dica: selecione um perfil na caixa à esquerda para visualizar os detalhes.")

geo_page()
