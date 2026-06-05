import io
import os
import re
import utm
import yaml
import math
import folium
import pycountry
import unicodedata
import numpy as np
import pandas as pd
from PIL import Image
from io import BytesIO
import streamlit as st
from textwrap import wrap
import statsmodels.api as sm
from fractions import Fraction
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import plotly.graph_objects as go
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
import matplotlib.patheffects as pe
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from streamlit_folium import st_folium
import streamlit_antd_components as sac
from datetime import datetime, timedelta
from plotly.subplots import make_subplots
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
from streamlit_pdf_viewer import pdf_viewer
from scipy.interpolate import griddata, Rbf
from scipy.ndimage import gaussian_filter1d
import streamlit.components.v1 as components
import matplotlib.patheffects as path_effects
from streamlit_js_eval import get_geolocation
from matplotlib.patches import Rectangle, Polygon
from matplotlib.offsetbox import OffsetImage, AnnotationBbox
from statsmodels.nonparametric.smoothers_lowess import lowess
from matplotlib.offsetbox import OffsetImage, AnnotationBbox, TextArea, VPacker
from matplotlib.ticker import FixedLocator, FuncFormatter, LogLocator, NullFormatter

# Configurações da página web
logo = 'logo.png'
img_logo = Image.open(logo)

cab = 'logo_syng.png'
img_cab = Image.open(cab)

image = Image.open(logo)

PAGE_CONFIG = {
    "page_title": "SYGA",
    "page_icon": image,
    "layout": "wide",
    "initial_sidebar_state": "expanded",
}

st.set_page_config(**PAGE_CONFIG)

st.markdown(
    """
    <style>
    [data-testid="stSidebar"][aria-expanded="true"] {
        width: 370px;
        min-width: 370px;
        max-width: 370px;
        overflow-x: hidden;
    }

    [data-testid="stSidebar"][aria-expanded="true"] > div:first-child {
        width: 370px;
        max-width: 370px;
        overflow-x: hidden;
        box-sizing: border-box;
    }

    [data-testid="stSidebar"][aria-expanded="false"] {
        width: 0 !important;
        min-width: 0 !important;
        max-width: 0 !important;
        overflow: visible;
    }

    [data-testid="stSidebar"][aria-expanded="false"] > div:first-child {
        width: 0 !important;
        min-width: 0 !important;
        max-width: 0 !important;
    }

    [data-testid="collapsedControl"] {
        display: flex !important;
        visibility: visible !important;
        z-index: 100000;
    }

    [data-testid="stSidebar"] * {
        box-sizing: border-box;
    }

    [data-testid="stSidebar"] .ant-tree,
    [data-testid="stSidebar"] .ant-tree-list,
    [data-testid="stSidebar"] .ant-tree-list-holder,
    [data-testid="stSidebar"] .ant-tree-list-holder-inner {
        width: 100% !important;
        max-width: 100% !important;
        overflow-x: hidden !important;
    }

    [data-testid="stSidebar"] .ant-tree-treenode,
    [data-testid="stSidebar"] .ant-tree-node-content-wrapper {
        max-width: 100% !important;
        overflow: hidden !important;
    }

    [data-testid="stSidebar"] .ant-tree-node-content-wrapper {
        padding-right: 6px !important;
    }

    [data-testid="stSidebar"] .ant-tree-title {
        display: inline-block;
        max-width: 100%;
        overflow-wrap: anywhere;
        white-space: normal;
        line-height: 1.25;
    }

    .info-card {
        border: 1px solid rgba(49, 51, 63, 0.18);
        border-radius: 8px;
        padding: 0.75rem 0.9rem;
        background: rgba(250, 250, 250, 0.75);
        min-height: 72px;
        margin-bottom: 0.75rem;
    }

    .info-card-label {
        color: #6b7280;
        font-size: 0.78rem;
        font-weight: 600;
        margin-bottom: 0.25rem;
    }

    .info-card-value {
        color: #111827;
        font-size: 0.98rem;
        font-weight: 600;
        overflow-wrap: anywhere;
    }
    
    </style>
    """,
    unsafe_allow_html=True
)


def instalar_persistencia_expanders():
    components.html(
        """
        <script>
        (() => {
            const STORAGE_KEY = "syga_expanders_state_v1";
            const root = window.parent.document;

            const normalize = (text) => (text || "").replace(/\\s+/g, " ").trim();

            const getPageKey = () => {
                const selected =
                    root.querySelector(".ant-tree-node-selected .ant-tree-title") ||
                    root.querySelector('[aria-selected="true"]') ||
                    root.querySelector('[data-testid="stSidebar"] [aria-current="page"]');
                const header = root.querySelector("main h1, main h2, main h3");

                return normalize(selected?.textContent) ||
                    normalize(header?.textContent) ||
                    window.parent.location.pathname ||
                    "SYGA";
            };

            const readState = () => {
                try {
                    return JSON.parse(window.parent.localStorage.getItem(STORAGE_KEY) || "{}");
                } catch (_) {
                    return {};
                }
            };

            const writeState = (state) => {
                try {
                    window.parent.localStorage.setItem(STORAGE_KEY, JSON.stringify(state));
                } catch (_) {}
            };

            const expanderLabel = (details) => {
                const summary = details.querySelector("summary");
                return normalize(summary?.innerText || summary?.textContent);
            };

            const expanderKey = (details, index) => {
                const label = expanderLabel(details) || `Expander ${index + 1}`;
                return `${getPageKey()}::${index}::${label}`;
            };

            const syncExpanders = () => {
                const expanders = [
                    ...new Set([
                        ...root.querySelectorAll('details[data-testid="stExpander"]'),
                        ...root.querySelectorAll('[data-testid="stExpander"] details'),
                    ])
                ];
                const state = readState();

                expanders.forEach((details, index) => {
                    const key = expanderKey(details, index);

                    if (Object.prototype.hasOwnProperty.call(state, key)) {
                        const shouldOpen = Boolean(state[key]);
                        if (details.open !== shouldOpen) {
                            details.open = shouldOpen;
                        }
                    }

                    if (details.dataset.sygaExpanderPersistent === "true") {
                        return;
                    }

                    details.dataset.sygaExpanderPersistent = "true";
                    details.addEventListener("toggle", () => {
                        const latestState = readState();
                        latestState[expanderKey(details, index)] = details.open;
                        writeState(latestState);
                    });
                });
            };

            window.parent.__sygaExpanderSync = syncExpanders;

            syncExpanders();
            setTimeout(syncExpanders, 100);
            setTimeout(syncExpanders, 500);

            if (!window.parent.__sygaExpanderPersistenceInstalled) {
                window.parent.__sygaExpanderPersistenceInstalled = true;
                new MutationObserver(() => {
                    window.parent.clearTimeout(window.parent.__sygaExpanderPersistenceTimer);
                    window.parent.__sygaExpanderPersistenceTimer = window.parent.setTimeout(() => {
                        window.parent.__sygaExpanderSync?.();
                    }, 80);
                }).observe(root.body, { childList: true, subtree: true });
            }
        })();
        </script>
        """,
        height=0,
        width=0,
    )


instalar_persistencia_expanders()

st.image(img_cab, width=2000)

paises = {
    "Afeganistão": "af",
    "África do Sul": "za",
    "Albânia": "al",
    "Alemanha": "de",
    "Andorra": "ad",
    "Angola": "ao",
    "Antígua e Barbuda": "ag",
    "Arábia Saudita": "sa",
    "Argélia": "dz",
    "Argentina": "ar",
    "Armênia": "am",
    "Austrália": "au",
    "Áustria": "at",
    "Azerbaijão": "az",
    "Bahamas": "bs",
    "Bahrein": "bh",
    "Bangladesh": "bd",
    "Barbados": "bb",
    "Bélgica": "be",
    "Belize": "bz",
    "Benim": "bj",
    "Bielorrússia": "by",
    "Bolívia": "bo",
    "Bósnia e Herzegovina": "ba",
    "Botsuana": "bw",
    "Brasil": "br",
    "Brunei": "bn",
    "Bulgária": "bg",
    "Burkina Faso": "bf",
    "Burundi": "bi",
    "Butão": "bt",
    "Cabo Verde": "cv",
    "Camarões": "cm",
    "Camboja": "kh",
    "Canadá": "ca",
    "Catar": "qa",
    "Cazaquistão": "kz",
    "Chade": "td",
    "Chile": "cl",
    "China": "cn",
    "Chipre": "cy",
    "Colômbia": "co",
    "Comores": "km",
    "Congo": "cg",
    "Coreia do Norte": "kp",
    "Coreia do Sul": "kr",
    "Costa do Marfim": "ci",
    "Costa Rica": "cr",
    "Croácia": "hr",
    "Cuba": "cu",
    "Dinamarca": "dk",
    "Djibuti": "dj",
    "Dominica": "dm",
    "Egito": "eg",
    "El Salvador": "sv",
    "Emirados Árabes Unidos": "ae",
    "Equador": "ec",
    "Eritreia": "er",
    "Eslováquia": "sk",
    "Eslovênia": "si",
    "Espanha": "es",
    "Estados Unidos": "us",
    "Estônia": "ee",
    "Etiópia": "et",
    "Fiji": "fj",
    "Filipinas": "ph",
    "Finlândia": "fi",
    "França": "fr",
    "Gabão": "ga",
    "Gâmbia": "gm",
    "Gana": "gh",
    "Geórgia": "ge",
    "Granada": "gd",
    "Grécia": "gr",
    "Guatemala": "gt",
    "Guiana": "gy",
    "Guiné": "gn",
    "Guiné-Bissau": "gw",
    "Guiné Equatorial": "gq",
    "Haiti": "ht",
    "Honduras": "hn",
    "Hungria": "hu",
    "Iêmen": "ye",
    "Ilhas Marshall": "mh",
    "Ilhas Salomão": "sb",
    "Índia": "in",
    "Indonésia": "id",
    "Irã": "ir",
    "Iraque": "iq",
    "Irlanda": "ie",
    "Islândia": "is",
    "Israel": "il",
    "Itália": "it",
    "Jamaica": "jm",
    "Japão": "jp",
    "Jordânia": "jo",
    "Kiribati": "ki",
    "Kuwait": "kw",
    "Laos": "la",
    "Lesoto": "ls",
    "Letônia": "lv",
    "Líbano": "lb",
    "Libéria": "lr",
    "Líbia": "ly",
    "Liechtenstein": "li",
    "Lituânia": "lt",
    "Luxemburgo": "lu",
    "Macedônia do Norte": "mk",
    "Madagascar": "mg",
    "Malásia": "my",
    "Malawi": "mw",
    "Maldivas": "mv",
    "Mali": "ml",
    "Malta": "mt",
    "Marrocos": "ma",
    "Maurício": "mu",
    "Mauritânia": "mr",
    "México": "mx",
    "Micronésia": "fm",
    "Moçambique": "mz",
    "Moldávia": "md",
    "Mônaco": "mc",
    "Mongólia": "mn",
    "Montenegro": "me",
    "Myanmar": "mm",
    "Namíbia": "na",
    "Nauru": "nr",
    "Nepal": "np",
    "Nicarágua": "ni",
    "Níger": "ne",
    "Nigéria": "ng",
    "Noruega": "no",
    "Nova Zelândia": "nz",
    "Omã": "om",
    "Países Baixos": "nl",
    "Palau": "pw",
    "Panamá": "pa",
    "Papua-Nova Guiné": "pg",
    "Paquistão": "pk",
    "Paraguai": "py",
    "Peru": "pe",
    "Polônia": "pl",
    "Portugal": "pt",
    "Quênia": "ke",
    "Quirguistão": "kg",
    "Reino Unido": "gb",
    "República Centro-Africana": "cf",
    "República Dominicana": "do",
    "Romênia": "ro",
    "Ruanda": "rw",
    "Rússia": "ru",
    "Samoa": "ws",
    "San Marino": "sm",
    "Santa Lúcia": "lc",
    "São Cristóvão e Nevis": "kn",
    "São Tomé e Príncipe": "st",
    "São Vicente e Granadinas": "vc",
    "Senegal": "sn",
    "Serra Leoa": "sl",
    "Sérvia": "rs",
    "Seychelles": "sc",
    "Singapura": "sg",
    "Síria": "sy",
    "Somália": "so",
    "Sri Lanka": "lk",
    "Sudão": "sd",
    "Sudão do Sul": "ss",
    "Suécia": "se",
    "Suíça": "ch",
    "Suriname": "sr",
    "Tailândia": "th",
    "Tajiquistão": "tj",
    "Tanzânia": "tz",
    "Timor-Leste": "tl",
    "Togo": "tg",
    "Tonga": "to",
    "Trinidad e Tobago": "tt",
    "Tunísia": "tn",
    "Turcomenistão": "tm",
    "Turquia": "tr",
    "Tuvalu": "tv",
    "Ucrânia": "ua",
    "Uganda": "ug",
    "Uruguai": "uy",
    "Uzbequistão": "uz",
    "Vanuatu": "vu",
    "Vaticano": "va",
    "Venezuela": "ve",
    "Vietnã": "vn",
    "Zâmbia": "zm",
    "Zimbábue": "zw"
}


@st.cache_data(show_spinner=False)
def carregar_workbook(file_bytes):
    buffer = io.BytesIO(file_bytes)
    wb = load_workbook(buffer, data_only=True)
    return wb


def exibir_campo_info(rotulo, valor):
    valor_exibido = valor if valor not in (None, "") else "Não informado"

    st.markdown(
        f"""
        <div class="info-card">
            <div class="info-card-label">{rotulo}</div>
            <div class="info-card-value">{valor_exibido}</div>
        </div>
        """,
        unsafe_allow_html=True
    )


def _ler_trajetoria_do_xlsm(wb, modo: str) -> pd.DataFrame:
    """
    Lê diretamente da aba Trajetória:
    - Planejada: MD/Incl/Azimute/TVD em B/C/D/E
    - Executada: MD/Incl/Azimute/TVD em K/L/M/N
    """

    if "Trajetória" not in wb.sheetnames:
        raise ValueError("A aba 'Trajetória' não existe no arquivo.")

    ws = wb["Trajetória"]

    if modo == "Executada":
        col_md, col_inc, col_azi, col_tvd = "K", "L", "M", "N"
    else:
        col_md, col_inc, col_azi, col_tvd = "B", "C", "D", "E"

    h_md = ws[f"{col_md}6"].value
    h_inc = ws[f"{col_inc}6"].value
    h_azi = ws[f"{col_azi}6"].value
    h_tvd = ws[f"{col_tvd}6"].value

    def _norm(x):
        return str(x).strip().lower() if x is not None else ""

    if not (
        "md" in _norm(h_md)
        and ("incl" in _norm(h_inc) or "inc" in _norm(h_inc))
        and ("az" in _norm(h_azi) or "azim" in _norm(h_azi))
        and "tvd" in _norm(h_tvd)
    ):
        raise ValueError(
            f"Header não bate no esperado. Lido em {col_md}6:{col_tvd}6 -> "
            f"{[h_md, h_inc, h_azi, h_tvd]}. "
            "Confirme se a aba Trajetória possui MD, Incl, Azimute e TVD."
        )

    rows = []
    r = 7

    while r <= ws.max_row:
        md = ws[f"{col_md}{r}"].value
        inc = ws[f"{col_inc}{r}"].value
        azi = ws[f"{col_azi}{r}"].value
        tvd = ws[f"{col_tvd}{r}"].value

        if md is None or md == "":
            break

        rows.append((md, tvd, inc, azi))
        r += 1

    df = pd.DataFrame(rows, columns=["MD", "TVD", "Inc", "Azi"])

    df["MD"] = pd.to_numeric(df["MD"], errors="coerce")
    df["TVD"] = pd.to_numeric(df["TVD"], errors="coerce")
    df["Inc"] = pd.to_numeric(df["Inc"], errors="coerce")
    df["Azi"] = pd.to_numeric(df["Azi"], errors="coerce")

    df = df.dropna(subset=["MD", "TVD"]).sort_values("MD").reset_index(drop=True)

    if df.empty:
        raise ValueError("Nenhum dado válido encontrado na aba Trajetória.")

    if (df["MD"].diff().fillna(1) <= 0).any():
        raise ValueError("Coluna MD deve ser estritamente crescente.")

    if (df["TVD"].diff().fillna(1) < 0).any():
        raise ValueError("Coluna TVD deve ser crescente.")

    return df


def _limitar_perfilagem_ao_tvd_final(
    df_full: pd.DataFrame,
    tvd_final: float,
    col_tvd: str = "Profundidade"
) -> pd.DataFrame:
    df = df_full.copy()
    df.columns = [str(c).strip() for c in df.columns]

    if col_tvd not in df.columns:
        raise ValueError(f"A coluna '{col_tvd}' não existe na aba Perfilagens.")

    df[col_tvd] = pd.to_numeric(df[col_tvd], errors="coerce")
    df = df.dropna(subset=[col_tvd]).sort_values(col_tvd).reset_index(drop=True)

    if df.empty:
        raise ValueError("A aba Perfilagens ficou vazia após limpar a coluna de profundidade.")

    tvd_final = float(tvd_final)

    if tvd_final <= df[col_tvd].min():
        raise ValueError(
            f"O TVD final do poço ({tvd_final:.2f} m) é menor ou igual ao início da perfilagem "
            f"({df[col_tvd].min():.2f} m)."
        )

    # Se a perfilagem já termina antes do poço, não corta nada
    if df[col_tvd].max() <= tvd_final:
        return df.reset_index(drop=True)

    # Tenta converter colunas numéricas que vieram como texto
    for col in df.columns:
        if col == col_tvd:
            continue

        serie_num = pd.to_numeric(df[col], errors="coerce")

        # Só substitui se houver pelo menos algum valor numérico real
        if serie_num.notna().sum() > 0:
            df[col] = serie_num

    df_cortado = df[df[col_tvd] <= tvd_final].copy()

    # Garante uma linha exatamente no TVD final
    prof_ultima = float(df_cortado[col_tvd].iloc[-1])

    if not np.isclose(prof_ultima, tvd_final):
        linha_final = {}

        for col in df.columns:
            if col == col_tvd:
                linha_final[col] = tvd_final

            elif pd.api.types.is_numeric_dtype(df[col]):
                df_valid = df[[col_tvd, col]].dropna()

                if len(df_valid) >= 2:
                    linha_final[col] = float(np.interp(
                        tvd_final,
                        df_valid[col_tvd].to_numpy(dtype=float),
                        df_valid[col].to_numpy(dtype=float)
                    ))
                elif len(df_valid) == 1:
                    linha_final[col] = df_valid[col].iloc[0]
                else:
                    linha_final[col] = np.nan

            else:
                valores_anteriores = df.loc[df[col_tvd] <= tvd_final, col].dropna()
                linha_final[col] = (
                    valores_anteriores.iloc[-1]
                    if not valores_anteriores.empty
                    else np.nan
                )

        df_cortado = pd.concat(
            [df_cortado, pd.DataFrame([linha_final])],
            ignore_index=True
        )

    df_cortado = df_cortado.sort_values(col_tvd).reset_index(drop=True)

    return df_cortado


def _gerar_df_interp_a_partir_df1_df2(df1: pd.DataFrame, df2: pd.DataFrame) -> pd.DataFrame:
    """
    Gera df_interp usando a trajetória calculada por mínima curvatura.
    Interpola Inc em função do TVD da perfilagem.
    Mantém Azi em degrau, usando o último azimute conhecido.
    """
    expand_from_zero = (st.session_state.get("ex", "Desativada") == "Ativada")

    col_prof1 = [col for col in df1.columns if "profund" in col.lower()]

    if not col_prof1:
        raise ValueError("Coluna de profundidade não encontrada em df1.")

    prof1_col = col_prof1[0]

    df_out = df2.copy()

    tvd_out = pd.to_numeric(df_out["TVD"], errors="coerce").to_numpy(dtype=float)
    md_out = pd.to_numeric(df_out["MD"], errors="coerce").to_numpy(dtype=float)
    inc_col = "Inclinação (°)" if "Inclinação (°)" in df_out.columns else "Inc"
    azi_col = "Azimute (°)" if "Azimute (°)" in df_out.columns else "Azi"

    inc_out = pd.to_numeric(df_out[inc_col], errors="coerce").to_numpy(dtype=float)
    azi_out = pd.to_numeric(df_out[azi_col], errors="coerce").to_numpy(dtype=float)

    sort_idx = np.argsort(tvd_out)
    tvd_sorted = tvd_out[sort_idx]
    md_sorted = md_out[sort_idx]
    inc_sorted = inc_out[sort_idx]
    azi_sorted = azi_out[sort_idx]

    valid = (
        np.isfinite(tvd_sorted)
        & np.isfinite(md_sorted)
        & np.isfinite(inc_sorted)
        & np.isfinite(azi_sorted)
    )

    tvd_sorted = tvd_sorted[valid]
    md_sorted = md_sorted[valid]
    inc_sorted = inc_sorted[valid]
    azi_sorted = azi_sorted[valid]

    if len(tvd_sorted) == 0:
        raise ValueError("df_out da trajetória ficou vazio após limpeza dos dados.")

    tvd1_original = pd.to_numeric(df1[prof1_col], errors="coerce").dropna().to_numpy(dtype=float)
    tvd1_original = np.sort(tvd1_original)

    if expand_from_zero:
        first_depth = tvd1_original[0]
        tvd1_extra = np.arange(0, first_depth, 1.0)
        tvd1 = np.concatenate((tvd1_extra, tvd1_original))
    else:
        tvd1 = tvd1_original.copy()

    md_interp = np.interp(tvd1, tvd_sorted, md_sorted)
    inc_interp = np.interp(tvd1, tvd_sorted, inc_sorted)

    idx = np.searchsorted(tvd_sorted, tvd1, side="right") - 1
    idx = np.clip(idx, 0, len(azi_sorted) - 1)
    azi_interp = azi_sorted[idx]

    df_interp = pd.DataFrame({
        "Profundidade": tvd1,
        "MD": md_interp,
        "Inc (°)": inc_interp,
        "Azi (°)": azi_interp
    })

    return df_interp


def _ler_inicio_do_xlsm(wb) -> dict:
    """
    Lê valores fixos na aba 'Início' via células:
      - D5  -> Nome do poço
      - D10 -> Objetivo (comments)
      - D7  -> Easting
      - D6  -> Northing
      - B9  -> Hemisfério
      - B8  -> Zona UTM
    """
    if "Início" not in wb.sheetnames and "Inicio" not in wb.sheetnames:
        raise ValueError("A aba 'Início' (ou 'Inicio') não existe no arquivo.")

    ws = wb["Início"] if "Início" in wb.sheetnames else wb["Inicio"]

    nome_poco = ws["D5"].value
    objetivo = ws["D10"].value
    easting = ws["D7"].value
    northing = ws["D6"].value
    zona = ws["D8"].value
    hem = ws["D9"].value

    def _to_float(v):
        try:
            if v is None or v == "":
                return None
            return float(v)
        except Exception:
            return None

    def _to_int(v):
        try:
            if v is None or v == "":
                return None
            return int(float(v))
        except Exception:
            return None

    def _to_str(v):
        if v is None or v == "":
            return None
        return str(v).strip()

    return {
        "poco": None if nome_poco in (None, "") else str(nome_poco).strip(),
        "comments": "" if objetivo in (None, "") else str(objetivo).strip(),
        "easting": _to_float(easting),
        "northing": _to_float(northing),
        "zona": _to_int(zona),
        "hem": _to_str(hem),
    }


def _ler_peso_fluido_do_xlsm(wb) -> pd.DataFrame:
    if "Fluido" in wb.sheetnames:
        ws = wb["Fluido"]
    elif "Geopressões" in wb.sheetnames:
        ws = wb["Geopressões"]
    elif "Geopressoes" in wb.sheetnames:
        ws = wb["Geopressoes"]
    else:
        raise ValueError("A aba 'Fluido' não existe no arquivo.")

    h_md = ws["B5"].value
    h_wp = ws["C5"].value
    h_we = ws["D5"].value

    def _norm(x):
        return str(x).strip().lower() if x is not None else ""

    if "prof" not in _norm(h_md):
        raise ValueError(f"Header inesperado em B5: {h_md}")
    if "planej" not in _norm(h_wp):
        raise ValueError(f"Header inesperado em C5: {h_wp}")
    if "execut" not in _norm(h_we):
        raise ValueError(f"Header inesperado em D5: {h_we}")

    rows = []
    vazio_seguidos = 0

    for md, wp, we in ws.iter_rows(
        min_row=6,
        max_row=min(ws.max_row, 10000),
        min_col=2,   # B
        max_col=4,   # D
        values_only=True
    ):
        if md in (None, ""):
            vazio_seguidos += 1
            if vazio_seguidos >= 3:
                break
            continue

        vazio_seguidos = 0
        rows.append((md, wp, we))

    if not rows:
        raise ValueError("Não encontrei dados de peso do fluido em B6/C6/D6...")

    df = pd.DataFrame(
        rows,
        columns=[
            "Profundidade (m)",
            "Peso do Fluido Planejado (lb/gal)",
            "Peso do Fluido Executado (lb/gal)",
        ],
    )

    df = df.apply(pd.to_numeric, errors="coerce")
    df = df.dropna(subset=["Profundidade (m)"]).sort_values("Profundidade (m)").reset_index(drop=True)

    if df.empty:
        raise ValueError("Não encontrei dados válidos de peso do fluido em B6/C6/D6...")

    return df


def _ler_sapatas_do_xlsm(wb) -> pd.DataFrame:
    """
    Lê as sapatas na aba 'Início' (ou 'Inicio').

    Células:
      - TVD: E13, E14, E15 e opcionalmente E16 e E17
      - OD : F13, F14, F15 e opcionalmente F16 e F17

    Retorna DataFrame com:
      - Ordem
      - Fase
      - Profundidade da sapata (m)
    """
    if "Início" not in wb.sheetnames and "Inicio" not in wb.sheetnames:
        raise ValueError("A aba 'Início' (ou 'Inicio') não existe no arquivo.")

    ws = wb["Início"] if "Início" in wb.sheetnames else wb["Inicio"]

    def _to_float(v):
        try:
            if v is None or str(v).strip() == "":
                return None
            return float(v)
        except Exception:
            return None

    rows = []

    for ordem, r in enumerate(range(13, 18), start=1):
        tvd = _to_float(ws[f"E{r}"].value)
        od = _to_float(ws[f"F{r}"].value)

        if tvd is not None and tvd > 0:
            rows.append((ordem, od, tvd))

    if not rows:
        raise ValueError("Nenhuma sapata válida foi encontrada em E13:E16 / F13:F16 da aba 'Início'.")

    sapatas_df = pd.DataFrame(
        rows,
        columns=["Ordem", "Fase", "Profundidade da sapata (m)"]
    )

    sapatas_df["Ordem"] = pd.to_numeric(sapatas_df["Ordem"], errors="coerce").astype(int)
    sapatas_df["Fase"] = pd.to_numeric(sapatas_df["Fase"], errors="coerce")
    sapatas_df["Profundidade da sapata (m)"] = pd.to_numeric(
        sapatas_df["Profundidade da sapata (m)"], errors="coerce"
    )

    sapatas_df = sapatas_df.dropna(subset=["Profundidade da sapata (m)"])
    sapatas_df = sapatas_df[sapatas_df["Profundidade da sapata (m)"] > 0].copy()

    sapatas_df["Fase"] = sapatas_df["Fase"].fillna(0).astype(float).round(3)
    sapatas_df["Profundidade da sapata (m)"] = sapatas_df["Profundidade da sapata (m)"].astype(float)

    return sapatas_df.reset_index(drop=True)


def _ler_fases_do_xlsm(wb) -> pd.DataFrame:
    """
    Lê as fases (diâmetro das brocas) na aba 'Início' (ou 'Inicio').

    Células:
      - C21, C22, C23, C24, C25

    Retorna DataFrame com:
      - Ordem
      - Fase
    """
    if "Início" not in wb.sheetnames and "Inicio" not in wb.sheetnames:
        raise ValueError("A aba 'Início' (ou 'Inicio') não existe no arquivo.")

    ws = wb["Início"] if "Início" in wb.sheetnames else wb["Inicio"]

    def _to_float(v):
        try:
            if v is None or str(v).strip() == "":
                return None
            return float(v)
        except Exception:
            return None

    rows = []

    for ordem, r in enumerate(range(21, 26), start=1):
        fase = _to_float(ws[f"C{r}"].value)

        if fase is not None and fase > 0:
            rows.append((ordem, fase))

    if not rows:
        raise ValueError("Nenhuma fase válida foi encontrada em C21:C24 da aba 'Início'.")

    fases_df = pd.DataFrame(rows, columns=["Ordem", "Fase"])
    fases_df["Ordem"] = pd.to_numeric(fases_df["Ordem"], errors="coerce").astype(int)
    fases_df["Fase"] = pd.to_numeric(fases_df["Fase"], errors="coerce")

    fases_df = fases_df.dropna(subset=["Fase"])
    fases_df = fases_df[fases_df["Fase"] > 0].copy()
    fases_df["Fase"] = fases_df["Fase"].astype(float).round(3)

    return fases_df.reset_index(drop=True)


def arquivo_carregado():
    return "main_xlsm" in st.session_state and st.session_state.main_xlsm is not None


def _ler_litologia_do_xlsm(wb) -> pd.DataFrame:
    if "Litologia" not in wb.sheetnames:
        raise ValueError("A aba 'Litologia' não existe no arquivo.")

    ws = wb["Litologia"]

    h_fm = ws["C3"].value
    h_lit = ws["E3"].value
    h_top = ws["F3"].value
    h_bas = ws["G3"].value

    def _norm(x):
        return str(x).strip().lower() if x is not None else ""

    if not any(k in _norm(h_fm) for k in ["forma", "formação", "formacao", "fm"]):
        raise ValueError(f"Header de Formação inválido em C3: {h_fm}")
    if "litolog" not in _norm(h_lit):
        raise ValueError(f"Header de Litologia inválido em E3: {h_lit}")
    if "top" not in _norm(h_top):
        raise ValueError(f"Header de Topo inválido em F3: {h_top}")
    if "bas" not in _norm(h_bas):
        raise ValueError(f"Header de Base inválido em G3: {h_bas}")

    rows = []
    r = 4
    while r <= ws.max_row:
        fm = ws[f"C{r}"].value
        lit = ws[f"E{r}"].value
        topo = ws[f"F{r}"].value
        base = ws[f"G{r}"].value

        if (fm in (None, "") and lit in (None, "") and topo in (None, "") and base in (None, "")):
            break

        rows.append((fm, lit, topo, base))
        r += 1

    df = pd.DataFrame(rows, columns=["Formação", "Litologia", "Topo", "Base"])

    df["Formação"] = df["Formação"].astype(str).str.strip()
    df["Litologia"] = df["Litologia"].astype(str).str.strip()
    df["Topo"] = pd.to_numeric(df["Topo"], errors="coerce")
    df["Base"] = pd.to_numeric(df["Base"], errors="coerce")

    df = df.dropna(subset=["Topo"]).reset_index(drop=True)
    df = df[(df["Litologia"] != "") | (df["Formação"] != "")].reset_index(drop=True)

    if df.empty:
        raise ValueError("Nenhuma linha válida encontrada na aba 'Litologia'.")

    return df


def _aplicar_litologia_no_state(poco_nome: str, df_lito: pd.DataFrame):
    if "pocos" not in st.session_state:
        st.session_state.pocos = {}
    if poco_nome not in st.session_state.pocos:
        st.session_state.pocos[poco_nome] = {}

    st.session_state.pocos[poco_nome]["profundidade"] = df_lito["Topo"].astype(float).tolist()
    st.session_state.pocos[poco_nome]["litologia"] = df_lito["Litologia"].astype(str).tolist()

    # ✅ agora vem do Excel
    st.session_state.pocos[poco_nome]["formation"] = df_lito["Formação"].astype(str).tolist()

    st.session_state.n_fm = len(df_lito)

    # Pré-preenche widgets
    for i in range(st.session_state.n_fm):
        st.session_state[f"prof_{i}"] = float(st.session_state.pocos[poco_nome]["profundidade"][i])
        st.session_state[f"fm_{i}"] = st.session_state.pocos[poco_nome]["formation"][i]

        lit = st.session_state.pocos[poco_nome]["litologia"][i]
        mapa_lit = {
            "argilito": "Argilito",
            "arenito": "Arenito",
            "folhelho": "Folhelho",
            "calcario": "Calcário",
            "calcário": "Calcário",
            "carbonato": "Carbonato",
            "siltito": "Siltito",
            "diamictito": "Diamictito",
            "conglomerado": "Conglomerado",
            "anidrita / gipsita": "Anidrita / Gipsita",
            "anidrita": "Anidrita / Gipsita",
            "gipsita": "Anidrita / Gipsita",
            "halita": "Halita",
            "calcilutito": "Calcilutito",
            "calculutito": "Calcilutito",
            "calcissiltito": "Calcissiltito",
            "calcarenito": "Calcarenito",
            "calcirrudito": "Calcirrudito",
            "marga": "Margas",
            "margas": "Margas",
            "coquina": "Coquina",
            "dolomito": "Dolomito",
            "basalto": "Basalto",
            "diabásio": "Diabásio",
            "diabasio": "Diabásio",
        }
        st.session_state[f"lit_{i}"] = mapa_lit.get(str(lit).strip().lower(), "Arenito")


def _garantir_litologia_importada(poco_nome: str) -> bool:
    if "pocos" not in st.session_state:
        st.session_state.pocos = {poco_nome: {}}
    if poco_nome not in st.session_state.pocos:
        st.session_state.pocos[poco_nome] = {}

    if st.session_state.get("lito_import_ok", False):
        df_lito_excel = st.session_state.get("df_lito_excel", pd.DataFrame())
        poco = st.session_state.pocos.get(poco_nome, {})
        if (
                isinstance(df_lito_excel, pd.DataFrame)
                and not df_lito_excel.empty
                and not poco.get("profundidade")
        ):
            _aplicar_litologia_no_state(poco_nome, df_lito_excel)
        return True

    if st.session_state.get("main_xlsm") is None or st.session_state.get("wb") is None:
        return False

    try:
        df_lito_excel = _ler_litologia_do_xlsm(st.session_state.wb)
        st.session_state.df_lito_excel = df_lito_excel
        _aplicar_litologia_no_state(poco_nome, df_lito_excel)
        st.session_state.lito_import_ok = True
        return True
    except Exception:
        st.session_state.lito_import_ok = False
        st.session_state.df_lito_excel = pd.DataFrame()
        return False


def plot_correlacao_com_logs(pocos, logs, description, lines, v_selected, tvd_labels, escala=(30, 300)):
    """
    Plota correlação entre múltiplos poços, cada um com seu log track no estilo clássico.
    """
    fig = go.Figure()
    paleta = {
        "Argilito": {"bg": "#9ACD32", "simbol": "|"},
        "Folhelho": {"bg": "#9ACD32", "simbol": "-"},
        "Siltito": {"bg": "#A67B5B", "simbol": "-"},
        "Arenito": {"bg": "#FFD580", "simbol": "."},
        "Diamictito": {"bg": "#E97451", "simbol": "."},
        "Conglomerado": {"bg": "#CD853F", "simbol": "."},
        "Anidrita / Gipsita": {"bg": "#E6E6FA", "simbol": "/"},
        "Halita": {"bg": "#FFFFFF", "simbol": "."},
        "Calcário": {"bg": "#B0C4DE", "simbol": "."},
        "Carbonato": {"bg": "#cfe8f3", "simbol": "x"},
        "Calcilutito": {"bg": "#C9B7D8", "simbol": "-"},
        "Margas": {"bg": "#C7D4B5", "simbol": "\\"},
        "Marga": {"bg": "#C7D4B5", "simbol": "\\"},
        "Calcissiltito": {"bg": "#D8BFD8", "simbol": "."},
        "Calcarenito": {"bg": "#F5DEB3", "simbol": "."},
        "Calcirrudito": {"bg": "#4682B4", "simbol": "."},
        "Coquina": {"bg": "#FFDEAD", "simbol": "."},
        "Dolomito": {"bg": "#C2B280", "simbol": "."},
        "Basalto": {"bg": "#2F4F4F", "simbol": "+"},
        "Diabásio": {"bg": "#556B2F", "simbol": "."},
    }

    nomes = list(pocos.keys())
    n_pocos = len(nomes)
    espacamento = 2  # espaço lateral entre poços

    for idx, nome in enumerate(nomes):
        if nome not in v_selected:
            pass
        else:
            x_center = idx * espacamento
            poco = pocos[nome]

            profundidade = poco["profundidade"]
            fm = poco['formation']
            litologias = poco["litologia"]
            perfil = pd.DataFrame(poco['perfil'])
            tvd = poco['tvd']
            syn = False
            if 'Syn Log' in poco:
                syn = True

            # === 1) Faixas de formation ===
            for i in range(len(profundidade)):
                if i + 1 == len(profundidade):
                    z_top, z_base = profundidade[i], tvd
                else:
                    z_top, z_base = profundidade[i], profundidade[i + 1]
                lit = litologias[i]

                fig.add_trace(go.Scatter(
                    x=[x_center - 0.5, x_center + 0.5, x_center + 0.5, x_center - 0.5, x_center - 0.5],
                    y=[z_top, z_top, z_base, z_base, z_top],
                    fill="toself",
                    line=dict(color="black", width=1),
                    fillpattern=dict(
                        shape=paleta[lit]["simbol"],  # padrão diagonal
                        fgcolor="black",  # cor do padrão
                        bgcolor=paleta[lit]["bg"],
                        size=4,  # <<< controla o tamanho/espacamento
                        solidity=0.05  # densidade do padrão (0 = ralo, 1 = cheio)
                    ),
                    showlegend=False
                ))

                if tvd_labels:
                    fig.add_annotation(
                        x=x_center + 0.6,  # posição no eixo x
                        y=z_top,  # posição no eixo y (topo da formação)
                        text=f'{profundidade[i]} m',  # nome da formação
                        showarrow=False,  # sem seta
                        font=dict(size=12, color="black", family="Arial"),
                        align="left",
                        bordercolor="black",
                        borderwidth=1,
                        borderpad=2,
                        bgcolor="white",  # fundo da caixa
                        opacity=0.8
                    )

                if description:
                    # nome da formation
                    fig.add_trace(go.Scatter(
                        x=[x_center],
                        y=[(z_top + z_base) / 2],
                        text=[f"<b>{fm[i]}</b> - {lit}"],  # <<< negrito com HTML
                        mode="text",
                        textfont=dict(size=12, color="black", family="Arial"),  # define a fonte
                        showlegend=False,
                        hoverinfo="skip"
                    ))

            if syn:
                k = poco['Syn Log'].keys()
                x = list(k)
                x.remove('Depth synthetic log')
                for line in range(3):
                    if logs[line]:
                        if line == 0:
                            log = 'Synthetic Sonic log'
                            color = 'red'
                            scale = st.session_state.s_scale
                        elif line == 1:
                            log = 'Synthetic Density log'
                            color = 'purple'
                            scale = st.session_state.d_scale
                        else:
                            log = 'Synthetic Gamma ray log'
                            color = 'black'
                            scale = st.session_state.g_scale
                        try:
                            # === 2) Curva de log ===
                            z_log = poco['Syn Log']["Depth synthetic log"]
                            valores = np.array(poco['Syn Log'][log])
                            # normalizar valores para caber dentro da faixa do poço
                            vmin = min(poco['Syn Log'][log])
                            vmax = max(poco['Syn Log'][log])
                            x_log = x_center - 0.5 + (valores - vmin) / (vmax - vmin) * 1.0

                            fig.add_trace(go.Scatter(
                                x=x_log,
                                y=z_log,
                                mode="lines",
                                line=dict(color=color, width=1),
                                name=f"Log {nome}",
                                showlegend=False,
                                opacity=st.session_state.op
                            ))
                            # ['Above', 'Below', 'Remove']
                            if scale == 'Above':
                                y_pos = -50
                                top = y_pos - 70
                            elif scale == 'Below':
                                y_pos = tvd + 50
                                top = y_pos + 70
                            else:
                                y_pos = False
                                top = False

                            if y_pos:
                                # === 3) Escala do log logo acima do topo (y=0) ===
                                for v in np.linspace(vmin, vmax, 6):
                                    xpos = x_center - 0.5 + (v - vmin) / (vmax - vmin) * 1.0
                                    fig.add_trace(go.Scatter(
                                        x=[xpos],
                                        y=[y_pos],
                                        text=[str(int(v))],
                                        mode="text",
                                        textfont=dict(size=9, color=color),
                                        showlegend=False,
                                        hoverinfo="skip"
                                    ))

                                    # título do perfil
                                    fig.add_trace(go.Scatter(
                                        x=[x_center],
                                        y=[top],
                                        text=log,
                                        mode="text",
                                        textfont=dict(size=11, color=color),
                                        showlegend=False,
                                        hoverinfo="skip"
                                    ))
                        except KeyError:
                            pass

            for line in range(3):
                if logs[line]:
                    if line == 0:
                        log = 'Sonic log'
                        color = 'red'
                        scale = st.session_state.s_scale
                    elif line == 1:
                        log = 'Density log'
                        color = 'purple'
                        scale = st.session_state.d_scale
                    else:
                        log = 'Gamma ray log'
                        color = 'black'
                        scale = st.session_state.g_scale

                    try:
                        # === 2) Curva de log ===
                        z_log = perfil["Profundidade"]
                        valores = perfil[log]
                        # normalizar valores para caber dentro da faixa do poço
                        vmin = min(perfil[log])
                        vmax = max(perfil[log])
                        x_log = x_center - 0.5 + (valores - vmin) / (vmax - vmin) * 1.0

                        fig.add_trace(go.Scatter(
                            x=x_log,
                            y=z_log,
                            mode="lines",
                            line=dict(color=color, width=1),
                            name=f"Log {nome}",
                            showlegend=False,
                            opacity=st.session_state.op
                        ))
                        # ['Above', 'Below', 'Remove']
                        if scale == 'Above':
                            y_pos = -50
                            top = y_pos - 70
                        elif scale == 'Below':
                            y_pos = tvd + 50
                            top = y_pos + 70
                        else:
                            y_pos = False
                            top = False

                        if y_pos:
                            # === 3) Escala do log logo acima do topo (y=0) ===
                            for v in np.linspace(vmin, vmax, 6):
                                xpos = x_center - 0.5 + (v - vmin) / (vmax - vmin) * 1.0
                                fig.add_trace(go.Scatter(
                                    x=[xpos],
                                    y=[y_pos],
                                    text=[str(int(v))],
                                    mode="text",
                                    textfont=dict(size=9, color=color),
                                    showlegend=False,
                                    hoverinfo="skip"
                                ))

                                # título do perfil
                                fig.add_trace(go.Scatter(
                                    x=[x_center],
                                    y=[top],
                                    text=log,
                                    mode="text",
                                    textfont=dict(size=11, color=color),
                                    showlegend=False,
                                    hoverinfo="skip"
                                ))
                    except KeyError:
                        pass

            if lines:
                if idx < n_pocos - 1:
                    prox = pocos[nomes[idx + 1]]
                    if nomes[idx + 1] in v_selected:
                        # cria dicionários formacao->profundidade
                        mapa_atual = dict(zip(pocos[nomes[idx]]["formation"], pocos[nomes[idx]]["profundidade"]))
                        mapa_prox = dict(zip(prox["formation"], prox["profundidade"]))

                        # percorre formações em comum
                        for formacao, z_top in mapa_atual.items():
                            if formacao in mapa_prox:
                                z_top_prox = mapa_prox[formacao]

                                # desenha linha tracejada
                                fig.add_trace(go.Scatter(
                                    x=[x_center + 0.5, (idx + 1) * espacamento - 0.5],
                                    y=[z_top, z_top_prox],
                                    mode="lines",
                                    line=dict(color="black", width=1, dash="dash"),
                                    showlegend=False
                                ))
                            if z_top == profundidade[-1]:
                                z_top = pocos[nomes[idx]]['tvd']
                                z_top_prox = prox['tvd']
                                fig.add_trace(go.Scatter(
                                    x=[x_center + 0.5, (idx + 1) * espacamento - 0.5],
                                    y=[z_top, z_top_prox],
                                    mode="lines",
                                    line=dict(color="black", width=1, dash="dash"),
                                    showlegend=False
                                ))

            # Well name above lithology layers
            fig.add_trace(go.Scatter(
                x=[x_center],
                y=[-250],
                text=[nome],
                mode="text",
                textfont=dict(size=14, color="blue"),
                showlegend=False,
                hoverinfo="skip"
            ))

    # === Ajustes ===
    fig.update_yaxes(dtick=200)  # passo da escala do eixo da profundidade

    fig.update_yaxes(
        autorange="reversed",
        title="Profundidade (m)",
        range=[0, max(max(p["profundidade"]) for p in pocos.values())]  # começa em 0
    )
    fig.update_xaxes(visible=False)  # remove eixo X

    fig.update_layout(
        title="Correlação estratigráfica",
        plot_bgcolor="white",
        width=1000,
        height=800,
    )

    return fig


def preparar_widget_persistente(chave_estado, chave_widget):
    if chave_widget not in st.session_state:
        st.session_state[chave_widget] = st.session_state.get(chave_estado)


def salvar_widget_persistente(chave_estado, chave_widget):
    st.session_state[chave_estado] = st.session_state.get(chave_widget)


def sincronizar_widgets_persistentes(pares):
    for chave_estado, chave_widget in pares:
        if chave_widget in st.session_state:
            st.session_state[chave_estado] = st.session_state.get(chave_widget)


def estado_expander_persistente(chave_estado, padrao=True):
    if chave_estado not in st.session_state:
        st.session_state[chave_estado] = padrao

    return bool(st.session_state.get(chave_estado, padrao))


def _buscar_coluna_por_partes(df, *partes):
    partes_norm = [str(parte).lower() for parte in partes]
    for coluna in df.columns:
        coluna_norm = str(coluna).lower()
        if all(parte in coluna_norm for parte in partes_norm):
            return coluna
    return None


def _inserir_agua_offshore_litologia(litho_tops):
    if st.session_state.get("tipo_poco") != "Offshore":
        return litho_tops

    rtkb = float(st.session_state.get("rtkb", 0.0))
    lda = float(st.session_state.get("lda", 0.0))
    if lda <= 0:
        return litho_tops

    base_agua = rtkb + lda
    tops_originais = sorted(
        [[float(prof), lit] for prof, lit in litho_tops],
        key=lambda item: item[0]
    )

    litologia_base = tops_originais[0][1] if tops_originais else "Fm. Permeável"
    for prof, lit in tops_originais:
        if prof <= base_agua:
            litologia_base = lit
        else:
            break

    novos_tops = [
        [prof, lit]
        for prof, lit in tops_originais
        if not (rtkb <= prof < base_agua)
    ]

    novos_tops.append([rtkb, "Água"])
    if not any(abs(prof - base_agua) < 1e-9 for prof, _ in novos_tops):
        novos_tops.append([base_agua, litologia_base])

    return sorted(novos_tops, key=lambda item: item[0])


def _desenhar_preenchimento_agua(ax, top_plot, base_plot, largura=0.5):
    altura = base_plot - top_plot

    rect = mpatches.Rectangle(
        (0, top_plot),
        width=largura,
        height=altura,
        facecolor="#d9f7ff",
        edgecolor="#1f8fcf",
        linewidth=0.9,
        alpha=0.92,
        zorder=1
    )
    ax.add_patch(rect)

    if altura <= 0:
        return

    rng = np.random.default_rng(7)

    n_ondas = int(np.clip(altura / 10, 8, 55))
    ys_ondas = np.linspace(top_plot, base_plot, n_ondas)

    for j, y0 in enumerate(ys_ondas):
        xs = np.linspace(0.0, largura, 160)
        amplitude = min(altura * 0.03, 8.0)
        fase = j * 0.9
        largura_eixo_px = max(
            ax.get_position().width * ax.figure.get_figwidth() * ax.figure.dpi,
            1.0
        )
        xs_norm = xs / largura if largura else xs
        xs_px = xs_norm * largura_eixo_px
        comprimento_onda_px = 34.0
        ys = y0 + amplitude * np.sin(((xs_px / comprimento_onda_px) * 2 * np.pi) + fase)

        linha, = ax.plot(
            xs,
            ys,
            color="#168fbd",
            linewidth=1.4,
            alpha=0.85,
            zorder=3
        )
        linha.set_clip_path(rect)

    n_reflexos = int(np.clip(altura / 35, 3, 14))
    for _ in range(n_reflexos):
        y0 = rng.uniform(top_plot + altura * 0.08, base_plot - altura * 0.08)
        x0 = rng.uniform(largura * 0.12, largura * 0.48)
        reflexo_largura = rng.uniform(largura * 0.20, largura * 0.44)

        reflexo, = ax.plot(
            [x0, min(x0 + reflexo_largura, largura * 0.92)],
            [y0, y0],
            color="white",
            linewidth=rng.uniform(0.8, 1.5),
            alpha=0.45,
            zorder=3
        )
        reflexo.set_clip_path(rect)

    n_bolhas = int(np.clip(altura / 14, 10, 70))
    for _ in range(n_bolhas):
        x_bolha = rng.uniform(largura * 0.07, largura * 0.93)
        y_bolha = rng.uniform(top_plot, base_plot)

        bolha = mpatches.Ellipse(
            (x_bolha, y_bolha),
            width=rng.uniform(largura * 0.07, largura * 0.15),
            height=rng.uniform(7.0, 16.0),
            facecolor="#e9fbff",
            edgecolor="#2aa7d8",
            linewidth=1.0,
            alpha=0.85,
            zorder=4
        )
        bolha.set_clip_path(rect)
        ax.add_patch(bolha)

        brilho = mpatches.Ellipse(
            (x_bolha - largura * 0.02, y_bolha - 2.0),
            width=largura * 0.035,
            height=3.0,
            facecolor="white",
            edgecolor="none",
            alpha=0.85,
            zorder=5
        )
        brilho.set_clip_path(rect)
        ax.add_patch(brilho)


def _atualizar_df_idade_do_state(prof_max=None):
    sincronizar_widgets_persistentes([
        ("idg", "_w_idg"),
        ("n_id", "_w_n_id"),
    ])

    if st.session_state.get("idg") != "Sim":
        return

    n_id = int(st.session_state.get("n_id", st.session_state.get("_w_n_id", 1)) or 1)
    tem_widget_idade = any(
        f"prof_inicial_2_{i}" in st.session_state
        or f"prof_final_2_{i}" in st.session_state
        or f"idg_{i}" in st.session_state
        for i in range(n_id)
    )
    if not tem_widget_idade:
        return

    prof_ini_id = []
    prof_fim_id = []
    idade_geo = []

    for i in range(n_id):
        if i == 0:
            p_ini = st.session_state.get(f"prof_inicial_2_{i}", 0.0)
        else:
            p_ini = st.session_state.get(f"prof_final_2_{i - 1}", 0.0)

        if i == n_id - 1 and prof_max is not None:
            p_fim = float(prof_max + 100)
        else:
            p_fim = st.session_state.get(f"prof_final_2_{i}", 0.0)

        prof_ini_id.append(p_ini)
        prof_fim_id.append(p_fim)
        idade_geo.append(st.session_state.get(f"idg_{i}", ""))

    st.session_state.df_idade = pd.DataFrame({
        "Topo (m)": prof_ini_id,
        "Base (m)": prof_fim_id,
        "Idade": idade_geo
    })


def idade_formacao(ax_idade, df_idade, y_max, y_min=0):
    if not isinstance(df_idade, pd.DataFrame) or df_idade.empty:
        return

    df_plot = df_idade.copy()
    df_plot["Topo (m)"] = pd.to_numeric(df_plot["Topo (m)"], errors="coerce")
    df_plot["Base (m)"] = pd.to_numeric(df_plot["Base (m)"], errors="coerce")
    df_plot["Idade"] = df_plot["Idade"].fillna("").astype(str)
    df_plot = df_plot.dropna(subset=["Topo (m)", "Base (m)"])
    if st.session_state.get("tipo_poco") == "Offshore":
        rtkb = float(st.session_state.get("rtkb", 0.0))
        lda = float(st.session_state.get("lda", 0.0))
        if lda > 0:
            base_agua = rtkb + lda
            df_plot = df_plot[df_plot["Base (m)"] > base_agua].copy()
            df_plot.loc[df_plot["Topo (m)"] < base_agua, "Topo (m)"] = base_agua
            df_plot = pd.concat([
                pd.DataFrame({
                    "Topo (m)": [rtkb],
                    "Base (m)": [base_agua],
                    "Idade": ["Água"]
                }),
                df_plot
            ], ignore_index=True)
    if df_plot.empty:
        return

    ax_idade.set_xlim(0, 0.5)
    ax_idade.set_xticks([])
    y_min_idade = float(y_min)
    y_max_idade = float(y_max)

    ax_idade.set_ylim(y_max_idade, y_min_idade)
    ax_idade.set_title("Idade", fontsize=10, fontweight="bold")

    idades_unicas = df_plot["Idade"].unique()
    cmap = plt.cm.get_cmap("Accent", len(idades_unicas))

    cores = {
        idade: cmap(i)
        for i, idade in enumerate(idades_unicas)
    }
    cores["Água"] = "#bfefff"

    for _, row in df_plot.iterrows():
        topo = max(float(row["Topo (m)"]), y_min_idade)
        base = min(float(row["Base (m)"]), y_max_idade)
        if base <= topo:
            continue

        if str(row["Idade"]).strip().lower().endswith("gua"):
            _desenhar_preenchimento_agua(ax_idade, topo, base, largura=0.5)
            continue

        ax_idade.fill_betweenx(
            [topo, base],
            0,
            0.5,
            color=cores[row["Idade"]],
            edgecolor="black"
        )
        ax_idade.text(
            0.25,
            (topo + base) / 2,
            row["Idade"],
            ha="center",
            va="center",
            fontsize=8,
            rotation=90,
            fontweight="bold",
            color="black",
            path_effects=[
                pe.Stroke(linewidth=3, foreground="white"),
                pe.Normal()
            ]
        )


def suavizar(x, y, frac=None):
    x = np.asarray(x)
    y = np.asarray(y)

    # Proteção básica
    if len(y) < 3:
        return y

    if frac is None:
        frac = 0.1

    y_suav = lowess(y, x, frac=frac, return_sorted=False)

    return y_suav


def add_watermark(ax, logo_path="logo2.png", xy=(0.80, 0.25), zoom=0.20, alpha=0.10, zorder=0):
    """
    Marca d'água robusta:
    - NÃO distorce em eixo log (usa ax.transAxes)
    - transparência SEMPRE muda (ajusta canal alpha do RGBA)
    """
    if not os.path.exists(logo_path):
        return

    # lê com PIL e força RGBA
    im = Image.open(logo_path).convert("RGBA")
    arr = np.asarray(im).astype(np.float32)

    # aplica alpha global multiplicando o canal A (0..255)
    arr[..., 3] = arr[..., 3] * float(alpha)
    arr[..., 3] = np.clip(arr[..., 3], 0, 255)

    arr = arr.astype(np.uint8)

    oi = OffsetImage(arr, zoom=zoom)  # aqui já vai com alpha aplicado
    ab = AnnotationBbox(
        oi,
        xy=xy,
        xycoords=ax.transAxes,   # independe de log/linear
        frameon=False,
        box_alignment=(0.5, 0.5),
        zorder=zorder
    )
    ax.add_artist(ab)


def lito(ax1, df_pp, profundidades, litologias, bases, y_min=None, y_max=None):
    try:
        label = True
        line_w = 0.8

        tipo_coluna_lito = st.session_state.get(
            "tipo_coluna_litologica_graficos",
            "Permeável / Não permeável"
        )

        if "s_gr" not in st.session_state:
            st.session_state.s_gr = False

        s_gr_ativo = st.session_state.get("s_gr") is True or st.session_state.get("s_gr") == "Sim"
        coluna_gr = "Raio Gama Suavizado" if s_gr_ativo else "Perfil Raio Gama"

        pode_gerar_perm_nao_perm = (
            tipo_coluna_lito == "Permeável / Não permeável"
            and "LBF_calc" in df_pp.columns
            and coluna_gr in df_pp.columns
            and "Profundidade (m)" in df_pp.columns
        )

        if pode_gerar_perm_nao_perm:
            df_lito_calc = df_pp[
                ["Profundidade (m)", coluna_gr, "LBF_calc"]
            ].copy()

            df_lito_calc["Profundidade (m)"] = pd.to_numeric(
                df_lito_calc["Profundidade (m)"],
                errors="coerce"
            )
            df_lito_calc[coluna_gr] = pd.to_numeric(
                df_lito_calc[coluna_gr],
                errors="coerce"
            )
            df_lito_calc["LBF_calc"] = pd.to_numeric(
                df_lito_calc["LBF_calc"],
                errors="coerce"
            )

            df_lito_calc = df_lito_calc.dropna(
                subset=["Profundidade (m)", coluna_gr, "LBF_calc"]
            ).sort_values("Profundidade (m)")

            if df_lito_calc.empty:
                return

            df_lito_calc["Classificação"] = np.where(
                df_lito_calc[coluna_gr] >= df_lito_calc["LBF_calc"],
                "Folhelho",
                "Fm. Permeável"
            )

            litho_tops = []

            primeira_prof = float(df_lito_calc["Profundidade (m)"].iloc[0])

            if st.session_state.tipo_poco == "Onshore":
                # Antes do primeiro dado de perfilagem, considerar formação permeável
                if primeira_prof > 0:
                    litho_tops.append([0.0, "Fm. Permeável"])


            else:
                rtkb = float(st.session_state.get("rtkb", 0.0))
                lda = float(st.session_state.get("lda", 0.0))
                base_agua = rtkb + lda

                if lda > 0:
                    litho_tops.append([rtkb, "Água"])
                    litho_tops.append([base_agua, "Fm. Permeável"])

                elif primeira_prof > rtkb:
                    litho_tops.append([primeira_prof, "Fm. Permeável"])

            classe_anterior = litho_tops[-1][1] if litho_tops else None

            for _, row in df_lito_calc.iterrows():
                prof_atual = float(row["Profundidade (m)"])
                classe_atual = row["Classificação"]

                if st.session_state.tipo_poco == "Offshore":
                    rtkb = float(st.session_state.get("rtkb", 0.0))
                    lda = float(st.session_state.get("lda", 0.0))
                    base_agua = rtkb + lda

                    if prof_atual < rtkb:
                        continue

                    if lda > 0 and rtkb <= prof_atual < base_agua:
                        continue

                if classe_atual != classe_anterior:
                    litho_tops.append([prof_atual, classe_atual])
                    classe_anterior = classe_atual

            label = False
            line_w = 0

        else:
            if not profundidades or not litologias:
                return

            litho_tops = [[x, y] for x, y in zip(profundidades, litologias)]
            litho_tops = _inserir_agua_offshore_litologia(litho_tops)

            if False and st.session_state.tipo_poco == "Offshore":
                rtkb = float(st.session_state.get("rtkb", 0.0))
                lda = float(st.session_state.get("lda", 0.0))
                base_agua = rtkb + lda

                if lda > 0:
                    litologia_apos_agua = next(
                        (lit for prof, lit in litho_tops if float(prof) >= base_agua),
                        litho_tops[-1][1] if litho_tops else "Fm. Permeável"
                    )
                    litho_tops = [
                        [float(prof), lit]
                        for prof, lit in litho_tops
                        if not (rtkb <= float(prof) < base_agua)
                    ]
                    litho_tops.append([rtkb, "Água"])
                    litho_tops.append([base_agua, litologia_apos_agua])
                    litho_tops = sorted(litho_tops, key=lambda item: item[0])

            label = True
            line_w = 0.8

        litho_intervals = []
        for i, (top, lit) in enumerate(litho_tops):
            if i < len(litho_tops) - 1:
                base = litho_tops[i + 1][0]
            else:
                base = bases

            litho_intervals.append((top, base, lit))

        df_perm_nao_perm = pd.DataFrame(
            litho_intervals,
            columns=["Topo (m)", "Base (m)", "Classificação"]
        )

        st.session_state.df_perm_nao_perm = df_perm_nao_perm

        litho_styles = {
            "Água": {"color": "#bfefff", "hatch": None, "edgecolor": "#2f80b7"},
            "Argilito": {"color": "#9ACD32", "hatch": "|||", "edgecolor": "black"},
            "Arenito": {"color": "#fff7a1", "hatch": "...", "edgecolor": "black"},
            "Fm. Permeável": {"color": "#fff7a1", "hatch": "...", "edgecolor": "black"},
            "Folhelho": {"color": "#2f4f4f", "hatch": None, "edgecolor": "black"},
            "Siltito": {"color": "#8b4513", "hatch": None, "edgecolor": "black"},
            "Diamictito": {"color": "#E97451", "hatch": "..", "edgecolor": "black"},
            "Conglomerado": {"color": "#ffb347", "hatch": "oo", "edgecolor": "black"},
            "Anidrita / Gipsita": {"color": "#E6E6FA", "hatch": "///", "edgecolor": "black"},
            "Halita": {"color": "#ffffff", "hatch": None, "edgecolor": "black"},
            "Calcário": {"color": "#a7c7e7", "hatch": "///", "edgecolor": "#003366"},
            "Carbonato": {"color": "#cfe8f3", "hatch": "xx", "edgecolor": "#003366"},
            "Calcilutito": {"color": "#C9B7D8", "hatch": "---", "edgecolor": "black"},
            "Margas": {"color": "#C7D4B5", "hatch": "\\\\", "edgecolor": "black"},
            "Calcissiltito": {"color": "#D8BFD8", "hatch": "---", "edgecolor": "black"},
            "Calcarenito": {"color": "#F5DEB3", "hatch": "...", "edgecolor": "black"},
            "Calcirrudito": {"color": "#4682B4", "hatch": "oo", "edgecolor": "black"},
            "Coquina": {"color": "#FFDEAD", "hatch": "oo", "edgecolor": "black"},
            "Dolomito": {"color": "#C2B280", "hatch": "xx", "edgecolor": "black"},
            "Basalto": {"color": "#2b2b2b", "hatch": None, "edgecolor": "black"},
            "Diabásio": {"color": "#556B2F", "hatch": "++", "edgecolor": "black"},
        }

        y_min_lito = float(st.session_state.y_min_s if y_min is None else y_min)
        y_max_lito = float(st.session_state.y_max_s if y_max is None else y_max)

        ax1.set_xlim(0, 0.5)
        ax1.set_ylim(y_max_lito, y_min_lito)
        ax1.set_xticks([])
        ax1.tick_params(axis='y', which='both', left=False, labelleft=False)
        ax1.set_title("Litologia", fontsize=10, fontweight='bold')

        for top, base, lit in litho_intervals:
            top_plot = max(float(top), y_min_lito)
            base_plot = min(float(base), y_max_lito)

            if base_plot <= top_plot:
                continue

            style = litho_styles.get(
                lit,
                {"color": "gray", "hatch": None, "edgecolor": "black"}
            )

            if lit == "Água":
                _desenhar_preenchimento_agua(ax1, top_plot, base_plot, largura=0.5)
                continue

                altura = base_plot - top_plot

                rect = mpatches.Rectangle(
                    (0, top_plot),
                    width=0.5,
                    height=altura,
                    facecolor="#d9f7ff",
                    edgecolor="#1f8fcf",
                    linewidth=0.9,
                    alpha=0.92,
                    zorder=1
                )
                ax1.add_patch(rect)

                if altura > 0:
                    rng = np.random.default_rng(7)

                    n_ondas = int(np.clip(altura / 10, 8, 55))
                    ys_ondas = np.linspace(top_plot, base_plot, n_ondas)

                    for j, y0 in enumerate(ys_ondas):
                        xs = np.linspace(0.0, 0.5, 160)
                        amplitude = min(altura * 0.03, 8.0)
                        fase = j * 0.9
                        ys = y0 + amplitude * np.sin((xs * 28) + fase)

                        linha, = ax1.plot(
                            xs,
                            ys,
                            color="#168fbd",
                            linewidth=1.4,
                            alpha=0.85,
                            zorder=3
                        )

                        linha.set_clip_path(rect)

                    # Reflexos claros
                    n_reflexos = int(np.clip(altura / 35, 3, 14))
                    for _ in range(n_reflexos):
                        y0 = rng.uniform(top_plot + altura * 0.08, base_plot - altura * 0.08)
                        x0 = rng.uniform(0.06, 0.24)
                        largura = rng.uniform(0.10, 0.22)

                        reflexo, = ax1.plot(
                            [x0, min(x0 + largura, 0.46)],
                            [y0, y0],
                            color="white",
                            linewidth=rng.uniform(0.8, 1.5),
                            alpha=0.45,
                            zorder=3
                        )
                        reflexo.set_clip_path(rect)

                    n_bolhas = int(np.clip(altura / 14, 10, 70))

                    for _ in range(n_bolhas):
                        x_bolha = rng.uniform(0.035, 0.465)
                        y_bolha = rng.uniform(top_plot, base_plot)

                        tamanho = rng.uniform(0.035, 0.075)
                        altura_bolha = rng.uniform(7.0, 16.0)

                        bolha = mpatches.Ellipse(
                            (x_bolha, y_bolha),
                            width=tamanho,
                            height=altura_bolha,
                            facecolor="#e9fbff",
                            edgecolor="#2aa7d8",
                            linewidth=1.0,
                            alpha=0.85,
                            zorder=4
                        )
                        bolha.set_clip_path(rect)
                        ax1.add_patch(bolha)

                        brilho = mpatches.Ellipse(
                            (x_bolha - tamanho * 0.20, y_bolha - altura_bolha * 0.20),
                            width=tamanho * 0.30,
                            height=altura_bolha * 0.30,
                            facecolor="white",
                            edgecolor="none",
                            alpha=0.85,
                            zorder=5
                        )
                        brilho.set_clip_path(rect)
                        ax1.add_patch(brilho)

                continue

            rect = mpatches.Rectangle(
                (0, top_plot),
                width=0.5,
                height=base_plot - top_plot,
                facecolor=style["color"],
                edgecolor=style["edgecolor"],
                hatch=style["hatch"],
                linewidth=line_w,
            )

            ax1.add_patch(rect)

            mid = (top_plot + base_plot) / 2

            if label:
                txt = ax1.text(
                    0.25,
                    mid,
                    lit,
                    ha="center",
                    va="center",
                    fontsize=9,
                    fontweight="bold",
                    color="black",
                    zorder=5,
                    rotation=0
                )

                txt.set_path_effects([
                    path_effects.Stroke(linewidth=1.5, foreground='white'),
                    path_effects.Normal()
                ])

    except Exception as e:
        st.error(f"Erro em lito(): {e}")
        raise


def normal(df, df_pp=None):
    prof_ini = float(st.session_state.get("rtkb", 0.0))
    prof_anormal = float(st.session_state.get("anormal", 0.0))

    val_ini = 0.0
    val_fim = float(st.session_state.get("gn", 8.5))

    coluna_prof = "Profundidade (m)"
    coluna_pp = "Gradiente de Pressão de Poros (lb/gal)"
    coluna_extrap = "Linha Extrapolada"

    df_ref = df_pp if df_pp is not None else df

    prof_fim = prof_anormal

    if (
        df_ref is not None
        and isinstance(df_ref, pd.DataFrame)
        and coluna_prof in df_ref.columns
    ):
        df_aux = df_ref.copy()

        df_aux[coluna_prof] = pd.to_numeric(
            df_aux[coluna_prof],
            errors="coerce"
        )

        df_aux = (
            df_aux
            .dropna(subset=[coluna_prof])
            .sort_values(coluna_prof)
            .reset_index(drop=True)
        )

        if not df_aux.empty:
            prof_primeira_df_pp = float(df_aux[coluna_prof].iloc[0])

            if prof_anormal < prof_primeira_df_pp:
                prof_fim = prof_anormal

                if coluna_pp in df_aux.columns:
                    df_pp_val = df_aux[[coluna_prof, coluna_pp]].copy()

                    df_pp_val[coluna_pp] = pd.to_numeric(
                        df_pp_val[coluna_pp],
                        errors="coerce"
                    )

                    df_pp_val = df_pp_val.dropna(
                        subset=[coluna_prof, coluna_pp]
                    )

                    if not df_pp_val.empty:
                        idx = (df_pp_val[coluna_prof] - prof_fim).abs().idxmin()
                        prof_fim = float(df_pp_val.loc[idx, coluna_prof])
                        val_fim = float(df_pp_val.loc[idx, coluna_pp])

            else:
                if coluna_extrap in df_aux.columns:
                    serie_extrap = df_aux[coluna_extrap]

                    if serie_extrap.dtype == bool:
                        mask_real = ~serie_extrap.fillna(False)

                    elif np.issubdtype(serie_extrap.dtype, np.number):
                        mask_real = ~(serie_extrap.fillna(0).astype(int).astype(bool))

                    else:
                        mask_extrap = (
                            serie_extrap
                            .fillna(False)
                            .astype(str)
                            .str.strip()
                            .str.lower()
                            .isin(["true", "1", "sim", "yes", "verdadeiro"])
                        )
                        mask_real = ~mask_extrap

                    df_real = df_aux[mask_real].copy()

                    if not df_real.empty:
                        prof_fim = float(df_real[coluna_prof].iloc[0])
                    else:
                        prof_fim = prof_primeira_df_pp

                else:
                    prof_fim = prof_primeira_df_pp

                val_fim = float(st.session_state.get("gn", 8.5))

    alpha = float(st.session_state.get("alfa_pp", 10))

    n = len(df)
    profundidade = np.linspace(prof_ini, prof_fim, n)

    denom = prof_fim - prof_ini

    if denom == 0:
        valores = np.full(n, np.nan)
    else:
        t = (profundidade - prof_ini) / denom
        t = np.clip(t, 0.0, 1.0)

        valores = val_ini + (val_fim - val_ini) * (
            (1 - np.exp(-alpha * t)) / (1 - np.exp(-alpha))
        )

    df_gfs = pd.DataFrame({
        "Profundidade (m)": profundidade,
        "Gradiente de Pressão de Poros (lb/gal)": valores
    })

    if prof_fim > prof_ini:
        mask_none = df_gfs["Profundidade (m)"] <= prof_ini
    else:
        mask_none = df_gfs["Profundidade (m)"] >= prof_ini

    df_gfs.loc[
        mask_none,
        "Gradiente de Pressão de Poros (lb/gal)"
    ] = None

    if (
        st.session_state.get("tipo_poco") == "Onshore"
        and st.session_state.get("ex") == "Ativada"
    ):
        mask_rtkb = df_gfs["Profundidade (m)"] <= prof_ini

        df_gfs.loc[
            mask_rtkb,
            "Gradiente de Pressão de Poros (lb/gal)"
        ] = None

    st.session_state.df_gfs = df_gfs

    return df_gfs


def montar_df_pp_base(df_sobrecarga_calc):
    if "Profundidade" in df_sobrecarga_calc.columns:
        col_prof = "Profundidade"
    elif "Profundidade em relação a mesa rotativa (m)" in df_sobrecarga_calc.columns:
        col_prof = "Profundidade em relação a mesa rotativa (m)"
    else:
        raise ValueError("Coluna de profundidade não encontrada no dataframe de sobrecarga.")

    if "Perfil de densidade" in df_sobrecarga_calc.columns:
        col_dens = "Perfil de densidade"
    elif "Densidade (g/cm³)" in df_sobrecarga_calc.columns:
        col_dens = "Densidade (g/cm³)"
    else:
        col_dens = None

    if "Perfil sônico" in df_sobrecarga_calc.columns:
        col_son = "Perfil sônico"
    elif "Sônico (µs/pé)" in df_sobrecarga_calc.columns:
        col_son = "Sônico (µs/pé)"
    else:
        col_son = None

    df_pp = pd.DataFrame({
        "Profundidade (m)": df_sobrecarga_calc[col_prof],
        "Perfil de densidade (g/cm³)": df_sobrecarga_calc[col_dens] if col_dens else np.nan,
        "Perfil sônico (µs/pé)": df_sobrecarga_calc[col_son] if col_son else np.nan,
    })

    if "Perfil Raio Gama" in df_sobrecarga_calc.columns:
        df_pp["Perfil Raio Gama"] = df_sobrecarga_calc["Perfil Raio Gama"]
    else:
        df_pp["Perfil Raio Gama"] = np.nan

    if "Raio Gama Suavizado" in df_sobrecarga_calc.columns:
        df_pp["Raio Gama Suavizado"] = df_sobrecarga_calc["Raio Gama Suavizado"]
    else:
        df_pp["Raio Gama Suavizado"] = np.nan

    if "Gradiente de Sobrecarga (lb/gal)" in df_sobrecarga_calc.columns:
        df_pp["Gradiente de Sobrecarga (lb/gal)"] = df_sobrecarga_calc["Gradiente de Sobrecarga (lb/gal)"]

    if "Pressão de Sobrecarga (psi)" in df_sobrecarga_calc.columns:
        df_pp["Pressão de Sobrecarga (psi)"] = df_sobrecarga_calc["Pressão de Sobrecarga (psi)"]

    if "Linha Extrapolada" in df_sobrecarga_calc.columns:
        df_pp["Linha Extrapolada"] = df_sobrecarga_calc["Linha Extrapolada"]
    else:
        df_pp["Linha Extrapolada"] = False

    return df_pp


TRENDING_COLORS = [
    "#FF8C00",
    "#8A2BE2",
    "#32CD32",
    "#FFD700",
    "#FF1493",
    "#ADFF2F",
]


def _valor_float_ou_none(valor):
    try:
        if valor is None:
            return None

        valor_float = float(valor)

        if np.isnan(valor_float):
            return None

        return valor_float

    except Exception:
        return None


def _coletar_lbfs_poros():
    lbfs = []

    n_lbf = int(st.session_state.get("n_lbf", 1))

    for i in range(n_lbf):
        if f"lbf_valor_{i}" not in st.session_state:
            continue

        valor_lbf = _valor_float_ou_none(st.session_state.get(f"lbf_valor_{i}"))
        inclinacao_lbf = _valor_float_ou_none(st.session_state.get(f"lbf_inclinacao_{i}"))

        if valor_lbf is None or inclinacao_lbf is None:
            continue

        prof_ini_lbf = None
        prof_fim_lbf = None

        if n_lbf > 1:
            prof_ini_lbf = _valor_float_ou_none(
                st.session_state.get(f"lbf_prof_ini_{i}", None)
            )
            prof_fim_lbf = _valor_float_ou_none(
                st.session_state.get(f"lbf_prof_fim_{i}", None)
            )

        lbfs.append({
            "lbf": valor_lbf,
            "inclbf": inclinacao_lbf,
            "prof_ini": prof_ini_lbf,
            "prof_fim": prof_fim_lbf,
        })

    return lbfs



def _remover_chaves_lbf(indice):
    for chave in (
        f"lbf_valor_{indice}",
        f"_w_lbf_valor_{indice}",
        f"lbf_inclinacao_{indice}",
        f"_w_lbf_inclinacao_{indice}",
        f"lbf_prof_ini_{indice}",
        f"_w_lbf_prof_ini_{indice}",
        f"lbf_prof_fim_{indice}",
        f"_w_lbf_prof_fim_{indice}",
        f"exp_lbf_{indice}",
    ):
        st.session_state.pop(chave, None)


def _renderizar_campos_lbf(preparar_widget, incluir_titulo=True, persistente=False):
    def kwargs_widget(chave_estado):
        chave_widget = preparar_widget(chave_estado)
        kwargs = {"key": chave_widget}

        if persistente:
            kwargs.update({
                "on_change": salvar_widget_persistente,
                "args": (chave_estado, chave_widget),
            })

        return kwargs

    if "n_lbf" not in st.session_state:
        st.session_state.n_lbf = 1

    if incluir_titulo:
        st.markdown("### Linhas Base de Folhelhos")

    for i in range(int(st.session_state.n_lbf)):
        with st.expander(
            f"LBF {i + 1}",
            expanded=estado_expander_persistente(f"exp_lbf_{i}", True)
        ):
            if st.session_state.n_lbf > 1:
                colun1, colun2 = st.columns(2)

                with colun1:
                    chave_lbf_prof_ini = f"lbf_prof_ini_{i}"
                    if chave_lbf_prof_ini not in st.session_state:
                        st.session_state[chave_lbf_prof_ini] = 0.0
                    st.number_input(
                        "Profundidade inicial da LBF",
                        step=1.0,
                        format="%.2f",
                        min_value=0.0,
                        **kwargs_widget(chave_lbf_prof_ini)
                    )

                with colun2:
                    chave_lbf_prof_fim = f"lbf_prof_fim_{i}"
                    if chave_lbf_prof_fim not in st.session_state:
                        st.session_state[chave_lbf_prof_fim] = 0.0
                    st.number_input(
                        "Profundidade final da LBF",
                        step=1.0,
                        format="%.2f",
                        min_value=0.0,
                        **kwargs_widget(chave_lbf_prof_fim)
                    )

            chave_lbf_valor = f"lbf_valor_{i}"
            chave_widget_lbf_valor = f"_w_{chave_lbf_valor}"
            valor_lbf_atual = _valor_float_ou_none(st.session_state.get(chave_lbf_valor))
            valor_widget_lbf_atual = _valor_float_ou_none(st.session_state.get(chave_widget_lbf_valor))

            if valor_lbf_atual is None or valor_lbf_atual <= 1.0:
                st.session_state[chave_lbf_valor] = 110.0

            if chave_widget_lbf_valor in st.session_state and (
                valor_widget_lbf_atual is None or valor_widget_lbf_atual <= 1.0
            ):
                st.session_state[chave_widget_lbf_valor] = st.session_state[chave_lbf_valor]

            st.number_input(
                "Ponto inicial da LBF",
                step=10.0,
                format="%.2f",
                min_value=1.0,
                help=(
                    "Linha Base de Folhelhos (LBF)\n\n"
                    "- Representa o comportamento esperado dos folhelhos normalmente compactados.\n"
                    "- Traçada no registro raio gama (GAPI) × profundidade (m)."
                ),
                **kwargs_widget(chave_lbf_valor)
            )

            chave_lbf_inclinacao = f"lbf_inclinacao_{i}"
            if chave_lbf_inclinacao not in st.session_state:
                st.session_state[chave_lbf_inclinacao] = 0.0
            st.number_input(
                "Inclinação da LBF",
                step=0.1,
                format="%f",
                **kwargs_widget(chave_lbf_inclinacao)
            )


def _sincronizar_widgets_lbf_do_estado():
    n_lbf = int(st.session_state.get("n_lbf", 1))

    for i in range(n_lbf):
        for chave in (
            f"lbf_valor_{i}",
            f"lbf_inclinacao_{i}",
            f"lbf_prof_ini_{i}",
            f"lbf_prof_fim_{i}",
        ):
            if chave in st.session_state:
                st.session_state[f"_w_{chave}"] = st.session_state[chave]


def _atualizar_df_pp_lito_por_lbf():
    df_base = st.session_state.get("df_sobrecarga", pd.DataFrame())

    if not isinstance(df_base, pd.DataFrame) or df_base.empty:
        df_base = st.session_state.get("df1", pd.DataFrame())

    if not isinstance(df_base, pd.DataFrame) or df_base.empty:
        return

    try:
        df_lito = montar_df_pp_base(df_base)
        df_lito = _calcular_lbf_pp(df_lito, _coletar_lbfs_poros())
        st.session_state.df_pp_lito = df_lito.copy()

    except Exception:
        pass


def _montar_df_lbf_litologia():
    df_base = st.session_state.get("df_sobrecarga", pd.DataFrame())

    if not isinstance(df_base, pd.DataFrame) or df_base.empty:
        df_base = st.session_state.get("df1", pd.DataFrame())

    if not isinstance(df_base, pd.DataFrame) or df_base.empty:
        return pd.DataFrame()

    df_lbf = montar_df_pp_base(df_base)

    if (
        "Raio Gama Suavizado" in df_lbf.columns
        and df_lbf["Raio Gama Suavizado"].notna().any()
    ):
        return df_lbf

    if (
        "Profundidade (m)" in df_lbf.columns
        and "Perfil Raio Gama" in df_lbf.columns
    ):
        try:
            prof = pd.to_numeric(df_lbf["Profundidade (m)"], errors="coerce")
            gr = pd.to_numeric(df_lbf["Perfil Raio Gama"], errors="coerce")
            mask = prof.notna() & gr.notna()

            if mask.any():
                df_lbf.loc[mask, "Raio Gama Suavizado"] = suavizar(
                    prof[mask],
                    gr[mask]
                )

        except Exception as e:
            st.warning(f"Erro ao suavizar raio gama para LBF: {e}")

    return df_lbf


def _calcular_lbf_pp(df_pp, lbfs):
    df_pp = df_pp.copy()

    if "Profundidade (m)" not in df_pp.columns:
        df_pp["LBF_calc"] = np.nan
        return df_pp

    prof = pd.to_numeric(
        df_pp["Profundidade (m)"],
        errors="coerce"
    )

    limitar_lbf_onshore = (
        st.session_state.get("tipo_poco") == "Onshore"
        and int(st.session_state.get("n_lbf", 1)) == 1
        and int(st.session_state.get("n_trending", 1)) == 1
    )

    prof_primeiro_gr = None

    if limitar_lbf_onshore:
        if (
            st.session_state.get("s_gr") == "Sim"
            and "Raio Gama Suavizado" in df_pp.columns
        ):
            coluna_gr_ref = "Raio Gama Suavizado"
        else:
            coluna_gr_ref = "Perfil Raio Gama"

        if coluna_gr_ref in df_pp.columns:
            curva_gr_ref = pd.to_numeric(
                df_pp[coluna_gr_ref],
                errors="coerce"
            )

            prof_validas_gr = prof[
                prof.notna()
                & curva_gr_ref.notna()
            ]

            if not prof_validas_gr.empty:
                prof_primeiro_gr = prof_validas_gr.min()

    df_pp["LBF_calc"] = np.nan

    for idx, lbf in enumerate(lbfs):
        try:
            prof_ref = (
                lbf["prof_ini"]
                if lbf["prof_ini"] is not None
                else prof.min()
            )

            lbf_line = lbf["inclbf"] * (prof - prof_ref) + lbf["lbf"]

            mask_base = prof.notna()

            if prof_primeiro_gr is not None and pd.notna(prof_primeiro_gr):
                mask_base = mask_base & (prof >= prof_primeiro_gr)

            if st.session_state.get("tipo_poco") == "Offshore":
                mask_base = mask_base & (df_pp.index >= 1)

            prof_ini = lbf.get("prof_ini")
            prof_fim = lbf.get("prof_fim")

            if (
                prof_ini is not None
                and prof_fim is not None
                and prof_fim > prof_ini
            ):
                mask_intervalo = (prof >= prof_ini) & (prof <= prof_fim)
                mask_final = mask_base & mask_intervalo

            else:
                mask_final = mask_base

            df_pp.loc[mask_final, "LBF_calc"] = lbf_line[mask_final]

        except Exception as e:
            st.warning(f"Erro ao calcular LBF {idx + 1}: {e}")

    if df_pp["LBF_calc"].isna().any() and lbfs:
        try:
            primeira_lbf = lbfs[0]

            prof_ref = (
                primeira_lbf["prof_ini"]
                if primeira_lbf["prof_ini"] is not None
                else prof.min()
            )

            lbf_padrao = (
                primeira_lbf["inclbf"] * (prof - prof_ref)
                + primeira_lbf["lbf"]
            )

            if st.session_state.get("tipo_poco") == "Offshore":
                mask_fill = df_pp.index >= 1

                df_pp.loc[mask_fill, "LBF_calc"] = df_pp.loc[
                    mask_fill,
                    "LBF_calc"
                ].fillna(
                    lbf_padrao[mask_fill]
                )

            else:
                mask_fill = prof.notna()

                if prof_primeiro_gr is not None and pd.notna(prof_primeiro_gr):
                    mask_fill = mask_fill & (prof >= prof_primeiro_gr)

                df_pp.loc[mask_fill, "LBF_calc"] = df_pp.loc[
                    mask_fill,
                    "LBF_calc"
                ].fillna(
                    lbf_padrao[mask_fill]
                )

        except Exception as e:
            st.warning(f"Erro no preenchimento padrão da LBF: {e}")

    return df_pp


def _calcular_reta_normal_trending(prof, tr):
    pp1 = tr.get("pp1")
    pp2 = tr.get("pp2")
    s1 = tr.get("s1")
    s2 = tr.get("s2")

    if (
        pp1 is None
        or pp2 is None
        or s1 is None
        or s2 is None
        or pp1 == pp2
        or s1 <= 0
        or s2 <= 0
    ):
        return None

    # Igual à versão 6.0:
    # a reta normal do Trending é calculada no espaço logarítmico.
    # Assim, quando o gráfico está em semilogx, ela aparece como reta
    # e não fica deformada pela escala log do eixo X.
    m = np.log10(s2 / s1) / (pp2 - pp1)
    s_normal = s1 * 10 ** (m * (prof - pp1))

    return s_normal


def _coletar_trendings_poros():
    trendings = []

    n_trending = int(st.session_state.get("n_trending", 1))

    for i in range(n_trending):
        if f"trend_pp1_{i}" not in st.session_state:
            continue

        pp1 = _valor_float_ou_none(st.session_state.get(f"trend_pp1_{i}"))
        pp2 = _valor_float_ou_none(st.session_state.get(f"trend_pp2_{i}"))
        s1 = _valor_float_ou_none(st.session_state.get(f"trend_s1_{i}"))
        s2 = _valor_float_ou_none(st.session_state.get(f"trend_s2_{i}"))

        if pp1 is None or pp2 is None or s1 is None or s2 is None:
            continue

        prof_ini_trending = None
        prof_fim_trending = None

        if n_trending > 1:
            prof_ini_trending = _valor_float_ou_none(
                st.session_state.get(f"trend_prof_ini_{i}", None)
            )
            prof_fim_trending = _valor_float_ou_none(
                st.session_state.get(f"trend_prof_fim_{i}", None)
            )

        trendings.append({
            "pp1": pp1,
            "pp2": pp2,
            "s1": s1,
            "s2": s2,
            "prof_ini": prof_ini_trending,
            "prof_fim": prof_fim_trending,
        })

    return trendings


def _calcular_trending_pp(df_pp, trendings):
    df_pp = df_pp.copy()

    if "Profundidade (m)" not in df_pp.columns:
        df_pp["Perfil sônico (µs/pé) Reta Normal"] = np.nan
        return df_pp

    prof = pd.to_numeric(
        df_pp["Profundidade (m)"],
        errors="coerce"
    )

    limitar_trending_onshore = (
        st.session_state.get("tipo_poco") == "Onshore"
        and int(st.session_state.get("n_lbf", 1)) == 1
        and int(st.session_state.get("n_trending", 1)) == 1
    )

    prof_primeiro_sonico = None

    if limitar_trending_onshore and "Perfil sônico (µs/pé)" in df_pp.columns:
        sonico_ref = pd.to_numeric(
            df_pp["Perfil sônico (µs/pé)"],
            errors="coerce"
        )

        prof_validas_sonico = prof[
            prof.notna()
            & sonico_ref.notna()
        ]

        if not prof_validas_sonico.empty:
            prof_primeiro_sonico = prof_validas_sonico.min()

    df_pp["Perfil sônico (µs/pé) Reta Normal"] = np.nan

    for idx, tr in enumerate(trendings):
        try:
            s_normal = _calcular_reta_normal_trending(prof, tr)

            if s_normal is None:
                continue

            s_normal = pd.to_numeric(
                pd.Series(s_normal, index=df_pp.index),
                errors="coerce"
            )

            mask_base = prof.notna()

            if (
                prof_primeiro_sonico is not None
                and pd.notna(prof_primeiro_sonico)
            ):
                mask_base = mask_base & (prof >= prof_primeiro_sonico)

            if st.session_state.get("tipo_poco") == "Offshore":
                mask_base = mask_base & (df_pp.index >= 1)

            prof_ini = tr.get("prof_ini")
            prof_fim = tr.get("prof_fim")

            if (
                prof_ini is not None
                and prof_fim is not None
                and prof_fim > prof_ini
            ):
                mask_intervalo = (prof >= prof_ini) & (prof <= prof_fim)
                mask_final = mask_base & mask_intervalo

            else:
                mask_final = mask_base

            df_pp.loc[
                mask_final,
                "Perfil sônico (µs/pé) Reta Normal"
            ] = s_normal[mask_final]

        except Exception as e:
            st.warning(f"Erro ao calcular Trending {idx + 1}: {e}")

    if (
        df_pp["Perfil sônico (µs/pé) Reta Normal"].isna().any()
        and trendings
    ):
        try:
            tr0 = trendings[0]
            s_normal_padrao = _calcular_reta_normal_trending(prof, tr0)

            if s_normal_padrao is not None:
                s_normal_padrao = pd.to_numeric(
                    pd.Series(s_normal_padrao, index=df_pp.index),
                    errors="coerce"
                )

                if st.session_state.get("tipo_poco") == "Offshore":
                    mask_fill = df_pp.index >= 1

                    df_pp.loc[
                        mask_fill,
                        "Perfil sônico (µs/pé) Reta Normal"
                    ] = df_pp.loc[
                        mask_fill,
                        "Perfil sônico (µs/pé) Reta Normal"
                    ].fillna(
                        s_normal_padrao[mask_fill]
                    )

                else:
                    mask_fill = prof.notna()

                    if (
                        prof_primeiro_sonico is not None
                        and pd.notna(prof_primeiro_sonico)
                    ):
                        mask_fill = mask_fill & (prof >= prof_primeiro_sonico)

                    df_pp.loc[
                        mask_fill,
                        "Perfil sônico (µs/pé) Reta Normal"
                    ] = df_pp.loc[
                        mask_fill,
                        "Perfil sônico (µs/pé) Reta Normal"
                    ].fillna(
                        s_normal_padrao[mask_fill]
                    )

        except Exception as e:
            st.warning(f"Erro no preenchimento padrão da reta normal: {e}")

    return df_pp


def _aplicar_suavizacao_pressao_poros(df_pp):
    df_pp = df_pp.copy()

    col_prof = "Profundidade (m)"
    col_gp = "Gradiente de Pressão de Poros (lb/gal)"
    col_gp_suav = "Gradiente de Pressão de Poros Suavizado (lb/gal)"
    col_pp_suav = "Pressão de Poros Suavizado (psi)"

    if col_prof not in df_pp.columns or col_gp not in df_pp.columns:
        return df_pp

    if st.session_state.get("spp", "Não") != "Sim":
        return df_pp

    janela_media = int(st.session_state.get("janela_spp", 20))
    limite_variacao = float(st.session_state.get("limite_spp", 0.10))

    janela_media = max(1, janela_media)
    limite_variacao = max(0.0, limite_variacao)

    grad_original = pd.to_numeric(
        df_pp[col_gp],
        errors="coerce"
    )

    prof = pd.to_numeric(
        df_pp[col_prof],
        errors="coerce"
    )

    grad_suav = grad_original.copy()

    anormal = float(st.session_state.get("anormal", 0.0))

    mask_suav = (
            grad_original.notna()
            & prof.notna()
            & (prof >= anormal)
    )

    if st.session_state.get("tipo_poco") == "Offshore":
        mask_suav = mask_suav & (df_pp.index >= 1)

    if not mask_suav.any():
        df_pp[col_gp_suav] = grad_suav
        df_pp[col_pp_suav] = 0.1704 * grad_suav * prof
        return df_pp

    media_movel = (
        grad_original.loc[mask_suav]
        .rolling(
            window=(2 * janela_media + 1),
            center=True,
            min_periods=1
        )
        .mean()
    )

    media_limitada = media_movel.copy()

    valor_anterior = None

    for idx in media_limitada.index:
        valor_atual = media_limitada.loc[idx]

        if pd.isna(valor_atual):
            continue

        if valor_anterior is None:
            media_limitada.loc[idx] = valor_atual
            valor_anterior = valor_atual
            continue

        if valor_atual > valor_anterior + limite_variacao:
            valor_atual = valor_anterior + limite_variacao

        elif valor_atual < valor_anterior - limite_variacao:
            valor_atual = valor_anterior - limite_variacao

        media_limitada.loc[idx] = valor_atual
        valor_anterior = valor_atual

    grad_suav.loc[mask_suav] = media_limitada

    df_pp[col_gp_suav] = grad_suav

    df_pp[col_pp_suav] = (
        0.1704
        * df_pp[col_gp_suav]
        * prof
    )

    return df_pp



def _coletar_boyances_poros():
    boyances = []

    n_boyance = int(st.session_state.get("n_boyance", 1))

    for i in range(n_boyance):
        fpr = st.session_state.get(f"fpr_{i}")

        if fpr is None:
            continue

        prof_inicial = None
        prof_final = None

        if n_boyance > 1:
            prof_inicial = st.session_state.get(f"prof_inicial_{i}")
            prof_final = st.session_state.get(f"prof_final_{i}")

        boyances.append({
            "fpr": fpr,
            "prof_inicial": prof_inicial,
            "prof_final": prof_final
        })

    return boyances


def _selecoes_boyance_poros():
    selecoes = st.session_state.get("o_boyance", [])

    if isinstance(selecoes, str):
        return [selecoes]

    if isinstance(selecoes, (list, tuple, set)):
        return list(selecoes)

    return []


def _calcular_boyance_pp(df_pp, boyances=None):
    df_pp = df_pp.copy()

    if st.session_state.get("boyance", "Não") != "Sim":
        return df_pp

    boyances = boyances if boyances is not None else _coletar_boyances_poros()

    col_prof = "Profundidade (m)"
    col_gr = "Perfil Raio Gama"
    col_gr_suav = "Raio Gama Suavizado"
    col_lbf = "LBF_calc"
    col_gp = "Gradiente de Pressão de Poros (lb/gal)"
    col_gp_suav = "Gradiente de Pressão de Poros Suavizado (lb/gal)"
    col_pp = "Pressão de Poros (psi)"

    col_fpr = "FPR_efetivo"
    col_formacao = "Formação"
    col_press_ta_bf = "Pressão Boyance (TA = BF)"
    col_boy_ta_bf = "Boyance (lb/gal) (TA = BF)"
    col_press_ba_tf = "Pressão Boyance (BA = TF)"
    col_boy_ba_tf = "Boyance (lb/gal) (BA = TF)"

    for col in (
        col_fpr,
        col_formacao,
        col_press_ta_bf,
        col_boy_ta_bf,
        col_press_ba_tf,
        col_boy_ba_tf,
    ):
        if col not in df_pp.columns:
            df_pp[col] = np.nan

    colunas_obrigatorias = [col_prof, col_gr, col_lbf, col_gp]
    if any(col not in df_pp.columns for col in colunas_obrigatorias):
        return df_pp

    prof = pd.to_numeric(df_pp[col_prof], errors="coerce")
    grad_pp = pd.to_numeric(df_pp[col_gp], errors="coerce")

    if col_gp_suav in df_pp.columns:
        grad_referencia = pd.to_numeric(df_pp[col_gp_suav], errors="coerce").combine_first(grad_pp)
    else:
        grad_referencia = grad_pp

    if col_pp in df_pp.columns:
        pressao_referencia = pd.to_numeric(df_pp[col_pp], errors="coerce")
    else:
        pressao_referencia = 0.1704 * grad_referencia * prof

    df_pp[col_fpr] = np.nan

    if len(boyances) == 1:
        try:
            df_pp[col_fpr] = float(boyances[0].get("fpr"))
        except (TypeError, ValueError):
            df_pp[col_fpr] = np.nan
    else:
        for boyance in boyances:
            try:
                fpr = float(boyance.get("fpr"))
                prof_inicial = float(boyance.get("prof_inicial"))
                prof_final = float(boyance.get("prof_final"))
            except (TypeError, ValueError):
                continue

            mask_intervalo = (
                prof.notna()
                & (prof >= prof_inicial)
                & (prof <= prof_final)
            )

            df_pp.loc[mask_intervalo, col_fpr] = fpr

    fpr_padrao = None
    if boyances:
        try:
            fpr_padrao = float(boyances[0].get("fpr"))
        except (TypeError, ValueError):
            fpr_padrao = None

    if fpr_padrao is not None:
        df_pp[col_fpr] = df_pp[col_fpr].ffill().fillna(fpr_padrao)

    fpr_efetivo = pd.to_numeric(df_pp[col_fpr], errors="coerce")

    incremento = (
        0.1704
        * fpr_efetivo
        * (prof - prof.shift(1))
    )

    if (
        st.session_state.get("s_gr") == "Sim"
        and col_gr_suav in df_pp.columns
    ):
        curva_gr = pd.to_numeric(df_pp[col_gr_suav], errors="coerce")
    else:
        curva_gr = pd.to_numeric(df_pp[col_gr], errors="coerce")

    lbf_calc = pd.to_numeric(df_pp[col_lbf], errors="coerce")

    df_pp[col_formacao] = np.where(
        curva_gr < lbf_calc,
        "Formação Permeável",
        "Formação Impermeável"
    )

    topo_permeavel = (
        (df_pp[col_formacao] == "Formação Permeável")
        & (df_pp[col_formacao].shift(1) != "Formação Permeável")
    )

    df_pp[col_press_ta_bf] = np.nan
    df_pp.loc[topo_permeavel, col_press_ta_bf] = (
        grad_referencia.shift(1)
        * 0.1704
        * prof
    )

    id_camada = topo_permeavel.cumsum()
    mask_perm = df_pp[col_formacao] == "Formação Permeável"

    serie_ta_bf = (
        df_pp.loc[mask_perm]
        .groupby(id_camada[mask_perm], group_keys=False)
        .apply(
            lambda g: (
                g[col_press_ta_bf].iloc[0]
                + incremento.loc[g.index].fillna(0).cumsum()
            )
        )
    )

    if not serie_ta_bf.empty:
        serie_ta_bf = serie_ta_bf.sort_index()
        df_pp.loc[serie_ta_bf.index, col_press_ta_bf] = serie_ta_bf

    denom = 0.1704 * prof.replace(0, np.nan)

    boyance_ta_bf = np.where(
        df_pp[col_formacao] == "Formação Impermeável",
        grad_referencia,
        pd.to_numeric(df_pp[col_press_ta_bf], errors="coerce") / denom
    )

    df_pp[col_boy_ta_bf] = (
        pd.Series(boyance_ta_bf, index=df_pp.index)
        .replace([np.inf, -np.inf], np.nan)
        .fillna(grad_referencia)
    )

    df_pp[col_press_ba_tf] = np.nan

    base_permeavel = (
        (df_pp[col_formacao] == "Formação Permeável")
        & (df_pp[col_formacao].shift(-1) != "Formação Permeável")
    )

    for idx in df_pp.index[base_permeavel]:
        prox_idx = idx + 1
        if prox_idx in df_pp.index and df_pp.loc[prox_idx, col_formacao] == "Formação Impermeável":
            df_pp.loc[idx, col_press_ba_tf] = pressao_referencia.loc[prox_idx]
        else:
            df_pp.loc[idx, col_press_ba_tf] = pressao_referencia.loc[idx]

    id_camada = (
        (df_pp[col_formacao] == "Formação Permeável")
        & (df_pp[col_formacao].shift(1) != "Formação Permeável")
    ).cumsum()

    serie_ba_tf = (
        df_pp.loc[mask_perm]
        .groupby(id_camada[mask_perm], group_keys=False)
        .apply(
            lambda g: (
                g[col_press_ba_tf].iloc[-1]
                - incremento.loc[g.index].fillna(0).iloc[::-1].cumsum().iloc[::-1]
            )
        )
    )

    if not serie_ba_tf.empty:
        serie_ba_tf = serie_ba_tf.sort_index()
        df_pp.loc[serie_ba_tf.index, col_press_ba_tf] = serie_ba_tf

    impermeavel = df_pp[col_formacao] == "Formação Impermeável"
    impermeavel_acima = impermeavel.shift(fill_value=False).cumsum() > 0

    boyance_ba_tf = np.where(
        (~impermeavel) & impermeavel_acima,
        pd.to_numeric(df_pp[col_press_ba_tf], errors="coerce") / denom,
        grad_referencia
    )

    df_pp[col_boy_ba_tf] = (
        pd.Series(boyance_ba_tf, index=df_pp.index)
        .replace([np.inf, -np.inf], np.nan)
        .fillna(grad_referencia)
    )

    return df_pp


def _calcular_pressao_poros_por_partes(df_pp):
    df_pp = df_pp.copy()

    col_prof = "Profundidade (m)"
    col_sonico = "Perfil sônico (µs/pé)"
    col_sonico_suav = "Perfil sônico suavizado (µs/pé)"
    col_sonico_normal = "Perfil sônico (µs/pé) Reta Normal"
    col_gr = "Perfil Raio Gama"
    col_gr_suav = "Raio Gama Suavizado"
    col_lbf = "LBF_calc"
    col_gs = "Gradiente de Sobrecarga (lb/gal)"
    col_gp = "Gradiente de Pressão de Poros (lb/gal)"
    col_gp_normal = "Gradiente de Pressão de Poros Normal (lb/gal)"
    col_pp = "Pressão de Poros (psi)"

    colunas_obrigatorias = [
        col_prof,
        col_sonico,
        col_sonico_normal,
        col_gr,
        col_lbf,
        col_gs
    ]

    colunas_faltantes = [
        col for col in colunas_obrigatorias
        if col not in df_pp.columns
    ]

    if colunas_faltantes:
        st.warning(
            "Não foi possível calcular a Pressão de Poros. "
            f"Colunas ausentes: {', '.join(colunas_faltantes)}"
        )
        return df_pp

    prof = pd.to_numeric(df_pp[col_prof], errors="coerce")
    sonico_bruto = pd.to_numeric(df_pp[col_sonico], errors="coerce")
    sonico_normal = pd.to_numeric(df_pp[col_sonico_normal], errors="coerce")
    grad_sobrecarga = pd.to_numeric(df_pp[col_gs], errors="coerce")
    lbf_calc = pd.to_numeric(df_pp[col_lbf], errors="coerce")

    if (
        st.session_state.get("suav_s") == "Sim"
        and col_sonico in df_pp.columns
    ):
        try:
            sonico_suavizado = suavizar(prof, sonico_bruto)

            df_pp[col_sonico_suav] = pd.to_numeric(
                pd.Series(sonico_suavizado, index=df_pp.index),
                errors="coerce"
            )

        except Exception as e:
            st.warning(f"Erro ao suavizar sônico para Eaton: {e}")

    if (
        st.session_state.get("suav_s") == "Sim"
        and col_sonico_suav in df_pp.columns
    ):
        sonico_usado = pd.to_numeric(df_pp[col_sonico_suav], errors="coerce")
        nome_sonico_usado = col_sonico_suav

    else:
        sonico_usado = sonico_bruto
        nome_sonico_usado = col_sonico

    if (
        st.session_state.get("s_gr") == "Sim"
        and col_gr_suav in df_pp.columns
    ):
        curva_gr = pd.to_numeric(df_pp[col_gr_suav], errors="coerce")
        nome_curva_gr = col_gr_suav

    else:
        curva_gr = pd.to_numeric(df_pp[col_gr], errors="coerce")
        nome_curva_gr = col_gr

    df_pp["Sônico usado Eaton"] = sonico_usado
    df_pp["Sônico normal usado Eaton"] = sonico_normal
    df_pp["Curva GR usada Eaton"] = curva_gr

    mask_folhelho = (
        curva_gr.notna()
        & lbf_calc.notna()
        & (curva_gr >= lbf_calc)
    )

    if st.session_state.get("tipo_poco") == "Offshore":
        mask_folhelho = mask_folhelho & (df_pp.index >= 1)

    df_pp["Ponto válido para Eaton"] = mask_folhelho
    df_pp["Trecho PP"] = np.where(
        mask_folhelho,
        "Eaton no folhelho",
        "Mantém PP do folhelho acima"
    )

    gp = []

    gn = float(st.session_state.get("gn", 8.5))
    expoente = float(st.session_state.get("expoente", 3.0))
    anormal = float(st.session_state.get("anormal", 0.0))

    if st.session_state.get("tipo_poco") == "Onshore":
        normal(df_pp)

        df_gfs_aux = st.session_state.df_gfs.copy()

        df_gfs_aux["Profundidade (m)"] = pd.to_numeric(
            df_gfs_aux["Profundidade (m)"],
            errors="coerce"
        )

        df_gfs_aux["Gradiente de Pressão de Poros (lb/gal)"] = pd.to_numeric(
            df_gfs_aux["Gradiente de Pressão de Poros (lb/gal)"],
            errors="coerce"
        )

        df_gfs_aux = df_gfs_aux.dropna(
            subset=[
                "Profundidade (m)",
                "Gradiente de Pressão de Poros (lb/gal)"
            ]
        )

        df_gfs_aux = (
            df_gfs_aux
            .sort_values("Profundidade (m)")
            .drop_duplicates(subset=["Profundidade (m)"])
        )

        if not df_gfs_aux.empty:
            df_pp[col_gp_normal] = np.interp(
                prof.astype(float),
                df_gfs_aux["Profundidade (m)"].astype(float),
                df_gfs_aux["Gradiente de Pressão de Poros (lb/gal)"].astype(float),
                left=np.nan,
                right=np.nan
            )
        else:
            df_pp[col_gp_normal] = np.nan

    else:
        df_pp[col_gp_normal] = np.nan

    ultimo_grad_folhelho = gn

    for i in range(len(df_pp)):
        prof_i = prof.iloc[i]
        son_i = sonico_usado.iloc[i]
        son_norm_i = sonico_normal.iloc[i]
        gs_i = grad_sobrecarga.iloc[i]
        ponto_valido = bool(mask_folhelho.iloc[i])

        if pd.isna(prof_i):
            gp.append(np.nan)
            continue

        if st.session_state.get("tipo_poco") == "Offshore" and i < 1:
            gp.append(np.nan)
            continue

        if prof_i < anormal:
            gp.append(gn)
            ultimo_grad_folhelho = gn
            continue

        if not ponto_valido:
            gp.append(ultimo_grad_folhelho)
            continue

        if (
            pd.isna(son_i)
            or pd.isna(son_norm_i)
            or pd.isna(gs_i)
            or son_i <= 0
            or son_norm_i <= 0
        ):
            gp.append(ultimo_grad_folhelho)
            continue

        grad_eaton = (
            gs_i
            - (
                (gs_i - gn)
                * ((son_i / son_norm_i) ** (-expoente))
            )
        )

        if pd.isna(grad_eaton):
            grad_final = ultimo_grad_folhelho

        elif grad_eaton < gn:
            grad_final = gn

        else:
            grad_final = grad_eaton

        gp.append(grad_final)
        ultimo_grad_folhelho = grad_final

    df_pp[col_gp] = pd.to_numeric(gp, errors="coerce")

    if (
        st.session_state.get("tipo_poco") == "Onshore"
        and "df_gfs" in st.session_state
        and isinstance(st.session_state.df_gfs, pd.DataFrame)
        and not st.session_state.df_gfs.empty
    ):
        prof_fim_normal = float(
            pd.to_numeric(
                st.session_state.df_gfs["Profundidade (m)"],
                errors="coerce"
            ).max()
        )

        mask_normal = prof <= prof_fim_normal

        df_pp.loc[mask_normal, col_gp] = (
            df_pp.loc[mask_normal, col_gp_normal]
            .combine_first(df_pp.loc[mask_normal, col_gp])
        )

        mask_ate_anormal = (
            (prof > prof_fim_normal)
            & (prof < anormal)
        )

        df_pp.loc[mask_ate_anormal, col_gp] = (
            df_pp.loc[mask_ate_anormal, col_gp]
            .fillna(gn)
        )

        if st.session_state.get("ex") == "Ativada":
            rtkb = float(st.session_state.get("rtkb", 0.0))
            mask_rtkb = prof <= rtkb

            df_pp.loc[mask_rtkb, col_gp_normal] = np.nan
            df_pp.loc[mask_rtkb, col_gp] = np.nan

    df_pp[col_pp] = (
            0.1704
            * df_pp[col_gp]
            * df_pp[col_prof]
    )

    if "Gradiente de Sobrecarga (lb/gal)" in df_pp.columns:
        df_pp["Pressão de Sobrecarga (psi)"] = (
                0.1704
                * pd.to_numeric(
            df_pp["Gradiente de Sobrecarga (lb/gal)"],
            errors="coerce"
        )
                * pd.to_numeric(
            df_pp[col_prof],
            errors="coerce"
        )
        )

    df_pp = _aplicar_suavizacao_pressao_poros(df_pp)
    df_pp = _calcular_boyance_pp(df_pp)

    return df_pp




def plotar_boyance_pp(ax, df_pp, modo_grafico="Gradiente (lb/gal)"):
    if st.session_state.get("boyance", "Não") != "Sim":
        return

    if not isinstance(df_pp, pd.DataFrame) or df_pp.empty:
        return

    if "Profundidade (m)" not in df_pp.columns:
        return

    selecoes = _selecoes_boyance_poros()
    col_prof = "Profundidade (m)"

    curvas = []

    if "Topo Aren. = Base Folh." in selecoes:
        curvas.append((
            "Pressão Boyance (TA = BF)" if modo_grafico == "Pressão (psi)" else "Boyance (lb/gal) (TA = BF)",
            "Boyance (TA = BF)",
            "red"
        ))

    if "Base Aren. = Topo Folh." in selecoes:
        curvas.append((
            "Pressão Boyance (BA = TF)" if modo_grafico == "Pressão (psi)" else "Boyance (lb/gal) (BA = TF)",
            "Boyance (BA = TF)",
            "green"
        ))

    for coluna, label, cor in curvas:
        if coluna not in df_pp.columns:
            continue

        df_boyance_plot = df_pp[[col_prof, coluna]].copy()
        df_boyance_plot[col_prof] = pd.to_numeric(df_boyance_plot[col_prof], errors="coerce")
        df_boyance_plot[coluna] = pd.to_numeric(df_boyance_plot[coluna], errors="coerce")
        df_boyance_plot = df_boyance_plot.dropna(subset=[col_prof, coluna])

        if df_boyance_plot.empty:
            continue

        ax.plot(
            df_boyance_plot[coluna],
            df_boyance_plot[col_prof],
            color=cor,
            linestyle="-",
            linewidth=2,
            label=label
        )


def plotar_rft(ax, modo_grafico="Gradiente (lb/gal)", label="Teste RFT"):
    if (
        "rft_pontos_pp" not in st.session_state
        or not isinstance(st.session_state.rft_pontos_pp, pd.DataFrame)
        or st.session_state.rft_pontos_pp.empty
    ):
        return

    df_rft_plot = st.session_state.rft_pontos_pp.copy()

    col_prof = "Profundidade (m)"
    col_rft = "Teste RFT (lb/gal)"

    if col_prof not in df_rft_plot.columns or col_rft not in df_rft_plot.columns:
        return

    df_rft_plot[col_prof] = pd.to_numeric(
        df_rft_plot[col_prof],
        errors="coerce"
    )

    df_rft_plot[col_rft] = pd.to_numeric(
        df_rft_plot[col_rft],
        errors="coerce"
    )

    df_rft_plot = df_rft_plot.dropna(
        subset=[col_prof, col_rft]
    )

    if df_rft_plot.empty:
        return

    if modo_grafico == "Gradiente (lb/gal)":
        x_rft = df_rft_plot[col_rft]
    else:
        x_rft = (
            df_rft_plot[col_rft]
            * 0.1704
            * df_rft_plot[col_prof]
        )

    ax.scatter(
        x_rft,
        df_rft_plot[col_prof],
        color="limegreen",
        edgecolors="black",
        marker="o",
        s=50,
        label=label,
        zorder=50
    )

    if st.session_state.get("mostrar_texto_rft_pp", "Sim") == "Sim":
        for x, prof, rft in zip(
            x_rft,
            df_rft_plot[col_prof],
            df_rft_plot[col_rft]
        ):
            ax.annotate(
                f"{rft:.2f} ppg",
                xy=(x, prof),
                xytext=(6, 0),
                textcoords="offset points",
                fontsize=8,
                color="black",
                va="center",
                ha="left",
                zorder=51
            )


def plotar_gradiente_colapso(ax, modo_grafico="Gradiente (lb/gal)", label="Gradiente de Colapso"):
    if (
        "colapso_pontos_pp" not in st.session_state
        or not isinstance(st.session_state.colapso_pontos_pp, pd.DataFrame)
        or st.session_state.colapso_pontos_pp.empty
    ):
        return

    df_colapso_plot = st.session_state.colapso_pontos_pp.copy()

    col_prof = "Profundidade (m)"
    col_colapso = "Gradiente de Colapso (lb/gal)"

    if col_prof not in df_colapso_plot.columns or col_colapso not in df_colapso_plot.columns:
        return

    df_colapso_plot[col_prof] = pd.to_numeric(
        df_colapso_plot[col_prof],
        errors="coerce"
    )

    df_colapso_plot[col_colapso] = pd.to_numeric(
        df_colapso_plot[col_colapso],
        errors="coerce"
    )

    df_colapso_plot = df_colapso_plot.dropna(
        subset=[col_prof, col_colapso]
    )

    if df_colapso_plot.empty:
        return

    df_colapso_plot = df_colapso_plot.sort_values(col_prof)

    if modo_grafico == "Gradiente (lb/gal)":
        x_colapso = df_colapso_plot[col_colapso]
    else:
        x_colapso = (
            df_colapso_plot[col_colapso]
            * 0.1704
            * df_colapso_plot[col_prof]
        )

    ax.scatter(
        x_colapso,
        df_colapso_plot[col_prof],
        color="purple",
        edgecolors="black",
        marker="s",
        s=50,
        label=label,
        zorder=45
    )

    if st.session_state.get("mostrar_texto_colapso_pp", "Sim") == "Sim":
        for x, prof, grad in zip(
            x_colapso,
            df_colapso_plot[col_prof],
            df_colapso_plot[col_colapso]
        ):
            ax.annotate(
                f"{grad:.2f} ppg",
                xy=(x, prof),
                xytext=(6, 0),
                textcoords="offset points",
                fontsize=8,
                color="black",
                va="center",
                ha="left",
                zorder=46
            )


def _plotar_pressao_poros_com_contexto(df_pp):
    if not isinstance(df_pp, pd.DataFrame) or df_pp.empty:
        st.warning("Calcule a Pressão de Poros antes de gerar o gráfico.")
        return

    col_prof = "Profundidade (m)"
    col_gp = "Gradiente de Pressão de Poros (lb/gal)"
    col_gp_suav = "Gradiente de Pressão de Poros Suavizado (lb/gal)"
    col_pp = "Pressão de Poros (psi)"
    col_pp_suav = "Pressão de Poros Suavizado (psi)"
    col_gs = "Gradiente de Sobrecarga (lb/gal)"
    col_ps = "Pressão de Sobrecarga (psi)"

    if col_prof not in df_pp.columns:
        st.warning("Coluna de profundidade não encontrada em df_pp.")
        return

    opcao_grafico_pp = st.session_state.get("ogp", "Gradiente (lb/gal)")

    usar_pp_suavizada = (
            st.session_state.get("spp", "Não") == "Sim"
    )

    if opcao_grafico_pp == "Pressão (psi)":
        coluna_principal = (
            col_pp_suav
            if usar_pp_suavizada and col_pp_suav in df_pp.columns
            else col_pp
        )

        coluna_sobrecarga = col_ps
        titulo = "Pressão de Poros (psi)"
        xlabel = "Pressão (psi)"

        legenda_principal = (
            "Pressão de Poros Suavizada"
            if coluna_principal == col_pp_suav
            else "Pressão de Poros"
        )

        legenda_sobrecarga = "Pressão de Sobrecarga"

    else:
        coluna_principal = (
            col_gp_suav
            if usar_pp_suavizada and col_gp_suav in df_pp.columns
            else col_gp
        )

        coluna_sobrecarga = col_gs
        titulo = "Gradiente de Pressão de Poros (lb/gal)"
        xlabel = "Gradiente (ppg)"

        legenda_principal = (
            "Gradiente de Pressão de Poros Suavizado"
            if coluna_principal == col_gp_suav
            else "Gradiente de Pressão de Poros"
        )

        legenda_sobrecarga = "Gradiente de Sobrecarga"

    plotar_curva_principal = coluna_principal in df_pp.columns

    df_plot = df_pp.copy()

    df_plot[col_prof] = pd.to_numeric(
        df_plot[col_prof],
        errors="coerce"
    )

    if plotar_curva_principal:
        df_plot[coluna_principal] = pd.to_numeric(
            df_plot[coluna_principal],
            errors="coerce"
        )

    if coluna_sobrecarga in df_plot.columns:
        df_plot[coluna_sobrecarga] = pd.to_numeric(
            df_plot[coluna_sobrecarga],
            errors="coerce"
        )

    if plotar_curva_principal:
        df_plot = df_plot.dropna(subset=[col_prof, coluna_principal])
    else:
        df_plot = df_plot.dropna(subset=[col_prof])

    if df_plot.empty:
        st.warning("N?o h? dados de profundidade v?lidos para montar o gr?fico.")
        return

    selected = st.session_state.get(
        "well_selected",
        st.session_state.get("poco", "Poço")
    )

    _garantir_litologia_importada(selected)

    poco = st.session_state.get("pocos", {}).get(selected, {})
    profundidades = poco.get("profundidade", [])
    litologias = poco.get("litologia", [])

    usar_coluna_idade = (
        st.session_state.get("idg") == "Sim"
        and "df_idade" in st.session_state
        and isinstance(st.session_state.df_idade, pd.DataFrame)
        and not st.session_state.df_idade.empty
    )

    fig = plt.figure(figsize=(8, 10))

    if usar_coluna_idade:
        gs = gridspec.GridSpec(
            1,
            4,
            width_ratios=[0.10, 0.18, 0.21, 1],
            wspace=0
        )

        ax_idade = fig.add_subplot(gs[0])
        ax1 = fig.add_subplot(gs[1], sharey=ax_idade)

        ax_gap = fig.add_subplot(gs[2])
        ax_gap.axis("off")

        ax = fig.add_subplot(gs[3], sharey=ax_idade)

        idade_formacao(
            ax_idade,
            st.session_state.df_idade,
            st.session_state.y_max_pp,
            st.session_state.y_min_pp
        )

        ax_idade.tick_params(
            axis="y",
            which="both",
            left=False,
            right=False,
            labelleft=False,
            labelright=False
        )

        ax_idade.set_ylabel("")

        plt.setp(ax1.get_yticklabels(), visible=False)
        plt.setp(ax.get_yticklabels(), visible=False)

    else:
        gs = gridspec.GridSpec(
            1,
            3,
            width_ratios=[0.18, 0.21, 1],
            wspace=0
        )

        ax1 = fig.add_subplot(gs[0])

        ax_gap = fig.add_subplot(gs[1])
        ax_gap.axis("off")

        ax = fig.add_subplot(gs[2], sharey=ax1)

        plt.setp(ax.get_yticklabels(), visible=False)

    df_lito_pp = df_pp

    if (
        st.session_state.get(
            "tipo_coluna_litologica_graficos",
            "Permeável / Não permeável"
        ) == "Permeável / Não permeável"
        and "df_pp_lito" in st.session_state
        and isinstance(st.session_state.df_pp_lito, pd.DataFrame)
        and not st.session_state.df_pp_lito.empty
    ):
        df_lito_pp = st.session_state.df_pp_lito.copy()

    if "Profundidade (m)" not in df_lito_pp.columns and "Profundidade" in df_lito_pp.columns:
        df_lito_pp["Profundidade (m)"] = df_lito_pp["Profundidade"]

    lito(
        ax1,
        df_lito_pp,
        profundidades,
        litologias,
        st.session_state.y_max_pp,
        y_min=st.session_state.y_min_pp,
        y_max=st.session_state.y_max_pp
    )

    if plotar_curva_principal:
        ax.plot(
            df_plot[coluna_principal],
            df_plot[col_prof],
            color="orange",
            linestyle="-",
            linewidth=2,
            label=legenda_principal
        )

    if (
        plotar_curva_principal
        and st.session_state.get("grafpp") == "Sim"
        and coluna_sobrecarga in df_pp.columns
    ):
        df_sob = df_pp[[col_prof, coluna_sobrecarga]].copy()

        df_sob[col_prof] = pd.to_numeric(
            df_sob[col_prof],
            errors="coerce"
        )

        df_sob[coluna_sobrecarga] = pd.to_numeric(
            df_sob[coluna_sobrecarga],
            errors="coerce"
        )

        df_sob = df_sob.dropna(subset=[col_prof, coluna_sobrecarga])

        if not df_sob.empty:
            ax.plot(
                df_sob[coluna_sobrecarga],
                df_sob[col_prof],
                color="black",
                linestyle="-",
                linewidth=2,
                label=legenda_sobrecarga
            )

    if (
        plotar_curva_principal
        and "df_mud" in st.session_state
        and isinstance(st.session_state.df_mud, pd.DataFrame)
        and not st.session_state.df_mud.empty
    ):
        df_mud = st.session_state.df_mud.copy()

        col_mud_prof = "Profundidade (m)"
        col_mud_plan = "Peso do Fluido Planejado (lb/gal)"
        col_mud_exec = "Peso do Fluido Executado (lb/gal)"

        if col_mud_prof in df_mud.columns:
            df_mud[col_mud_prof] = pd.to_numeric(
                df_mud[col_mud_prof],
                errors="coerce"
            )

            mostrar_planejado = st.session_state.get("fpl", "Não") == "Sim"
            mostrar_executado = (
                st.session_state.get("option") == "Retroanálise"
                and st.session_state.get("fex", "Não") == "Sim"
            )

            if mostrar_planejado and col_mud_plan in df_mud.columns:
                df_mud[col_mud_plan] = pd.to_numeric(
                    df_mud[col_mud_plan],
                    errors="coerce"
                )

                df_mud_plan = df_mud.dropna(
                    subset=[col_mud_prof, col_mud_plan]
                )

                if not df_mud_plan.empty:
                    ax.plot(
                        df_mud_plan[col_mud_plan],
                        df_mud_plan[col_mud_prof],
                        linestyle="-",
                        color="green",
                        linewidth=2,
                        label="Peso do Fluido (Planejado)",
                        zorder=5
                    )

            if mostrar_executado and col_mud_exec in df_mud.columns:
                df_mud[col_mud_exec] = pd.to_numeric(
                    df_mud[col_mud_exec],
                    errors="coerce"
                )

                df_mud_exec = df_mud.dropna(
                    subset=[col_mud_prof, col_mud_exec]
                )

                if not df_mud_exec.empty:
                    ax.plot(
                        df_mud_exec[col_mud_exec],
                        df_mud_exec[col_mud_prof],
                        linestyle="-",
                        color="mediumvioletred",
                        linewidth=2,
                        label="Peso do Fluido (Executado)",
                        zorder=5
                    )

    plotar_boyance_pp(ax, df_pp, modo_grafico=opcao_grafico_pp)
    plotar_rft(ax, modo_grafico=opcao_grafico_pp)
    plotar_gradiente_colapso(ax, modo_grafico=opcao_grafico_pp)

    ax.set_title(
        titulo,
        fontsize=14,
        fontweight="bold"
    )

    ax.set_xlabel(xlabel, fontsize=12)
    ax.set_ylabel("Profundidade TVD (m)", fontsize=12)

    ax.invert_yaxis()
    ax.tick_params(axis="y", which="both", left=True, labelleft=True)

    y_min_pp = float(st.session_state.get("y_min_pp", 0.0))
    y_max_pp = float(st.session_state.get("y_max_pp", 1000.0))
    y_step_pp = max(0.01, float(st.session_state.get("y_step_pp", 200.0)))

    x_min_pp = float(st.session_state.get("x_min_pp", 7.0))
    x_max_pp = float(st.session_state.get("x_max_pp", 21.0))
    x_step_pp = max(0.01, float(st.session_state.get("x_step_pp", 2.0)))

    if x_max_pp <= x_min_pp:
        x_max_pp = x_min_pp + x_step_pp

    if y_max_pp <= y_min_pp:
        y_max_pp = y_min_pp + y_step_pp

    ax.set_yticks(
        np.arange(y_min_pp, y_max_pp + y_step_pp * 0.5, y_step_pp)
    )

    ax.set_ylim(y_max_pp, y_min_pp)

    x_tick_ini_pp = math.ceil(x_min_pp)
    x_tick_fim_pp = math.floor(x_max_pp)
    x_tick_step_pp = max(1.0, x_step_pp)
    if x_tick_fim_pp >= x_tick_ini_pp:
        ax.set_xticks(
            np.arange(x_tick_ini_pp, x_tick_fim_pp + x_tick_step_pp * 0.5, x_tick_step_pp)
        )
    if np.isclose(x_step_pp, 1.0):
        ax.set_xticks(
            np.arange(x_min_pp, x_max_pp + 0.25, 0.5),
            minor=True
        )
    ax.xaxis.set_major_formatter(FuncFormatter(lambda valor, pos: f"{valor:.0f}"))

    ax.set_xlim(x_min_pp, x_max_pp)

    ax.grid(True, which="major", linestyle="--", alpha=0.5)
    if np.isclose(x_step_pp, 1.0):
        ax.grid(True, which="minor", axis="x", linestyle="--", alpha=0.5)

    if ax.get_legend_handles_labels()[0]:
        ax.legend(
            loc="upper right",
            fontsize=8,
            frameon=True,
            shadow=True,
            fancybox=True,
            framealpha=1,
            facecolor="white",
            edgecolor="gray"
        )

    add_watermark(
        ax,
        logo_path="logo2.png",
        xy=(0.50, 0.5),
        zoom=0.2,
        alpha=0.2,
        zorder=0
    )

    st.session_state.fig_pp = fig
    st.session_state.fig_pressao_poros = fig

    st.pyplot(fig)


def _plotar_gr_poros_com_contexto(df_pp, modo_grafico, lbfs=None, trendings=None):
    if not isinstance(df_pp, pd.DataFrame) or df_pp.empty:
        st.warning("Calcule o Gradiente de Sobrecarga antes de gerar o gráfico.")
        return

    lbfs = lbfs or []
    trendings = trendings or []

    if modo_grafico == "Trending":
        coluna_base = "Perfil sônico (µs/pé)"
        coluna_suavizada = None
        label_base = "Sônico"
        label_suavizada = "Sônico suavizado"
        xlabel_plot = "Perfil sônico (µs/pé)"

    else:
        coluna_base = "Perfil Raio Gama"
        coluna_suavizada = "Raio Gama Suavizado"
        label_base = "Perfil Raio Gama"
        label_suavizada = "Raio Gama Suavizado"
        xlabel_plot = "Perfil Raio Gama (GAPI)"

    if coluna_base not in df_pp.columns or "Profundidade (m)" not in df_pp.columns:
        st.warning(f"Dados de {label_base}/profundidade não disponíveis para gerar o gráfico.")
        return

    selected = st.session_state.get("well_selected", st.session_state.get("poco", "Poço"))
    _garantir_litologia_importada(selected)

    poco = st.session_state.get("pocos", {}).get(selected, {})
    profundidades = poco.get("profundidade", [])
    litologias = poco.get("litologia", [])

    usar_coluna_idade = (
        st.session_state.get("idg") == "Sim"
        and "df_idade" in st.session_state
        and isinstance(st.session_state.df_idade, pd.DataFrame)
        and not st.session_state.df_idade.empty
    )

    fig = plt.figure(figsize=(8, 10))

    if usar_coluna_idade:
        gs = gridspec.GridSpec(
            1,
            4,
            width_ratios=[0.10, 0.18, 0.21, 1],
            wspace=0
        )

        ax_idade = fig.add_subplot(gs[0])
        ax1 = fig.add_subplot(gs[1], sharey=ax_idade)

        ax_gap = fig.add_subplot(gs[2])
        ax_gap.axis("off")

        ax = fig.add_subplot(gs[3], sharey=ax_idade)

        idade_formacao(
            ax_idade,
            st.session_state.df_idade,
            st.session_state.y_max_pp,
            st.session_state.y_min_pp
        )

        ax_idade.tick_params(
            axis="y",
            which="both",
            left=False,
            right=False,
            labelleft=False,
            labelright=False
        )

        ax_idade.set_ylabel("")

        plt.setp(ax1.get_yticklabels(), visible=False)
        plt.setp(ax.get_yticklabels(), visible=False)

    else:
        gs = gridspec.GridSpec(
            1,
            3,
            width_ratios=[0.18, 0.21, 1],
            wspace=0
        )

        ax1 = fig.add_subplot(gs[0])

        ax_gap = fig.add_subplot(gs[1])
        ax_gap.axis("off")

        ax = fig.add_subplot(gs[2], sharey=ax1)

        plt.setp(ax.get_yticklabels(), visible=False)

    df_lito_gr = df_pp

    if (
        st.session_state.get(
            "tipo_coluna_litologica_graficos",
            "Permeável / Não permeável"
        ) == "Permeável / Não permeável"
        and "df_pp_lito" in st.session_state
        and isinstance(st.session_state.df_pp_lito, pd.DataFrame)
        and not st.session_state.df_pp_lito.empty
    ):
        df_lito_gr = st.session_state.df_pp_lito.copy()

    if "Profundidade (m)" not in df_lito_gr.columns and "Profundidade" in df_lito_gr.columns:
        df_lito_gr["Profundidade (m)"] = df_lito_gr["Profundidade"]

    lito(
        ax1,
        df_lito_gr,
        profundidades,
        litologias,
        st.session_state.y_max_pp,
        y_min=st.session_state.y_min_pp,
        y_max=st.session_state.y_max_pp
    )

    colunas_plot = ["Profundidade (m)", coluna_base]

    if coluna_suavizada is not None and coluna_suavizada in df_pp.columns:
        colunas_plot.append(coluna_suavizada)

    for col_extra in ["Perfil Raio Gama", "Raio Gama Suavizado", "LBF_calc"]:
        if col_extra in df_pp.columns and col_extra not in colunas_plot:
            colunas_plot.append(col_extra)

    df_plot = df_pp[colunas_plot].copy()

    df_plot["Profundidade (m)"] = pd.to_numeric(
        df_plot["Profundidade (m)"],
        errors="coerce"
    )

    df_plot[coluna_base] = pd.to_numeric(
        df_plot[coluna_base],
        errors="coerce"
    )

    if coluna_suavizada is not None and coluna_suavizada in df_plot.columns:
        df_plot[coluna_suavizada] = pd.to_numeric(
            df_plot[coluna_suavizada],
            errors="coerce"
        )

    for col_num in ["Perfil Raio Gama", "Raio Gama Suavizado", "LBF_calc"]:
        if col_num in df_plot.columns:
            df_plot[col_num] = pd.to_numeric(
                df_plot[col_num],
                errors="coerce"
            )

    df_plot = df_plot.dropna(subset=["Profundidade (m)", coluna_base])

    y_min_pp = int(st.session_state.get("y_min_pp", 0))
    y_max_pp = int(st.session_state.get("y_max_pp", 1000))
    y_step_pp = max(1, int(st.session_state.get("y_step_pp", 200)))

    if modo_grafico == "Trending":
        df_plot = df_plot[df_plot[coluna_base] > 0]

        if df_plot.empty:
            st.warning("Não há valores positivos de perfil sônico para plotar em escala semilog.")
            return

        ax.semilogx(
            df_plot[coluna_base],
            df_plot["Profundidade (m)"],
            color="red",
            linewidth=2,
            label=label_base
        )

        if st.session_state.get("suav_s") == "Sim":
            x_suav = suavizar(
                df_plot["Profundidade (m)"],
                df_plot[coluna_base]
            )

            x_suav = pd.to_numeric(
                pd.Series(x_suav, index=df_plot.index),
                errors="coerce"
            )

            mask_suav = x_suav > 0

            ax.semilogx(
                x_suav[mask_suav],
                df_plot.loc[mask_suav, "Profundidade (m)"],
                color="blue",
                linewidth=2,
                label=label_suavizada,
                zorder=3
            )

        if "LBF_calc" in df_plot.columns:
            if (
                st.session_state.get("s_gr") == "Sim"
                and "Raio Gama Suavizado" in df_plot.columns
            ):
                coluna_rg_lbf = "Raio Gama Suavizado"
            else:
                coluna_rg_lbf = "Perfil Raio Gama"

            if coluna_rg_lbf in df_plot.columns:
                df_pontos_lbf = df_plot.dropna(
                    subset=[
                        "Profundidade (m)",
                        coluna_base,
                        coluna_rg_lbf,
                        "LBF_calc"
                    ]
                ).copy()

                df_pontos_lbf = df_pontos_lbf[
                    (df_pontos_lbf[coluna_rg_lbf] >= df_pontos_lbf["LBF_calc"])
                    & (df_pontos_lbf[coluna_base] > 0)
                ]

                if not df_pontos_lbf.empty:
                    ax.semilogx(
                        df_pontos_lbf[coluna_base],
                        df_pontos_lbf["Profundidade (m)"],
                        marker="o",
                        linestyle="None",
                        color="cyan",
                        markersize=1,
                        label="Sônico ≥ LBF"
                    )

        prof = pd.to_numeric(
            df_pp["Profundidade (m)"],
            errors="coerce"
        )

        for idx, tr in enumerate(trendings):
            try:
                cor_trending = TRENDING_COLORS[idx % len(TRENDING_COLORS)]

                s_normal = _calcular_reta_normal_trending(prof, tr)

                if s_normal is None:
                    continue

                s_normal = pd.to_numeric(
                    pd.Series(s_normal, index=df_pp.index),
                    errors="coerce"
                )

                mask_base = (
                        prof.notna()
                        & s_normal.notna()
                        & (s_normal > 0)
                )

                if (
                    st.session_state.get("tipo_poco") == "Onshore"
                    and int(st.session_state.get("n_lbf", 1)) == 1
                    and int(st.session_state.get("n_trending", 1)) == 1
                    and "Perfil sônico (µs/pé)" in df_pp.columns
                ):
                    sonico_ref = pd.to_numeric(
                        df_pp["Perfil sônico (µs/pé)"],
                        errors="coerce"
                    )

                    prof_validas_sonico = prof[
                        prof.notna()
                        & sonico_ref.notna()
                    ]

                    if not prof_validas_sonico.empty:
                        mask_base = mask_base & (
                            prof >= prof_validas_sonico.min()
                        )

                if st.session_state.get("tipo_poco") == "Offshore":
                    mask_base = mask_base & (df_pp.index >= 1)

                prof_ini = tr.get("prof_ini")
                prof_fim = tr.get("prof_fim")

                if (
                        prof_ini is not None
                        and prof_fim is not None
                        and prof_fim > prof_ini
                ):
                    mask_intervalo = (prof >= prof_ini) & (prof <= prof_fim)
                    mask_final = mask_base & mask_intervalo

                else:
                    mask_final = mask_base

                if mask_final.any():
                    ax.semilogx(
                        s_normal[mask_final],
                        prof[mask_final],
                        linestyle="--",
                        linewidth=3,
                        color=cor_trending,
                        label=f"Trending {idx + 1}",
                        zorder=5
                    )

            except Exception as e:
                st.warning(f"Erro ao plotar Trending {idx + 1}: {e}")

        ax.xaxis.set_major_locator(FixedLocator([10, 100, 1000]))
        ax.xaxis.set_major_formatter(
            FuncFormatter(lambda valor, pos: f"{int(valor)}" if valor in (10, 100, 1000) else "")
        )

        ax.xaxis.set_minor_locator(
            LogLocator(base=10.0, subs=np.arange(2, 10) * 0.1)
        )
        ax.xaxis.set_minor_formatter(NullFormatter())

        ax.grid(True, which="major", linestyle="--", alpha=0.6)
        ax.grid(True, which="minor", linestyle=":", alpha=0.4)

        maior_x_positivo = float(df_plot[coluna_base].max())

        x_min_plot = 10.0
        x_max_plot = max(
            1000.0,
            maior_x_positivo * 1.05
        )

        ax.set_xlim(x_min_plot, x_max_plot)

    else:
        ax.plot(
            df_plot[coluna_base],
            df_plot["Profundidade (m)"],
            color="red",
            linewidth=2,
            label=label_base
        )

        if (
            st.session_state.get("s_gr") == "Sim"
            and coluna_suavizada in df_plot.columns
        ):
            df_suav_gr = df_plot.dropna(
                subset=["Profundidade (m)", coluna_suavizada]
            )

            if not df_suav_gr.empty:
                ax.plot(
                    df_suav_gr[coluna_suavizada],
                    df_suav_gr["Profundidade (m)"],
                    color="blue",
                    linewidth=2,
                    label=label_suavizada
                )

        prof = pd.to_numeric(
            df_pp["Profundidade (m)"],
            errors="coerce"
        )

        for idx, lbf in enumerate(lbfs):
            try:
                cor_lbf = TRENDING_COLORS[idx % len(TRENDING_COLORS)]

                prof_ref = (
                    lbf["prof_ini"]
                    if lbf["prof_ini"] is not None
                    else prof.min()
                )

                lbf_line = lbf["inclbf"] * (prof - prof_ref) + lbf["lbf"]

                mask_base = prof.notna()

                if (
                    st.session_state.get("tipo_poco") == "Onshore"
                    and int(st.session_state.get("n_lbf", 1)) == 1
                    and int(st.session_state.get("n_trending", 1)) == 1
                ):
                    if (
                        st.session_state.get("s_gr") == "Sim"
                        and "Raio Gama Suavizado" in df_pp.columns
                    ):
                        coluna_gr_ref = "Raio Gama Suavizado"
                    else:
                        coluna_gr_ref = "Perfil Raio Gama"

                    if coluna_gr_ref in df_pp.columns:
                        curva_gr_ref = pd.to_numeric(
                            df_pp[coluna_gr_ref],
                            errors="coerce"
                        )

                        prof_validas_gr = prof[
                            prof.notna()
                            & curva_gr_ref.notna()
                        ]

                        if not prof_validas_gr.empty:
                            mask_base = mask_base & (
                                prof >= prof_validas_gr.min()
                            )

                if st.session_state.get("tipo_poco") == "Offshore":
                    mask_base = mask_base & (df_pp.index >= 1)

                prof_ini = lbf.get("prof_ini")
                prof_fim = lbf.get("prof_fim")

                if (
                    prof_ini is not None
                    and prof_fim is not None
                    and prof_fim > prof_ini
                ):
                    mask_intervalo = (prof >= prof_ini) & (prof <= prof_fim)
                    mask_final = mask_base & mask_intervalo

                else:
                    mask_final = mask_base

                if mask_final.any():
                    ax.plot(
                        lbf_line[mask_final],
                        prof[mask_final],
                        color=cor_lbf,
                        linestyle="--",
                        linewidth=3,
                        label=f"LBF {idx + 1}"
                    )

            except Exception as e:
                st.warning(f"Erro ao plotar LBF {idx + 1}: {e}")

        valores_x = []

        if coluna_base in df_plot.columns:
            valores_x.append(df_plot[coluna_base])

        if (
            st.session_state.get("s_gr") == "Sim"
            and coluna_suavizada in df_plot.columns
        ):
            valores_x.append(df_plot[coluna_suavizada])

        for linha in ax.get_lines():
            x_data = pd.to_numeric(
                pd.Series(linha.get_xdata()),
                errors="coerce"
            )

            if not x_data.empty:
                valores_x.append(x_data)

        if valores_x:
            serie_x = pd.concat(valores_x, ignore_index=True).dropna()
            serie_x = serie_x[np.isfinite(serie_x)]

            if not serie_x.empty:
                x_min_plot = float(serie_x.min())
                x_max_plot = float(serie_x.max())

                if x_max_plot <= x_min_plot:
                    margem_x = max(abs(x_min_plot) * 0.05, 1.0)
                else:
                    margem_x = (x_max_plot - x_min_plot) * 0.05

                ax.set_xlim(x_min_plot - margem_x, x_max_plot + margem_x)

        ax.grid(True, linestyle="--", alpha=0.5)

    ax.set_title(
        f"{label_base} x Profundidade (m)",
        fontsize=14,
        fontweight="bold"
    )

    ax.set_xlabel(xlabel_plot, fontsize=12)
    ax.set_ylabel("Profundidade TVD (m)", fontsize=12)

    ax.invert_yaxis()
    ax.tick_params(axis="y", which="both", left=True, labelleft=True)

    ax.set_yticks(range(y_min_pp, y_max_pp + y_step_pp, y_step_pp))
    ax.set_ylim(y_max_pp, y_min_pp)

    ax.legend(
        loc="upper right",
        fontsize=8,
        frameon=True,
        shadow=True,
        fancybox=True,
        framealpha=1,
        facecolor="white",
        edgecolor="gray"
    )

    add_watermark(
        ax,
        logo_path="logo2.png",
        xy=(0.50, 0.5),
        zoom=0.2,
        alpha=0.2,
        zorder=0
    )

    if modo_grafico == "Trending":
        st.session_state.fig_trending_poros = fig
        st.session_state.fig1 = fig

    else:
        st.session_state.fig_lbf_poros = fig
        st.session_state.fig2 = fig

    st.pyplot(fig)


def _df_pontos_lot_padrao():
    return pd.DataFrame({
        "Tipo": ["LOT", "LOT"],
        "Profundidade (m)": [0.0, 0.0],
        "Peso Eq. (lb/gal)": [0.0, 0.0],
    })


def _normalizar_pontos_lot_fratura(df_lot):
    if not isinstance(df_lot, pd.DataFrame) or df_lot.empty:
        return _df_pontos_lot_padrao().iloc[0:0].copy()

    df_lot = df_lot.copy()

    for coluna in ["Tipo", "Profundidade (m)", "Peso Eq. (lb/gal)"]:
        if coluna not in df_lot.columns:
            df_lot[coluna] = np.nan

    df_lot = df_lot[["Tipo", "Profundidade (m)", "Peso Eq. (lb/gal)"]]
    df_lot["Tipo"] = df_lot["Tipo"].astype(str).str.upper().str.strip()
    df_lot.loc[~df_lot["Tipo"].isin(["LOT", "FIT"]), "Tipo"] = "LOT"
    df_lot["Profundidade (m)"] = pd.to_numeric(df_lot["Profundidade (m)"], errors="coerce")
    df_lot["Peso Eq. (lb/gal)"] = pd.to_numeric(df_lot["Peso Eq. (lb/gal)"], errors="coerce")

    df_lot = df_lot.dropna(subset=["Profundidade (m)", "Peso Eq. (lb/gal)"])
    df_lot = df_lot[(df_lot["Profundidade (m)"] > 0) & (df_lot["Peso Eq. (lb/gal)"] > 0)]

    return df_lot.sort_values("Profundidade (m)").reset_index(drop=True)


def _serie_interpolada_fratura(df_ref, coluna, profundidades):
    if coluna not in df_ref.columns or "Profundidade (m)" not in df_ref.columns:
        return pd.Series(np.nan, index=profundidades.index)

    df_interp = df_ref[["Profundidade (m)", coluna]].copy()
    df_interp["Profundidade (m)"] = pd.to_numeric(df_interp["Profundidade (m)"], errors="coerce")
    df_interp[coluna] = pd.to_numeric(df_interp[coluna], errors="coerce")
    df_interp = (
        df_interp
        .dropna(subset=["Profundidade (m)", coluna])
        .sort_values("Profundidade (m)")
        .drop_duplicates(subset=["Profundidade (m)"])
    )

    if df_interp.empty:
        return pd.Series(np.nan, index=profundidades.index)

    return pd.Series(
        np.interp(
            profundidades.astype(float),
            df_interp["Profundidade (m)"].astype(float),
            df_interp[coluna].astype(float),
            left=np.nan,
            right=np.nan
        ),
        index=profundidades.index
    )



def _carregar_lots_yaml_fratura():
    try:
        with open("pocos.yaml", "r", encoding="utf-8") as f:
            dados_yaml = yaml.safe_load(f) or {}
    except Exception:
        return pd.DataFrame(columns=["Nome", "Distância (km)", "Profundidade Vertical (m)", "Peso Eq. (lb/gal)"])

    pocos = dados_yaml.get("pocos", [])

    easting_base = st.session_state.get("easting")
    northing_base = st.session_state.get("northing")
    zona_base = st.session_state.get("zona")
    hem_base = st.session_state.get("hem", "Sul")
    raio_km = float(st.session_state.get("raio", 0.1))

    if easting_base is None or northing_base is None or zona_base is None:
        return pd.DataFrame(columns=["Nome", "Distância (km)", "Profundidade Vertical (m)", "Peso Eq. (lb/gal)"])

    def _hemisferio_norte(valor):
        txt = str(valor).strip().lower()
        return txt in ("n", "norte", "north")

    def _haversine(lat1, lon1, lat2, lon2):
        r = 6371000
        phi1 = math.radians(lat1)
        phi2 = math.radians(lat2)
        dphi = math.radians(lat2 - lat1)
        dlambda = math.radians(lon2 - lon1)
        a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        return r * c

    try:
        lat_base, lon_base = utm.to_latlon(
            float(easting_base),
            float(northing_base),
            int(zona_base),
            northern=(str(hem_base).strip().lower() == "norte")
        )
    except Exception:
        return pd.DataFrame(columns=["Nome", "Distância (km)", "Profundidade Vertical (m)", "Peso Eq. (lb/gal)"])

    profundidade_maxima = st.session_state.get("profundidade_maxima", None)
    dados_pontos = []

    for poco in pocos:
        prof_yaml = poco.get("profundidade_vertical_m", None)
        peso_yaml = poco.get("peso_eq_lb_gal", None)

        if prof_yaml is None or peso_yaml is None:
            continue

        if profundidade_maxima is not None and prof_yaml > profundidade_maxima:
            continue

        try:
            e = poco["coordenadas"]["easting"]
            n = poco["coordenadas"]["northing"]
            lat_p, lon_p = utm.to_latlon(
                float(e),
                float(n),
                int(poco["zona_utm"]),
                northern=_hemisferio_norte(poco.get("hem", "Sul"))
            )
        except Exception:
            continue

        dist = _haversine(lat_base, lon_base, lat_p, lon_p)

        if dist > raio_km * 1000:
            continue

        dados_pontos.append({
            "Nome": poco.get("nome", ""),
            "Distância (km)": round(dist / 1000, 2),
            "Profundidade Vertical (m)": prof_yaml,
            "Peso Eq. (lb/gal)": peso_yaml,
        })

    df_dentro_exibir = pd.DataFrame(dados_pontos)

    if df_dentro_exibir.empty:
        return pd.DataFrame(columns=["Nome", "Distância (km)", "Profundidade Vertical (m)", "Peso Eq. (lb/gal)"])

    return df_dentro_exibir.sort_values("Distância (km)").reset_index(drop=True)


def _converter_lots_yaml_para_fratura(df_lots_yaml):
    if not isinstance(df_lots_yaml, pd.DataFrame) or df_lots_yaml.empty:
        return pd.DataFrame(columns=["Tipo", "Profundidade (m)", "Peso Eq. (lb/gal)"])

    df_lot = pd.DataFrame({
        "Tipo": "LOT",
        "Profundidade (m)": pd.to_numeric(
            df_lots_yaml.get("Profundidade Vertical (m)", pd.Series(dtype=float)),
            errors="coerce"
        ),
        "Peso Eq. (lb/gal)": pd.to_numeric(
            df_lots_yaml.get("Peso Eq. (lb/gal)", pd.Series(dtype=float)),
            errors="coerce"
        ),
    })

    return _normalizar_pontos_lot_fratura(df_lot)


def _calibrar_k_fratura(df_pp, df_lot, usar_auxiliar=False):
    df_lot = _normalizar_pontos_lot_fratura(df_lot)

    if df_lot.empty:
        return pd.DataFrame(columns=[
            "Tipo",
            "Profundidade (m)",
            "Peso Eq. (lb/gal)",
            "Gradiente de P. de poros (lb/gal)",
            "P. Poros (psi)",
            "P. Absorção (psi)",
            "P. Sobrecarga (psi)",
            "K",
        ])

    df_ref = df_pp.copy()
    prof_lot = pd.to_numeric(df_lot["Profundidade (m)"], errors="coerce")

    col_gp = (
        "Gradiente de Pressão de Poros Suavizado (lb/gal)"
        if st.session_state.get("spp", "Não") == "Sim"
        and "Gradiente de Pressão de Poros Suavizado (lb/gal)" in df_ref.columns
        else "Gradiente de Pressão de Poros (lb/gal)"
    )

    grad_pp = _serie_interpolada_fratura(df_ref, col_gp, prof_lot)

    if "Pressão de Sobrecarga (psi)" not in df_ref.columns and "Gradiente de Sobrecarga (lb/gal)" in df_ref.columns:
        df_ref["Pressão de Sobrecarga (psi)"] = (
            0.1704
            * pd.to_numeric(df_ref["Gradiente de Sobrecarga (lb/gal)"], errors="coerce")
            * pd.to_numeric(df_ref["Profundidade (m)"], errors="coerce")
        )

    pressao_sob = _serie_interpolada_fratura(df_ref, "Pressão de Sobrecarga (psi)", prof_lot)

    peso_eq_ajustado = pd.to_numeric(df_lot["Peso Eq. (lb/gal)"], errors="coerce").copy()
    peso_eq_ajustado.loc[df_lot["Tipo"] == "FIT"] = peso_eq_ajustado.loc[df_lot["Tipo"] == "FIT"] + 0.5

    gf = pd.DataFrame({
        "Tipo": df_lot["Tipo"].values,
        "Profundidade (m)": prof_lot.values,
        "Peso Eq. (lb/gal)": pd.to_numeric(df_lot["Peso Eq. (lb/gal)"], errors="coerce").values,
        "Gradiente de P. de poros (lb/gal)": grad_pp.values,
        "P. Poros (psi)": 0.1704 * prof_lot.values * grad_pp.values,
        "P. Absorção (psi)": 0.1704 * prof_lot.values * peso_eq_ajustado.values,
        "P. Sobrecarga (psi)": pressao_sob.values,
    })

    gf["K"] = (
        (gf["P. Absorção (psi)"] - gf["P. Poros (psi)"])
        / (gf["P. Sobrecarga (psi)"] - gf["P. Poros (psi)"])
    )

    if usar_auxiliar:
        linha_aux = pd.DataFrame([{
            "Tipo": "AUX",
            "Profundidade (m)": 1.0,
            "Peso Eq. (lb/gal)": np.nan,
            "Gradiente de P. de poros (lb/gal)": np.nan,
            "P. Poros (psi)": np.nan,
            "P. Absorção (psi)": np.nan,
            "P. Sobrecarga (psi)": np.nan,
            "K": 0.01,
        }])
        gf = pd.concat([linha_aux, gf], ignore_index=True)

    return gf.sort_values("Profundidade (m)").reset_index(drop=True)


def _ajustar_tendencia_k_fratura(gf):
    if not isinstance(gf, pd.DataFrame) or gf.empty:
        return None, None

    df_fit = gf[["Profundidade (m)", "K"]].copy()
    df_fit["Profundidade (m)"] = pd.to_numeric(df_fit["Profundidade (m)"], errors="coerce")
    df_fit["K"] = pd.to_numeric(df_fit["K"], errors="coerce")
    df_fit = df_fit.dropna(subset=["Profundidade (m)", "K"])
    df_fit = df_fit[(df_fit["Profundidade (m)"] > 0) & np.isfinite(df_fit["K"])]

    if len(df_fit) < 2:
        return None, None

    try:
        b, log_a = np.polyfit(df_fit["K"], np.log(df_fit["Profundidade (m)"]), 1)
        a = float(np.exp(log_a))
        b = float(b)
    except Exception:
        return None, None

    if not np.isfinite(a) or not np.isfinite(b) or b == 0:
        return None, None

    return a, b


def _calcular_df_f_fratura(df_pp, gf=None):
    df_pp = df_pp.copy()

    col_prof = "Profundidade (m)"
    col_gs = "Gradiente de Sobrecarga (lb/gal)"
    col_gp = (
        "Gradiente de Pressão de Poros Suavizado (lb/gal)"
        if st.session_state.get("spp", "Não") == "Sim"
        and "Gradiente de Pressão de Poros Suavizado (lb/gal)" in df_pp.columns
        else "Gradiente de Pressão de Poros (lb/gal)"
    )

    if any(col not in df_pp.columns for col in [col_prof, col_gs, col_gp]):
        return pd.DataFrame()

    df_f = pd.DataFrame({
        "Profundidade (m)": pd.to_numeric(df_pp[col_prof], errors="coerce"),
        "MD": pd.to_numeric(df_pp["MD"], errors="coerce") if "MD" in df_pp.columns else pd.to_numeric(df_pp[col_prof], errors="coerce"),
        "Gradiente de Sobrecarga (lb/gal)": pd.to_numeric(df_pp[col_gs], errors="coerce"),
        "Gradiente de Pressão de Poros (lb/gal)": pd.to_numeric(df_pp[col_gp], errors="coerce"),
    })

    gf_ref = gf if gf is not None else st.session_state.get("edited_gf", st.session_state.get("gf", pd.DataFrame()))
    a, b = _ajustar_tendencia_k_fratura(gf_ref)

    if a is None or b is None:
        return df_f

    prof = pd.to_numeric(df_f["Profundidade (m)"], errors="coerce")
    df_f["K"] = (np.log(prof) - np.log(a)) / b

    df_f["Gradiente de Fratura (lb/gal)"] = (
        df_f["Gradiente de Pressão de Poros (lb/gal)"]
        + df_f["K"]
        * (
            df_f["Gradiente de Sobrecarga (lb/gal)"]
            - df_f["Gradiente de Pressão de Poros (lb/gal)"]
        )
    )

    df_f["Gradiente de Fratura (lb/gal)"] = pd.to_numeric(
        df_f["Gradiente de Fratura (lb/gal)"],
        errors="coerce"
    ).clip(lower=0)

    df_f["Pressão de Fratura (psi)"] = (
        0.1704
        * df_f["Gradiente de Fratura (lb/gal)"]
        * df_f["Profundidade (m)"]
    )

    st.session_state.a_k = a
    st.session_state.b_k = b

    return df_f


def _plotar_gradiente_fratura(df_f, df_pp):
    modo = st.session_state.get("ogf", "Gradiente (lb/gal)")
    col_prof = "Profundidade (m)"

    plotar_curva_fratura = (
        isinstance(df_f, pd.DataFrame)
        and not df_f.empty
        and col_prof in df_f.columns
        and "Gradiente de Fratura (lb/gal)" in df_f.columns
    )

    if not plotar_curva_fratura:
        df_f = df_pp.copy() if isinstance(df_pp, pd.DataFrame) else pd.DataFrame()

    y_min_f = float(st.session_state.get("y_min_f", 0.0))
    y_max_f = float(st.session_state.get("y_max_f", 1000.0))
    y_step_f = max(0.01, float(st.session_state.get("y_step_f", 200.0)))
    x_min_f = float(st.session_state.get("x_min_f", 7.0))
    x_max_f = float(st.session_state.get("x_max_f", 21.0))
    x_step_f = max(0.01, float(st.session_state.get("x_step_f", 2.0)))

    if y_max_f <= y_min_f:
        y_max_f = y_min_f + y_step_f
    if x_max_f <= x_min_f:
        x_max_f = x_min_f + x_step_f

    usar_coluna_idade = (
        st.session_state.get("idg") == "Sim"
        and "df_idade" in st.session_state
        and isinstance(st.session_state.df_idade, pd.DataFrame)
        and not st.session_state.df_idade.empty
    )

    fig = plt.figure(figsize=(8, 10))

    if usar_coluna_idade:
        gs = gridspec.GridSpec(1, 4, width_ratios=[0.10, 0.18, 0.21, 1], wspace=0)
        ax_idade = fig.add_subplot(gs[0])
        ax1 = fig.add_subplot(gs[1], sharey=ax_idade)
        ax_gap = fig.add_subplot(gs[2])
        ax_gap.axis("off")
        ax = fig.add_subplot(gs[3], sharey=ax_idade)
        idade_formacao(ax_idade, st.session_state.df_idade, y_max_f, y_min_f)
        ax_idade.tick_params(axis="y", which="both", left=False, right=False, labelleft=False, labelright=False)
        ax_idade.set_ylabel("")
        plt.setp(ax1.get_yticklabels(), visible=False)
        plt.setp(ax.get_yticklabels(), visible=False)
    else:
        gs = gridspec.GridSpec(1, 3, width_ratios=[0.18, 0.21, 1], wspace=0)
        ax1 = fig.add_subplot(gs[0])
        ax_gap = fig.add_subplot(gs[1])
        ax_gap.axis("off")
        ax = fig.add_subplot(gs[2], sharey=ax1)
        plt.setp(ax.get_yticklabels(), visible=False)

    selected = st.session_state.get("well_selected", st.session_state.get("poco", "Poço"))
    _garantir_litologia_importada(selected)
    poco = st.session_state.get("pocos", {}).get(selected, {})

    df_lito_f = df_pp if isinstance(df_pp, pd.DataFrame) and not df_pp.empty else df_f

    if (
        st.session_state.get(
            "tipo_coluna_litologica_graficos",
            "Permeável / Não permeável"
        ) == "Permeável / Não permeável"
        and "df_pp_lito" in st.session_state
        and isinstance(st.session_state.df_pp_lito, pd.DataFrame)
        and not st.session_state.df_pp_lito.empty
    ):
        df_lito_f = st.session_state.df_pp_lito.copy()

    if "Profundidade (m)" not in df_lito_f.columns and "Profundidade" in df_lito_f.columns:
        df_lito_f["Profundidade (m)"] = df_lito_f["Profundidade"]

    lito(
        ax1,
        df_lito_f,
        poco.get("profundidade", []),
        poco.get("litologia", []),
        y_max_f,
        y_min=y_min_f,
        y_max=y_max_f
    )

    if modo == "Pressão (psi)":
        col_fratura = "Pressão de Fratura (psi)"
        xlabel = "Pressão (psi)"
        titulo = "Pressão de Fratura (psi)"
        if plotar_curva_fratura and col_fratura not in df_f.columns:
            df_f[col_fratura] = 0.1704 * df_f["Gradiente de Fratura (lb/gal)"] * df_f[col_prof]
    else:
        col_fratura = "Gradiente de Fratura (lb/gal)"
        xlabel = "Gradiente (ppg)"
        titulo = "Gradiente de Fratura (lb/gal)"

    if plotar_curva_fratura:
        df_plot = df_f[[col_prof, col_fratura]].copy()
        df_plot[col_prof] = pd.to_numeric(df_plot[col_prof], errors="coerce")
        df_plot[col_fratura] = pd.to_numeric(df_plot[col_fratura], errors="coerce")
        df_plot = df_plot.dropna(subset=[col_prof, col_fratura])
    else:
        df_plot = pd.DataFrame(columns=[col_prof, col_fratura])

    if plotar_curva_fratura and not df_plot.empty:
        ax.plot(
            df_plot[col_fratura],
            df_plot[col_prof],
            color="brown",
            linestyle="-",
            linewidth=2,
            label="Gradiente de Fratura" if modo == "Gradiente (lb/gal)" else "Pressão de Fratura"
        )

    if plotar_curva_fratura and st.session_state.get("grap", "Sim") == "Sim" and "Gradiente de Pressão de Poros (lb/gal)" in df_f.columns:
        if modo == "Pressão (psi)":
            x_pp = 0.1704 * pd.to_numeric(df_f["Gradiente de Pressão de Poros (lb/gal)"], errors="coerce") * pd.to_numeric(df_f[col_prof], errors="coerce")
            label_pp = "Pressão de Poros"
        else:
            x_pp = pd.to_numeric(df_f["Gradiente de Pressão de Poros (lb/gal)"], errors="coerce")
            label_pp = "Gradiente de Pressão de Poros"

        ax.plot(x_pp, df_f[col_prof], color="orange", linestyle="-", linewidth=2, label=label_pp)

    if plotar_curva_fratura and st.session_state.get("gras", "Sim") == "Sim" and "Gradiente de Sobrecarga (lb/gal)" in df_f.columns:
        if modo == "Pressão (psi)":
            x_gs = 0.1704 * pd.to_numeric(df_f["Gradiente de Sobrecarga (lb/gal)"], errors="coerce") * pd.to_numeric(df_f[col_prof], errors="coerce")
            label_gs = "Pressão de Sobrecarga"
        else:
            x_gs = pd.to_numeric(df_f["Gradiente de Sobrecarga (lb/gal)"], errors="coerce")
            label_gs = "Gradiente de Sobrecarga"

        ax.plot(x_gs, df_f[col_prof], color="black", linestyle="-", linewidth=2, label=label_gs)

    if plotar_curva_fratura and st.session_state.get("janela_fratura", "Sim") == "Sim" and "Gradiente de Pressão de Poros (lb/gal)" in df_f.columns:
        if modo == "Pressão (psi)":
            x1 = 0.1704 * pd.to_numeric(df_f["Gradiente de Pressão de Poros (lb/gal)"], errors="coerce") * pd.to_numeric(df_f[col_prof], errors="coerce")
            x2 = pd.to_numeric(df_f.get("Pressão de Fratura (psi)"), errors="coerce")
        else:
            x1 = pd.to_numeric(df_f["Gradiente de Pressão de Poros (lb/gal)"], errors="coerce")
            x2 = pd.to_numeric(df_f["Gradiente de Fratura (lb/gal)"], errors="coerce")

        y = pd.to_numeric(df_f[col_prof], errors="coerce")
        ax.fill_betweenx(y, x1, x2, where=(x2 > x1), color="lightgreen", alpha=0.2, label="Janela Operacional", interpolate=True)

    df_lot = _normalizar_pontos_lot_fratura(
        st.session_state.get("lot_pontos_fratura", pd.DataFrame())
        if st.session_state.get("lot_fratura", "Sim") == "Sim"
        else st.session_state.get("lot_pontos_yaml_fratura", pd.DataFrame())
    )
    if plotar_curva_fratura and not df_lot.empty:
        for tipo, cor, marcador, label in [("LOT", "red", "D", "LOT's"), ("FIT", "blue", "^", "FIT's")]:
            df_tipo = df_lot[df_lot["Tipo"] == tipo]
            if df_tipo.empty:
                continue

            if modo == "Pressão (psi)":
                x_ponto = df_tipo["Peso Eq. (lb/gal)"] * 0.1704 * df_tipo["Profundidade (m)"]
            else:
                x_ponto = df_tipo["Peso Eq. (lb/gal)"]

            ax.scatter(
                x_ponto,
                df_tipo["Profundidade (m)"],
                color=cor,
                edgecolors="black",
                linewidths=0.8,
                label=label,
                zorder=5,
                marker=marcador,
                s=50
            )

            chave_texto = "mostrar_texto_lot_tabs4" if tipo == "LOT" else "mostrar_texto_fit_tabs4"
            if st.session_state.get(chave_texto, "Sim") == "Sim":
                for x_item, y_item, peso_item in zip(x_ponto, df_tipo["Profundidade (m)"], df_tipo["Peso Eq. (lb/gal)"]):
                    ax.annotate(
                        f"{peso_item:.2f} ppg",
                        xy=(x_item, y_item),
                        xytext=(6, 0),
                        textcoords="offset points",
                        fontsize=8,
                        color="black",
                        va="center",
                        ha="left",
                        zorder=6
                    )

    if plotar_curva_fratura:
        plotar_rft(ax, modo_grafico=modo)
        plotar_gradiente_colapso(ax, modo_grafico=modo)

    ax.set_title(titulo, fontsize=14, fontweight="bold")
    ax.set_xlabel(xlabel, fontsize=12)
    ax.set_ylabel("Profundidade TVD (m)", fontsize=12)
    ax.invert_yaxis()
    ax.tick_params(axis="y", which="both", left=True, labelleft=True)
    ax.set_yticks(np.arange(y_min_f, y_max_f + y_step_f * 0.5, y_step_f))
    ax.set_ylim(y_max_f, y_min_f)
    ax.set_xlim(x_min_f, x_max_f)

    x_tick_ini = math.ceil(x_min_f)
    x_tick_fim = math.floor(x_max_f)
    x_tick_step = max(1.0, x_step_f)
    if x_tick_fim >= x_tick_ini:
        ax.set_xticks(np.arange(x_tick_ini, x_tick_fim + x_tick_step * 0.5, x_tick_step))
    if np.isclose(x_step_f, 1.0):
        ax.set_xticks(np.arange(x_min_f, x_max_f + 0.25, 0.5), minor=True)

    ax.xaxis.set_major_formatter(FuncFormatter(lambda valor, pos: f"{valor:.0f}"))
    ax.grid(True, which="major", linestyle="--", alpha=0.5)
    if np.isclose(x_step_f, 1.0):
        ax.grid(True, which="minor", axis="x", linestyle="--", alpha=0.5)

    if ax.get_legend_handles_labels()[0]:
        ax.legend(loc="upper right", fontsize=8, frameon=True, shadow=True, fancybox=True, framealpha=1, facecolor="white", edgecolor="gray")

    add_watermark(ax, logo_path="logo2.png", xy=(0.50, 0.5), zoom=0.2, alpha=0.2, zorder=0)

    st.session_state.fig_fratura = fig
    st.pyplot(fig)


def _plotar_k_fratura(gf):
    if isinstance(gf, pd.DataFrame) and not gf.empty and "K" in gf.columns and "Profundidade (m)" in gf.columns:
        gf_plot = gf[["Profundidade (m)", "K"]].copy()
        gf_plot["Profundidade (m)"] = pd.to_numeric(gf_plot["Profundidade (m)"], errors="coerce")
        gf_plot["K"] = pd.to_numeric(gf_plot["K"], errors="coerce")
        gf_plot = gf_plot.dropna(subset=["Profundidade (m)", "K"])
    else:
        gf_plot = pd.DataFrame(columns=["Profundidade (m)", "K"])

    fig, ax = plt.subplots(figsize=(8, 10))

    if not gf_plot.empty:
        ax.plot(gf_plot["K"], gf_plot["Profundidade (m)"], color="blue", linestyle="None", marker="o", markersize=8, markerfacecolor="black", markeredgecolor="red", label="K")

    a, b = _ajustar_tendencia_k_fratura(gf_plot)
    if not gf_plot.empty and a is not None and b is not None and gf_plot["K"].max() > gf_plot["K"].min():
        k_values = np.linspace(gf_plot["K"].min(), gf_plot["K"].max(), 200)
        depth_trend = a * np.exp(b * k_values)
        ax.plot(k_values, depth_trend, color="red", linestyle="--", linewidth=2, label="Tendência exponencial de K")

    y_min_f = float(st.session_state.get("y_min_f", 0.0))
    y_max_padrao = gf_plot["Profundidade (m)"].max() + 100 if not gf_plot.empty else 1000.0
    y_max_f = float(st.session_state.get("y_max_f", y_max_padrao))
    y_step_f = max(0.01, float(st.session_state.get("y_step_f", 200.0)))

    ax.set_title("K x Profundidade", fontsize=14, fontweight="bold")
    ax.set_xlabel("K", fontsize=12)
    ax.set_ylabel("Profundidade TVD (m)", fontsize=12)
    ax.invert_yaxis()
    ax.set_yticks(np.arange(y_min_f, y_max_f + y_step_f * 0.5, y_step_f))
    ax.set_ylim(y_max_f, y_min_f)
    ax.grid(True, linestyle="--", alpha=0.5)
    if ax.get_legend_handles_labels()[0]:
        ax.legend(loc="upper right", fontsize=8, frameon=True, shadow=True, fancybox=True, framealpha=1, facecolor="white", edgecolor="gray")
    add_watermark(ax, logo_path="logo2.png", xy=(0.50, 0.5), zoom=0.2, alpha=0.2, zorder=0)

    st.session_state.fig_k_prof = fig
    st.pyplot(fig)


def _coluna_existente(df, opcoes):
    if not isinstance(df, pd.DataFrame):
        return None

    for coluna in opcoes:
        if coluna in df.columns:
            return coluna

    return None


def _interpolar_coluna_por_profundidade(df_ref, col_prof_ref, col_valor, profundidades_destino):
    if (
        not isinstance(df_ref, pd.DataFrame)
        or df_ref.empty
        or col_prof_ref not in df_ref.columns
        or col_valor not in df_ref.columns
    ):
        return pd.Series(np.nan, index=profundidades_destino.index)

    df_interp = df_ref[[col_prof_ref, col_valor]].copy()
    df_interp[col_prof_ref] = pd.to_numeric(df_interp[col_prof_ref], errors="coerce")
    df_interp[col_valor] = pd.to_numeric(df_interp[col_valor], errors="coerce")
    df_interp = (
        df_interp
        .dropna(subset=[col_prof_ref, col_valor])
        .sort_values(col_prof_ref)
        .drop_duplicates(subset=[col_prof_ref])
    )

    if df_interp.empty:
        return pd.Series(np.nan, index=profundidades_destino.index)

    return pd.Series(
        np.interp(
            pd.to_numeric(profundidades_destino, errors="coerce").astype(float),
            df_interp[col_prof_ref].astype(float),
            df_interp[col_valor].astype(float),
            left=np.nan,
            right=np.nan
        ),
        index=profundidades_destino.index
    )


def _montar_df_estabilidade(df_pp):
    if not isinstance(df_pp, pd.DataFrame) or df_pp.empty:
        return pd.DataFrame()

    if "Profundidade (m)" not in df_pp.columns:
        return pd.DataFrame()

    prof = pd.to_numeric(df_pp["Profundidade (m)"], errors="coerce")

    df_est = pd.DataFrame({
        "Profundidade (m)": prof,
        "MD": (
            pd.to_numeric(df_pp["MD"], errors="coerce")
            if "MD" in df_pp.columns
            else pd.Series(np.nan, index=df_pp.index)
        ),
        "Perfil de densidade (g/cm³)": (
            pd.to_numeric(df_pp["Perfil de densidade (g/cm³)"], errors="coerce")
            if "Perfil de densidade (g/cm³)" in df_pp.columns
            else pd.Series(np.nan, index=df_pp.index)
        ),
        "Perfil sônico (µs/pé)": (
            pd.to_numeric(df_pp["Perfil sônico (µs/pé)"], errors="coerce")
            if "Perfil sônico (µs/pé)" in df_pp.columns
            else pd.Series(np.nan, index=df_pp.index)
        ),
        "Gradiente de Sobrecarga (lb/gal)": (
            pd.to_numeric(df_pp["Gradiente de Sobrecarga (lb/gal)"], errors="coerce")
            if "Gradiente de Sobrecarga (lb/gal)" in df_pp.columns
            else pd.Series(np.nan, index=df_pp.index)
        ),
        "Gradiente de Pressão de Poros (lb/gal)": (
            pd.to_numeric(df_pp["Gradiente de Pressão de Poros (lb/gal)"], errors="coerce")
            if "Gradiente de Pressão de Poros (lb/gal)" in df_pp.columns
            else pd.Series(np.nan, index=df_pp.index)
        ),
    })

    df_f = st.session_state.get("df_f", pd.DataFrame())
    if (
        isinstance(df_f, pd.DataFrame)
        and not df_f.empty
        and "Gradiente de Fratura (lb/gal)" in df_f.columns
    ):
        df_est["Gradiente de Fratura (lb/gal)"] = _interpolar_coluna_por_profundidade(
            df_f,
            "Profundidade (m)",
            "Gradiente de Fratura (lb/gal)",
            prof
        )
    else:
        df_est["Gradiente de Fratura (lb/gal)"] = np.nan

    fontes_trajetoria = [
        st.session_state.get("df_interp", pd.DataFrame()),
        st.session_state.get("df1", pd.DataFrame()),
        st.session_state.get("df_out_traj", pd.DataFrame()),
        st.session_state.get("df2", pd.DataFrame()),
    ]

    for df_traj in fontes_trajetoria:
        if not isinstance(df_traj, pd.DataFrame) or df_traj.empty:
            continue

        col_prof_traj = _coluna_existente(
            df_traj,
            ["Profundidade (m)", "Profundidade", "TVD"]
        )
        col_md = _coluna_existente(df_traj, ["MD"])
        col_incl = _coluna_existente(
            df_traj,
            ["Incl", "Inclinação (°)", "Inc (°)", "Inc"]
        )
        col_azi = _coluna_existente(
            df_traj,
            ["Azi", "Azi (°)", "Azimute (°)", "Azimute"]
        )

        if col_prof_traj is None:
            continue

        if df_est["MD"].isna().all() and col_md is not None:
            df_est["MD"] = _interpolar_coluna_por_profundidade(
                df_traj,
                col_prof_traj,
                col_md,
                prof
            )

        if "Incl (°)" not in df_est.columns and col_incl is not None:
            df_est["Incl (°)"] = _interpolar_coluna_por_profundidade(
                df_traj,
                col_prof_traj,
                col_incl,
                prof
            )

        if "Azi (°)" not in df_est.columns and col_azi is not None:
            df_est["Azi (°)"] = _interpolar_coluna_por_profundidade(
                df_traj,
                col_prof_traj,
                col_azi,
                prof
            )

        if (
            not df_est["MD"].isna().all()
            and "Incl (°)" in df_est.columns
            and "Azi (°)" in df_est.columns
        ):
            break

    if "Incl (°)" not in df_est.columns:
        df_est["Incl (°)"] = np.nan

    if "Azi (°)" not in df_est.columns:
        df_est["Azi (°)"] = np.nan

    df_est = df_est[
        [
            "Profundidade (m)",
            "MD",
            "Perfil de densidade (g/cm³)",
            "Perfil sônico (µs/pé)",
            "Gradiente de Sobrecarga (lb/gal)",
            "Gradiente de Pressão de Poros (lb/gal)",
            "Gradiente de Fratura (lb/gal)",
            "Incl (°)",
            "Azi (°)",
        ]
    ]

    return df_est.dropna(subset=["Profundidade (m)"]).reset_index(drop=True)


def _df_direcoes_tensoes_padrao():
    return pd.DataFrame({
        "Profundidade (m)": [],
        "Direção SH": [],
    })


def _df_relacao_tensoes_padrao():
    return pd.DataFrame({
        "Profundidade (m)": [],
        "SH% Sobrecarga": [],
        "Sh% Sobrecarga": [],
    })


def _normalizar_direcoes_tensoes(df_direcoes):
    if not isinstance(df_direcoes, pd.DataFrame) or df_direcoes.empty:
        return _df_direcoes_tensoes_padrao()

    df_direcoes = df_direcoes.copy()

    for coluna in ["Profundidade (m)", "Direção SH"]:
        if coluna not in df_direcoes.columns:
            df_direcoes[coluna] = np.nan

    df_direcoes = df_direcoes[["Profundidade (m)", "Direção SH"]]
    df_direcoes["Profundidade (m)"] = pd.to_numeric(df_direcoes["Profundidade (m)"], errors="coerce")
    df_direcoes["Direção SH"] = pd.to_numeric(df_direcoes["Direção SH"], errors="coerce")

    return (
        df_direcoes
        .dropna(subset=["Profundidade (m)", "Direção SH"])
        .sort_values("Profundidade (m)")
        .drop_duplicates(subset=["Profundidade (m)"], keep="last")
        .reset_index(drop=True)
    )


def _normalizar_relacao_tensoes(df_relacao):
    if not isinstance(df_relacao, pd.DataFrame) or df_relacao.empty:
        return _df_relacao_tensoes_padrao()

    df_relacao = df_relacao.copy()

    for coluna in ["Profundidade (m)", "SH% Sobrecarga", "Sh% Sobrecarga"]:
        if coluna not in df_relacao.columns:
            df_relacao[coluna] = np.nan

    df_relacao = df_relacao[["Profundidade (m)", "SH% Sobrecarga", "Sh% Sobrecarga"]]
    df_relacao["Profundidade (m)"] = pd.to_numeric(df_relacao["Profundidade (m)"], errors="coerce")
    df_relacao["SH% Sobrecarga"] = pd.to_numeric(df_relacao["SH% Sobrecarga"], errors="coerce")
    df_relacao["Sh% Sobrecarga"] = pd.to_numeric(df_relacao["Sh% Sobrecarga"], errors="coerce")

    return (
        df_relacao
        .dropna(subset=["Profundidade (m)", "SH% Sobrecarga", "Sh% Sobrecarga"])
        .sort_values("Profundidade (m)")
        .drop_duplicates(subset=["Profundidade (m)"], keep="last")
        .reset_index(drop=True)
    )


def _interpolar_direcao_sh(df_direcoes, profundidades):
    df_direcoes = _normalizar_direcoes_tensoes(df_direcoes)
    prof = pd.to_numeric(profundidades, errors="coerce").to_numpy(dtype=float)

    if df_direcoes.empty:
        return pd.Series(np.zeros(len(prof)), index=profundidades.index)

    if len(df_direcoes) == 1:
        valor = float(df_direcoes["Direção SH"].iloc[0]) % 360
        return pd.Series(np.full(len(prof), valor), index=profundidades.index)

    x = df_direcoes["Profundidade (m)"].to_numpy(dtype=float)
    ang = np.radians(df_direcoes["Direção SH"].to_numpy(dtype=float) % 360)
    sin_i = np.interp(prof, x, np.sin(ang))
    cos_i = np.interp(prof, x, np.cos(ang))

    return pd.Series(
        np.degrees(np.arctan2(sin_i, cos_i)) % 360,
        index=profundidades.index
    )


def _interpolar_relacao_tensoes(df_relacao, profundidades):
    df_relacao = _normalizar_relacao_tensoes(df_relacao)
    prof = pd.to_numeric(profundidades, errors="coerce").to_numpy(dtype=float)

    if df_relacao.empty:
        return pd.DataFrame({
            "SH% Sobrecarga": np.full(len(prof), 0.70),
            "Sh% Sobrecarga": np.full(len(prof), 0.68),
        }, index=profundidades.index)

    if len(df_relacao) == 1:
        return pd.DataFrame({
            "SH% Sobrecarga": np.full(len(prof), float(df_relacao["SH% Sobrecarga"].iloc[0])),
            "Sh% Sobrecarga": np.full(len(prof), float(df_relacao["Sh% Sobrecarga"].iloc[0])),
        }, index=profundidades.index)

    x = df_relacao["Profundidade (m)"].to_numpy(dtype=float)

    return pd.DataFrame({
        "SH% Sobrecarga": np.interp(
            prof,
            x,
            df_relacao["SH% Sobrecarga"].to_numpy(dtype=float)
        ),
        "Sh% Sobrecarga": np.interp(
            prof,
            x,
            df_relacao["Sh% Sobrecarga"].to_numpy(dtype=float)
        ),
    }, index=profundidades.index)


def _aplicar_parametros_mecanicos_df_est(df_est):
    if not isinstance(df_est, pd.DataFrame) or df_est.empty:
        return df_est

    df_est = df_est.copy()

    col_sonico = "Perfil sônico (µs/pé)"
    col_dens = "Perfil de densidade (g/cm³)"

    sonico = (
        pd.to_numeric(df_est[col_sonico], errors="coerce")
        if col_sonico in df_est.columns
        else pd.Series(np.nan, index=df_est.index)
    )
    densidade = (
        pd.to_numeric(df_est[col_dens], errors="coerce")
        if col_dens in df_est.columns
        else pd.Series(np.nan, index=df_est.index)
    )

    if st.session_state.get("lft", "Calculado") == "Constante":
        if st.session_state.get("tipo_poco") == "Offshore":
            df_est["Φ (°)"] = np.nan
            df_est.loc[df_est.index[1:], "Φ (°)"] = float(st.session_state.get("phi_constante", 30.0))

        else:
            df_est["Φ (°)"] = float(st.session_state.get("phi_constante", 30.0))
    else:
        with np.errstate(divide="ignore", invalid="ignore"):
            vp = 304800 / sonico
            razao_phi = (vp - 1000) / (vp + 1000)
            df_est["Φ (°)"] = np.round(
                np.degrees(np.arcsin(np.clip(razao_phi, -1, 1))),
                2
            )

    with np.errstate(divide="ignore", invalid="ignore"):
        dts = ((1 / (((0.8042 * (((1000000 / sonico) / 3.281) / 1000)) - 0.8559) * 1000)) * 1000000) / 3.281
        df_est["DTS"] = np.round(dts, 2)

        razao_dts_sonico = df_est["DTS"] / sonico
        poisson = (
            (0.5 * razao_dts_sonico ** 2 - 1)
            / (razao_dts_sonico ** 2 - 1)
        )
        df_est["Poisson"] = np.round(poisson, 2)

    if st.session_state.get("ucs", "Mechpro") == "Lacy":
        with np.errstate(divide="ignore", invalid="ignore"):
            df_est["G dinam (MMpsi)"] = np.round(
                (1.34 * 10 ** 10 * densidade / (df_est["DTS"] ** 2)) / 10 ** 6,
                2
            )
            df_est["E dinâmico (MMpsi)"] = np.round(
                2 * df_est["G dinam (MMpsi)"] * (1 + df_est["Poisson"]),
                2
            )
            df_est["E estático (MMpsi)"] = np.round(
                0.018 * (df_est["E dinâmico (MMpsi)"] ** 2)
                + 0.422 * df_est["E dinâmico (MMpsi)"],
                2
            )
            df_est["UCS (psi)"] = np.round(
                (
                    0.2787 * df_est["E estático (MMpsi)"] ** 2
                    + 2.458 * df_est["E estático (MMpsi)"]
                ) * 1000,
                2
            )

    else:
        df_est["Vsh"] = 0.5
        with np.errstate(divide="ignore", invalid="ignore"):
            df_est["UCS (psi)"] = np.round(
                145.0377
                * 1.9e-20
                * (1000 * densidade) ** 2
                * (304800 / sonico) ** 4
                * ((1 + df_est["Poisson"]) / (1 - df_est["Poisson"])) ** 2
                * (1 - 2 * df_est["Poisson"])
                * (1 + 0.79 * df_est["Vsh"]),
                2
            )

    with np.errstate(divide="ignore", invalid="ignore"):
        phi_rad = np.radians(pd.to_numeric(df_est["Φ (°)"], errors="coerce"))
        df_est["So (psi)"] = np.round(
            (
                pd.to_numeric(df_est["UCS (psi)"], errors="coerce")
                * (1 - np.sin(phi_rad))
            )
            / (2 * np.cos(phi_rad)),
            2
        )

    df_est["Sv (psi)"] = np.round(
        pd.to_numeric(df_est["Gradiente de Sobrecarga (lb/gal)"], errors="coerce")
        * 0.1704
        * pd.to_numeric(df_est["Profundidade (m)"], errors="coerce"),
        2
    )

    if st.session_state.get("usar_relacao_tensoes", False):
        relacao_tensoes = _interpolar_relacao_tensoes(
            st.session_state.get("relacao_tensoes_df", pd.DataFrame()),
            df_est["Profundidade (m)"]
        )
    else:
        relacao_tensoes = pd.DataFrame({
            "SH% Sobrecarga": np.full(len(df_est), 0.70),
            "Sh% Sobrecarga": np.full(len(df_est), 0.68),
        }, index=df_est.index)

    if st.session_state.get("usar_direcoes_tensoes", False):
        direcao_sh = _interpolar_direcao_sh(
            st.session_state.get("direcoes_tensoes_df", pd.DataFrame()),
            df_est["Profundidade (m)"]
        )
    else:
        direcao_sh = pd.Series(np.zeros(len(df_est)), index=df_est.index)

    df_est["SH (psi)"] = np.round(
        pd.to_numeric(relacao_tensoes["SH% Sobrecarga"], errors="coerce")
        * df_est["Sv (psi)"],
        2
    )
    df_est["Direção de SH (°)"] = np.round(direcao_sh % 360, 2)
    df_est["Sh (psi)"] = np.round(
        pd.to_numeric(relacao_tensoes["Sh% Sobrecarga"], errors="coerce")
        * df_est["Sv (psi)"],
        2
    )
    df_est["Direção de Sh (°)"] = np.round((df_est["Direção de SH (°)"] + 90) % 360, 2)

    lxxl = np.cos(np.radians(df_est['Azi (°)'] - df_est['Direção de SH (°)'])) * np.cos(np.radians(df_est['Incl (°)']))
    lyxl = -np.sin(np.radians(df_est['Azi (°)'] - df_est['Direção de SH (°)']))
    lzxl = np.cos(np.radians(df_est['Azi (°)'] - df_est['Direção de SH (°)'])) * np.sin(np.radians(df_est['Incl (°)']))

    lxyl = np.sin(np.radians(df_est['Azi (°)'] - df_est['Direção de SH (°)'])) * np.cos(np.radians(df_est['Incl (°)']))
    lyyl = np.cos(np.radians(df_est['Azi (°)'] - df_est['Direção de SH (°)']))
    lzyl = np.sin(np.radians(df_est['Azi (°)'] - df_est['Direção de SH (°)'])) * np.sin(np.radians(df_est['Incl (°)']))

    lxzl = -np.sin(np.radians(df_est['Incl (°)']))
    lyzl = 0
    lzzl = np.cos(np.radians(df_est['Incl (°)']))

    df_est.insert(
        loc=df_est.columns.get_loc('Sv (psi)') + 1,
        column='lxxl',
        value=round(lxxl, 2)
    )
    df_est.insert(
        loc=df_est.columns.get_loc('lxxl') + 1,
        column='lyxl',
        value=round(lyxl, 2)
    )
    df_est.insert(
        loc=df_est.columns.get_loc('lyxl') + 1,
        column='lzxl',
        value=round(lzxl, 2)
    )
    df_est.insert(
        loc=df_est.columns.get_loc('lzxl') + 1,
        column='lxyl',
        value=round(lxyl, 2)
    )
    df_est.insert(
        loc=df_est.columns.get_loc('lxyl') + 1,
        column='lyyl',
        value=round(lyyl, 2)
    )
    df_est.insert(
        loc=df_est.columns.get_loc('lyyl') + 1,
        column='lzyl',
        value=round(lzyl, 2)
    )
    df_est.insert(
        loc=df_est.columns.get_loc('lzyl') + 1,
        column='lxzl',
        value=round(lxzl, 2)
    )
    df_est.insert(
        loc=df_est.columns.get_loc('lxzl') + 1,
        column='lyzl',
        value=round(lyzl, 2)
    )
    df_est.insert(
        loc=df_est.columns.get_loc('lyzl') + 1,
        column='lzzl',
        value=round(lzzl, 2)
    )

    df_est.insert(
        loc=df_est.columns.get_loc('lzzl') + 1,
        column='τxy',
        value=round((lxxl * lyxl * df_est['SH (psi)']) + (lxyl * lyyl * df_est['Sh (psi)']) + (
                    lxzl * lyzl * df_est['Sv (psi)']), 2)
    )

    df_est.insert(
        loc=df_est.columns.get_loc('τxy') + 1,
        column='τyz',
        value=round((lyxl * lzxl * df_est['SH (psi)']) + (lyyl * lzyl * df_est['Sh (psi)']) + (
                    lyzl * lzzl * df_est['Sv (psi)']), 2)
    )

    df_est.insert(
        loc=df_est.columns.get_loc('τyz') + 1,
        column='τzx',
        value=round((lzxl * lxxl * df_est['SH (psi)']) + (lzyl * lxyl * df_est['Sh (psi)']) + (
                    lzzl * lxzl * df_est['Sv (psi)']), 2)
    )

    # === Direções θA e θB a partir dos cisalhamentos τyz e τzx ===
    tau_yz = pd.to_numeric(df_est["τyz"], errors="coerce")
    tau_zx = pd.to_numeric(df_est["τzx"], errors="coerce")

    thetaA_raw = np.degrees(np.arctan2(tau_yz, tau_zx))
    thetaA_cont = pd.Series(thetaA_raw, index=df_est.index, dtype=float)

    # Em vez de usar apenas τzx, usa o módulo do cisalhamento no plano yz-zx.
    # Isso evita apagar ângulos válidos quando τzx é pequeno, mas τyz é significativo.
    modulo_cisalhamento = np.sqrt(tau_yz ** 2 + tau_zx ** 2)

    limiar_cisalhamento_theta = 0.0
    thetaA_cont = thetaA_cont.mask(modulo_cisalhamento < limiar_cisalhamento_theta)

    # Corrige saltos artificiais de 180° mantendo continuidade angular.
    valores_theta = thetaA_cont.to_numpy(dtype=float)
    indices_validos = np.where(np.isfinite(valores_theta))[0]

    if len(indices_validos) > 0:
        anterior = valores_theta[indices_validos[0]]

        for idx in indices_validos[1:]:
            atual = valores_theta[idx]

            candidatos = np.array([
                atual - 180.0,
                atual,
                atual + 180.0
            ])

            valores_theta[idx] = candidatos[
                np.argmin(np.abs(candidatos - anterior))
            ]

            anterior = valores_theta[idx]

    thetaA_cont = pd.Series(valores_theta, index=df_est.index, dtype=float)

    # Para exibição na tabela, normaliza para 0–360°.
    thetaA_exibir = np.round(thetaA_cont % 360, 2)
    thetaB_exibir = np.round((thetaA_cont + 90) % 360, 2)

    df_est.insert(
        loc=df_est.columns.get_loc("τzx") + 1,
        column="θA (°)",
        value=thetaA_exibir
    )

    df_est.insert(
        loc=df_est.columns.get_loc("θA (°)") + 1,
        column="θB (°)",
        value=thetaB_exibir
    )

    if st.session_state.get("tipo_poco") == "Offshore" and not df_est.empty:
        colunas_primeira_linha_offshore = [
            "Vsh",
            "Direção de SH (°)",
            "Direção de Sh (°)",
            "lyzl",
            "θA (°)",
            "θB (°)"
        ]

        colunas_primeira_linha_offshore = [
            col for col in colunas_primeira_linha_offshore
            if col in df_est.columns
        ]

        df_est.loc[
            df_est.index[0],
            colunas_primeira_linha_offshore
        ] = np.nan

    ordem_colunas = [
        "Profundidade (m)",
        "MD",
        "Perfil de densidade (g/cm³)",
        "Perfil sônico (µs/pé)",
        "Gradiente de Sobrecarga (lb/gal)",
        "Gradiente de Pressão de Poros (lb/gal)",
        "Gradiente de Fratura (lb/gal)",
        "Incl (°)",
        "Azi (°)",
        "Φ (°)",
        "DTS",
        "Poisson",
        "G dinam (MMpsi)",
        "E dinâmico (MMpsi)",
        "E estático (MMpsi)",
        "Vsh",
        "UCS (psi)",
        "So (psi)",
        "Sv (psi)",
        "SH (psi)",
        "Direção de SH (°)",
        "Sh (psi)",
        "Direção de Sh (°)",
        "lxxl",
        "lyxl",
        "lzxl",
        "lxyl",
        "lyyl",
        "lzyl",
        "lxzl",
        "lyzl",
        "lzzl",
        "τxy",
        "τyz",
        "τzx",
        "θA (°)",
        "θB (°)",
        ]

    colunas_ordenadas = [col for col in ordem_colunas if col in df_est.columns]
    colunas_restantes = [col for col in df_est.columns if col not in colunas_ordenadas]
    df_est = df_est[colunas_ordenadas + colunas_restantes]


    return df_est


def pagina_entrada_dados():
    st.header("Entrada de Dados")
    c1, c2, c3 = st.columns((1, 1, 1))

    with c1:
        container = st.container(border=True)
        with container:
            st.markdown("### Upload de Arquivo Excel")

            col1, col2 = st.columns(2)
            with col1:
                preparar_widget_persistente("read_step", "_w_read_step")
                step = st.number_input(
                    "Intervalo entre linhas para leitura",
                    min_value=1,
                    step=1,
                    key="_w_read_step",
                    on_change=salvar_widget_persistente,
                    args=("read_step", "_w_read_step"),
                )

            preparar_widget_persistente("traj_modo", "_w_traj_modo")
            traj_modo = st.selectbox(
                "Trajetória utilizada para os cálculos",
                ["Planejada", "Executada"],
                key="_w_traj_modo",
                on_change=salvar_widget_persistente,
                args=("traj_modo", "_w_traj_modo"),
            )
            preparar_widget_persistente("option", "_w_option")
            st.selectbox(
                "Objetivo do estudo",
                ["Retroanálise", "Previsão de Geopressões"],
                key="_w_option",
                on_change=salvar_widget_persistente,
                args=("option", "_w_option"),
            )

            uploaded_file = st.file_uploader(
                "***Envie o seu arquivo Excel***",
                type=["xlsx", "xls", "xlsm"],
                key="_w_uploaded_xlsm"
            )

            if uploaded_file:
                file_bytes = uploaded_file.getvalue()
            elif arquivo_carregado():
                file_bytes = st.session_state.main_xlsm
            else:
                file_bytes = None

            if file_bytes:
                arquivo_hash = hash(file_bytes)

                import_key = (
                    arquivo_hash,
                    int(st.session_state.read_step),
                    st.session_state.traj_modo,
                )

                arquivo_novo = st.session_state.get("main_xlsm_hash") != arquivo_hash

                deve_importar = (
                        st.session_state.get("main_xlsm_import_key") != import_key
                        or not isinstance(st.session_state.get("df1"), pd.DataFrame)
                        or st.session_state.df1.empty
                        or not isinstance(st.session_state.get("df2"), pd.DataFrame)
                        or st.session_state.df2.empty
                        or not isinstance(st.session_state.get("df_interp"), pd.DataFrame)
                        or st.session_state.df_interp.empty
                )

                if arquivo_novo:
                    st.session_state.main_xlsm_hash = arquivo_hash
                    st.session_state.lito_import_ok = False
                    st.session_state.wb = carregar_workbook(file_bytes)

                if st.session_state.get("wb") is None:
                    st.session_state.wb = carregar_workbook(file_bytes)

                st.session_state.main_xlsm = file_bytes

                if deve_importar:
                    st.session_state.main_xlsm_import_key = import_key

                    try:
                        # Selecionar a aba desejada
                        sheet_name = "Perfilagens"

                        # Carregar os dados da aba selecionada
                        df_full = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)
                        df_full.columns = [str(c).strip() for c in df_full.columns]

                        # Colunas obrigatórias
                        colunas_obrigatorias = [
                            "Profundidade",
                            "MD",
                            "Perfil de densidade",
                            "Perfil sônico",
                            "Perfil Raio Gama"
                        ]

                        colunas_faltantes = [c for c in colunas_obrigatorias if c not in df_full.columns]

                        if colunas_faltantes:
                            st.error(
                                "❌ O arquivo enviado não contém todas as colunas obrigatórias.\n\n"
                                f"**Colunas ausentes:** {', '.join(colunas_faltantes)}"
                            )
                            st.stop()

                        # Lê a trajetória antes de cortar a perfilagem
                        df2 = _ler_trajetoria_do_xlsm(
                            st.session_state.wb,
                            st.session_state.traj_modo
                        )
                        st.session_state.df2 = df2

                        # Usa a trajetória lida diretamente da aba Trajetória
                        df_out_traj = df2.copy()

                        df_out_traj["Inclinação (°)"] = df_out_traj["Inc"]
                        df_out_traj["Azimute (°)"] = df_out_traj["Azi"]

                        st.session_state.df_out_traj = df_out_traj.copy()

                        # TVD final do poço lido da aba Trajetória
                        tvd_final_poco = float(df_out_traj["TVD"].iloc[-1])
                        st.session_state.tvd_final_poco = tvd_final_poco

                        linhas_antes = len(df_full)

                        # Limita a perfilagem ao TVD final do poço
                        df_full = _limitar_perfilagem_ao_tvd_final(
                            df_full,
                            tvd_final=tvd_final_poco,
                            col_tvd="Profundidade"
                        )

                        linhas_depois = len(df_full)

                        # Aplicar o intervalo step depois do corte
                        df = df_full.iloc[::step].reset_index(drop=True)

                        # Garantir que o último registro cortado está no df
                        if not df_full.iloc[-1].equals(df.iloc[-1]):
                            df = pd.concat([df, df_full.iloc[[-1]]], ignore_index=True)

                        st.session_state.df1 = df

                        with col2:
                            st.write("")
                            st.write("")
                            st.markdown(f"**Total de linhas carregadas:** {len(df)}")

                        st.write("Dados Importados:")
                        st.dataframe(df, use_container_width=True, hide_index=True)

                        # Garantir que o último registro original está no df
                        if not df_full.iloc[-1].equals(df.iloc[-1]):
                            # Adiciona o último registro ao dataframe
                            df = pd.concat([df, df_full.iloc[[-1]]], ignore_index=True)

                        # Colunas obrigatórias
                        colunas_obrigatorias = [
                            "Profundidade",
                            "MD",
                            "Perfil de densidade",
                            "Perfil sônico",
                            "Perfil Raio Gama"
                        ]

                        # Verificação
                        colunas_faltantes = [c for c in colunas_obrigatorias if c not in df.columns]

                        if colunas_faltantes:
                            st.error(
                                "❌ O arquivo enviado não contém todas as colunas obrigatórias.\n\n"
                                f"**Colunas ausentes:** {', '.join(colunas_faltantes)}"
                            )
                            st.stop()
                        else:
                            st.session_state.df1 = df

                    except Exception as e:
                        st.error(f"Erro ao processar o arquivo: {e}")

                    # lê a aba Início e preenche session_state
                    try:
                        info_ini = _ler_inicio_do_xlsm(st.session_state.wb)

                        if info_ini["poco"] is not None:
                            st.session_state.poco = info_ini["poco"]

                        # Objetivo -> comments
                        st.session_state.comments = info_ini["comments"]

                        # Coordenadas
                        if info_ini["easting"] is not None:
                            st.session_state.easting = info_ini["easting"]

                        if info_ini["northing"] is not None:
                            st.session_state.northing = info_ini["northing"]

                        if info_ini["zona"] is not None:
                            st.session_state.zona = info_ini["zona"]

                        if info_ini["hem"] is not None:
                            st.session_state.hem = info_ini["hem"]

                    except Exception as e:
                        st.error(f"Não foi possível ler a aba 'Início' para preencher dados do poço: {e}")

                    try:
                        df_mud = _ler_peso_fluido_do_xlsm(st.session_state.wb)
                        st.session_state.df_mud = df_mud

                    except Exception as e:
                        if st.session_state.option == "Retroanálise":
                            st.error(f"Não foi possível ler pesos do fluido na aba 'Geopressões': {e}")
                        else:
                            pass

                    try:
                        df_eventos = pd.read_excel(BytesIO(file_bytes), sheet_name="Eventos")

                        # limpa nomes de colunas (resolve "MD Final " com espaço)
                        df_eventos.columns = [str(c).strip() for c in df_eventos.columns]

                        # garante as colunas esperadas
                        col_req_evt = ["MD Inicial", "MD Final", "Evento"]
                        falt = [c for c in col_req_evt if c not in df_eventos.columns]
                        if falt:
                            st.warning(f"A aba 'Eventos' existe, mas faltam colunas: {', '.join(falt)}")
                            st.session_state.df_eventos = pd.DataFrame(columns=col_req_evt)
                        else:
                            # normaliza tipos e strings
                            df_eventos = df_eventos[col_req_evt].copy()
                            df_eventos["MD Inicial"] = pd.to_numeric(df_eventos["MD Inicial"], errors="coerce")
                            df_eventos["MD Final"] = pd.to_numeric(df_eventos["MD Final"], errors="coerce")
                            df_eventos["Evento"] = (
                                df_eventos["Evento"]
                                .astype("string")
                                .fillna("")
                                .str.strip()
                                .str.replace(r"[,\.;:]+$", "", regex=True)  # remove vírgula no final (ex: "Overpull,")
                            )

                            # remove linhas inválidas
                            df_eventos = df_eventos.dropna(subset=["MD Inicial"])
                            df_eventos = df_eventos[df_eventos["Evento"] != ""].reset_index(drop=True)

                            st.session_state.df_eventos = df_eventos

                    except Exception:
                        # se não tiver a aba, só deixa vazio
                        st.session_state.df_eventos = pd.DataFrame(columns=["MD Inicial", "MD Final", "Evento"])

                    try:
                        if "df2" not in st.session_state or not isinstance(st.session_state.df2, pd.DataFrame):
                            df2 = _ler_trajetoria_do_xlsm(
                                st.session_state.wb,
                                st.session_state.traj_modo
                            )
                            st.session_state.df2 = df2

                        st.session_state.df_interp = _gerar_df_interp_a_partir_df1_df2(
                            st.session_state.df1,
                            st.session_state.df2
                        )

                    except Exception as e:
                        st.error(f"Não foi possível carregar/interpolar trajetória pela aba 'Trajetória': {e}")

                    try:
                        sapatas_df = _ler_sapatas_do_xlsm(st.session_state.wb)
                        st.session_state.sapatas_df = sapatas_df

                    except Exception as e:
                        if st.session_state.option == "Retroanálise":
                            st.error(f"Não foi possível ler as sapatas do Excel: {e}")
                        else:
                            pass

                    try:
                        fases_df = _ler_fases_do_xlsm(st.session_state.wb)
                        st.session_state.fases_df = fases_df

                    except Exception as e:
                        if st.session_state.option == "Retroanálise":
                            st.error(f"Não foi possível ler as fases do Excel: {e}")
                        else:
                            pass

                else:
                    df = st.session_state.df1

                    with col2:
                        st.write("")
                        st.write("")
                        st.markdown(f"**Total de linhas carregadas:** {len(df)}")

                    st.write("Dados Importados:")
                    st.dataframe(df, use_container_width=True, hide_index=True)

    with c2:
        with st.container(border=True):
            st.markdown('#### Informações básicas do poço')
            preparar_widget_persistente("tipo_poco", "_w_tipo_poco")
            st.segmented_control(
                "Ambiente do poço",
                options=["Onshore", "Offshore"],
                key="_w_tipo_poco",
                width="stretch",
                on_change=salvar_widget_persistente,
                args=("tipo_poco", "_w_tipo_poco"),
            )

            tipo_poco = st.session_state.tipo_poco
            if st.session_state.get("_tipo_poco_anterior") != tipo_poco:
                if st.session_state.get("datum") in (None, "", "RTKB", "RTKB + LDA"):
                    st.session_state.datum = "RTKB" if tipo_poco == "Onshore" else "RTKB + LDA"
                    st.session_state["_w_datum"] = st.session_state.datum

                st.session_state._tipo_poco_anterior = tipo_poco

            preparar_widget_persistente("datum", "_w_datum")
            st.text_input(
                "Datum",
                key="_w_datum",
                on_change=salvar_widget_persistente,
                args=("datum", "_w_datum"),
            )
            exibir_campo_info("Nome do Poço", st.session_state.get("poco"))
            exibir_campo_info("Objetivo do Poço", st.session_state.get("comments"))
            lista_paises = list(paises.keys())

            preparar_widget_persistente("country_name", "_w_country_name")
            st.selectbox(
                "País",
                options=lista_paises,
                key="_w_country_name",
                on_change=salvar_widget_persistente,
                args=("country_name", "_w_country_name"),
            )
            codigo_pais = paises.get(st.session_state.country_name)
            flag_path = f"Flag/{codigo_pais}.png" if codigo_pais else None

            preparar_widget_persistente("user_name", "_w_user_name")
            st.text_input(
                'Nome do Usuário',
                max_chars=None,
                key='_w_user_name',
                type="default",
                on_change=salvar_widget_persistente,
                args=("user_name", "_w_user_name"),
            )

            preparar_widget_persistente("company_name", "_w_company_name")
            st.text_input(
                'Nome da Companhia',
                max_chars=None,
                key='_w_company_name',
                type="default",
                on_change=salvar_widget_persistente,
                args=("company_name", "_w_company_name"),
            )

            preparar_widget_persistente("field_name", "_w_field_name")
            st.text_input(
                'Nome do Campo',
                max_chars=None,
                key='_w_field_name',
                type="default",
                on_change=salvar_widget_persistente,
                args=("field_name", "_w_field_name"),
            )

    # Função para calcular distância geodésica (em metros) entre dois pontos lat/lon
    def haversine(lat1, lon1, lat2, lon2):
        R = 6371000
        phi1 = math.radians(lat1)
        phi2 = math.radians(lat2)
        dphi = math.radians(lat2 - lat1)
        dlambda = math.radians(lon2 - lon1)
        a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        return R * c
    # --- CARREGAR POÇOS DO YAML ---
    with open("pocos.yaml", "r", encoding="utf-8") as f:
        dados_yaml = yaml.safe_load(f)

    pocos = dados_yaml['pocos']

    # ------------------------------
    # GEOLOCALIZAÇÃO AUTOMÁTICA
    # ------------------------------
    geo = get_geolocation()

    if geo is not None:
        st.session_state.geo_auto_tentado = True

        if isinstance(geo, dict) and "coords" in geo:
            lat = geo["coords"].get("latitude")
            lon = geo["coords"].get("longitude")

            if lat is not None and lon is not None and not st.session_state.geo_auto_ok:
                try:
                    easting_auto, northing_auto, zona_auto, _ = utm.from_latlon(float(lat), float(lon))

                    st.session_state.easting = float(easting_auto)
                    st.session_state.northing = float(northing_auto)
                    st.session_state.zona = int(zona_auto)
                    st.session_state.hem = "Norte" if float(lat) >= 0 else "Sul"
                    st.session_state.geo_auto_ok = True
                    st.rerun()
                except Exception:
                    pass
    with c3:
        with st.container(border=True, height=610):
            st.markdown(f"#### Coordenadas da cabeça do poço {st.session_state.get('poco', '')}")

            col_zona, col_hem = st.columns(2)

            with col_zona:
                exibir_campo_info("Zona UTM", st.session_state.get("zona"))

            with col_hem:
                exibir_campo_info("Hemisfério", st.session_state.get("hem"))

            col_easting, col_northing = st.columns(2)

            with col_easting:
                exibir_campo_info(
                    "Coordenada Leste (Easting)",
                    f"{st.session_state.get('easting', 0):.2f}"
                )

            with col_northing:
                exibir_campo_info(
                    "Coordenada Norte (Northing)",
                    f"{st.session_state.get('northing', 0):.2f}"
                )

            preparar_widget_persistente("raio", "_w_raio")
            st.number_input(
                "Raio de investigação (km)",
                min_value=0.1,
                step=0.1,
                format="%.2f",
                key="_w_raio",
                on_change=salvar_widget_persistente,
                args=("raio", "_w_raio"),
            )

            lat_base, lon_base = utm.to_latlon(
                st.session_state.easting,
                st.session_state.northing,
                st.session_state.zona,
                northern=(st.session_state.hem == "Norte")
            )

            # --- CRIA MAPA COM POÇO BASE E CÍRCULO ---
            m = folium.Map(
                location=[lat_base, lon_base],
                zoom_start=7,
                zoom_control=False,
                attributionControl=False,
                tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
                attr='Esri'
            )

            folium.Circle(
                location=[lat_base, lon_base],
                radius=st.session_state.raio * 1000,
                color="green",
                fill=True,
                fill_opacity=0.2,
                popup=f"Raio: {st.session_state.raio:.2f} km"
            ).add_to(m)

            if uploaded_file:
                st.session_state.profundidade_maxima = max(df['Profundidade'])

            profundidade_maxima = st.session_state.get('profundidade_maxima', None)

            if profundidade_maxima is not None:
                pocos_filtrados = [
                    poco for poco in pocos
                    if poco.get("profundidade_vertical_m") is not None and
                       poco["profundidade_vertical_m"] <= profundidade_maxima
                ]
            else:
                pocos_filtrados = pocos

            # --- PLOTAR POÇOS E CALCULAR DISTÂNCIAS ---
            dados_pontos = []
            def _hemisferio_norte(valor):
                txt = str(valor).strip().lower()
                return txt in ("n", "norte", "north")
            for poco in pocos_filtrados:
                e = poco['coordenadas']['easting']
                n = poco['coordenadas']['northing']

                hem_poco = poco.get("hem", "Sul")

                lat_p, lon_p = utm.to_latlon(
                    e,
                    n,
                    poco['zona_utm'],
                    northern=_hemisferio_norte(hem_poco)
                )

                dist = haversine(lat_base, lon_base, lat_p, lon_p)
                dentro_do_raio = dist <= (st.session_state.raio * 1000)

                cor = "red"
                icone = "map-marker"

                popup_text = f"{poco['nome']}<br>Distância: {dist / 1000:.2f} km"

                folium.Marker(
                    location=[lat_p, lon_p],
                    popup=folium.Popup(popup_text, max_width=300),
                    icon=folium.Icon(color=cor, icon=icone)
                ).add_to(m)

                dados_pontos.append({
                    "Nome": poco['nome'],
                    "Origem": poco.get("origem", "yaml"),
                    "Easting": e,
                    "Northing": n,
                    "Distância (km)": round(dist / 1000, 2),
                    "Profundidade Vertical (m)": poco.get("profundidade_vertical_m", None),
                    "Peso Eq. (lb/gal)": poco.get("peso_eq_lb_gal", None)
                })

            # --- PLOTAR POÇO PRINCIPAL POR CIMA DOS DEMAIS ---
            folium.Marker(
                location=[lat_base, lon_base],
                popup=st.session_state.get('poco', ''),
                icon=folium.CustomIcon('poço.png', icon_size=(30, 30)),
                z_index_offset=10000
            ).add_to(m)

            # --- MOSTRAR MAPA ---
            altura_mapa = 240
            st.session_state["mapa_folium_pdf"] = m
            st_folium(
                m,
                use_container_width=True,
                height=altura_mapa,
                returned_objects=[],
            )
            m.save('filename.png')


def pagina_coluna_litologica():
    st.header("Litologia")

    if not arquivo_carregado():
        st.error('Por favor, insira um documento!', icon="🚨")
        return

    df = st.session_state.df1.copy()
    prof_max_idade = float(df["Profundidade"].max()) if "Profundidade" in df.columns else None
    _atualizar_df_idade_do_state(prof_max_idade)

    y_max_padrao_sobrecarga = int(df['Profundidade'].max()) + 100
    y_max_padrao_anterior = st.session_state.get("_sobrecarga_y_max_padrao")
    y_max_atual = st.session_state.get("y_max_s")
    y_max_parece_padrao_antigo = y_max_atual in (None, 1000, y_max_padrao_anterior)
    y_max_parece_lda = (
        st.session_state.tipo_poco == "Offshore"
        and y_max_atual is not None
        and y_max_atual <= st.session_state.get("lda", 0) + 100
        and y_max_atual < y_max_padrao_sobrecarga
    )

    if y_max_parece_padrao_antigo or y_max_parece_lda:
        st.session_state.y_max_s = y_max_padrao_sobrecarga
        st.session_state["_w_y_max_s"] = y_max_padrao_sobrecarga

    st.session_state._sobrecarga_y_max_padrao = y_max_padrao_sobrecarga

    profundidades = []
    formacoes = []
    litologias = []

    st.session_state.well_name = st.session_state.get("poco", "Poço")

    if "n_lbf" not in st.session_state:
        st.session_state.n_lbf = 1

    for chave, valor_padrao in (
        ("s_gr", "Não"),
        ("y_min_pp", 0),
        ("y_max_pp", y_max_padrao_sobrecarga),
        ("y_step_pp", 200),
    ):
        if chave not in st.session_state:
            st.session_state[chave] = valor_padrao

    col_outros, col_lbf, col_grafico = st.columns(3)

    with col_outros:
        with st.expander(
                "Idades geológicas",
                expanded=estado_expander_persistente("exp_lito_idades", True)
        ):
            preparar_widget_persistente("idg", "_w_idg")
            st.selectbox(
                "Inserir Idade Geológica",
                ["Não", "Sim"],
                key="_w_idg",
                on_change=salvar_widget_persistente,
                args=("idg", "_w_idg")
            )

            if st.session_state.idg == "Sim":
                df_idade_salvo = st.session_state.get("df_idade", pd.DataFrame())
                if not isinstance(df_idade_salvo, pd.DataFrame):
                    df_idade_salvo = pd.DataFrame(columns=["Topo (m)", "Base (m)", "Idade"])
                if not df_idade_salvo.empty and st.session_state.get("n_id", 1) == 1:
                    st.session_state.n_id = len(df_idade_salvo)
                preparar_widget_persistente("n_id", "_w_n_id")
                st.number_input(
                    "Quantidade de idades geológicas",
                    min_value=1,
                    step=1,
                    key="_w_n_id",
                    on_change=salvar_widget_persistente,
                    args=("n_id", "_w_n_id")
                )

                prof_ini_id = []
                prof_fim_id = []
                idade_geo = []

                prof_max = float(df["Profundidade"].max()) if "Profundidade" in df.columns else 0.0

                for i in range(int(st.session_state.n_id)):
                    col1, col2, col3 = st.columns(3)

                    with col1:
                        topo_salvo = (
                            float(df_idade_salvo["Topo (m)"].iloc[i])
                            if i < len(df_idade_salvo) and pd.notna(df_idade_salvo["Topo (m)"].iloc[i])
                            else 0.0
                        )
                        if i == 0:
                            if f"prof_inicial_2_{i}" not in st.session_state:
                                st.session_state[f"prof_inicial_2_{i}"] = topo_salvo
                            p_ini = st.number_input(
                                f"Profundidade inicial **Intervalo {i + 1}**",
                                step=1.0,
                                format="%f",
                                min_value=0.0,
                                key=f"prof_inicial_2_{i}"
                            )
                        else:
                            st.session_state[f"prof_inicial_2_{i}"] = st.session_state.get(f"prof_final_2_{i - 1}", 0.0)
                            p_ini = st.number_input(
                                f"Profundidade inicial **Intervalo {i + 1}**",
                                disabled=True,
                                key=f"prof_inicial_2_{i}"
                            )

                    with col2:
                        base_salva = (
                            float(df_idade_salvo["Base (m)"].iloc[i])
                            if i < len(df_idade_salvo) and pd.notna(df_idade_salvo["Base (m)"].iloc[i])
                            else 0.0
                        )
                        if i == int(st.session_state.n_id) - 1:
                            st.session_state[f"prof_final_2_{i}"] = float(prof_max + 100)
                            p_fim = st.number_input(
                                f"Profundidade final **Intervalo {i + 1}**",
                                disabled=True,
                                key=f"prof_final_2_{i}"
                            )
                        else:
                            if f"prof_final_2_{i}" not in st.session_state:
                                st.session_state[f"prof_final_2_{i}"] = base_salva
                            p_fim = st.number_input(
                                f"Profundidade final **Intervalo {i + 1}**",
                                step=1.0,
                                format="%f",
                                min_value=0.0,
                                key=f"prof_final_2_{i}"
                            )

                    with col3:
                        idade_salva = (
                            str(df_idade_salvo["Idade"].iloc[i])
                            if i < len(df_idade_salvo) and pd.notna(df_idade_salvo["Idade"].iloc[i])
                            else ""
                        )
                        if f"idg_{i}" not in st.session_state:
                            st.session_state[f"idg_{i}"] = idade_salva
                        idade = st.text_input(
                            f"Idade Geológica {i + 1}",
                            key=f"idg_{i}"
                        )

                    prof_ini_id.append(p_ini)
                    prof_fim_id.append(p_fim)
                    idade_geo.append(idade)

                st.session_state.df_idade = pd.DataFrame({
                    "Topo (m)": prof_ini_id,
                    "Base (m)": prof_fim_id,
                    "Idade": idade_geo
                })

        with st.expander(
                "Descrição das camadas litológicas",
                expanded=estado_expander_persistente("exp_lito_descricao", True)
        ):
            if "pocos" not in st.session_state:
                st.session_state.pocos = {"Poço": {}}

            selected = st.session_state.get("well_selected", "Poço")
            st.session_state.well_selected = selected

            if selected not in st.session_state.pocos:
                st.session_state.pocos[selected] = {}

            _garantir_litologia_importada(selected)

            excel_carregado = st.session_state.get("lito_import_ok", False)

            if not excel_carregado:
                st.info("Nenhum dado de litologia foi encontrado no arquivo Excel.")
                return

            lithology = [
                "Argilito", "Arenito", "Folhelho", "Calcário", "Carbonato",
                "Calcilutito", "Margas", "Siltito", "Diamictito", "Conglomerado",
                "Anidrita / Gipsita", "Halita", "Calcissiltito", "Calcarenito",
                "Calcirrudito", "Coquina", "Dolomito", "Basalto", "Diabásio", "Marga",
            ]

            n_fm_padrao = len(st.session_state.pocos[selected].get("formation", []))

            if n_fm_padrao == 0:
                st.info("Nenhum dado de litologia foi encontrado no arquivo Excel.")
                return

            st.session_state.n_fm = n_fm_padrao

            dados_formacoes = []

            for i in range(int(st.session_state.n_fm)):
                p = st.session_state.pocos[selected].get("profundidade", [])
                f = st.session_state.pocos[selected].get("formation", [])
                y = st.session_state.pocos[selected].get("litologia", [])

                prof_default = p[i] if i < len(p) else 0.0
                fm_default = f[i] if i < len(f) else ""
                lit_default = y[i] if i < len(y) and y[i] in lithology else ""

                dados_formacoes.append({
                    "Topo (TVD)": float(prof_default),
                    "Formação": str(fm_default),
                    "Litologia": lit_default
                })

            df_formacoes = pd.DataFrame(dados_formacoes)

            df_formacoes["Topo (TVD)"] = pd.to_numeric(
                df_formacoes["Topo (TVD)"],
                errors="coerce"
            )

            st.dataframe(
                df_formacoes.style
                .format({
                    "Topo (TVD)": "{:.2f}"
                })
                .set_properties(**{
                    "text-align": "center"
                })
                .set_table_styles([
                    {
                        "selector": "th",
                        "props": [("text-align", "center")]
                    },
                    {
                        "selector": "td",
                        "props": [("text-align", "center")]
                    }
                ]),
                use_container_width=True,
                hide_index=True
            )

            profundidades = df_formacoes["Topo (TVD)"].astype(float).tolist()
            formacoes = df_formacoes["Formação"].astype(str).tolist()
            litologias = df_formacoes["Litologia"].astype(str).tolist()

            st.session_state.pocos[selected]["profundidade"] = profundidades
            st.session_state.pocos[selected]["formation"] = formacoes
            st.session_state.pocos[selected]["litologia"] = litologias

            if "perfil" not in st.session_state.pocos[selected]:
                st.session_state.pocos[selected]["perfil"] = {}

            if (
                    "df_idade" in st.session_state
                    and isinstance(st.session_state.df_idade, pd.DataFrame)
                    and not st.session_state.df_idade.empty
            ):
                base_final = st.session_state.df_idade["Base (m)"].max()
            else:
                base_final = max(profundidades) if profundidades else 0

            st.session_state.pocos[selected]["tvd"] = base_final + 100

        with st.expander(
                "Visualização",
                expanded=estado_expander_persistente("exp_lito_visualizacao", True)
        ):
            if not st.session_state.get("lito_import_ok", False):
                st.info("A visualização da coluna litológica será exibida após carregar dados de litologia no Excel.")
                return

            fig = plot_correlacao_com_logs(
                st.session_state.pocos,
                [False, False, False],
                True,
                True,
                list(st.session_state.pocos.keys()),
                False,
                escala=(30, 300)
            )

            st.session_state.fig_coluna_lito = fig
            st.plotly_chart(fig, use_container_width=True)

    with col_lbf:
        with st.expander(
            "Linhas Base de Folhelhos",
            expanded=estado_expander_persistente("exp_lito_lbf", True)
        ):
            col_add_lbf_lito, col_rem_lbf_lito = st.columns(2)

            with col_add_lbf_lito:
                if st.button(
                        "➕ LBF",
                        type="primary",
                        use_container_width=True,
                        key="b_add_lbf_litologia"
                ):
                    st.session_state.n_lbf = int(st.session_state.get("n_lbf", 1)) + 1
                    st.rerun()

            with col_rem_lbf_lito:
                if st.button(
                        "➖ LBF",
                        type="primary",
                        use_container_width=True,
                        key="b_rem_lbf_litologia"
                ):
                    if int(st.session_state.get("n_lbf", 1)) > 1:
                        idx = int(st.session_state.n_lbf) - 1
                        _remover_chaves_lbf(idx)
                        st.session_state.n_lbf -= 1
                        st.rerun()

            _sincronizar_widgets_lbf_do_estado()

            def preparar_widget_lbf_litologia(chave_estado):
                chave_widget = f"_w_{chave_estado}"
                preparar_widget_persistente(chave_estado, chave_widget)
                return chave_widget

            _renderizar_campos_lbf(preparar_widget_lbf_litologia, incluir_titulo=False, persistente=True)
            _atualizar_df_pp_lito_por_lbf()

            opcoes_sim_nao = ["Sim", "Não"]
            if st.session_state.get("s_gr") not in opcoes_sim_nao:
                st.session_state.s_gr = "Não"
            preparar_widget_persistente("s_gr", "_w_s_gr")
            st.selectbox(
                "Suavizar Raio Gama",
                opcoes_sim_nao,
                key="_w_s_gr",
                on_change=salvar_widget_persistente,
                args=("s_gr", "_w_s_gr")
            )

            opcoes_lito_graficos = ["Permeável / Não permeável", "Litologia do Excel"]
            if st.session_state.get("tipo_coluna_litologica_graficos") not in opcoes_lito_graficos:
                st.session_state.tipo_coluna_litologica_graficos = "Permeável / Não permeável"
            preparar_widget_persistente("tipo_coluna_litologica_graficos", "_w_tipo_coluna_litologica_graficos")
            st.selectbox(
                "Litologia nos subplots",
                opcoes_lito_graficos,
                key="_w_tipo_coluna_litologica_graficos",
                on_change=salvar_widget_persistente,
                args=("tipo_coluna_litologica_graficos", "_w_tipo_coluna_litologica_graficos")
            )

    with col_grafico:
        with st.container(border=True):
            st.markdown("### LBF")

            df_lbf_lito = _montar_df_lbf_litologia()

            if isinstance(df_lbf_lito, pd.DataFrame) and not df_lbf_lito.empty:
                lbfs_lito = _coletar_lbfs_poros()
                df_lbf_lito = _calcular_lbf_pp(df_lbf_lito, lbfs_lito)
                st.session_state.df_pp_lito = df_lbf_lito.copy()
                st.session_state.df_lbf_litologia = df_lbf_lito.copy()

                _plotar_gr_poros_com_contexto(
                    df_lbf_lito,
                    "LBF",
                    lbfs=lbfs_lito
                )

            else:
                st.info("Calcule o Gradiente de Sobrecarga para gerar o gráfico da LBF.")


def pagina_sobrecarga():
    st.header("Gradiente de Sobrecarga")

    if not arquivo_carregado():
        st.error('Por favor, insira um documento!', icon="🚨")
        return

    df = st.session_state.df1.copy()
    prof_max_idade = float(df["Profundidade"].max()) if "Profundidade" in df.columns else None
    _atualizar_df_idade_do_state(prof_max_idade)

    y_max_padrao_sobrecarga = int(df['Profundidade'].max()) + 100
    y_max_padrao_anterior = st.session_state.get("_sobrecarga_y_max_padrao")
    y_max_atual = st.session_state.get("y_max_s")
    y_max_parece_padrao_antigo = y_max_atual in (None, 1000, y_max_padrao_anterior)
    y_max_parece_lda = (
        st.session_state.tipo_poco == "Offshore"
        and y_max_atual is not None
        and y_max_atual <= st.session_state.get("lda", 0) + 100
        and y_max_atual < y_max_padrao_sobrecarga
    )

    if y_max_parece_padrao_antigo or y_max_parece_lda:
        st.session_state.y_max_s = y_max_padrao_sobrecarga
        st.session_state["_w_y_max_s"] = y_max_padrao_sobrecarga

    st.session_state._sobrecarga_y_max_padrao = y_max_padrao_sobrecarga

    selected = st.session_state.get("well_selected", st.session_state.get("poco", "Poço"))
    st.session_state.well_selected = selected
    _garantir_litologia_importada(selected)
    poco = st.session_state.get("pocos", {}).get(selected, {})

    profundidades = poco.get("profundidade", [])
    litologias = poco.get("litologia", [])
    formacoes = poco.get("formation", [])


    tipo_sobrecarga_anterior = st.session_state.get("_sobrecarga_tipo_poco_anterior")
    if tipo_sobrecarga_anterior != st.session_state.tipo_poco:
        if st.session_state.get("rtkb") in (None, 9.4, 25.0):
            st.session_state.rtkb = 9.4 if st.session_state.tipo_poco == "Onshore" else 25.0
            st.session_state["_w_rtkb"] = st.session_state.rtkb
        if st.session_state.tipo_poco == "Onshore" and st.session_state.get("es") in (None, 0.0):
            st.session_state.es = 110.0
            st.session_state["_w_es"] = st.session_state.es
        if st.session_state.tipo_poco == "Offshore" and st.session_state.get("lda") in (None, 0.0):
            st.session_state.lda = 978.0
            st.session_state["_w_lda"] = st.session_state.lda
        st.session_state._sobrecarga_tipo_poco_anterior = st.session_state.tipo_poco

    tab_dados, tab_tabela = st.tabs(["Dados", "Tabela de dados calculados"])

    with tab_dados:
        st.subheader("Dados")
        c1, c2, c3 = st.columns((1, 1, 1))
        with c1:
            with st.container(border=True):
                if st.session_state.tipo_poco == "Onshore":
                    x = False
                else:
                    x = True
                preparar_widget_persistente("gard", "_w_gard")
                st.segmented_control("***Correlação para estimativa da densidade da formação***",
                                     ['Perfil de Densidade', 'Gardner'],
                                     selection_mode="multi",
                                     key='_w_gard',
                                     width="stretch",
                                     on_change=salvar_widget_persistente,
                                     args=("gard", "_w_gard"))

                preparar_widget_persistente("ex", "_w_ex")
                st.segmented_control("***Extrapolação***",
                                     ['Desativada', 'Ativada'],
                                     selection_mode="single",
                                     key='_w_ex',
                                     width="stretch",
                                     disabled=False,
                                     on_change=salvar_widget_persistente,
                                     args=("ex", "_w_ex"))

                # MÉDIA DAS DENSIDADES CASO FOR FAZER A EXTRAPOLAÇÃO
                if st.session_state.ex == 'Ativada':
                    st.session_state.md = round(df['Perfil de densidade'].dropna().head(20).mean(), 2)
                    if not st.session_state.get("_ds_inicializado_por_media", False):
                        st.session_state.ds = st.session_state.md
                        st.session_state["_w_ds"] = st.session_state.ds
                        st.session_state._ds_inicializado_por_media = True
                    preparar_widget_persistente("ds", "_w_ds")
                    st.number_input('Insira o ***Valor médio da densidade das camadas superiores***',
                                    step=0.1, format='%.2f', key='_w_ds', min_value=0.0,
                                    on_change=salvar_widget_persistente,
                                    args=("ds", "_w_ds"))

                # CÁLCULOS
                with st.form("gs_form", border=False):
                    # INSERIR DADOS
                    st.markdown('***Dados de elevação***')
                    preparar_widget_persistente("rtkb", "_w_rtkb")
                    st.number_input('***Air Gap (m)***', step=1.0, format='%.2f', key='_w_rtkb', min_value=0.0)
                    if st.session_state.tipo_poco == "Onshore":
                        preparar_widget_persistente("es", "_w_es")
                        st.number_input('***Elevação do DATUM (m)***', step=1.0, format='%.2f', key='_w_es',
                                        min_value=0.0)
                    else:
                        preparar_widget_persistente("lda", "_w_lda")
                        st.number_input("Insira o valor da ***Lâmina D'água (m)***", step=1.0, format='%.2f',
                                        key='_w_lda', min_value=0.0)

                    sincronizar_widgets_persistentes([
                        ("rtkb", "_w_rtkb"),
                        ("es", "_w_es"),
                        ("lda", "_w_lda"),
                    ])
                    st.session_state.nf = st.session_state.rtkb + st.session_state.es
                    float_sonic = df['Perfil sônico'].apply(lambda x: isinstance(x, float)).any()
                    float_dens = df['Perfil de densidade'].apply(lambda x: isinstance(x, float)).any()

                    df.insert(
                        loc=5,
                        column='Raio Gama Suavizado',
                        value=suavizar(
                            df['Profundidade'],
                            df['Perfil Raio Gama']
                        )
                    )

                    if float_dens or float_sonic:
                        # OFFSHORE
                        if st.session_state.tipo_poco == "Offshore":
                            if st.session_state.ex == 'Desativada':
                                df_sup = pd.DataFrame({
                                    'Profundidade': [st.session_state.rtkb,
                                                     st.session_state.rtkb + st.session_state.lda],
                                    'Perfil de densidade': [0, 1.03],
                                    'Perfil sônico': [None, None]
                                })
                            else:
                                prof_start = (st.session_state.rtkb + st.session_state.lda) + 1
                                prof_end = df['Profundidade'].iloc[0]

                                if prof_end > prof_start:
                                    prof = list(range(int(prof_start), int(prof_end) + 1))
                                else:
                                    prof = []

                                df_sup = pd.DataFrame({
                                    'Profundidade': [st.session_state.rtkb,
                                                     st.session_state.rtkb + st.session_state.lda] + prof,
                                    'Perfil de densidade': [None, 1.04356087080324] + [st.session_state.md] * len(
                                        prof),
                                })
                            # Concatena dataframe superior com o dataframe principal
                            df = pd.concat([df_sup, df], ignore_index=True)

                            dt = [0]
                            dd, dm = [], []

                            if 'Gardner' in st.session_state.gard:
                                mask = df['Perfil de densidade'].isna()

                                sonico = df.loc[mask, 'Perfil sônico']

                                df.loc[mask, 'Perfil de densidade'] = (
                                        0.23 * ((1_000_000 / sonico) ** 0.25)
                                )
                            # Cálculo de ΔD (m)
                            dd.append(df['Profundidade'].iloc[0])
                            for i in range(1, len(df)):
                                dd.append(
                                    df['Profundidade'].iloc[i] - df['Profundidade'].iloc[i - 1])
                            df['ΔD (m)'] = dd

                            # Cálculo de Densidade x ΔD x 1,422 (psi)
                            for i in range(len(df)):
                                dm.append(df['Perfil de densidade'].iloc[i] * df['ΔD (m)'].iloc[i] * 1.422)

                            df['Densidade x ΔD x 1,422 (psi)'] = dm
                            df['Pressão de Sobrecarga (psi)'] = df['Densidade x ΔD x 1,422 (psi)'].cumsum()

                            gs = []
                            for i in range(len(df)):
                                if df['Profundidade'].iloc[i] <= st.session_state.rtkb:
                                    gs.append(None)
                                else:
                                    gs.append(df['Pressão de Sobrecarga (psi)'].iloc[i] / (
                                            0.1704 * df['Profundidade'].iloc[i]))

                            df['Gradiente de Sobrecarga (lb/gal)'] = gs

                            st.session_state.df_sobrecarga = df.copy()
                            st.session_state.df = st.session_state.df_sobrecarga

                        # ONSHORE
                        else:
                            # SE EXTRAPOLAR E NÃO TIVER DADOS DE DENSIDADE, DENSIDADE DAS CAMADAS EXTRAPOLADAS IGUAL A 2
                            # SE NÃO EXTRAPOLAR, NÃO FAZ NADA
                            if st.session_state.ex == 'Ativada':
                                prof_min = df['Profundidade'][0]
                                ext = np.arange(0, prof_min, 1)
                                ext_list = list(ext)
                                prof_list = df['Profundidade'].tolist()
                                dens_list = df['Perfil de densidade'].tolist()
                                son_list = df['Perfil sônico'].tolist()
                                gr_list = df['Perfil Raio Gama'].tolist()
                                gr_list_s = df['Raio Gama Suavizado'].tolist()
                                profundidade = ext_list + prof_list
                                linha_extrapolada = [True] * len(ext_list) + [False] * len(prof_list)
                                densidade = [st.session_state.ds if prof > st.session_state.rtkb else 0
                                             for prof in ext_list
                                             ] + dens_list

                                sonico = [None] * len(ext_list) + son_list
                                gr = [None] * len(ext_list) + gr_list
                                gr_s = [None] * len(ext_list) + gr_list_s

                            else:
                                profundidade = df['Profundidade'].tolist()
                                densidade = df['Perfil de densidade'].tolist()
                                sonico = df['Perfil sônico'].tolist()
                                gr = df['Perfil Raio Gama'].tolist()
                                gr_s = df['Raio Gama Suavizado'].tolist()
                                linha_extrapolada = [False] * len(profundidade)

                            profundidade_solo = []
                            for i in range(len(profundidade)):
                                if st.session_state.ex == 'Ativada':
                                    if profundidade[i] <= st.session_state.rtkb:
                                        profundidade_solo.append(0)
                                    else:
                                        profundidade_solo.append(profundidade[i] - st.session_state.rtkb)
                                else:
                                    profundidade_solo.append(profundidade[i])

                            st.session_state.ext_df = pd.DataFrame({
                                'Profundidade em relação a mesa rotativa (m)': profundidade,
                                'Profundidade em relação ao solo (m)': profundidade_solo,
                                'Densidade (g/cm³)': densidade,
                                'Sônico (µs/pé)': sonico,
                                'Perfil Raio Gama': gr,
                                'Raio Gama Suavizado': gr_s,
                                'Linha Extrapolada': linha_extrapolada
                            })

                            if st.session_state.ex == 'Ativada':
                                st.session_state.ext_df.loc[0, 'Densidade (g/cm³)'] = 0

                            if 'Gardner' in st.session_state.gard:
                                mask = st.session_state.ext_df['Densidade (g/cm³)'].isna()

                                sonico = st.session_state.ext_df.loc[mask, 'Sônico (µs/pé)']

                                st.session_state.ext_df.loc[mask, 'Densidade (g/cm³)'] = (
                                        0.23 * ((1_000_000 / sonico) ** 0.25)
                                )

                            dd = []

                            for i in range(len(st.session_state.ext_df)):
                                if i == 0:
                                    if st.session_state.ex == 'Desativada':
                                        dd.append(st.session_state.ext_df[
                                                      'Profundidade em relação a mesa rotativa (m)'].iloc[0])
                                    else:
                                        dd.append(0)
                                else:
                                    if st.session_state.ex == 'Ativada':
                                        if st.session_state.ext_df[
                                            'Profundidade em relação a mesa rotativa (m)'].iloc[
                                            i] <= st.session_state.rtkb:
                                            dd.append(0)
                                        else:
                                            dd.append(st.session_state.ext_df[
                                                          'Profundidade em relação a mesa rotativa (m)'].iloc[
                                                          i] - st.session_state.ext_df[
                                                          'Profundidade em relação a mesa rotativa (m)'].iloc[
                                                          i - 1])
                                    else:
                                        dd.append(st.session_state.ext_df[
                                                      'Profundidade em relação a mesa rotativa (m)'].iloc[i] -
                                                  st.session_state.ext_df[
                                                      'Profundidade em relação a mesa rotativa (m)'].iloc[
                                                      i - 1])
                            st.session_state.ext_df['ΔD (m)'] = dd

                            dm = []
                            for i in range(len(st.session_state.ext_df)):
                                if st.session_state.ext_df['Profundidade em relação a mesa rotativa (m)'].iloc[
                                    i] <= st.session_state.rtkb:
                                    dm.append(0)
                                else:
                                    dm.append(st.session_state.ext_df['Densidade (g/cm³)'].iloc[i] *
                                              st.session_state.ext_df['ΔD (m)'].iloc[i] * 1.422)

                            st.session_state.ext_df['Densidade (g/cm³) x ΔD x 1,422 (psi)'] = dm

                            st.session_state.ext_df['Pressão de Sobrecarga (psi)'] = st.session_state.ext_df[
                                'Densidade (g/cm³) x ΔD x 1,422 (psi)'].cumsum()

                            gs = []
                            for i in range(len(st.session_state.ext_df)):
                                if st.session_state.ext_df['Profundidade em relação a mesa rotativa (m)'].iloc[
                                    i] < st.session_state.rtkb:
                                    gs.append(0)
                                else:
                                    gs.append(
                                        st.session_state.ext_df['Pressão de Sobrecarga (psi)'].iloc[i] /
                                        (0.1704 * st.session_state.ext_df[
                                            'Profundidade em relação a mesa rotativa (m)'].iloc[i]))

                            st.session_state.ext_df['Gradiente de Sobrecarga (lb/gal)'] = gs
                            st.session_state.df_sobrecarga = st.session_state.ext_df.copy()
                        # FALTA IMPLEMENTAR
                        # MILLER
                        # BOURGOYNE

                    else:
                        st.warning('Não foram encontrados dados válidos de perfilagem')

                    st.form_submit_button("Calcular Gradiente de Sobrecarga", use_container_width=True,
                                          type='primary')

        with c2:
            with st.container(border=True):
                if st.session_state.tipo_poco == "Onshore":
                    image = Image.open("rig.png")
                    st.image(image, width=200)
                    st.markdown(
                        f"""
                        <div style='position: absolute; top: -240px; left: 200px;
                                    padding: 10px; border-radius: 5px;'>
                            <strong>Datum: </strong> {st.session_state.datum} <br> 
                        </div>
                        <div style='position: absolute; top: -197px; left: 200px;
                                    padding: 10px; border-radius: 5px;'>
                            <strong>Elevação DATUM: </strong> {st.session_state.es:.2f} m <br>
                        </div>
                        <div style='position: absolute; top: -170px; left: 200px;
                                    padding: 10px; border-radius: 5px;'>
                            <strong>Air gap: </strong> {st.session_state.rtkb:.2f} m <br>
                        </div>
                        <div style='position: absolute; top: -140px; left: 200px;
                                    padding: 10px; border-radius: 5px;'>
                            <strong>Solo <br>
                        </div>
                        <div style='position: absolute; top: -90px; left: 200px;
                                    padding: 10px; border-radius: 5px;'>
                            <strong>Elevação do solo: </strong> {st.session_state.es - st.session_state.rtkb:.2f} m <br>
                        </div>
                        <div style='position: absolute; top: -43px; left: 200px;
                                    padding: 10px; border-radius: 5px;'>
                            <strong>Nível do Mar </strong>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

                else:
                    # if not st.session_state.bop:
                    image = Image.open("rig_offshore.png")
                    st.image(image, width=200)
                    st.markdown(
                        f"""
                        <div style='position: absolute; top: -240px; left: 200px;
                                    padding: 10px; border-radius: 5px;'>
                            <strong>Datum: </strong> {st.session_state.datum} <br> 
                        </div>
                        <div style='position: absolute; top: -170px; left: 200px;
                                    padding: 10px; border-radius: 5px;'>
                            <strong>Air gap: </strong> {st.session_state.rtkb:.2f} m <br>
                        </div>
                        <div style='position: absolute; top: -90px; left: 200px;
                                    padding: 10px; border-radius: 5px;'>
                            <strong>Lâmina D'água: </strong> {st.session_state.lda:.2f} m <br>
                        </div>
                        <div style='position: absolute; top: -43px; left: 200px;
                                    padding: 10px; border-radius: 5px;'>
                            <strong>Leito Marinho </strong>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

        with c3:
            with st.container(border=True):
                def reset_config():
                    df_calc = st.session_state.get("df_sobrecarga", pd.DataFrame())
                    if st.session_state.ogs == "Gradiente (lb/gal)":
                        st.session_state.x_min_s = 7.0
                        st.session_state.x_max_s = 21.0
                        st.session_state.x_step_s = 2.0
                    else:
                        st.session_state.x_min_s = 0.0
                        if isinstance(df_calc, pd.DataFrame) and "Pressão de Sobrecarga (psi)" in df_calc.columns:
                            st.session_state.x_max_s = float(int(df_calc['Pressão de Sobrecarga (psi)'].max()) + 200)
                        else:
                            st.session_state.x_max_s = 1000.0
                        st.session_state.x_step_s = 500.0

                    st.session_state.y_min_s = 0.0
                    st.session_state.y_max_s = float(int(df['Profundidade'].max()) + 100)
                    st.session_state.y_step_s = 200.0
                    for chave in ("x_min_s", "x_max_s", "x_step_s", "y_min_s", "y_max_s", "y_step_s"):
                        st.session_state[f"_w_{chave}"] = st.session_state[chave]
                with st.expander("****Configurações do Gráfico****", expanded=False):
                    preparar_widget_persistente("ogs", "_w_ogs")
                    st.segmented_control("***Opção de Gráfico***", ['Gradiente (lb/gal)',
                                                                    'Pressão (psi)'],
                                         selection_mode="single",
                                         key='_w_ogs', width="stretch",
                                         on_change=salvar_widget_persistente,
                                         args=("ogs", "_w_ogs"))
                    preparar_widget_persistente("x_min_s", "_w_x_min_s")
                    st.number_input("Eixo X - mínimo", step=0.5, format="%.2f", key="_w_x_min_s",
                                    on_change=salvar_widget_persistente,
                                    args=("x_min_s", "_w_x_min_s"))
                    preparar_widget_persistente("x_max_s", "_w_x_max_s")
                    st.number_input("Eixo X - máximo", step=0.5, format="%.2f", key="_w_x_max_s",
                                    on_change=salvar_widget_persistente,
                                    args=("x_max_s", "_w_x_max_s"))
                    preparar_widget_persistente("x_step_s", "_w_x_step_s")
                    st.number_input("Passo do eixo X", min_value=0.01, step=0.5, format="%.2f", key="_w_x_step_s",
                                    on_change=salvar_widget_persistente,
                                    args=("x_step_s", "_w_x_step_s"))

                    preparar_widget_persistente("y_min_s", "_w_y_min_s")
                    st.number_input("Eixo Y - mínimo", step=100.0, format="%.2f", key="_w_y_min_s",
                                    on_change=salvar_widget_persistente,
                                    args=("y_min_s", "_w_y_min_s"))
                    preparar_widget_persistente("y_max_s", "_w_y_max_s")
                    st.number_input("Eixo Y - máximo", step=100.0, format="%.2f", key="_w_y_max_s",
                                    on_change=salvar_widget_persistente,
                                    args=("y_max_s", "_w_y_max_s"))
                    preparar_widget_persistente("y_step_s", "_w_y_step_s")
                    st.number_input("Passo do eixo Y", min_value=0.01, step=50.0, format="%.2f", key="_w_y_step_s",
                                    on_change=salvar_widget_persistente,
                                    args=("y_step_s", "_w_y_step_s"))
                    st.button("Resetar Configurações Gráficas - Gradiente de Sobrecarga", on_click=reset_config,
                              type="primary", use_container_width=True)
                df_calc_plot = st.session_state.get("df_sobrecarga", pd.DataFrame())
                if not isinstance(df_calc_plot, pd.DataFrame) or df_calc_plot.empty:
                    st.info("Calcule o Gradiente de Sobrecarga para gerar o grÃ¡fico.")
                    return

                col_prof_plot = (
                    "Profundidade"
                    if "Profundidade" in df_calc_plot.columns
                    else _buscar_coluna_por_partes(df_calc_plot, "Profundidade", "mesa rotativa")
                )
                if col_prof_plot is None:
                    st.info("Calcule o Gradiente de Sobrecarga para gerar o grÃ¡fico.")
                    return

                if st.session_state.ogs == "Gradiente (lb/gal)":
                    if "Gradiente de Sobrecarga (lb/gal)" not in df_calc_plot.columns:
                        st.info("Calcule o Gradiente de Sobrecarga para gerar o grÃ¡fico.")
                        return
                    if st.session_state.tipo_poco == "Onshore":
                        st.session_state.oes = df_calc_plot['Gradiente de Sobrecarga (lb/gal)']
                        st.session_state.profs = df_calc_plot[col_prof_plot]
                    else:
                        st.session_state.oes = df_calc_plot['Gradiente de Sobrecarga (lb/gal)']
                        st.session_state.profs = df_calc_plot[col_prof_plot]
                    st.session_state.oesl = "G. de Sobrecarga"

                else:
                    if "Pressão de Sobrecarga (psi)" not in df_calc_plot.columns:
                        st.info("Calcule o Gradiente de Sobrecarga para gerar o grÃ¡fico.")
                        return
                    st.session_state.oes = df_calc_plot['Pressão de Sobrecarga (psi)']
                    st.session_state.profs = df_calc_plot[col_prof_plot]
                    st.session_state.oesl = "P. de Sobrecarga"

                # Ajuste da figura com coluna de idade + coluna litológica
                st.session_state.fig_gs = plt.figure(figsize=(8, 10))

                usar_coluna_idade = (
                        st.session_state.get("idg") == "Sim"
                        and "df_idade" in st.session_state
                        and isinstance(st.session_state.df_idade, pd.DataFrame)
                        and not st.session_state.df_idade.empty
                )

                if usar_coluna_idade:
                    # COM coluna de idade
                    gs = gridspec.GridSpec(
                        1, 4,
                        width_ratios=[0.10, 0.18, 0.21, 1],
                        wspace=0
                    )

                    ax_idade = st.session_state.fig_gs.add_subplot(gs[0])
                    ax1 = st.session_state.fig_gs.add_subplot(gs[1], sharey=ax_idade)

                    ax_gap = st.session_state.fig_gs.add_subplot(gs[2])
                    ax_gap.axis("off")

                    ax = st.session_state.fig_gs.add_subplot(gs[3], sharey=ax_idade)

                    idade_formacao(
                        ax_idade,
                        st.session_state.df_idade,
                        st.session_state.y_max_s,
                        st.session_state.y_min_s
                    )

                    ax_idade.tick_params(
                        axis="y",
                        which="both",
                        left=False,
                        right=False,
                        labelleft=False,
                        labelright=False
                    )

                    ax_idade.set_ylabel("")

                    plt.setp(ax1.get_yticklabels(), visible=False)
                    plt.setp(ax.get_yticklabels(), visible=False)

                else:
                    # SEM coluna de idade
                    gs = gridspec.GridSpec(
                        1, 3,
                        width_ratios=[0.18, 0.21, 1],
                        wspace=0
                    )

                    ax1 = st.session_state.fig_gs.add_subplot(gs[0])

                    ax_gap = st.session_state.fig_gs.add_subplot(gs[1])
                    ax_gap.axis("off")

                    ax = st.session_state.fig_gs.add_subplot(gs[2], sharey=ax1)

                    plt.setp(ax.get_yticklabels(), visible=False)

                # DataFrame usado pela coluna litológica
                # Se o usuário escolheu "Permeável / Não permeável", tenta usar o df_pp salvo.
                # Se ainda não existir LBF_calc, cai automaticamente para a litologia do Excel.
                df_lito_gs = df.copy()

                if (
                        st.session_state.get(
                            "tipo_coluna_litologica_graficos",
                            "Permeável / Não permeável"
                        ) == "Permeável / Não permeável"
                        and "df_pp_lito" in st.session_state
                        and isinstance(st.session_state.df_pp_lito, pd.DataFrame)
                        and not st.session_state.df_pp_lito.empty
                ):
                    df_lito_gs = st.session_state.df_pp_lito.copy()

                # Garante compatibilidade caso o dataframe usado seja o df original
                if "Profundidade (m)" not in df_lito_gs.columns and "Profundidade" in df_lito_gs.columns:
                    df_lito_gs["Profundidade (m)"] = df_lito_gs["Profundidade"]

                lito(
                    ax1,
                    df_lito_gs,
                    profundidades,
                    litologias,
                    st.session_state.y_max_s
                )

                if st.session_state.tipo_poco == "Onshore":
                    if st.session_state.rtkb and st.session_state.es != 0:
                        if st.session_state.ex == 'Desativada':
                            ax.plot(st.session_state.oes, st.session_state.profs,
                                    color='black', linestyle='-', linewidth=2, label=st.session_state.oesl)
                        else:
                            ax.plot(st.session_state.oes,
                                    st.session_state.ext_df['Profundidade em relação a mesa rotativa (m)'],
                                    color='black', linestyle='-', linewidth=2, label=st.session_state.oesl)

                else:
                    if st.session_state.rtkb and st.session_state.lda != 0:
                        ax.plot(st.session_state.oes, st.session_state.profs,
                                color='black', linestyle='-', linewidth=2, label=st.session_state.oesl)

                if st.session_state.ogs == "Gradiente (lb/gal)":
                    ax.set_title('Gradiente de Sobrecarga (lb/gal)', fontsize=14, fontweight='bold')
                    ax.set_xlabel('Gradiente (ppg)', fontsize=12)
                    ax.set_ylabel('Profundidade em relação à mesa rotativa', fontsize=12)
                else:
                    ax.set_title('Pressão de Sobrecarga (psi)', fontsize=14, fontweight='bold')
                    ax.set_xlabel('Pressão (psi)', fontsize=12)
                    ax.set_ylabel('Profundidade em relação à mesa rotativa', fontsize=12)
                ax.set_ylabel('Profundidade em relação à mesa rotativa', fontsize=12)
                ax.invert_yaxis()
                max_depth = int(df['Profundidade'].max()) + 100
                margin = 20
                y_min_s = float(st.session_state.y_min_s)
                y_max_s = float(st.session_state.y_max_s)
                y_step_s = max(0.01, float(st.session_state.y_step_s))
                x_min_s = float(st.session_state.x_min_s)
                x_max_s = float(st.session_state.x_max_s)
                x_step_s = max(0.01, float(st.session_state.x_step_s))
                ax.set_yticks(
                    np.arange(y_min_s, y_max_s + y_step_s * 0.5, y_step_s))
                ax.set_ylim(y_max_s, y_min_s)
                x_tick_ini_s = math.ceil(x_min_s)
                x_tick_fim_s = math.floor(x_max_s)
                x_tick_step_s = max(1.0, x_step_s)
                if x_tick_fim_s >= x_tick_ini_s:
                    ax.set_xticks(
                        np.arange(x_tick_ini_s, x_tick_fim_s + x_tick_step_s * 0.5, x_tick_step_s))
                if np.isclose(x_step_s, 1.0):
                    ax.set_xticks(
                        np.arange(x_min_s, x_max_s + 0.25, 0.5),
                        minor=True)
                ax.xaxis.set_major_formatter(FuncFormatter(lambda valor, pos: f"{valor:.0f}"))
                ax.set_xlim(x_min_s, x_max_s)
                ax.tick_params(axis='y', which='both', left=True, labelleft=True)
                ax.grid(True, which='major', linestyle='--', alpha=0.5)
                if np.isclose(x_step_s, 1.0):
                    ax.grid(True, which='minor', axis='x', linestyle='--', alpha=0.5)
                ax.legend(
                    loc='upper right',
                    fontsize=8,
                    frameon=True,
                    shadow=True,
                    fancybox=True,
                    framealpha=1,
                    facecolor='white',
                    edgecolor='gray'
                )
                add_watermark(
                    ax,
                    logo_path="logo2.png",
                    xy=(0.50, 0.5),
                    zoom=0.2,
                    alpha=0.2,
                    zorder=0
                )

                st.pyplot(st.session_state.fig_gs)

    with tab_tabela:
        st.subheader("Tabela de dados calculados")
        df_sobrecarga_calc = st.session_state.get("df_sobrecarga", pd.DataFrame())
        if not isinstance(df_sobrecarga_calc, pd.DataFrame) or df_sobrecarga_calc.empty:
            st.warning("Calcule o Gradiente de Sobrecarga na aba Dados antes de visualizar a tabela.")
            return

        try:
            st.dataframe(df_sobrecarga_calc, use_container_width=True, hide_index=True)

        except Exception as e:
            pass


def pagina_poros():
    st.header("Gradiente de Pressão de Poros")

    tab_dados, tab_tabela = st.tabs(["Dados", "Tabela de dados calculados"])

    with tab_dados:
        st.subheader("Dados")
        df_sobrecarga_calc = st.session_state.get("df_sobrecarga", pd.DataFrame())

        if not isinstance(df_sobrecarga_calc, pd.DataFrame) or df_sobrecarga_calc.empty:
            st.error('Calcule o Gradiente de Sobrecarga antes de calcular Pressão de Poros.', icon="🚨")
            return

        df_pp = montar_df_pp_base(df_sobrecarga_calc)

        if "ogp" not in st.session_state:
            st.session_state.ogp = "Gradiente (lb/gal)"

        if "x_min_pp" not in st.session_state:
            if st.session_state.get("ogp", "Gradiente (lb/gal)") == "Pressão (psi)":
                st.session_state.x_min_pp = 0
            else:
                st.session_state.x_min_pp = 7

        if "x_max_pp" not in st.session_state:
            if st.session_state.get("ogp", "Gradiente (lb/gal)") == "Pressão (psi)":
                max_x_pp = pd.to_numeric(
                    df_pp.get("Pressão de Poros (psi)", pd.Series(dtype=float)),
                    errors="coerce"
                ).max()

                st.session_state.x_max_pp = int(max_x_pp) + 500 if pd.notna(max_x_pp) else 5000
            else:
                st.session_state.x_max_pp = 21

        if "x_step_pp" not in st.session_state:
            if st.session_state.get("ogp", "Gradiente (lb/gal)") == "Pressão (psi)":
                st.session_state.x_step_pp = 500
            else:
                st.session_state.x_step_pp = 2

        if "y_min_pp" not in st.session_state:
            st.session_state.y_min_pp = 0

        if "y_max_pp" not in st.session_state:
            st.session_state.y_max_pp = int(df_pp["Profundidade (m)"].max()) + 100

        if "y_step_pp" not in st.session_state:
            st.session_state.y_step_pp = 200

        if "n_trending" not in st.session_state:
            st.session_state.n_trending = 1

        if "n_lbf" not in st.session_state:
            st.session_state.n_lbf = 1

        if "n_boyance" not in st.session_state:
            st.session_state.n_boyance = 1

        padrao_fluido_retro = (
            "Sim"
            if st.session_state.get("option") == "Retroanálise"
            else "Não"
        )

        for chave, valor_padrao in (
                ("graf", "Trending"),
                ("spp", "Não"),
                ("grafpp", "Não"),
                ("suav_s", "Não"),
                ("s_gr", "Não"),
                ("fpl", padrao_fluido_retro),
                ("fex", padrao_fluido_retro),
                ("janela_spp", 20),
                ("limite_spp", 0.10),
                ("boyance", "Não"),
                ("o_boyance", ["Base Aren. = Topo Folh."]),
        ):
            if chave not in st.session_state:
                st.session_state[chave] = valor_padrao

        def reset_config_pp():
            st.session_state.x_min_pp = 0

            if st.session_state.get("ogp", "Gradiente (lb/gal)") == "Pressão (psi)":
                max_x_pp = pd.to_numeric(
                    df_pp.get("Pressão de Poros (psi)", pd.Series(dtype=float)),
                    errors="coerce"
                ).max()

                st.session_state.x_max_pp = int(max_x_pp) + 500 if pd.notna(max_x_pp) else 5000
                st.session_state.x_step_pp = 500

            else:
                max_x_pp = pd.to_numeric(
                    df_pp.get("Gradiente de Sobrecarga (lb/gal)", pd.Series(dtype=float)),
                    errors="coerce"
                ).max()

                st.session_state.x_max_pp = 21
                st.session_state.x_step_pp = 2

            st.session_state.y_min_pp = 0
            st.session_state.y_max_pp = int(df_pp["Profundidade (m)"].max()) + 100
            st.session_state.y_step_pp = 200

            for chave in (
                    "x_min_pp",
                    "x_max_pp",
                    "x_step_pp",
                    "y_min_pp",
                    "y_max_pp",
                    "y_step_pp"
            ):
                st.session_state[f"_w_{chave}"] = st.session_state[chave]

        def chaves_form_poros():
            chaves = ["expoente", "anormal", "gn"]

            n_lbf_form = int(st.session_state.get("n_lbf", 1))
            n_trending_form = int(st.session_state.get("n_trending", 1))

            for i in range(n_lbf_form):
                if n_lbf_form > 1:
                    chaves.extend([
                        f"lbf_prof_ini_{i}",
                        f"lbf_prof_fim_{i}",
                    ])

                chaves.extend([
                    f"lbf_valor_{i}",
                    f"lbf_inclinacao_{i}",
                ])

            for i in range(n_trending_form):
                if n_trending_form > 1:
                    chaves.extend([
                        f"trend_prof_ini_{i}",
                        f"trend_prof_fim_{i}",
                    ])

                chaves.extend([
                    f"trend_pp1_{i}",
                    f"trend_pp2_{i}",
                    f"trend_s1_{i}",
                    f"trend_s2_{i}",
                ])

            return chaves

        def restaurar_widgets_form_poros():
            for chave in chaves_form_poros():
                chave_widget = f"_w_{chave}"

                if chave in st.session_state:
                    st.session_state[chave_widget] = st.session_state[chave]

                elif chave_widget in st.session_state:
                    del st.session_state[chave_widget]

        def marcar_submit_poros(pares):
            sincronizar_widgets_persistentes(pares)
            st.session_state["_submit_poros_agora"] = True

        if not st.session_state.get("_submit_poros_agora", False):
            restaurar_widgets_form_poros()

        c1, c2, c3 = st.columns((1, 1, 1))
        with c1:
            with st.container(border=True):
                col_add_tr, col_rem_tr, col_add_lbf, col_rem_lbf = st.columns(4)

                with col_add_tr:
                    if st.button(
                            "➕ Trending",
                            type="primary",
                            use_container_width=True,
                            key="b_add_trending"
                    ):
                        st.session_state.n_trending += 1

                with col_rem_tr:
                    if st.button(
                            "➖ Trending",
                            type="primary",
                            use_container_width=True,
                            key="b_rem_trending"
                    ):
                        if st.session_state.n_trending > 1:
                            idx = st.session_state.n_trending - 1

                            keys_to_remove = [
                                f"trend_pp1_{idx}",
                                f"_w_trend_pp1_{idx}",
                                f"trend_pp2_{idx}",
                                f"_w_trend_pp2_{idx}",
                                f"trend_s1_{idx}",
                                f"_w_trend_s1_{idx}",
                                f"trend_s2_{idx}",
                                f"_w_trend_s2_{idx}",
                                f"trend_prof_ini_{idx}",
                                f"_w_trend_prof_ini_{idx}",
                                f"trend_prof_fim_{idx}",
                                f"_w_trend_prof_fim_{idx}",
                            ]

                            for k in keys_to_remove:
                                if k in st.session_state:
                                    del st.session_state[k]

                            st.session_state.n_trending -= 1

                with col_add_lbf:
                    if st.button(
                            "➕ LBF",
                            type="primary",
                            use_container_width=True,
                            key="b_add_lbf"
                    ):
                        st.session_state.n_lbf += 1

                with col_rem_lbf:
                    if st.button(
                            "➖ LBF",
                            type="primary",
                            use_container_width=True,
                            key="b_rem_lbf"
                    ):
                        if st.session_state.n_lbf > 1:
                            idx = st.session_state.n_lbf - 1

                            _remover_chaves_lbf(idx)

                            st.session_state.n_lbf -= 1

                with st.form("p_poros", border=False):
                    pares_form_poros = []

                    def preparar_widget_form_poros(chave_estado):
                        chave_widget = f"_w_{chave_estado}"
                        preparar_widget_persistente(chave_estado, chave_widget)
                        pares_form_poros.append((chave_estado, chave_widget))
                        return chave_widget

                    with st.expander("Informações Gerais", expanded=True):
                        st.number_input(
                            "Expoente de Eaton",
                            step=1.0,
                            format="%.2f",
                            key=preparar_widget_form_poros("expoente"),
                        )

                        st.number_input(
                            "Profundidade da zona normal (m)",
                            step=100.0,
                            format="%.2f",
                            key=preparar_widget_form_poros("anormal"),
                        )

                        st.number_input(
                            "Gradiente Normal (lb/gal)",
                            step=1.0,
                            format="%.2f",
                            key=preparar_widget_form_poros("gn"),
                            value=8.5
                        )

                    st.markdown("### Trendings")

                    for i in range(st.session_state.n_trending):
                        with st.expander(f"Trending {i + 1}", expanded=True):
                            if st.session_state.n_trending > 1:
                                colun1, colun2 = st.columns(2)

                                with colun1:
                                    chave_trend_prof_ini = f"trend_prof_ini_{i}"
                                    if chave_trend_prof_ini not in st.session_state:
                                        st.session_state[chave_trend_prof_ini] = 0.0
                                    st.number_input(
                                        "Profundidade inicial",
                                        step=1.0,
                                        format="%.2f",
                                        min_value=0.0,
                                        key=preparar_widget_form_poros(chave_trend_prof_ini)
                                    )

                                with colun2:
                                    chave_trend_prof_fim = f"trend_prof_fim_{i}"
                                    if chave_trend_prof_fim not in st.session_state:
                                        st.session_state[chave_trend_prof_fim] = 0.0
                                    st.number_input(
                                        "Profundidade final",
                                        step=1.0,
                                        format="%.2f",
                                        min_value=0.0,
                                        key=preparar_widget_form_poros(chave_trend_prof_fim)
                                    )

                            col_tr1, col_tr2 = st.columns(2)

                            with col_tr1:
                                chave_trend_pp1 = f"trend_pp1_{i}"
                                if chave_trend_pp1 not in st.session_state:
                                    st.session_state[chave_trend_pp1] = 400.0
                                st.number_input(
                                    "Profundidade 1",
                                    step=1.0,
                                    format="%.2f",
                                    min_value=0.0,
                                    value=400.0,
                                    key=preparar_widget_form_poros(chave_trend_pp1)
                                )

                                chave_trend_pp2 = f"trend_pp2_{i}"
                                if chave_trend_pp2 not in st.session_state:
                                    st.session_state[chave_trend_pp2] = 1000.0
                                st.number_input(
                                    "Profundidade 2",
                                    step=1.0,
                                    format="%.2f",
                                    min_value=0.0,
                                    value=1000.0,
                                    key=preparar_widget_form_poros(chave_trend_pp2)
                                )

                            with col_tr2:
                                chave_trend_s1 = f"trend_s1_{i}"
                                if chave_trend_s1 not in st.session_state:
                                    st.session_state[chave_trend_s1] = 110.0
                                st.number_input(
                                    "Leitura 1 do Sônico",
                                    step=1.0,
                                    format="%.2f",
                                    min_value=0.0,
                                    value=110.0,
                                    key=preparar_widget_form_poros(chave_trend_s1)
                                )

                                chave_trend_s2 = f"trend_s2_{i}"
                                if chave_trend_s2 not in st.session_state:
                                    st.session_state[chave_trend_s2] = 85.0
                                st.number_input(
                                    "Leitura 2 do Sônico",
                                    step=1.0,
                                    format="%.2f",
                                    min_value=0.0,
                                    value=85.0,
                                    key=preparar_widget_form_poros(chave_trend_s2)
                                )

                    _renderizar_campos_lbf(preparar_widget_form_poros)

                    calcular_poros_agora = st.form_submit_button(
                        "Calcular Gradiente de Pressão de Poros",
                        type="primary",
                        use_container_width=True,
                        on_click=marcar_submit_poros,
                        args=(pares_form_poros,)
                    )

                if calcular_poros_agora:
                    sincronizar_widgets_persistentes(pares_form_poros)

                    trendings = _coletar_trendings_poros()
                    lbfs = _coletar_lbfs_poros()

                    df_pp_calc = montar_df_pp_base(df_sobrecarga_calc)
                    df_pp_calc = _calcular_trending_pp(df_pp_calc, trendings)
                    df_pp_calc = _calcular_lbf_pp(df_pp_calc, lbfs)
                    df_pp_calc = _calcular_pressao_poros_por_partes(df_pp_calc)

                    st.session_state.df_pp = df_pp_calc.copy()
                    st.session_state.df_pp_lito = df_pp_calc.copy()
                    df_pp = st.session_state.df_pp.copy()

                else:
                    trendings = _coletar_trendings_poros()
                    lbfs = _coletar_lbfs_poros()
                    df_pp = st.session_state.get("df_pp", df_pp)
                    
        with c2:
            with st.container(border=True):
                preparar_widget_persistente("graf", "_w_graf")
                st.segmented_control(
                    "Gráficos",
                    ["LBF", "Trending"],
                    selection_mode="single",
                    default="Trending",
                    key="_w_graf",
                    width="stretch",
                    on_change=salvar_widget_persistente,
                    args=("graf", "_w_graf"),
                )

                df_gr_poros = df_pp

                if isinstance(df_sobrecarga_calc, pd.DataFrame) and not df_sobrecarga_calc.empty:
                    try:
                        df_gr_poros = montar_df_pp_base(df_sobrecarga_calc)
                        df_gr_poros = _calcular_lbf_pp(df_gr_poros, lbfs)
                        df_gr_poros = _calcular_trending_pp(df_gr_poros, trendings)
                        st.session_state.df_pp_lito = df_gr_poros.copy()

                    except Exception:
                        df_gr_poros = df_pp

                _plotar_gr_poros_com_contexto(
                    df_gr_poros,
                    st.session_state.get("graf", "Trending"),
                    lbfs=lbfs,
                    trendings=trendings
                )

        with c3:
            with st.container(border=True):
                st.write("")
                with st.expander("****Configurações do Gráfico****", expanded=False):
                    with st.expander("***Curvas e Suavizações***", expanded=False):
                        colu1, colu2 = st.columns(2)
                        opcoes_sim_nao = ["Sim", "Não"]
                        with colu1:
                            preparar_widget_persistente("grafpp", "_w_grafpp")
                            st.selectbox(
                                "Gradiente de Sobrecarga",
                                ["Sim", "Não"],
                                index=1,
                                key="_w_grafpp",
                                on_change=salvar_widget_persistente,
                                args=("grafpp", "_w_grafpp")
                            )
                            preparar_widget_persistente("spp", "_w_spp")
                            st.selectbox(
                                "Suavizar Pressão de Poros",
                                ["Sim", "Não"],
                                index=1,
                                key="_w_spp",
                                on_change=salvar_widget_persistente,
                                args=("spp", "_w_spp")
                            )

                            preparar_widget_persistente("janela_spp", "_w_janela_spp")
                            st.number_input(
                                "Média móvel da pressão de poros",
                                min_value=1,
                                step=1,
                                format="%d",
                                key="_w_janela_spp",
                                on_change=salvar_widget_persistente,
                                args=("janela_spp", "_w_janela_spp"),
                                help=(
                                    "Define quantos elementos acima e abaixo entram na média móvel.\n\n"
                                    "Exemplo: 20 significa média de x-20 até x+20."
                                )
                            )

                            if st.session_state.get("fpl") not in opcoes_sim_nao:
                                st.session_state.fpl = padrao_fluido_retro
                            preparar_widget_persistente("fpl", "_w_fpl")
                            st.selectbox(
                                "Visualizar peso do fluido planejado",
                                opcoes_sim_nao,
                                key="_w_fpl",
                                on_change=salvar_widget_persistente,
                                args=("fpl", "_w_fpl")
                            )

                        with colu2:
                            if st.session_state.get("suav_s") not in opcoes_sim_nao:
                                st.session_state.suav_s = "Não"
                            preparar_widget_persistente("suav_s", "_w_suav_s")
                            st.selectbox(
                                "Suavizar Sônico",
                                opcoes_sim_nao,
                                key="_w_suav_s",
                                on_change=salvar_widget_persistente,
                                args=("suav_s", "_w_suav_s")
                            )

                            if st.session_state.get("s_gr") not in opcoes_sim_nao:
                                st.session_state.s_gr = "Não"
                            preparar_widget_persistente("s_gr", "_w_s_gr")
                            st.selectbox(
                                "Suavizar Raio Gama",
                                opcoes_sim_nao,
                                key="_w_s_gr",
                                on_change=salvar_widget_persistente,
                                args=("s_gr", "_w_s_gr")
                            )

                            preparar_widget_persistente("limite_spp", "_w_limite_spp")
                            st.number_input(
                                "Filtro da pressão de poros",
                                min_value=0.0001,
                                step=0.01,
                                format="%f",
                                key="_w_limite_spp",
                                on_change=salvar_widget_persistente,
                                args=("limite_spp", "_w_limite_spp"),
                                help=(
                                    "Limita o quanto o próximo ponto suavizado pode subir em relação ao anterior.\n\n"
                                    "Exemplo: 0,10 significa que, se x+1 > x+0,10, então x+1 recebe x+0,10."
                                )
                            )

                            if st.session_state.get("fex") not in opcoes_sim_nao:
                                st.session_state.fex = padrao_fluido_retro
                            preparar_widget_persistente("fex", "_w_fex")
                            st.selectbox(
                                "Visualizar peso do fluido executado",
                                opcoes_sim_nao,
                                key="_w_fex",
                                disabled=(
                                    st.session_state.get("option") != "Retroanálise"
                                ),
                                on_change=salvar_widget_persistente,
                                args=("fex", "_w_fex")
                            )

                    with st.expander("***Boyance***", expanded=False):
                        if st.session_state.get("boyance") not in ["Não", "Sim"]:
                            st.session_state.boyance = "Não"

                        preparar_widget_persistente("boyance", "_w_boyance")
                        st.selectbox(
                            "Boyance Ativado",
                            ["Não", "Sim"],
                            key="_w_boyance",
                            on_change=salvar_widget_persistente,
                            args=("boyance", "_w_boyance")
                        )

                        if st.session_state.get("boyance", "Não") == "Sim":
                            col_add_boyance, col_rem_boyance = st.columns(2)

                            with col_add_boyance:
                                if st.button(
                                    "Adicionar Intervalo Boyance",
                                    type="primary",
                                    use_container_width=True,
                                    key="b_add_boyance"
                                ):
                                    st.session_state.n_boyance += 1

                            with col_rem_boyance:
                                if st.button(
                                    "Remover Intervalo Boyance",
                                    type="primary",
                                    use_container_width=True,
                                    key="b_rem_boyance"
                                ):
                                    if st.session_state.n_boyance > 1:
                                        idx = st.session_state.n_boyance - 1

                                        keys_to_remove = [
                                            f"fpr_{idx}",
                                            f"_w_fpr_{idx}",
                                            f"prof_inicial_{idx}",
                                            f"_w_prof_inicial_{idx}",
                                            f"prof_final_{idx}",
                                            f"_w_prof_final_{idx}",
                                        ]

                                        for k in keys_to_remove:
                                            if k in st.session_state:
                                                del st.session_state[k]

                                        st.session_state.n_boyance -= 1

                            if not isinstance(st.session_state.get("o_boyance"), list):
                                st.session_state.o_boyance = ["Base Aren. = Topo Folh."]

                            preparar_widget_persistente("o_boyance", "_w_o_boyance")
                            st.segmented_control(
                                "***Pressão***",
                                [
                                    "Base Aren. = Topo Folh.",
                                    "Topo Aren. = Base Folh."
                                ],
                                selection_mode="multi",
                                key="_w_o_boyance",
                                width="stretch",
                                on_change=salvar_widget_persistente,
                                args=("o_boyance", "_w_o_boyance")
                            )

                            for i in range(st.session_state.n_boyance):
                                with st.expander(f"Boyance - Intervalo {i + 1}", expanded=True):
                                    if st.session_state.n_boyance > 1:
                                        colun1, colun2 = st.columns(2)

                                        with colun1:
                                            chave_prof_inicial = f"prof_inicial_{i}"
                                            chave_w_prof_inicial = f"_w_prof_inicial_{i}"
                                            if chave_prof_inicial not in st.session_state:
                                                st.session_state[chave_prof_inicial] = 0.0
                                            preparar_widget_persistente(chave_prof_inicial, chave_w_prof_inicial)
                                            st.number_input(
                                                "Profundidade inicial",
                                                step=1.0,
                                                format="%.2f",
                                                min_value=0.0,
                                                key=chave_w_prof_inicial,
                                                on_change=salvar_widget_persistente,
                                                args=(chave_prof_inicial, chave_w_prof_inicial)
                                            )

                                        with colun2:
                                            chave_prof_final = f"prof_final_{i}"
                                            chave_w_prof_final = f"_w_prof_final_{i}"
                                            if chave_prof_final not in st.session_state:
                                                st.session_state[chave_prof_final] = 0.0
                                            preparar_widget_persistente(chave_prof_final, chave_w_prof_final)
                                            st.number_input(
                                                "Profundidade final",
                                                step=1.0,
                                                format="%.2f",
                                                min_value=0.0,
                                                key=chave_w_prof_final,
                                                on_change=salvar_widget_persistente,
                                                args=(chave_prof_final, chave_w_prof_final)
                                            )

                                    chave_fpr = f"fpr_{i}"
                                    chave_w_fpr = f"_w_fpr_{i}"
                                    if chave_fpr not in st.session_state:
                                        st.session_state[chave_fpr] = 8.5
                                    preparar_widget_persistente(chave_fpr, chave_w_fpr)
                                    st.number_input(
                                        "Peso do Fluido Contido nos Poros das Rochas",
                                        min_value=1.0,
                                        step=0.5,
                                        format="%.2f",
                                        value=8.5,
                                        key=chave_w_fpr,
                                        on_change=salvar_widget_persistente,
                                        args=(chave_fpr, chave_w_fpr)
                                    )


                    with st.expander("***Testes de Formação***", expanded=False):
                        opcoes_sim_nao_testes = ["Sim", "Não"]
                        col_txt_rft, col_txt_colapso = st.columns(2)

                        with col_txt_rft:
                            if st.session_state.get("mostrar_texto_rft_pp") not in opcoes_sim_nao_testes:
                                st.session_state.mostrar_texto_rft_pp = "Sim"
                            preparar_widget_persistente("mostrar_texto_rft_pp", "_w_mostrar_texto_rft_pp")
                            st.selectbox(
                                "Exibir texto RFT",
                                opcoes_sim_nao_testes,
                                key="_w_mostrar_texto_rft_pp",
                                on_change=salvar_widget_persistente,
                                args=("mostrar_texto_rft_pp", "_w_mostrar_texto_rft_pp")
                            )

                        with col_txt_colapso:
                            if st.session_state.get("mostrar_texto_colapso_pp") not in opcoes_sim_nao_testes:
                                st.session_state.mostrar_texto_colapso_pp = "Sim"
                            preparar_widget_persistente("mostrar_texto_colapso_pp", "_w_mostrar_texto_colapso_pp")
                            st.selectbox(
                                "Exibir texto gradiente de colapso",
                                opcoes_sim_nao_testes,
                                key="_w_mostrar_texto_colapso_pp",
                                on_change=salvar_widget_persistente,
                                args=("mostrar_texto_colapso_pp", "_w_mostrar_texto_colapso_pp")
                            )

                        st.markdown("##### Pontos de RFT")

                        if (
                            "rft_pontos_pp" not in st.session_state
                            or not isinstance(st.session_state.rft_pontos_pp, pd.DataFrame)
                        ):
                            st.session_state.rft_pontos_pp = pd.DataFrame({
                                "Profundidade (m)": [],
                                "Teste RFT (lb/gal)": []
                            })

                        rft_pontos_editado = st.data_editor(
                            st.session_state.rft_pontos_pp,
                            num_rows="dynamic",
                            use_container_width=True,
                            hide_index=True,
                            key="editor_rft_pontos_pp",
                            column_config={
                                "Profundidade (m)": st.column_config.NumberColumn(
                                    "Profundidade (m)",
                                    min_value=0.0,
                                    format="%.2f"
                                ),
                                "Teste RFT (lb/gal)": st.column_config.NumberColumn(
                                    "Teste RFT (lb/gal)",
                                    min_value=0.0,
                                    format="%.4f"
                                ),
                            }
                        )

                        col_rft1, col_rft2 = st.columns(2)

                        with col_rft1:
                            if st.button(
                                "Atualizar pontos RFT",
                                use_container_width=True,
                                type="primary",
                                key="btn_atualizar_rft_pp"
                            ):
                                st.session_state.rft_pontos_pp = rft_pontos_editado.copy()

                        with col_rft2:
                            if st.button(
                                "Limpar pontos RFT",
                                use_container_width=True,
                                type="primary",
                                key="btn_limpar_rft_pp"
                            ):
                                st.session_state.rft_pontos_pp = pd.DataFrame({
                                    "Profundidade (m)": [],
                                    "Teste RFT (lb/gal)": []
                                })
                                st.rerun()

                        st.markdown("##### Gradiente de Colapso")

                        if (
                            "colapso_pontos_pp" not in st.session_state
                            or not isinstance(st.session_state.colapso_pontos_pp, pd.DataFrame)
                        ):
                            st.session_state.colapso_pontos_pp = pd.DataFrame({
                                "Profundidade (m)": [],
                                "Gradiente de Colapso (lb/gal)": []
                            })

                        colapso_pontos_editado = st.data_editor(
                            st.session_state.colapso_pontos_pp,
                            num_rows="dynamic",
                            use_container_width=True,
                            hide_index=True,
                            key="editor_colapso_pontos_pp",
                            column_config={
                                "Profundidade (m)": st.column_config.NumberColumn(
                                    "Profundidade (m)",
                                    min_value=0.0,
                                    format="%.2f"
                                ),
                                "Gradiente de Colapso (lb/gal)": st.column_config.NumberColumn(
                                    "Gradiente de Colapso (lb/gal)",
                                    min_value=0.0,
                                    format="%.4f"
                                ),
                            }
                        )

                        col_colapso1, col_colapso2 = st.columns(2)

                        with col_colapso1:
                            if st.button(
                                "Atualizar gradiente de colapso",
                                use_container_width=True,
                                type="primary",
                                key="btn_atualizar_colapso_pp"
                            ):
                                st.session_state.colapso_pontos_pp = colapso_pontos_editado.copy()

                        with col_colapso2:
                            if st.button(
                                "Limpar gradiente de colapso",
                                use_container_width=True,
                                type="primary",
                                key="btn_limpar_colapso_pp"
                            ):
                                st.session_state.colapso_pontos_pp = pd.DataFrame({
                                    "Profundidade (m)": [],
                                    "Gradiente de Colapso (lb/gal)": []
                                })
                                st.rerun()

                    with st.expander("***Eixos***", expanded=False):
                        preparar_widget_persistente("ogp", "_w_ogp")
                        st.segmented_control(
                            "Opção de Gráfico",
                            ["Gradiente (lb/gal)", "Pressão (psi)"],
                            selection_mode="single",
                            key="_w_ogp",
                            width="stretch",
                            on_change=salvar_widget_persistente,
                            args=("ogp", "_w_ogp")
                        )

                        preparar_widget_persistente("x_min_pp", "_w_x_min_pp")
                        st.number_input(
                            "Eixo X - mínimo",
                            step=0.5,
                            format="%.2f",
                            key="_w_x_min_pp",
                            on_change=salvar_widget_persistente,
                            args=("x_min_pp", "_w_x_min_pp")
                        )

                        preparar_widget_persistente("x_max_pp", "_w_x_max_pp")
                        st.number_input(
                            "Eixo X - máximo",
                            step=0.5,
                            format="%.2f",
                            key="_w_x_max_pp",
                            on_change=salvar_widget_persistente,
                            args=("x_max_pp", "_w_x_max_pp")
                        )

                        preparar_widget_persistente("x_step_pp", "_w_x_step_pp")
                        st.number_input(
                            "Passo do eixo X",
                            min_value=0.01,
                            step=0.5,
                            format="%.2f",
                            key="_w_x_step_pp",
                            on_change=salvar_widget_persistente,
                            args=("x_step_pp", "_w_x_step_pp")
                        )

                        preparar_widget_persistente("y_min_pp", "_w_y_min_pp")
                        st.number_input(
                            "Eixo Y - mínimo",
                            step=100.0,
                            format="%.2f",
                            key="_w_y_min_pp",
                            on_change=salvar_widget_persistente,
                            args=("y_min_pp", "_w_y_min_pp")
                        )

                        preparar_widget_persistente("y_max_pp", "_w_y_max_pp")
                        st.number_input(
                            "Eixo Y - máximo",
                            step=100.0,
                            format="%.2f",
                            key="_w_y_max_pp",
                            on_change=salvar_widget_persistente,
                            args=("y_max_pp", "_w_y_max_pp")
                        )

                        preparar_widget_persistente("y_step_pp", "_w_y_step_pp")
                        st.number_input(
                            "Passo do eixo Y",
                            min_value=0.01,
                            step=50.0,
                            format="%.2f",
                            key="_w_y_step_pp",
                            on_change=salvar_widget_persistente,
                            args=("y_step_pp", "_w_y_step_pp")
                        )

                        st.button(
                            "Resetar Eixos - Gradiente de Pressão de Poros",
                            on_click=reset_config_pp,
                            type="primary",
                            use_container_width=True
                        )

                df_pp_plot = st.session_state.get("df_pp", df_pp)

                _plotar_pressao_poros_com_contexto(df_pp_plot)

    with tab_tabela:
        st.subheader("Tabela de dados calculados")

        df_pp_tabela = st.session_state.get("df_pp", df_pp)

        st.dataframe(
            df_pp_tabela,
            use_container_width=True,
            hide_index=True
        )


def pagina_estabilidade():
    st.header("Estabilidade de Poço")
    df_pp = st.session_state.get("df_pp", pd.DataFrame())

    if st.session_state.submenu_estabilidade == "Tensões em Volta do Poço":
        st.subheader("Tensões em Volta do Poço")

        df_est = _montar_df_estabilidade(df_pp)

        tab_dados, tab_tabela = st.tabs(["Dados", "Tabela de dados calculados"])

        with tab_dados:
            st.markdown("### Dados")
            c1, c2 = st.columns((0.5, 1))
            with c1:
                with st.container(border=True):
                    opcoes_lft = ["Calculado", "Constante"]

                    if st.session_state.get("lft") not in opcoes_lft:
                        st.session_state.lft = "Calculado"

                    preparar_widget_persistente("lft", "_w_lft")

                    if st.session_state.get("_w_lft") not in opcoes_lft:
                        st.session_state["_w_lft"] = st.session_state.lft

                    st.selectbox(
                        "Definir ângulo de fricção (Φ)",
                        opcoes_lft,
                        key="_w_lft",
                        on_change=salvar_widget_persistente,
                        args=("lft", "_w_lft")
                    )

                    sincronizar_widgets_persistentes([
                        ("lft", "_w_lft"),
                    ])

                    if st.session_state.lft == "Constante":
                        preparar_widget_persistente("phi_constante", "_w_phi_constante")

                        st.number_input(
                            "Ângulo de fricção constante (°)",
                            min_value=0.0,
                            max_value=90.0,
                            step=0.5,
                            format="%.2f",
                            key="_w_phi_constante",
                            on_change=salvar_widget_persistente,
                            args=("phi_constante", "_w_phi_constante")
                        )

                        sincronizar_widgets_persistentes([
                            ("phi_constante", "_w_phi_constante"),
                        ])

                    if "ucs" not in st.session_state:
                        st.session_state.ucs = "Mechpro"

                    preparar_widget_persistente("ucs", "_w_ucs")
                    if st.session_state.get("ucs") not in ["Lacy", "Mechpro"]:
                        st.session_state.ucs = "Mechpro"
                    st.session_state["_w_ucs"] = st.session_state.ucs
                    ucs_atual = st.selectbox(
                        "Método de cálculo do UCS",
                        ["Lacy", "Mechpro"],
                        key="_w_ucs",
                        on_change=salvar_widget_persistente,
                        args=("ucs", "_w_ucs")
                    )
                    st.session_state.ucs = ucs_atual or st.session_state.get("ucs", "Mechpro")

                    if "direcoes_tensoes_df" not in st.session_state:
                        st.session_state.direcoes_tensoes_df = _df_direcoes_tensoes_padrao()
                    if "relacao_tensoes_df" not in st.session_state:
                        st.session_state.relacao_tensoes_df = _df_relacao_tensoes_padrao()
                    if "usar_direcoes_tensoes" not in st.session_state:
                        st.session_state.usar_direcoes_tensoes = False
                    if "usar_relacao_tensoes" not in st.session_state:
                        st.session_state.usar_relacao_tensoes = False

                    with st.expander("Direções das tensões in situ", expanded=False):
                        direcoes_tensoes_editado = st.data_editor(
                            st.session_state.direcoes_tensoes_df,
                            num_rows="dynamic",
                            use_container_width=True,
                            hide_index=True,
                            key="editor_direcoes_tensoes",
                            column_config={
                                "Profundidade (m)": st.column_config.NumberColumn(
                                    "Profundidade (m)",
                                    min_value=0.0,
                                    format="%.2f"
                                ),
                                "Direção SH": st.column_config.NumberColumn(
                                    "Direção SH",
                                    min_value=0.0,
                                    max_value=360.0,
                                    format="%.2f"
                                ),
                            }
                        )

                        col_dir1, col_dir2 = st.columns(2)
                        with col_dir1:
                            if st.button(
                                "Inserir direções",
                                use_container_width=True,
                                type="primary",
                                key="btn_inserir_direcoes_tensoes"
                            ):
                                st.session_state.direcoes_tensoes_df = _normalizar_direcoes_tensoes(
                                    direcoes_tensoes_editado
                                )
                                st.session_state.usar_direcoes_tensoes = not st.session_state.direcoes_tensoes_df.empty
                                st.rerun()

                        with col_dir2:
                            if st.button(
                                "Resetar direções",
                                use_container_width=True,
                                type="primary",
                                key="btn_resetar_direcoes_tensoes"
                            ):
                                st.session_state.direcoes_tensoes_df = _df_direcoes_tensoes_padrao()
                                st.session_state.usar_direcoes_tensoes = False
                                st.rerun()

                    with st.expander("Relação das tensões in situ", expanded=False):
                        relacao_tensoes_editado = st.data_editor(
                            st.session_state.relacao_tensoes_df,
                            num_rows="dynamic",
                            use_container_width=True,
                            hide_index=True,
                            key="editor_relacao_tensoes",
                            column_config={
                                "Profundidade (m)": st.column_config.NumberColumn(
                                    "Profundidade (m)",
                                    min_value=0.0,
                                    format="%.2f"
                                ),
                                "SH% Sobrecarga": st.column_config.NumberColumn(
                                    "SH% Sobrecarga",
                                    min_value=0.0,
                                    format="%.2f"
                                ),
                                "Sh% Sobrecarga": st.column_config.NumberColumn(
                                    "Sh% Sobrecarga",
                                    min_value=0.0,
                                    format="%.2f"
                                ),
                            }
                        )

                        col_rel1, col_rel2 = st.columns(2)
                        with col_rel1:
                            if st.button(
                                "Inserir relação",
                                use_container_width=True,
                                type="primary",
                                key="btn_inserir_relacao_tensoes"
                            ):
                                st.session_state.relacao_tensoes_df = _normalizar_relacao_tensoes(
                                    relacao_tensoes_editado
                                )
                                st.session_state.usar_relacao_tensoes = not st.session_state.relacao_tensoes_df.empty
                                st.rerun()

                        with col_rel2:
                            if st.button(
                                "Resetar relação",
                                use_container_width=True,
                                type="primary",
                                key="btn_resetar_relacao_tensoes"
                            ):
                                st.session_state.relacao_tensoes_df = _df_relacao_tensoes_padrao()
                                st.session_state.usar_relacao_tensoes = False
                                st.rerun()

                    df_est = _aplicar_parametros_mecanicos_df_est(df_est)
                    st.session_state.df_est = df_est.copy()

            with c2:
                with st.container(border=True):
                    st.dataframe(df_est, use_container_width=True, hide_index=True)
            # with c3:
            #     with st.container(border=True):
            #         pass

        with tab_tabela:
            st.markdown("### Tabela de dados calculados")
            if df_est.empty:
                st.info("Calcule o Gradiente de Pressão de Poros para gerar df_est.")
            else:
                st.dataframe(df_est, use_container_width=True, hide_index=True)

    elif st.session_state.submenu_estabilidade == "Gradiente de Fratura":
        st.subheader("Gradiente de Fratura")

        if not isinstance(df_pp, pd.DataFrame) or df_pp.empty:
            st.warning("Calcule o Gradiente de Pressão de Poros antes de calcular o Gradiente de Fratura.")
            return

        tab_dados, tab_tabela = st.tabs(["Dados", "Tabela de dados calculados"])

        col_prof = "Profundidade (m)"
        if col_prof in df_pp.columns:
            prof_max_fratura = pd.to_numeric(df_pp[col_prof], errors="coerce").max()
        else:
            prof_max_fratura = np.nan

        for chave, valor_padrao in (
            ("lot_fratura", "Sim"),
            ("auxiliar_fratura", "Não"),
            ("ogf", "Gradiente (lb/gal)"),
            ("gras", "Sim"),
            ("grap", "Sim"),
            ("janela_fratura", "Sim"),
            ("mostrar_texto_lot_tabs4", "Sim"),
            ("mostrar_texto_fit_tabs4", "Sim"),
            ("x_min_f", 7.0),
            ("x_max_f", 21.0),
            ("x_step_f", 2.0),
            ("y_min_f", 0.0),
            ("y_max_f", float(int(prof_max_fratura) + 100) if pd.notna(prof_max_fratura) else 1000.0),
            ("y_step_f", 200.0),
        ):
            if chave not in st.session_state:
                st.session_state[chave] = valor_padrao

        if "lot_pontos_fratura" not in st.session_state or not isinstance(st.session_state.lot_pontos_fratura, pd.DataFrame):
            st.session_state.lot_pontos_fratura = _df_pontos_lot_padrao()

        with tab_dados:
            st.markdown("### Dados")
            c1, c2, c3 = st.columns((1, 1, 1))

            with c1:
                with st.container(border=True):
                    with st.expander("Leak Off Test", expanded=True):
                        opcoes_sim_nao = ["Sim", "Não"]
                        col1, col2 = st.columns(2)
                        with col1:
                            if st.session_state.get("lot_fratura") not in opcoes_sim_nao:
                                st.session_state.lot_fratura = "Sim"
                            preparar_widget_persistente("lot_fratura", "_w_lot_fratura")
                            st.selectbox(
                                "Inserir LOTs",
                                opcoes_sim_nao,
                                key="_w_lot_fratura",
                                on_change=salvar_widget_persistente,
                                args=("lot_fratura", "_w_lot_fratura")
                            )

                        with col2:
                            if st.session_state.get("auxiliar_fratura") not in opcoes_sim_nao:
                                st.session_state.auxiliar_fratura = "Não"
                            preparar_widget_persistente("auxiliar_fratura", "_w_auxiliar_fratura")
                            st.selectbox(
                                "K auxiliar",
                                opcoes_sim_nao,
                                key="_w_auxiliar_fratura",
                                on_change=salvar_widget_persistente,
                                args=("auxiliar_fratura", "_w_auxiliar_fratura"),
                                help=(
                                    "Adiciona um ponto auxiliar em 1 m com K = 0.01. "
                                    "Útil quando há poucos dados de LOT/FIT."
                                )
                            )

                        pontos_lot_calibracao = st.session_state.lot_pontos_fratura

                        if st.session_state.get("lot_fratura", "Sim") == "Sim":
                            if (
                                    "lot_pontos_fratura" not in st.session_state
                                    or not isinstance(st.session_state.lot_pontos_fratura, pd.DataFrame)
                            ):
                                st.session_state.lot_pontos_fratura = pd.DataFrame({
                                    "Tipo": [],
                                    "Profundidade (m)": [],
                                    "Peso Eq. (lb/gal)": []
                                })

                            if "Tipo" not in st.session_state.lot_pontos_fratura.columns:
                                st.session_state.lot_pontos_fratura.insert(0, "Tipo", "LOT")

                            st.session_state.lot_pontos_fratura["Tipo"] = (
                                st.session_state.lot_pontos_fratura["Tipo"]
                                .fillna("LOT")
                                .replace("", "LOT")
                            )
                            pontos_lot_editado = st.data_editor(
                                st.session_state.lot_pontos_fratura,
                                num_rows="dynamic",
                                use_container_width=True,
                                hide_index=True,
                                key="editor_lot_pontos_fratura",
                                column_config={
                                    "Tipo": st.column_config.SelectboxColumn(
                                        "Tipo",
                                        options=["LOT", "FIT"],
                                        required=True
                                    ),
                                    "Profundidade (m)": st.column_config.NumberColumn(
                                        "Profundidade (m)",
                                        min_value=0.0,
                                        format="%.2f"
                                    ),
                                    "Peso Eq. (lb/gal)": st.column_config.NumberColumn(
                                        "Peso Eq. (lb/gal)",
                                        min_value=0.0,
                                        format="%.2f"
                                    ),
                                }
                            )

                            pontos_lot_calibracao = pontos_lot_editado

                            col_lot1, col_lot2 = st.columns(2)

                            with col_lot1:
                                if st.button(
                                        "Atualizar LOT/FIT",
                                        type="primary",
                                        use_container_width=True,
                                        key="btn_atualizar_lot_fratura"
                                ):
                                    st.session_state.lot_pontos_fratura = _normalizar_pontos_lot_fratura(
                                        pontos_lot_editado)
                                    st.session_state.pop("editor_lot_pontos_fratura", None)
                                    st.rerun()

                            with col_lot2:
                                if st.button(
                                    "Limpar LOT/FIT",
                                    type="primary",
                                    use_container_width=True,
                                    key="btn_limpar_lot_fratura"
                                ):
                                    st.session_state.lot_pontos_fratura = _df_pontos_lot_padrao()
                                    st.rerun()

                        else:
                            df_lots_yaml = _carregar_lots_yaml_fratura()
                            st.markdown("### LOTs de poços próximos")

                            if df_lots_yaml.empty:
                                st.info("Nenhum LOT de poço próximo encontrado no YAML para o raio configurado.")
                                pontos_lot_calibracao = pd.DataFrame(columns=["Tipo", "Profundidade (m)", "Peso Eq. (lb/gal)"])
                            else:
                                st.dataframe(df_lots_yaml, use_container_width=True, hide_index=True)
                                pontos_lot_calibracao = _converter_lots_yaml_para_fratura(df_lots_yaml)
                                st.session_state.lot_pontos_yaml_fratura = pontos_lot_calibracao.copy()

                        if st.button(
                            "Calibrar curva de K",
                            type="primary",
                            use_container_width=True,
                            key="btn_calibrar_k_fratura"
                        ):
                            df_lot_calibrar = _normalizar_pontos_lot_fratura(pontos_lot_calibracao)
                            if st.session_state.get("lot_fratura", "Sim") == "Sim":
                                st.session_state.lot_pontos_fratura = df_lot_calibrar.copy()
                            st.session_state.gf = _calibrar_k_fratura(
                                df_pp,
                                df_lot_calibrar,
                                usar_auxiliar=(st.session_state.get("auxiliar_fratura") == "Sim")
                            )
                            st.session_state.edited_gf = st.session_state.gf.copy()
                            st.session_state.pop("editor_gf_fratura", None)

                    with st.expander("Relação das tensões", expanded=True):
                        if "edited_gf" not in st.session_state:
                            st.session_state.edited_gf = st.session_state.get(
                                "gf",
                                pd.DataFrame(columns=["Profundidade (m)", "K"])
                            ).copy()

                        gf_base = st.session_state.edited_gf.copy()
                        for coluna in ["Profundidade (m)", "K"]:
                            if coluna not in gf_base.columns:
                                gf_base[coluna] = np.nan

                        edited_gf_fratura = st.data_editor(
                            gf_base[["Profundidade (m)", "K"]],
                            num_rows="dynamic",
                            use_container_width=True,
                            hide_index=True,
                            key="editor_gf_fratura",
                            column_config={
                                "Profundidade (m)": st.column_config.NumberColumn(
                                    "Profundidade (m)",
                                    min_value=0.0,
                                    format="%.2f"
                                ),
                                "K": st.column_config.NumberColumn(
                                    "K",
                                    format="%.2f"
                                ),
                            }
                        )

                        st.session_state.edited_gf = edited_gf_fratura.copy()
                        a_k, b_k = _ajustar_tendencia_k_fratura(edited_gf_fratura)
                        st.session_state.a_k = a_k
                        st.session_state.b_k = b_k

                        if a_k is None or b_k is None:
                            st.caption("Informe ao menos dois pontos válidos de K para ajustar a curva.")
                        else:
                            st.caption(f"Curva ajustada: profundidade = {a_k:.2f} * exp({b_k:.4f} * K)")

                        if st.button(
                            "Calcular Gradiente de Fratura",
                            type="primary",
                            use_container_width=True,
                            key="btn_calcular_gradiente_fratura"
                        ):
                            df_f_calc = _calcular_df_f_fratura(df_pp, edited_gf_fratura)

                            if df_f_calc.empty or "Gradiente de Fratura (lb/gal)" not in df_f_calc.columns:
                                st.warning("Não foi possível calcular o Gradiente de Fratura. Verifique a curva de K.")
                            else:
                                st.session_state.df_f = df_f_calc.copy()
                                st.session_state.tem_gradiente_fratura_plotado = True

            with c2:
                with st.container(border=True):
                    st.markdown("### Método das Tensões Mínimas")
                    _plotar_k_fratura(st.session_state.get("edited_gf", st.session_state.get("gf", pd.DataFrame())))

            with c3:
                with st.container(border=True):
                    with st.expander("****Configurações do Gráfico****", expanded=False):
                        with st.expander("****Curvas e textos****", expanded=False):
                            opcoes_sim_nao = ["Sim", "Não"]

                            preparar_widget_persistente("ogf", "_w_ogf")
                            st.segmented_control(
                                "Opção de Gráfico",
                                ["Gradiente (lb/gal)", "Pressão (psi)"],
                                selection_mode="single",
                                key="_w_ogf",
                                width="stretch",
                                on_change=salvar_widget_persistente,
                                args=("ogf", "_w_ogf")
                            )

                            col_cfg1, col_cfg2 = st.columns(2)

                            for idx, (chave, label) in enumerate(
                                    (
                                            ("gras", "Gradiente de Sobrecarga"),
                                            ("grap", "Gradiente de Pressão de Poros"),
                                            ("janela_fratura", "Janela Operacional"),
                                            ("mostrar_texto_lot_tabs4", "Exibir texto LOT"),
                                            ("mostrar_texto_fit_tabs4", "Exibir texto FIT"),
                                    )
                            ):
                                coluna = col_cfg1 if idx % 2 == 0 else col_cfg2
                                with coluna:
                                    if st.session_state.get(chave) not in opcoes_sim_nao:
                                        st.session_state[chave] = "Sim"

                                    preparar_widget_persistente(chave, f"_w_{chave}")

                                    st.selectbox(
                                        label,
                                        opcoes_sim_nao,
                                        key=f"_w_{chave}",
                                        on_change=salvar_widget_persistente,
                                        args=(chave, f"_w_{chave}")
                                    )

                        with st.expander("***Eixos***", expanded=False):
                            for chave, label, step, min_value in (
                                ("x_min_f", "Eixo X - mínimo", 0.5, None),
                                ("x_max_f", "Eixo X - máximo", 0.5, None),
                                ("x_step_f", "Passo do eixo X", 0.5, 0.01),
                                ("y_min_f", "Eixo Y - mínimo", 100.0, None),
                                ("y_max_f", "Eixo Y - máximo", 100.0, None),
                                ("y_step_f", "Passo do eixo Y", 50.0, 0.01),
                            ):
                                preparar_widget_persistente(chave, f"_w_{chave}")
                                kwargs = {
                                    "label": label,
                                    "step": step,
                                    "format": "%.2f",
                                    "key": f"_w_{chave}",
                                    "on_change": salvar_widget_persistente,
                                    "args": (chave, f"_w_{chave}"),
                                }
                                if min_value is not None:
                                    kwargs["min_value"] = min_value
                                st.number_input(**kwargs)

                            if st.button(
                                "Resetar Eixos - Gradiente de Fratura",
                                type="primary",
                                use_container_width=True,
                                key="btn_reset_eixos_fratura"
                            ):
                                if st.session_state.get("ogf", "Gradiente (lb/gal)") == "Pressão (psi)":
                                    st.session_state.x_min_f = 0.0
                                    max_pressao = pd.to_numeric(
                                        st.session_state.get("df_f", pd.DataFrame()).get("Pressão de Fratura (psi)", pd.Series(dtype=float)),
                                        errors="coerce"
                                    ).max()
                                    st.session_state.x_max_f = float(int(max_pressao) + 500) if pd.notna(max_pressao) else 5000.0
                                    st.session_state.x_step_f = 500.0
                                else:
                                    st.session_state.x_min_f = 7.0
                                    st.session_state.x_max_f = 21.0
                                    st.session_state.x_step_f = 2.0

                                st.session_state.y_min_f = 0.0
                                st.session_state.y_max_f = float(int(prof_max_fratura) + 100) if pd.notna(prof_max_fratura) else 1000.0
                                st.session_state.y_step_f = 200.0

                                for chave in ("x_min_f", "x_max_f", "x_step_f", "y_min_f", "y_max_f", "y_step_f"):
                                    st.session_state[f"_w_{chave}"] = st.session_state[chave]

                                st.rerun()

                    _plotar_gradiente_fratura(st.session_state.get("df_f", pd.DataFrame()), df_pp)

        with tab_tabela:
            st.markdown("### Tabela de dados calculados")
            df_f_tabela = st.session_state.get("df_f", pd.DataFrame())

            if isinstance(df_f_tabela, pd.DataFrame) and not df_f_tabela.empty:
                st.dataframe(df_f_tabela, use_container_width=True, hide_index=True)
            else:
                st.info("Calcule o Gradiente de Fratura para visualizar a tabela.")


def pagina_sapatas():
    st.header("Assentamento de Sapatas")

    tab_dados, tab_tabela = st.tabs(["Dados", "Tabela de dados calculados"])

    with tab_dados:
        st.subheader("Dados")
        c1, c2, c3 = st.columns((1, 1, 1))
        with c1:
            with st.container(border=True):
                pass
        with c2:
            with st.container(border=True):
                pass
        with c3:
            with st.container(border=True):
                pass

    with tab_tabela:
        st.subheader("Tabela de dados calculados")
        pass


def pagina_fluido():
    st.header("Fluido de Perfuração")

    tab_dados, tab_tabela = st.tabs(["Dados", "Tabela de dados calculados"])

    with tab_dados:
        st.subheader("Dados")
        c1, c2, c3 = st.columns((1, 1, 1))
        with c1:
            with st.container(border=True):
                pass
        with c2:
            with st.container(border=True):
                pass
        with c3:
            with st.container(border=True):
                pass

    with tab_tabela:
        st.subheader("Tabela de dados calculados")
        pass


def pagina_anotacoes():
    st.header("Anotações")
    c1, c2 = st.columns((1, 1))

    with c1:
        with st.container(border=True):
            pass

    with c2:
        with st.container(border=True):
            pass


def pagina_relatorio():
    st.header("Relatório")
    c1, c2 = st.columns((1, 1))

    with c1:
        with st.container(border=True):
            pass

    with c2:
        with st.container(border=True):
            pass


def pagina_info():
    st.header("Informações sobre o SYGA")
    pass


def configurar_estado_inicial():
    defaults = {
        "pagina": "Entrada de Dados",
        "submenu_sobrecarga": "Dados",
        "submenu_poros": "Dados",
        "submenu_estabilidade": "Tensões em Volta do Poço",
        "submenu_estabilidade_tensoes": "Dados",
        "submenu_estabilidade_fratura": "Dados",
        "submenu_sapatas": "Dados",
        "submenu_fluido": "Dados",
        "tree_menu_indice": 0,

        "read_step": 1,
        "traj_modo": "Planejada",
        "option": "Previsão de Geopressões",

        "tipo_poco": "Onshore",
        "datum": "RTKB",
        "_tipo_poco_anterior": "Onshore",

        "country_name": "Brasil",
        "user_name": "",
        "company_name": "",
        "field_name": "",

        "poco": "Não informado",
        "comments": "",

        "zona": 24,
        "hem": "Sul",
        "easting": 857718.96,
        "northing": 8933902.28,
        "raio": 0.1,

        "geo_auto_ok": False,
        "geo_auto_tentado": False,

        "main_xlsm": None,
        "main_xlsm_hash": None,
        "main_xlsm_import_key": None,
        "wb": None,

        "df1": pd.DataFrame(),
        "df2": pd.DataFrame(),
        "df_interp": pd.DataFrame(),
        "df_out_traj": pd.DataFrame(),
        "df_mud": pd.DataFrame(),
        "df_eventos": pd.DataFrame(columns=["MD Inicial", "MD Final", "Evento"]),
        "sapatas_df": pd.DataFrame(),
        "fases_df": pd.DataFrame(),

        "lito_import_ok": False,
        "df_lito_excel": pd.DataFrame(),
        "pocos": {"Poço": {}},
        "well_selected": "Poço",
        "well_name": "Poço",

        "idg": "Não",
        "n_id": 1,
        "df_idade": pd.DataFrame(columns=["Topo (m)", "Base (m)", "Idade"]),

        "profundidade_maxima": None,

        "gard": ["Perfil de Densidade"],
        "gard_2": ["Miller"],
        "ex": "Desativada",
        "md": 0.0,
        "ds": 0.0,
        "rtkb": 9.4,
        "es": 110.0,
        "lda": 978.0,
        "nf": 119.4,
        "ogs": "Gradiente (lb/gal)",
        "x_min_s": 7,
        "x_max_s": 21,
        "x_step_s": 2,
        "y_min_s": 0,
        "y_max_s": 1000,
        "y_step_s": 200,
        "ext_df": pd.DataFrame(),
        "df_sobrecarga": pd.DataFrame(),
        "fig_gs": None,
        "oes": pd.Series(dtype=float),
        "profs": pd.Series(dtype=float),
        "oesl": "G. de Sobrecarga",

        "expoente": 3.0,
        "anormal": 400.0,
        "gn": 8.5,

        "lft": "Calculado",
        "phi_constante": 30.0,
        "ucs": "Mechpro",
        "usar_direcoes_tensoes": False,
        "usar_relacao_tensoes": False,
        "direcoes_tensoes_df": pd.DataFrame(columns=["Profundidade (m)", "Direção SH"]),
        "relacao_tensoes_df": pd.DataFrame(columns=["Profundidade (m)", "SH% Sobrecarga", "Sh% Sobrecarga"]),


    }

    for chave, valor in defaults.items():
        if chave not in st.session_state:
            if isinstance(valor, pd.DataFrame):
                st.session_state[chave] = valor.copy()
            elif isinstance(valor, dict):
                st.session_state[chave] = valor.copy()
            else:
                st.session_state[chave] = valor


def texto_negrito(texto):
    texto = unicodedata.normalize("NFD", texto)

    resultado = []

    for ch in texto:
        if "A" <= ch <= "Z":
            resultado.append(chr(0x1D5D4 + ord(ch) - ord("A")))
        elif "a" <= ch <= "z":
            resultado.append(chr(0x1D5EE + ord(ch) - ord("a")))
        elif "0" <= ch <= "9":
            resultado.append(chr(0x1D7EC + ord(ch) - ord("0")))
        else:
            resultado.append(ch)

    return "".join(resultado)


def montar_sidebar():
    with st.sidebar:
        st.title("Menu SYGA")

        items = [
            sac.TreeItem(texto_negrito("Entrada de Dados")),
            sac.TreeItem(texto_negrito("Litologia")),
            sac.TreeItem(texto_negrito("Gradiente de Sobrecarga")),
            sac.TreeItem(texto_negrito("Gradiente de Pressão de Poros")),
            sac.TreeItem(texto_negrito("Estabilidade de Poço"), children=[
                sac.TreeItem(texto_negrito("Tensões em Volta do Poço")),
                sac.TreeItem(texto_negrito("Gradiente de Fratura")),
            ]),
            sac.TreeItem(texto_negrito("Assentamento de Sapatas")),
            sac.TreeItem(texto_negrito("Fluido de Perfuração")),
            sac.TreeItem(texto_negrito("Anotações")),
            sac.TreeItem(texto_negrito("Relatório")),
            sac.TreeItem(texto_negrito("Informações sobre o SYGA")),
        ]

        if st.session_state.tree_menu_indice > 11:
            st.session_state.tree_menu_indice = 0

        item = sac.tree(
            items=items,
            index=st.session_state.tree_menu_indice,
            open_all=True,
            checkbox=False,
            show_line=True,
            return_index=True,
            key="tree_menu_syga"
        )

        if item is None:
            return

        if isinstance(item, list):
            if not item:
                return
            item = item[-1]

        st.session_state.tree_menu_indice = item

        mapa = {
            0: ("Entrada de Dados", {}),
            1: ("Litologia", {}),
            2: ("Gradiente de Sobrecarga", {"submenu_sobrecarga": "Dados"}),
            3: ("Gradiente de Pressão de Poros", {"submenu_poros": "Dados"}),
            4: (
                "Estabilidade de Poço",
                {
                    "submenu_estabilidade": "Tensões em Volta do Poço",
                    "submenu_estabilidade_tensoes": "Dados",
                }
            ),
            5: (
                "Estabilidade de Poço",
                {
                    "submenu_estabilidade": "Tensões em Volta do Poço",
                    "submenu_estabilidade_tensoes": "Dados",
                }
            ),
            6: (
                "Estabilidade de Poço",
                {
                    "submenu_estabilidade": "Gradiente de Fratura",
                    "submenu_estabilidade_fratura": "Dados",
                }
            ),
            7: ("Assentamento de Sapatas", {"submenu_sapatas": "Dados"}),
            8: ("Fluido de Perfuração", {"submenu_fluido": "Dados"}),
            9: ("Anotações", {}),
            10: ("Relatório", {}),
            11: ("Informações sobre o SYGA", {}),
        }

        if item in mapa:
            pagina, submenus = mapa[item]
            st.session_state.pagina = pagina

            for chave, valor in submenus.items():
                st.session_state[chave] = valor

        poco_atual = st.session_state.get("poco", "Poço não definido")

        st.sidebar.markdown(
            f"""
            <style>
                [data-testid="stSidebar"] .sidebar-footer {{
                    position: fixed;
                    bottom: 0;
                    left: 0;
                    width: 21rem;
                    max-width: 21rem;
                    box-sizing: border-box;
                    padding: 0.35rem 1rem 0.45rem 1rem;
                    border-top: 1px solid rgba(49, 51, 63, 0.18);
                    background: var(--background-color);
                    color: rgba(49, 51, 63, 0.72);
                    font-size: 0.78rem;
                    line-height: 1.25;
                    z-index: 10;
                }}

                [data-testid="stSidebar"] .sidebar-footer p {{
                    margin: 0.08rem 0;
                }}
            </style>
            <div class="sidebar-footer">
                <p>Poço: {poco_atual}</p>
                <p>Developed by Adriel Oliveira - 2025</p>
            </div>
            """,
            unsafe_allow_html=True
        )


def renderizar_pagina():
    paginas_app = {
        "Entrada de Dados": pagina_entrada_dados,
        "Litologia": pagina_coluna_litologica,
        "Coluna Litológica": pagina_coluna_litologica,
        "Gradiente de Sobrecarga": pagina_sobrecarga,
        "Gradiente de Pressão de Poros": pagina_poros,
        "Estabilidade de Poço": pagina_estabilidade,
        "Assentamento de Sapatas": pagina_sapatas,
        "Fluido de Perfuração": pagina_fluido,
        "Anotações": pagina_anotacoes,
        "Relatório": pagina_relatorio,
        "Informações sobre o SYGA": pagina_info,
    }

    pagina_atual = st.session_state.get("pagina", "Entrada de Dados")

    if pagina_atual != "Entrada de Dados" and not arquivo_carregado():
        st.error('Por favor, insira um documento!', icon="🚨")
        return

    paginas_app[pagina_atual]()


def geo_page():
    st.title('Syngular Geopressure Analysis - SYGA')

    configurar_estado_inicial()
    montar_sidebar()
    renderizar_pagina()


geo_page()

